"""KETI 시뮬레이션데이터통합 양식(과기대) 을 FAB 용으로 채운다.

원본 서식·열은 그대로 두고 데이터 행만 비운 뒤 **제공 자료에 실제로 있는 값만** 넣는다.
코드 체계(prefix·item_code·process_code·resource_id)와 분류·단위는 FAB 이 정할 몫이라 비워 둔다.
"""
import openpyxl, re, shutil
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

ASK = PatternFill("solid", fgColor="FFF2CC")           # 확인 요청 = 노란색
AUTHOR = "KETI 시뮬레이션"

# ── 하드코딩 ────────────────────────────────────────────────────────────
SRC = r"C:/Users/KangTaehui/KG/keti/CPRO_조립공정/원본자료/과기대정리자료/KETI.시뮬레이션데이터통합.xlsx"
RUNSHEET = r"C:/Users/KangTaehui/KG/keti/keti-fab/[KETI]전북본부 FAB 장비 및 공정 정보/3. TFT backplane runsheet_스마트.xlsx"
OUT = r"C:/Users/KangTaehui/KG/keti/CPRO_조립공정/시뮬레이션/Package/docs/FAB/FAB_시뮬레이션데이터통합.xlsx"

HEAD_ROW = {"CATEGORY": 3, "BOM": 3, "BOM_STRUCTURE": 3, "PROCESS_FLOW": 6, "RESOURCE": 6}

MODEL = "PI"
SUBSTRATE_SPEC = "370 x 470 mm"                        # 장비 사양서에 공통 기재

# 공정흐름도 2~5번 — 박스·그룹 명칭은 원문 그대로
FLOW_OTHER = [
    ("Bottom emission OLED 소재 평가용 TEG-cell", [
        ("Anode ITO 증착", "투명 전극(Electrode) 형성 및 패터닝 공정"),
        ("포토레지스트 패터닝 공정", "투명 전극(Electrode) 형성 및 패터닝 공정"),
        ("습식 식각", "투명 전극(Electrode) 형성 및 패터닝 공정"),
        ("포토레지스트 제거", "투명 전극(Electrode) 형성 및 패터닝 공정"),
        ("유기절연막 패터닝 공정", "유기 절연막(Insulator) 패터닝 공정"),
    ]),
    ("Top emission OLED 소재 평가용 TEG-cell", [
        ("cathode 증착", "고전도 반사 전극(Electrode) 형성 및 패터닝 공정"),
        ("포토레지스트 패터닝 공정", "고전도 반사 전극(Electrode) 형성 및 패터닝 공정"),
        ("습식 식각", "고전도 반사 전극(Electrode) 형성 및 패터닝 공정"),
        ("포토레지스트 제거", "고전도 반사 전극(Electrode) 형성 및 패터닝 공정"),
        ("유기절연막 패터닝 공정", "유기 절연막(Insulator) 패터닝 공정"),
    ]),
    ("전극 스크린 프린팅", [
        ("제판 설계 및 제작", "Screen Printing 공정"),
        ("Screen Printing", "Screen Printing 공정"),
        ("건조", "Screen Printing 공정"),
    ]),
    ("잉크젯 프린팅", [
        ("Ink 및 기판 준비", "Ink-jet Printing 공정"),
        ("인쇄적합성 평가", "Ink-jet Printing 공정"),
        ("인쇄", "Ink-jet Printing 공정"),
        ("건조", "Ink-jet Printing 공정"),
    ]),
]

# 공정흐름도 2~5번의 박스별 투입 재료 (흐름도에 화살표로 붙어 있는 것)
FLOW_STEP_MATERIALS = {
    ("Bottom emission OLED 소재 평가용 TEG-cell", "Anode ITO 증착"):
        ["Indium Tin Oxide", "Ar", "O2"],
    ("Bottom emission OLED 소재 평가용 TEG-cell", "포토레지스트 패터닝 공정"):
        ["Photoresist", "TMAH 2.38% 수용액"],
    ("Bottom emission OLED 소재 평가용 TEG-cell", "습식 식각"):
        ["ITO etchant"],
    ("Bottom emission OLED 소재 평가용 TEG-cell", "포토레지스트 제거"):
        ["Organic strip chemical"],
    ("Bottom emission OLED 소재 평가용 TEG-cell", "유기절연막 패터닝 공정"):
        ["유기절연막", "TMAH 2.38% 수용액"],

    ("Top emission OLED 소재 평가용 TEG-cell", "cathode 증착"):
        ["ITO/Ag/ITO 다층박막", "Ar"],
    ("Top emission OLED 소재 평가용 TEG-cell", "포토레지스트 패터닝 공정"):
        ["Photoresist", "TMAH 2.38% 수용액"],
    ("Top emission OLED 소재 평가용 TEG-cell", "습식 식각"):
        ["금속 etchant"],
    ("Top emission OLED 소재 평가용 TEG-cell", "포토레지스트 제거"):
        ["Organic strip chemical"],
    ("Top emission OLED 소재 평가용 TEG-cell", "유기절연막 패터닝 공정"):
        ["유기절연막", "TMAH 2.38% 수용액"],

    ("전극 스크린 프린팅", "제판 설계 및 제작"): ["제판"],
    ("전극 스크린 프린팅", "Screen Printing"): ["Glass", "Flim", "Paste"],

    ("잉크젯 프린팅", "Ink 및 기판 준비"): ["Ink", "Patterned glass or Flat glass"],
}

# 공정흐름도 2~5번에만 나오는 재료 (1번 TFT backplane 것은 runsheet 에서 읽는다)
FLOW_MATERIALS = [
    ("Indium Tin Oxide", ""), ("ITO/Ag/ITO 다층박막", ""), ("Photoresist", ""),
    ("TMAH 2.38% 수용액", ""), ("금속 etchant", ""), ("유기절연막", ""), ("PDL", ""),
    ("제판", "SUS mesh / 100~400 mesh / 900*900 mm 프레임"),
    ("Paste", "Ag or Carbon etc. / 점도 1×10⁴ ~ 1×10⁵ cPs @25 °C"),
    ("Ink", "점도 1 ~ 20 cPs @ 25 °C / 표면장력 약 25~50 mN/m / 입자 < 200 nm"),
    ("Flim", ""), ("Patterned glass or Flat glass", ""),
]

# runsheet Material·Strip chemical 원문 (품명은 원문 그대로 둔다)
MAT_ORDER = ["SiO₂/Si₃H₄", "Molybdenum(Mo)", "ITO", "SiO2", "IGZO",
             "ZPP1700PG-30", "ZPP-1700G", "ITO etchant",
             "Organic strip chemical", "NMP"]
# runsheet 가스 파라미터에 등장하는 가스 원문
GAS_ORDER = ["Ar", "N₂", "SiH4", "N2O", "NH₃", "O2", "CF4", "Cl2"]
# 기판 — runsheet Sheet1 및 장비 사양서
SUBSTRATE = [("PI", ""), ("Glass", SUBSTRATE_SPEC)]

# 담당 설비 — 장비리스트 국문명 · 로그로 확인된 챔버
CORE_DEV = {
    "PECVD": "박막증착장비 chA",
    "Sputter": "박막증착장비 chD·chE",
    "Dry etcher": "박막증착장비 chC",
    "Wet etcher": "엣쳐/스트리퍼",
    "Coater/Developer/Mask Aligner": "현상장비 / 마스크얼라이너",
    "Wet stripper": "유기스트리퍼",
    "Manual stripper": "유기스트리퍼",
}
MEMO = {
    "박막증착장비 chA": "로그 확보(72런) · 모델명 없음",
    "박막증착장비 chD·chE": "로그 미확보 · 모델명 없음",
    "박막증착장비 chC": "로그 확보(128런) · 모델명 없음",
    "엣쳐/스트리퍼": "모델명 없음",
    "현상장비 / 마스크얼라이너": "현상장비 모델명 없음 / 마스크얼라이너 MA-5501ML",
    "유기스트리퍼": "모델명 없음",
    "고온진공오븐": "runsheet 기재 명칭 · 보유장비리스트 미등재",
    "기판세정기(수용액크리너)": "모델명 없음 · runsheet 사용장비 열에 기재",
    "박막증착장비 TM": "로그 확보(GLASS_MOVEMENT 1,653건) · 모델명 없음",
    "스크린프린터": "Seria SSA-GL4737-RST · 공정흐름도 4번",
    "잉크젯 프린터": "4대 보유(Litrex 120L·Ulvac·Fuji Film DMP-2831·Dimatix) · 공정흐름도 5번",
}


# ── runsheet 파싱 ───────────────────────────────────────────────────────
def read_runsheet():
    ws = openpyxl.load_workbook(RUNSHEET, data_only=True)["PI Run sheet"]
    rows = [r + (None,) * 6 for r in ws.iter_rows(values_only=True)]
    grp = step = None
    recs, cur = [], None
    for r in rows:
        if r[0]:
            grp = str(r[0]).replace("\n", " ").strip()
        p, v = r[2], r[3]
        if p and v is None:
            step = str(p).strip()
            continue
        if p == "Process parameter":
            cur = {"grp": grp, "step": step, "eq": None, "power": None,
                   "mats": [], "gases": [], "times": {}}
            recs.append(cur)
            continue
        if cur is None or not p:
            continue
        k = str(p).strip()
        if k == "Equipment":
            cur["eq"] = str(v).strip()
        elif k in ("Material", "Strip chemical"):
            cur["mats"].append(str(v).strip())
        elif k.startswith("Power"):
            cur["power"] = str(v).strip()
        elif "gas" in k.lower():
            cur["gases"].append(k)
        elif re.search(r"time|Time|Annealing|Treatment", k):
            cur["times"][k] = v
    return recs


recs = read_runsheet()


def anneal_sec(c):
    """Annealing 은 오븐이 점유되는 시간이므로 cycle_time_sec 자리에 넣는다.

    CPRO 의 dep_wait_hr 은 본드 경화처럼 설비를 잡지 않는 대기라 성격이 다르다.
    """
    if c["step"] != "Annealing":
        return None
    tot = 0
    for v in c["times"].values():
        m = re.search(r"(\d+(?:\.\d+)?)\s*hrs?", str(v))
        if m:
            tot += float(m.group(1)) * 3600
    return int(tot) or None


# ── 원본 복사 후 데이터 비우기 ──────────────────────────────────────────
shutil.copyfile(SRC, OUT)
wb = openpyxl.load_workbook(OUT)
for name, hr in HEAD_ROW.items():
    ws = wb[name]
    if ws.max_row > hr:
        ws.delete_rows(hr + 1, ws.max_row - hr)


# ── CATEGORY — 분류 체계는 FAB 이 정할 몫. 비워 둔다 ────────────────────
wb["CATEGORY"].cell(4, 2).value = ""

# ── BOM — 자료에 나온 품명만 ────────────────────────────────────────────
ws = wb["BOM"]
for nm, spec in SUBSTRATE:
    ws.append([None, None, nm, spec])
for nm in MAT_ORDER + GAS_ORDER:
    ws.append([None, None, nm])
for nm, spec in FLOW_MATERIALS:
    ws.append([None, None, nm, spec])

# ── BOM_STRUCTURE — 공정 × 투입 재료. 투입량만 미확보 ───────────────────
ws = wb["BOM_STRUCTURE"]
for i, c in enumerate(recs, 1):                  # ① runsheet 31공정
    names = list(c["mats"])
    for key in c["gases"]:
        for nm in re.findall(r"[A-Za-z][A-Za-z0-9₂₃₄]*", key.split("(")[0]):
            if nm in GAS_ORDER and nm not in names:
                names.append(nm)
    for nm in names:
        ws.append([None, MODEL, nm, f"{MODEL}-{i}"])
for model, steps in FLOW_OTHER:                  # ②~⑤ 공정흐름도 박스
    for j, (step, grp) in enumerate(steps, 1):
        for nm in FLOW_STEP_MATERIALS.get((model, step), []):
            ws.append([None, model, nm, f"{model}-{j}"])

# ── PROCESS_FLOW — 공정흐름도 1~5번 전부 ────────────────────────────────
POWER_COL = 18                                   # 원본에 없는 열이라 뒤에 붙인다
ws = wb["PROCESS_FLOW"]
hr = HEAD_ROW["PROCESS_FLOW"]
ws.cell(hr, POWER_COL).value = "power_w"
ws.cell(hr, POWER_COL).fill = ASK
ws.cell(hr, POWER_COL).comment = Comment(
    "runsheet 에 기재된 Power( W ) 값을 그대로 옮긴 것입니다.\n\n"
    "플라즈마 발생용 RF 파워로 보이는데,\n"
    "① 공정에서 실제 소비하는 전력 전체인지\n"
    "② 일부이고 진공펌프·히터·칠러 등이 따로 있는지\n"
    "확인 부탁드립니다.\n\n"
    "②라면 설비별 정격전력 또는 분전반 계측값이 별도로 필요합니다.", AUTHOR)
ws.column_dimensions["R"].width = 12

for c in recs:                                   # ① TFT backplane = runsheet 31공정
    ws.append([None, MODEL, None, None, c["step"], c["grp"], None, None,
               None, None, None, anneal_sec(c), None, None, None, None, None,
               c["power"]])
    if c["power"]:
        ws.cell(ws.max_row, POWER_COL).fill = ASK
    if anneal_sec(c):
        cell = ws.cell(ws.max_row, 12)
        cell.fill = ASK
        cell.comment = Comment(
            "runsheet 의 Annealing 시간을 초로 환산한 값입니다.\n\n"
            "오븐이 그 시간 동안 점유되는 것이 맞는지,\n"
            "그동안 작업자는 다른 공정을 진행하시는지 확인 부탁드립니다.", AUTHOR)

for model, steps in FLOW_OTHER:                  # ②~⑤ 공정흐름도 박스
    for step, grp in steps:
        ws.append([None, model, None, None, step, grp, None, None,
                   None, None, None, None, None, None, None, None, None, None])

# ── RESOURCE — 설비명과 상태 메모만 ─────────────────────────────────────
ws = wb["RESOURCE"]
order, seen = [], set()
for c in recs:
    dev = CORE_DEV.get(c["eq"], "고온진공오븐" if c["step"] == "Annealing" else None)
    if dev and dev not in seen:
        seen.add(dev)
        order.append(dev)
order += ["기판세정기(수용액크리너)", "박막증착장비 TM", "스크린프린터", "잉크젯 프린터"]
for dev in order:
    ws.append([None, None, dev, None, None, None, None, None, None, None, None, MEMO.get(dev, "")])

wb.save(OUT)

for name, hr in HEAD_ROW.items():
    print(f"{name:16s} 데이터 {max(0, wb[name].max_row - hr)}행")
print("→", OUT)
