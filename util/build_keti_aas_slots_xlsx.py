"""12개 AAS 분류별로 확보 파라미터를 모아 엑셀로 낸다.

모든 행에 근거 URL 을 단다. PDF 는 그 파일을 내려받은 곳의 URL 을 쓴다.
파라미터명은 출처 표기를 그대로 옮긴다 (임의 개명 금지).
"""

import glob
import html
import json
import os
import re
import sys
import zipfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = r"C:\Users\KangTaehui\KG\keti\CPRO_조립공정\시뮬레이션\Package"
R = os.path.join(BASE, r"docs\원본자료\keti-fab\포털수집")
OUT = os.path.join(BASE, r"docs\원본자료\keti-fab\KETI_FAB_AAS_파라미터슬롯.xlsx")
KOSMO_ZIP = os.path.join(BASE, r"docs\원본자료\keti-fab\kasmo_후보데이터.zip")

ZEUS_READ = "https://www.zeus.go.kr/search/equip/read/"
ZEUS_RESV = "https://www.zeus.go.kr/resv/equip/read/"
ITUBE_VIEW = "https://www.itube.or.kr/aplct/equipSrch/sharingView.do?g_menu_id=MNID210100&equip_no="
ITUBE_FILE = "https://www.itube.or.kr/unitc/equipuse/myequip/fileDownMyEquip.do?g_menu_id=&equip_file_no="

LEDGER = {"가동상태", "i-Tube No.", "NTIS No", "설치기관", "주소", "담당자", "취득일", "취득금액",
          "내용연수", "구분", "용도", "표준 분류", "장비활용범위", "설치형태", "사용료 형태",
          "장비 사용료", "인증정보", "기능", "장비 상세설명", "제작사", "모델 명", "매뉴얼",
          "사용형태", "한글명", "영문명", "제작사명 | 모델명"}

# ── 12개 AAS 정의 ────────────────────────────────────────────────────
# 챔버는 (설비폴더, 챔버 판정 정규식, 그 챔버로 볼 유사장비 파일 키워드)
AAS = [
    ("증착", "박막증착장비 — PECVD", "증착/26_박막증착장비",
     r"pecvd|화학기상|SiO2|Si3N4|실리콘질화", ["화학기상", "PECVD", "TEOS"]),
    ("증착", "박막증착장비 — Sputter", "증착/26_박막증착장비",
     r"sputter|스퍼터|타깃|target", ["스퍼터", "Sputter"]),
    ("증착", "박막증착장비 — Dry Etcher", "증착/26_박막증착장비",
     r"etch|식각", ["식각", "에칭", "Etch"]),
    ("증착", "박막증착장비 — Thermal Evaporator", "증착/26_박막증착장비",
     r"thermal\s*evapor|열증착|CIGS|전자빔|e-?beam", ["열증착", "전자빔"]),
    ("증착", "유기증착기 — Plasma Chamber", "증착/24_유기증착기",
     r"plasma|플라즈마", ["플라즈마"]),
    ("증착", "유기증착기 — Organic Chamber", "증착/24_유기증착기",
     r"organic|유기|OLED|Low\s*Temp", ["유기", "OLED", "Sunicel"]),
    ("증착", "유기증착기 — Metal Chamber", "증착/24_유기증착기",
     r"metal|금속|BN\s*Boat|High\s*Temp|열증착", ["열증착"]),
    ("증착", "PEALD", "증착/44_PEALD", r".*", ["*"]),
    # 포토 3종 — 위치정리 PPT 의 구역 구분(포토레지스트 패터닝 / 식각설비(습식))을 따른다.
    # 세정설비(습식)는 별도 구역이라 포토에 넣지 않는다.
    ("포토", "코터·디벨로퍼 트랙", ["포토/43_스핀 트랙 시스템", "포토/18_현상장비",
                            "포토/34_스핀디벨로퍼"], r".*", ["*"]),
    ("포토", "노광 (마스크 얼라이너)", ["포토/42_마스크 얼라이너(8인치)", "포토/21_마스크얼라이너"],
     r".*", ["*"]),
    ("포토", "식각·스트립", ["포토/19_엣쳐_스트리퍼", "포토/35_유기스트리퍼"], r".*", ["*"]),
    ("참고", "화학 습식 증착(CBD)", "증착/27_화학 습식 증착(CBD)", r".*", ["*"]),
    ("프린터", "프린팅 (6대 통합)", ["프린터/20_스크린프린터", "프린터/09_잉크젯 프린터 for PLED #1",
                             "프린터/13_잉크젯 프린터 for PLED #2", "프린터/23_잉크젯 프린터(lab)",
                             "프린터/30_잉크젯프린터(lab #2)", "프린터/31_리버스 옵셋 프린터"],
     r".*", ["*"]),
]

# KOSMO 자료 — 사용자가 받아온 것이라 공개 URL 이 없다
KOSMO_NOTE = "사용자 제공 (kasmo_후보데이터.zip) — 공개 URL 없음"
# KOSMO 슬롯 배정 — 기능이 실제로 같은 곳에만 붙인다.
# 표면처리기는 플라즈마 '표면처리' 장비라 막을 쌓지 않는다. PECVD·식각과는 다른 장비다.
KOSMO_PLASMA_TARGETS = {"유기증착기 — Plasma Chamber"}
# 에칭장비는 건식 진공 식각이라 Dry Etcher 와 기능이 같다. TechnicalData + OperationalData 둘 다 쓴다.
KOSMO_ETCH_TARGETS = {"박막증착장비 — Dry Etcher"}
# 전해도금조는 상압 용액 공정이라 CBD(화학 습식 증착)와 욕조·약액 파라미터가 겹친다.
KOSMO_BATH_TARGETS = {"화학 습식 증착(CBD)"}
# 어느 장비에나 쓰는 공통 슬롯 (표면처리기 TechnicalData 에서 이것만 살린다)
KOSMO_COMMON = ["StandardCycleTime", "HourlyProductionCapacity"]


def pair(line):
    """'항목 : 값' 을 나눈다. 콜론이 여러 개면 마지막 콜론 앞까지를 항목명으로 본다.
    원문이 '○ Chamber : 재질 : STS 304' 처럼 계층을 콜론으로 이어 쓰는 경우가 있어서다."""
    line = html.unescape(line).strip(" -+·*○◦●⦁")
    if not line or "：" not in line and ":" not in line:
        return None
    line = line.replace("：", ":")
    head, _, val = line.rpartition(":")
    head = head.strip(" :·-")
    val = val.strip()
    if not head or not val or len(head) > 60:
        return None
    return (head, val)


def seg(text, start, ends):
    i = text.find(start)
    if i < 0:
        return ""
    j = min([text.find(e, i + 1) for e in ends if text.find(e, i + 1) > 0] or [len(text)])
    return text[i + len(start):j]


def load_kosmo():
    """KOSMO AASX 에서 슬롯 이름을 뽑는다 (참고용 슬롯 어휘)."""
    import xml.etree.ElementTree as ET
    NS = "{https://admin-shell.io/aas/3/0}"
    out = {"plasma": [], "etch_tech": [], "etch_oper": [], "bath_tech": [], "bath_oper": []}
    if not os.path.exists(KOSMO_ZIP):
        return out
    z = zipfile.ZipFile(KOSMO_ZIP)
    for member, key, want in [("표면처리기.aasx", "plasma", "OperationData"),
                              ("[일.6]SemiconductorCircuitEtchingEquipment.aasx", "etch_tech", "TechnicalData"),
                              ("[일.6]SemiconductorCircuitEtchingEquipment.aasx", "etch_oper", "OperationalData"),
                              ("[신규.2]ElectroplatingBath.aasx", "bath_tech", "TechnicalData"),
                              ("[신규.2]ElectroplatingBath.aasx", "bath_oper", "OperationalData")]:
        if member not in z.namelist():
            continue
        inner = zipfile.ZipFile(__import__("io").BytesIO(z.read(member)))
        xn = [n for n in inner.namelist() if n.endswith(".aas.xml")][0]
        root = ET.fromstring(inner.read(xn))
        for sm in root.iter():
            if sm.tag.split("}")[-1] != "submodel":
                continue
            ids = sm.find(NS + "idShort")
            if ids is None or ids.text != want:
                continue
            out.setdefault(key, [])
            for e in sm.iter():
                if e.tag.split("}")[-1] in ("property", "multiLanguageProperty", "range"):
                    n = e.find(NS + "idShort")
                    if n is not None and n.text and n.text not in out[key]:
                        out[key].append(n.text)
    return out


def collect(folder_rel, zeus_map, itube_map, manual_map):
    """설비 폴더 하나에서 (파라미터명, 값, 출처, URL) 목록을 만든다."""
    d = os.path.join(R, folder_rel)
    name = os.path.basename(folder_rel)
    no = int(name[:2])
    rows = []

    ze = zeus_map.get(no, {})
    zid = ze.get("zeus_id")
    if zid:
        url = ZEUS_READ + zid
        for ln in seg(ze.get("zeus_본문", ""), "구성 및 성능",
                      ["사용/활용 예", "시설장비 문의번호"]).split("\n")[1:]:
            p = pair(ln)
            if p and p[0] not in LEDGER:
                rows.append((p[0], p[1], "ZEUS 등록장비", url, name))

    rv = os.path.join(d, "_zeus_resv.txt")
    rid = None
    rh = os.path.join(d, "_zeus_resv.html")
    if os.path.exists(rh):
        m = re.search(r"/resv/equip/read/([A-Za-z0-9\-]+)", open(rh, encoding="utf-8").read())
        rid = m.group(1) if m else None
    if os.path.exists(rv):
        url = ZEUS_RESV + rid if rid else (ZEUS_READ + zid if zid else "")
        for ln in seg(open(rv, encoding="utf-8").read(), "특성", ["용도설명"]).split("\n"):
            p = pair(ln)
            if p and p[0] not in LEDGER:
                rows.append((p[0], p[1], "ZEUS 장비예약", url, name))

    te = itube_map.get(no, {})
    epn = te.get("itube_epn")
    if epn:
        url = ITUBE_VIEW + epn
        for k, v in (te.get("itube_정보") or {}).items():
            if str(v).strip() and k not in LEDGER and k not in ("국문명", "영문명", "온라인예약가능여부"):
                rows.append((k, str(v).strip(), "i-Tube", url, name))
        for f in te.get("매뉴얼_첨부", []):
            fu = ITUBE_FILE + f.get("파일번호", "")
            for g in manual_map.get(name, []):
                rows.append((g["파라미터명"], g["값"], f"설비 매뉴얼 PDF p{g['쪽']}", fu, name))
            break

    # 유사장비 — 파일명에 담긴 라벨과 폴더의 ZEUS id 를 함께 쓴다
    for f in sorted(glob.glob(os.path.join(d, "유사장비", "*.txt"))):
        b = os.path.basename(f)
        t = open(f, encoding="utf-8").read()
        m = re.search(r"시설장비활용번호\s*\n\s*([A-Za-z0-9\-]+)", t)
        url = ZEUS_RESV + m.group(1) if m else "https://www.zeus.go.kr/search"
        for head, ends in [("특성", ["용도설명"]),
                           ("구성 및 성능", ["사용/활용 예", "시설장비 문의번호"])]:
            for ln in seg(t, head, ends).split("\n"):
                p = pair(ln)
                if p and p[0] not in LEDGER:
                    rows.append((p[0], p[1], f"유사장비 · {b[:44]}", url, name))
    return rows


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    zeus_map = {e["no"]: e for e in json.load(open(os.path.join(R, "_zeus_mapping.json"), encoding="utf-8"))}
    itube_map = {e["no"]: e for e in json.load(open(os.path.join(R, "_itube_mapping.json"), encoding="utf-8"))}
    manual_map = json.load(open(os.path.join(R, "_manual_params.json"), encoding="utf-8"))
    kosmo = load_kosmo()

    wb = Workbook()
    HF = PatternFill("solid", fgColor="DDEBF7")
    summary = []
    first = True

    for group, aas_name, folders, chamber_re, sim_keys in AAS:
        fl = folders if isinstance(folders, list) else [folders]
        raw = []
        for fr in fl:
            raw += collect(fr, zeus_map, itube_map, manual_map)

        # 챔버 필터 (단일 설비를 챔버로 쪼개는 경우에만 적용)
        if chamber_re != r".*":
            pat = re.compile(chamber_re, re.I)
            keep = []
            for nm, v, src, url, eq in raw:
                is_sim = src.startswith("유사장비")
                if is_sim:
                    if any(k in src for k in sim_keys):
                        keep.append((nm, v, src, url, eq))
                elif pat.search(f"{nm} {v}"):
                    keep.append((nm, v, src, url, eq))
            raw = keep

        # KOSMO 슬롯 어휘 덧붙이기
        if aas_name in KOSMO_PLASMA_TARGETS:
            for n in kosmo["plasma"]:
                raw.append((n, "", "KOSMO 표면처리기 AAS · OperationData", KOSMO_NOTE, "-"))
        if aas_name in KOSMO_ETCH_TARGETS:
            for n in kosmo["etch_tech"]:
                raw.append((n, "", "KOSMO 반도체회로 에칭장비 AAS · TechnicalData", KOSMO_NOTE, "-"))
            for n in kosmo["etch_oper"]:
                raw.append((n, "", "KOSMO 반도체회로 에칭장비 AAS · OperationalData", KOSMO_NOTE, "-"))
        if aas_name in KOSMO_BATH_TARGETS:
            for n in kosmo["bath_tech"]:
                raw.append((n, "", "KOSMO 전해도금조 AAS · TechnicalData", KOSMO_NOTE, "-"))
            for n in kosmo["bath_oper"]:
                raw.append((n, "", "KOSMO 전해도금조 AAS · OperationalData", KOSMO_NOTE, "-"))
        for n in KOSMO_COMMON:
            raw.append((n, "", "KOSMO 표면처리기 AAS · TechnicalData(공통)", KOSMO_NOTE, "-"))

        # 같은 이름은 한 줄로 묶고 근거를 합친다
        merged = {}
        for nm, v, src, url, eq in raw:
            m = merged.setdefault(nm, {"vals": [], "srcs": [], "urls": [], "eqs": []})
            if v and v not in m["vals"]:
                m["vals"].append(v)
            if src not in m["srcs"]:
                m["srcs"].append(src)
            if url and url not in m["urls"]:
                m["urls"].append(url)
            if eq != "-" and eq not in m["eqs"]:
                m["eqs"].append(eq)

        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = aas_name[:31].replace("/", "-")
        ws.append(["분류", "AAS", "파라미터 슬롯", "확보된 값", "출처",
                   "근거 링크1", "근거 링크2", "근거 링크3", "설비"])
        for c in range(1, 10):
            ws.cell(1, c).fill = HF
            ws.cell(1, c).font = Font(bold=True)
        for nm, m in merged.items():
            urls = m["urls"][:3]
            ws.append([group, aas_name, nm,
                       " / ".join(m["vals"])[:250],
                       " / ".join(m["srcs"])[:200],
                       "", "", "",
                       ", ".join(m["eqs"])[:80]])
            r = ws.max_row
            for k in range(3):
                cell = ws.cell(r, 6 + k)
                if k < len(urls):
                    u = urls[k]
                    if u.startswith("http"):
                        cell.value = u.split("/")[2] + " 열기"
                        cell.hyperlink = u
                        cell.font = Font(color="0563C1", underline="single")
                    else:
                        cell.value = u            # KOSMO 처럼 URL 이 없는 출처
        for i, w in enumerate([8, 26, 34, 46, 40, 26, 26, 26, 26], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        summary.append((group, aas_name, len(merged), ws.title))
        print(f"{group:5} {aas_name:34} 슬롯 {len(merged):>4}")

    ws = wb.create_sheet("요약", 0)
    ws.append(["분류", "AAS", "파라미터 슬롯 수", "시트"])
    for c in range(1, 5):
        ws.cell(1, c).fill = HF
        ws.cell(1, c).font = Font(bold=True)
    for g, n, c, t in summary:
        ws.append([g, n, c, t])
    ws.append([])
    ws.append(["합계", "", sum(s[2] for s in summary), ""])
    for i, w in enumerate([10, 34, 16, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(OUT)
    print("\n생성:", OUT)


if __name__ == "__main__":
    main()
