"""FAB 시뮬레이션 요청 자료 — 공정정보·작업자운반·제품BOM계획 3시트.

확보분은 채워 두고 미확보분은 빈칸(노란색)으로 남겨 FAB이 채우도록 한다.
"""
import openpyxl, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 하드코딩 ────────────────────────────────────────────────────────────
RUNSHEET = r"C:/Users/KangTaehui/KG/keti/keti-fab/[KETI]전북본부 FAB 장비 및 공정 정보/3. TFT backplane runsheet_스마트.xlsx"
OUT = r"C:/Users/KangTaehui/KG/keti/CPRO_조립공정/시뮬레이션/Package/docs/FAB/FAB_시뮬_요청자료.xlsx"

HEAD_FILL = PatternFill("solid", fgColor="D9E2F3")
FILL_NEED = PatternFill("solid", fgColor="FFF2CC")     # 채워주셔야 할 칸
FILL_REF = PatternFill("solid", fgColor="F2F2F2")      # 참고용 (우리가 채움)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 핵심공정 8종 (Equipment → 핵심공정)
CORE = {
    "PECVD": "박막 증착", "Sputter": "박막 증착",
    "Dry etcher": "식각", "Wet etcher": "식각",
    "Coater/Developer/Mask Aligner": "감광막 코팅·현상 / 노광",
    "Wet stripper": "감광막 박리", "Manual stripper": "감광막 박리",
}
CORE_DEV = {
    "PECVD": "박막증착장비 chA", "Sputter": "박막증착장비 chD·chE",
    "Dry etcher": "박막증착장비 chC", "Wet etcher": "엣쳐/스트리퍼",
    "Coater/Developer/Mask Aligner": "현상장비 + 마스크얼라이너",
    "Wet stripper": "유기스트리퍼", "Manual stripper": "유기스트리퍼",
}
# 로그로 실측된 설비
MEASURED = {"박막증착장비 chA", "박막증착장비 chC"}

NUM = r"\d+(?:\.\d+)?"


def to_sec(key, raw):
    if raw is None:
        return None
    s = str(raw)
    if "?" in s:
        return None
    for pat, mul in ((r"(%s)\s*(?:hrs?|hours?)" % NUM, 3600),
                     (r"(%s)\s*min" % NUM, 60), (r"(%s)\s*(?:sec|Sec)" % NUM, 1)):
        m = re.search(pat, s)
        if m:
            return int(float(m.group(1)) * mul)
    nums = re.findall(NUM, s)
    if not nums:
        return None
    if "Deposition time" in key and "/" in s:
        return int(sum(float(x) for x in nums))
    if "/" in s:
        tail = re.findall(NUM, s.split("/", 1)[1])
        if not tail:
            return None
        n = float(tail[0])
    else:
        n = float(nums[0])
    return int(n * 60) if "(min)" in key else int(n)


# ── runsheet 파싱 ───────────────────────────────────────────────────────
wb = openpyxl.load_workbook(RUNSHEET, data_only=True)
ws = wb["PI Run sheet"]
rows = [r + (None,) * 6 for r in ws.iter_rows(values_only=True)]
grp = eqline = step = None
recs, cur = [], None
for r in rows:
    if r[0]:
        grp = str(r[0]).replace("\n", " ").strip()
    if r[1]:
        eqline = str(r[1]).replace("\n", " ").strip()
    p, v = r[2], r[3]
    if p and v is None:
        step = str(p).strip()
        continue
    if p == "Process parameter":
        cur = {"grp": grp, "line": eqline, "step": step, "eq": None, "times": {}}
        recs.append(cur)
        continue
    if cur is None or not p:
        continue
    k = str(p).strip()
    if k == "Equipment":
        cur["eq"] = str(v).strip()
    elif re.search(r"time|Time|Annealing|Treatment|^DI ", k):
        cur["times"][k] = v


def proc_time(c):
    """runsheet 에 적힌 시간 파라미터를 원문 그대로 나열한다.

    합산하면 스컴애싱을 넣을지 같은 판단이 끼어들므로 계산하지 않는다.
    """
    out = []
    for k, v in c["times"].items():
        if v is None:
            continue
        name = k.replace("Temp.(℃)/", "").replace("RPM/", "").replace("(sec)", "").strip()
        out.append(f"{name} {v}")
    return " · ".join(out)


# ── 엑셀 ────────────────────────────────────────────────────────────────
wbo = openpyxl.Workbook()


def setup(ws, cols, widths):
    ws.append(cols)
    for i, c in enumerate(ws[1], 1):
        c.font = Font(bold=True)
        c.fill = HEAD_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = widths[i - 1]
    ws.freeze_panes = "A2"


def style(ws, need_cols, ref_cols=()):
    for r in range(2, ws.max_row + 1):
        for c in ws[r]:
            c.border = BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for col in need_cols:
            ws.cell(r, col).fill = FILL_NEED
        for col in ref_cols:
            ws.cell(r, col).fill = FILL_REF


# 시트 1 — 공정 정보
s1 = wbo.active
s1.title = "1. 공정정보"
setup(s1, ["#", "공정그룹(층)", "공정명", "핵심공정", "담당 설비",
           "runsheet 기재 조건\n※원문 그대로", "설비 점유시간(초)\n※기판 넣고 빼는 시간 포함",
           "선행 공정", "동시 진행 가능 여부", "불량률(%)", "비고"],
      [4, 20, 20, 17, 21, 46, 16, 8, 12, 9, 16])
for i, c in enumerate(recs, 1):
    core = CORE.get(c["eq"], "열처리" if c["step"] == "Annealing" else "")
    dev = CORE_DEV.get(c["eq"], "고온진공오븐" if c["step"] == "Annealing" else "")
    pt = proc_time(c)
    s1.append([i, c["grp"], c["step"], core, dev,
               pt, "", f"#{i-1}" if i > 1 else "(없음)", "", "",
               "로그 실측 있음" if dev in MEASURED else ""])
style(s1, need_cols=(7, 9, 10), ref_cols=(6,))

# 시트 2 — 작업자·운반
s2 = wbo.create_sheet("2. 작업자·운반")
setup(s2, ["구분", "핵심공정", "담당 설비", "작업 인원(명)", "담당 방식\n(전담/순회)",
           "1인이 동시에 보는 설비 수", "다음 공정으로 옮기는 방법\n(사람/컨베이어/자동반송)",
           "1회 운반 매수", "비고"],
      [10, 20, 24, 12, 14, 16, 22, 12, 20])
CORE8 = [("기판 세정", "기판세정기"), ("박막 증착", "박막증착장비 chA·chD·chE"),
         ("감광막 코팅·현상", "현상장비 (스핀디벨로퍼·스핀트랙)"), ("노광", "마스크얼라이너"),
         ("식각", "박막증착장비 chC + 엣쳐/스트리퍼"), ("감광막 박리", "유기스트리퍼"),
         ("열처리·결정화", "고온진공오븐"), ("기판 반송", "박막증착장비 TM")]
for nm, dev in CORE8:
    s2.append(["공정별", nm, dev, "", "", "", "", "", ""])
s2.append([])
for label, memo in [("근무 시작 시각", "로그상 08~09시 추정"), ("근무 종료 시각", "로그상 16~17시 추정"),
                    ("휴게 시작·종료", "로그상 12시대 공백"), ("야간 운전 여부", "로그상 야간 실행 0건"),
                    ("주말 운전 여부", "로그상 주말 실행 0건"),
                    ("총 작업 인원", ""), ("작업 지시 방법", "누가 어떤 기준으로 다음 작업을 정하는지")]:
    s2.append(["전역", label, memo, "", "", "", "", "", ""])
style(s2, need_cols=(4, 5, 6, 7, 8))

# 시트 3 — 제품·BOM·계획
s3 = wbo.create_sheet("3. 제품·BOM·계획")
setup(s3, ["구분", "항목", "현재 파악", "값", "단위", "비고"],
      [12, 30, 34, 14, 12, 26])
BLOCK = [
    ("제품", "제작 대상 종류", "TFT backplane (PI 기판 / 유리 기판)", "", "", "시뮬 대상을 무엇으로 잡을지"),
    ("제품", "1회 제작 단위", "제작일정상 7매 → 6매 → 3매", "", "매", "의뢰 1건이 몇 매인지"),
    ("제품", "기판 규격", "370 × 470 mm", "", "mm", "확인만"),
    ("제품", "카세트 단위", "25매", "", "매", "확인만"),
    ("BOM", "기판 (유리/PI) 소요", "1장 투입 → 1장 산출", "", "매/1매", ""),
    ("BOM", "Mo 타깃 교체 주기", "미상", "", "매 또는 시간", "몇 매 처리 후 교체하는지"),
    ("BOM", "ITO 타깃 교체 주기", "미상", "", "매 또는 시간", ""),
    ("BOM", "IGZO 타깃 교체 주기", "미상", "", "매 또는 시간", ""),
    ("BOM", "Ag 타깃 교체 주기", "미상", "", "매 또는 시간", ""),
    ("BOM", "타깃 교체 소요시간", "미상", "", "시간", "교체 중 설비 정지시간"),
    ("BOM", "PR(ZPP1700PG-30) 1회 도포량", "미상", "", "mL/매", ""),
    ("BOM", "현상액 TMAH 2.38% 배스 용량", "미상", "", "L", ""),
    ("BOM", "현상액 교체 전 처리 매수", "미상", "", "매", ""),
    ("BOM", "ITO etchant 배스 용량 / 교체 매수", "미상", "", "L / 매", ""),
    ("BOM", "Organic strip chemical 배스 용량 / 교체 매수", "미상", "", "L / 매", ""),
    ("BOM", "NMP 배스 용량 / 교체 매수", "미상", "", "L / 매", ""),
    ("BOM", "DI water 사용량", "미상", "", "L/매", ""),
    ("BOM", "가스 사용량", "로그에서 산출 가능", "—", "—", "요청 불필요"),
    ("BOM", "포토마스크 보유 목록", "자료에 언급 없음", "", "", "층별 사용 마스크"),
    ("BOM", "포토마스크 제작 리드타임", "미상", "", "일", ""),
    ("재고", "자재 발주점 / 최대재고", "미상", "", "", "품목별"),
    ("재고", "자재 조달 리드타임", "미상", "", "일", "특히 ITO·IGZO 타깃"),
    ("재고", "자재 재고 관리 방법", "미상", "", "", "장부·시스템 유무"),
    ("계획", "제작 의뢰 단위", "미상", "", "", "건별 / 과제별"),
    ("계획", "납기 기준", "runsheet 일정란 공백", "", "일", "의뢰 후 며칠 내"),
    ("계획", "기판 1장 전체 리드타임", "미상", "", "일", "31공정 완주 소요일"),
    ("계획", "하루 진행 공정 수", "미상", "", "공정/일", ""),
    ("계획", "생산계획 문서 유무", "제작일정 양식은 있으나 날짜 미기입", "", "", ""),
    ("계획", "동시 의뢰 시 우선순위", "미상", "", "", ""),
    ("계획", "제품 5종 생산 비중", "TFT backplane·TEG-cell 2종·프린팅 2종", "", "%", ""),
]
for b in BLOCK:
    s3.append(list(b))
style(s3, need_cols=(4, 5))

wbo.save(OUT)

print(f"1. 공정정보    {s1.max_row-1}행  (공정 31 × 채울칸 3: 점유시간·동시진행·불량률)")
print(f"2. 작업자·운반  {s2.max_row-1}행  (핵심공정 8 + 전역 7)")
print(f"3. 제품·BOM·계획 {s3.max_row-1}행")
print("→", OUT)
