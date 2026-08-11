"""KETI FAB 잔여 자료조사 계획을 엑셀로 만든다.

직접 요청(KETI 담당자 연락)은 이미 진행 중이라 제외하고,
사람이 사이트에서 직접 찾아야 하는 것만 담는다.
"""

import glob
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

R = r"C:\Users\KangTaehui\KG\keti\CPRO_조립공정\시뮬레이션\Package\docs\원본자료\keti-fab\포털수집"
OUT = r"C:\Users\KangTaehui\KG\keti\CPRO_조립공정\시뮬레이션\Package\docs\원본자료\keti-fab\KETI_FAB_잔여자료조사.xlsx"

HEAD_FILL = PatternFill("solid", fgColor="DDEBF7")
HEAD_FONT = Font(bold=True)


def sheet(wb, title, cols, rows, widths, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


# ── 1. 사이트 요약 ───────────────────────────────────────────────────
SITES = [
    ("1", "NTIS 국가과학기술지식정보서비스", "https://www.ntis.go.kr",
     "장비 구축 과제의 최종보고서 — 도입 장비 사양표가 들어 있다",
     "회원가입 → 과제고유번호로 검색 → 과제상세 → 연구보고서 탭 → 원문 신청",
     "높음 — 이 설비 그 자체의 문서", "회원가입 필요, 기관 승인 있을 수 있음",
     "NTIS_과제 시트의 고유번호 20개"),
    ("2", "KIPRIS 특허정보검색서비스", "https://www.kipris.or.kr",
     "제조사 특허 명세서 원문 — Google Patents 판은 번역·요약이라 잘린다",
     "검색창에 KIPRIS_검색식 시트의 검색식을 그대로 입력",
     "중간 — 파라미터 종류는 얻지만 이 설비 값은 아님", "무료, 로그인 없이 검색 가능",
     "KIPRIS_검색식 시트의 검색식 9개"),
    ("3", "KIPRIS Plus (오픈API)", "https://plus.kipris.or.kr",
     "위 특허를 대량·자동으로 받기 위한 API 키",
     "회원가입 → API 신청 → 키 발급 → 키를 전달하면 자동 수집 가능",
     "중간 — 2번을 자동화하는 수단", "무료 발급, 회원가입 필요",
     "키 발급 1건"),
    ("4", "ScienceON (KISTI)", "https://scienceon.kisti.re.kr",
     "논문 원문 — 출판사 벽에 막힌 68편",
     "DOI 로 검색 → 원문 보기. 국내 계정으로 상당수 열림",
     "높음 — 실제 공정 조건이 실험 파트에 있음", "기관 계정 권장",
     "논문_미확보 시트의 68편"),
    ("5", "RISS 학술연구정보서비스", "https://www.riss.kr",
     "위와 동일 — ScienceON 에서 안 열린 것",
     "DOI 또는 제목으로 검색 → 대학도서관 연계 원문",
     "높음", "대학 계정 있으면 유리", "논문_미확보 시트 중 4번에서 실패분"),
    ("6", "Google Patents", "https://patents.google.com",
     "특허 검색·본문 (브라우저로는 정상, 자동 수집만 요청량 제한에 걸림)",
     "검색창에 assignee:에프엔에스테크 형식으로 입력",
     "낮음 — 이미 43건 수집 완료, 보완용", "무료",
     "KIPRIS 로 안 나오는 건만"),
    ("7", "제조사 웹사이트·기술문의", "제조사 시트 참조",
     "장비 사양서(Spec sheet) — 국산 장비는 웹 공개가 없어 문의가 빠름",
     "모델명·납품연도를 대고 사양서 요청",
     "높음 — 정확한 사양", "요청형이라 회신 대기 발생",
     "제조사 시트의 9개사"),
]


def build_ntis():
    rows, seen = [], set()
    for f in sorted(glob.glob(os.path.join(R, "*", "NTIS", "과제정보.json"))):
        d = json.load(open(f, encoding="utf-8"))
        p = d["ZEUS_과제정보"]
        eq = os.path.basename(os.path.dirname(os.path.dirname(f)))
        for h in d["NTIS_검색결과"][:5]:
            key = h["과제고유번호"]
            if not key or key in seen:
                continue
            seen.add(key)
            rows.append([p["세부과제명"], h["과제고유번호"], h["연도"], h["과제명"][:60],
                         p["주관기관"], p["연구책임자"], p["과제수행기간"], eq,
                         f"https://www.ntis.go.kr/project/pjtInfo.do?pjtId={h['과제고유번호']}",
                         "연구보고서 탭에서 최종보고서 원문 신청 → 도입 장비 사양표 확인"])
    return rows


MAKERS = [
    ("에프엔에스테크", "AP=[에프엔에스테크]", "기판세정기, 현상장비, 엣쳐/스트리퍼, 유기스트리퍼",
     "세정 bath 구성·온도·분사압력·이송속도"),
    ("선익시스템", "AP=[선익시스템]*AB=[증착]", "유기증착기",
     "챔버별 셀 온도 상한, 증발원 구조, 기판 이송속도"),
    ("테스", "AP=[테스]*AB=[증착]", "박막증착장비",
     "sputter/PECVD/RTA/dry etcher 챔버별 공정 온도·압력·파워"),
    ("제우스", "AP=[제우스]*AB=[식각]", "화학 습식 증착(CBD)",
     "식각액 온도, 회전속도, 노즐 토출 조건"),
    ("씨엔원", "AP=[씨엔원]*AB=[원자층]", "PEALD",
     "웨이퍼/분말 챔버 온도, 플라즈마 주파수·파워, 전구체 캐니스터"),
    ("나래나노텍", "AP=[나래나노텍]", "리버스 옵셋 프린터",
     "블랭킷 주행속도, Nip 조절범위, 얼라인 정밀도"),
    ("코디엠", "AP=[코디엠]*AB=[노광]", "마스크얼라이너",
     "노광량, 갭 제어, proximity pin 높이"),
    ("에스브이에스", "AP=[에스브이에스]*AB=[반도체]", "스핀 트랙 시스템",
     "코터/디벨로퍼 rpm, hot/cool plate 온도 (매뉴얼로 일부 확보됨)"),
    ("마이다스시스템", "AP=[마이다스시스템]", "스핀디벨로퍼",
     "spin speed·가속, 디스펜서 노즐 구성"),
]


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    wb = Workbook()

    sheet(wb, "사이트_요약",
          ["순위", "사이트", "주소", "여기서 얻는 것", "방법", "효과", "제약", "조사 대상"],
          SITES, [5, 26, 30, 40, 44, 26, 24, 26], first=True)

    ntis = build_ntis()
    sheet(wb, "NTIS_과제",
          ["세부과제명", "과제고유번호", "연도", "NTIS 과제명", "주관기관", "연구책임자",
           "수행기간", "대표 설비", "링크", "확인할 것"],
          ntis, [30, 14, 7, 34, 22, 11, 20, 22, 46, 40])

    sheet(wb, "KIPRIS_검색식",
          ["제조사", "검색식", "해당 설비", "찾을 파라미터", "링크"],
          [[m[0], m[1], m[2], m[3], "https://www.kipris.or.kr"] for m in MAKERS],
          [16, 26, 34, 40, 26])

    miss_path = os.path.join(R, "_논문_미확보.json")
    rows = []
    for meta in sorted(glob.glob(os.path.join(R, "*", "논문", "*.txt"))):
        b = os.path.basename(meta)
        if b.startswith("_") or b.endswith("_본문.txt"):
            continue
        if os.path.exists(meta.replace(".txt", "_본문.txt")):
            continue
        eq = os.path.basename(os.path.dirname(os.path.dirname(meta)))
        ln = open(meta, encoding="utf-8").read().split("\n")
        doi = next((l.strip() for l in ln[:5] if "doi.org" in l), "")
        rows.append([eq, ln[0][:110], ln[1][:70], doi,
                     "실험/Methods 파트의 공정 조건 (온도·rpm·시간·유량·압력)"])
    json.dump(rows, open(miss_path, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    sheet(wb, "논문_미확보",
          ["설비", "제목", "저널·연도·인용", "DOI", "확인할 것"],
          rows, [22, 60, 34, 40, 40])

    sheet(wb, "제조사",
          ["제조사", "해당 설비", "요청할 자료", "찾을 파라미터", "검색어(웹)"],
          [[m[0], m[2], "장비 사양서(Spec sheet), 매뉴얼 사양 장", m[3],
            f"{m[0]} 사양서 OR spec sheet"] for m in MAKERS],
          [16, 34, 32, 40, 30])

    wb.save(OUT)
    print("생성:", OUT)
    for ws in wb:
        print(f"  {ws.title:14} {ws.max_row - 1}행")


if __name__ == "__main__":
    main()
