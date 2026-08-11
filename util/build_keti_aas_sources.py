"""AAS 12개별로 '사람이 열어볼 자료'만 모은다.

넣는 것   : 원본 PDF, 사이트 화면 캡처, 클릭 가능한 링크 목록(엑셀)
넣지 않는 것 : md·html·txt·json — 전부 내가 파싱하거나 만들어낸 중간 산출물이다.

본설비(KETI 보유)와 유사장비(타 기관)를 한 목록에 담되 '구분' 열로 나눈다.
"""

import csv
import glob
import json
import os
import re
import shutil
import sys
import zipfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import bundle_keti_aas_rawtext as B
from keti_similar_verdict import verdict_of
from keti_source_similarity import similarity_of
from keti_bullet import bullets, bullets_list, bullets_loc, search_keywords
import keti_zeus_slots

BASE = r"C:\Users\KangTaehui\KG\keti\CPRO_조립공정\시뮬레이션\Package"
OUT = os.path.join(BASE, r"docs\원본자료\keti-fab\AAS자료모음")
R = B.R

ZEUS_READ = "https://www.zeus.go.kr/search/equip/read/"
ZEUS_RESV = "https://www.zeus.go.kr/resv/equip/read/"
ITUBE_VIEW = "https://www.itube.or.kr/aplct/equipSrch/sharingView.do?g_menu_id=MNID210100&equip_no="
ITUBE_FILE = "https://www.itube.or.kr/unitc/equipuse/myequip/fileDownMyEquip.do?g_menu_id=&equip_file_no="

KOSMO_ZIP = os.path.join(BASE, r"docs\원본자료\keti-fab\kasmo_후보데이터.zip")
KOSMO_NOTE = "사용자 제공 (kasmo_후보데이터.zip) — 공개 URL 없음"

# KOSMO 참고자료를 붙일 AAS. 파라미터 슬롯을 가져온 곳과 같아야 한다.
# 중복본((1) 이 붙은 것)과 구버전(260128)은 뺐다.
# 논문·특허는 설비 사양서가 아니라 '그 공정을 돌린 조건'이다.
# 값은 그날 실험치라 못 쓰지만 **조정하는 파라미터의 이름**은 슬롯 후보가 된다.
PAPER_DIR = os.path.join(BASE, r"docs\원본자료\keti-fab\논문수집\본문")
PATENT_DIR = os.path.join(BASE, r"docs\원본자료\keti-fab\특허수집\명세서")
PAPER_SLOT = os.path.join(BASE, r"docs\원본자료\keti-fab\논문수집\_논문_슬롯후보.tsv")
PATENT_SLOT = os.path.join(BASE, r"docs\원본자료\keti-fab\특허수집\_특허_슬롯후보.tsv")


def patent_slots_by_pub(with_quote=False):
    """특허에서 뽑아 둔 슬롯 후보를 공개번호별로 모은다."""
    out = {}
    if not os.path.exists(PATENT_SLOT):
        return out
    txt = open(PATENT_SLOT, encoding="utf-8-sig").read().replace(chr(13) + chr(10), chr(10))
    for r in csv.DictReader(txt.split(chr(10)), delimiter="	"):
        pub = (r.get("공개번호") or "").strip()
        nm = (r.get("Property후보") or "").strip()
        if pub and nm:
            out.setdefault(pub, []).append(
                (nm, (r.get("원문표현") or "").strip()) if with_quote else nm)
    return out


def paper_slots_by_equip():
    """논문에서 뽑아 둔 슬롯 후보를 설비별로 모은다."""
    out = {}
    if not os.path.exists(PAPER_SLOT):
        return out
    txt = open(PAPER_SLOT, encoding="utf-8-sig").read().replace(chr(13) + chr(10), chr(10))
    for r in csv.DictReader(txt.split(chr(10)), delimiter="	"):
        eq = (r.get("설비") or "").strip()
        nm = (r.get("Property후보") or "").strip()
        if not eq or not nm:
            continue
        out.setdefault(eq, {}).setdefault(nm, []).append((r.get("DOI", ""), r.get("절위치", "")))
    return out

# 제조사 → 그 제조사가 만든 KETI 설비가 속한 AAS
MAKER_AAS = {
    "선익시스템": ["유기증착기-OrganicChamber", "유기증착기-MetalChamber", "유기증착기-PlasmaChamber"],
    "테스": ["박막증착장비-PECVD", "박막증착장비-Sputter", "박막증착장비-DryEtcher",
             "박막증착장비-ThermalEvaporator"],
    "씨엔원": ["PEALD"],
    "에프엔에스테크": ["식각/스트립"],
    "제우스": ["CBD"],
    "마이다스시스템": ["현상장비"],
    "에스브이에스": ["현상장비"],
    "코디엠": ["마스크 얼라이너"],
    "나래나노텍": ["프린팅"],
}

KOSMO_FILES = {
    "유기증착기-PlasmaChamber": [
        ("표면처리기.aasx", "KOSMO 표면처리기 AAS"),
        ("제조데이터 표준 활용 가이드-표면처리기.docx", "KOSMO 표면처리기 활용 가이드"),
        ("사전검증보고서-표면처리기.docx", "KOSMO 표면처리기 사전검증보고서"),
        ("AAS_체크리스트-표면처리기.xlsx", "KOSMO 표면처리기 체크리스트"),
    ],
    "박막증착장비-DryEtcher": [
        ("[일.6]SemiconductorCircuitEtchingEquipment.aasx", "KOSMO 반도체회로 에칭장비 AAS"),
        ("제조데이터 표준 활용 가이드_반도체회로 에칭장비.docx",
         "KOSMO 반도체회로 에칭장비 활용 가이드 — 2-(나) 가 슬롯 선별 근거"),
        ("반도체회로 에칭장비 사전 검증 보고서.docx", "KOSMO 반도체회로 에칭장비 사전검증보고서"),
        ("반도체회로 에칭장비_체크리스트.xlsx", "KOSMO 반도체회로 에칭장비 체크리스트"),
    ],
    "CBD": [
        ("[신규.2]ElectroplatingBath.aasx", "KOSMO 전해도금조 AAS"),
        ("가이드 전해도금조_260311.docx.docx", "KOSMO 전해도금조 활용 가이드 (260311 최신본)"),
        ("사전검증보고서-전해도금조.docx", "KOSMO 전해도금조 사전검증보고서"),
        ("전해도금조_체크리스트.xlsx", "KOSMO 전해도금조 체크리스트"),
    ],
}

# 장비가 조정하는 값을 나타내는 표현 (결과 측정치는 노린 것이 아니다)
SPEC_RE = re.compile(
    r"(?:온도|압력|속도|파워|전력|주파수|유량|간격|각도|두께|시간|농도|회전수|전압|전류)"
    r"[^.]{0,20}?(?:범위|제어|조절|유지|설정|이내|이상|이하|내지)"
    # 마이크로미터는 특허 본문에서 그리스 뮤(μm)로 쓰는 경우가 압도적이다.
    # 완성형 기호(㎛)만 넣으면 대부분을 놓친다 — 실측 μm 56회 vs ㎛ 1회.
    r"|\d[\d.,]*\s*(?:℃|℉|㎜|mm|㎝|cm|㎛|μm|um|㎚|nm|Torr|mTorr|sccm|slm"
    r"|kW|W|MHz|kHz|Hz|rpm|kPa|MPa|Pa|bar|psi|%|°)")

LINK = Font(color="0563C1", underline="single")
HEAD = PatternFill("solid", fgColor="DDEBF7")


def safe(name, n=60):
    return re.sub(r'[\\/:*?"<>|]', "_", name)[:n].strip()


def meta_of(text):
    def one(pat):
        m = re.search(pat + r"\s*\n\s*([^\n]+)", text)
        return m.group(1).strip() if m else ""
    i, j = text.find("특성"), text.find("용도설명")
    lines = [l for l in text[i + 2:j].split("\n") if l.strip()] if 0 < i < j else []
    return one(r"제작사명\(모델명\)") or one(r"제작사명 \| 모델명"), one(r"보유기관명"), len(lines)



# 논문 본문에서 절 제목을 뽑아 '파라미터가 어디에 있는지' 를 적는다.
_SEC = re.compile(
    r"^[ 	]*("
    r"\d+(?:\.\d+)*\.?[ 	]+[A-Z][^" + chr(10) + r"]{3,60}"
    r"|(?:Materials and Methods|Experimental(?: Section| Details| Methods| Procedure)?|Methods)[^" + chr(10) + r"]{0,40}"
    r")[ 	]*$", re.M)


def paper_location(path):
    """논문 txt 에서 파라미터가 실린 절 제목을 모아 위치 문자열을 만든다."""
    raw = open(path, encoding="utf-8").read()
    head = raw.split(chr(10))[:8]
    mode = next((l for l in head if l.startswith("[추출모드")), "")
    body = raw.split(chr(10) + chr(10), 1)[-1]
    secs, seen = [], set()
    for m in _SEC.finditer(body[:12000]):
        t = " ".join(m.group(1).split())[:52]
        if t.lower() in seen:
            continue
        seen.add(t.lower())
        secs.append(t)
        if len(secs) >= 6:
            break
    if secs:
        return " → ".join(secs)
    if "fallback" in mode:
        return "실험 절이 없어 파라미터 언급 문단만 발췌 (리뷰 논문 가능성)"
    if "초록" in mode or "abstract" in mode.lower():
        return "초록만 — 본문 접근 불가"
    return mode.strip("[]") or "본문 전체"



# 논문이 다루는 공정을 본문에서 찾아, 그 AAS 설비와 무엇이 닮았는지 적는다.
_PROC = re.compile(
    r"(?<![A-Za-z])"
    r"(PEALD|ALD|PECVD|LPCVD|CVD|sputter\w*|evaporat\w*|spin[- ]coat\w*|screen[- ]print\w*"
    r"|inkjet|EHD|DRIE|RIE|etch\w*|photolith\w*|develop\w*|bake|anneal\w*|hydrothermal"
    r"|electrospin\w*|CMP|lift[- ]off|IPL|curing)", re.I)

# 그 AAS 가 실제로 하는 공정 (논문 공정이 여기 걸리면 '같은 공정'으로 본다)
_AAS_PROC = {
    "박막증착장비-PECVD": ["pecvd", "cvd", "lpcvd"],
    "박막증착장비-Sputter": ["sputter"],
    "박막증착장비-DryEtcher": ["rie", "drie", "etch"],
    "박막증착장비-ThermalEvaporator": ["evaporat"],
    "유기증착기-PlasmaChamber": ["rie", "etch"],
    "유기증착기-OrganicChamber": ["evaporat"],
    "유기증착기-MetalChamber": ["evaporat"],
    "PEALD": ["peald", "ald"],
    "현상장비": ["spin coat", "spin-coat", "develop", "bake", "photolith"],
    "마스크 얼라이너": ["photolith"],
    "식각/스트립": ["etch", "develop", "lift-off", "lift off"],
    "프린팅": ["screen print", "inkjet", "ehd", "curing"],
    "CBD": ["hydrothermal", "electrospin", "curing"],
}


def paper_similarity(path, aas):
    """논문 본문의 공정과 그 AAS 의 공정을 대조해 유사성 문장을 만든다."""
    body = open(path, encoding="utf-8").read().split(chr(10) + chr(10), 1)[-1]
    found = sorted({m.group(1).lower() for m in _PROC.finditer(body)})
    want = _AAS_PROC.get(aas, [])
    hit = sorted({w for w in want for f in found if w in f})
    shown = ", ".join(found[:8])
    if hit:
        return (f"이 AAS 의 공정({', '.join(hit)})을 실제로 돌린 논문 / "
                f"본문 공정: {shown} / "
                f"값은 그날 실험 조건이라 사용 불가 / 조정 파라미터의 이름만 Property 후보")
    if not found:
        return "본문에서 이 AAS 의 공정 미확인 / Property 후보로 쓰기 어려움"
    return (f"이 AAS 의 공정과 직접 겹치지 않음 / 본문 공정: {shown} / "
            f"같은 계통 장비면 파라미터 이름만 참고 가능")



# KETI FAB 12종에 없는 공정이거나 소모품 물성만 다루는 특허. 수치가 많아도 슬롯이 안 나온다.
OFF_TOPIC = re.compile(r"연마\s*패드|연마패드|CMP|폴리싱\s*패드|전자액자|모니터링|다이싱|본딩\s*와이어")
# 소모품 물성 용어 — 이것만 잔뜩 나오면 장비 운전 조건이 아니다
MATERIAL = re.compile(r"쇼어|경도|인장\s*강도|인열\s*강도|연신율|압축률|비중|밀도|탄성\s*계수|모듈러스")


# 등급은 다섯으로만 쓴다. 유사장비는 계통이 온전히 겹치면 '유사장비',
# 일부만 겹치면 '부분유사' 로 나눈다.
GRADE = {
    "최우선": "유사장비", "사용가능": "유사장비",
    "부분참고": "부분유사", "참고만": "부분유사",
}


def grade_of(구분, 등급, 장비명=""):
    if 구분 == "Fab장비":
        return "Fab장비"
    if 구분 == "논문":
        return "논문"
    if 구분 == "특허":
        return "특허"
    if 구분 == "KOSMO 참고":
        # AAS 파일 자체는 유사장비 사양, 가이드·보고서·체크리스트는 부분 참고
        return "유사장비" if "AAS" in 장비명 else "부분유사"
    return GRADE.get(등급, "부분유사")


def patent_verdict(text, maker):
    """특허 본문에 실제로 파라미터가 있는지 판정한다.

    국산 장비 특허는 구조(부품 배치·형상) 위주가 많아 수치 사양이 없는 경우가 흔하다.
    그런 특허는 슬롯을 주지 못하므로 목록에서 구분해 둔다.
    """
    n_num = len(SPEC_RE.findall(text))
    n_ctrl = len(re.findall(r"(?:제어|조절|가변|설정)(?:할 수|하는|되는|된다)", text))
    n_mat = len(MATERIAL.findall(text))
    head = text[:1500]
    if OFF_TOPIC.search(head) or (n_mat >= 5 and n_mat * 2 >= n_num):
        return (f"{maker} 명의 특허 / KETI FAB 12종에 없는 공정이거나 소모품 물성 위주 / "
                f"수치 {n_num}건 중 재료물성 용어 {n_mat}건 / Property 후보 아님",
                "해당 없음 (공정 무관 · 소모품 물성)")
    # 대형 특허는 수집 때 앞 1만자만 저장했다. 그 경우 실제 수치는 이보다 많을 수 있다.
    # 브라우저 수집분만 앞 1만자로 잘랐다. 딱 그 언저리일 때만 표시한다.
    cut = (" (대형 특허라 앞부분만 수집 — 실제로는 더 많을 수 있음)"
           if 9500 <= len(text) <= 11000 else "")
    # 실제로 숫자+단위가 붙은 것만 따로 센다.
    # SPEC_RE 의 첫 대안('온도를 제어' 같은 서술)은 수치가 아니라 정성 표현이라 구분해야 한다.
    n_val = len(re.findall(
        r"\d[\d.,]*\s*(?:℃|℉|㎜|mm|㎝|cm|㎛|μm|um|㎚|nm|Torr|mTorr|sccm|slm"
        r"|kW|W|MHz|kHz|Hz|rpm|kPa|MPa|Pa|bar|psi|%)", text))
    if n_num >= 8 and n_val == 0:
        return (f"{maker} 명의 특허 / 제어·조절 서술 {n_num}건 / 수치가 붙은 표현 없음 "
                f"(정성 서술뿐) / Property 이름 후보로만 참고",
                "본문 전체 (수치 없음)")
    if n_num >= 8:
        return (f"{maker} 명의 특허 / 수치 사양 {n_num}건 / Property 후보 확보" + cut,
                f"본문 + 청구항 / 수치 사양 {n_num}건" + cut)
    if n_num >= 1:
        return (f"{maker} 명의 특허 / 수치 사양 {n_num}건뿐이라 Property 기여 작음 / "
                f"부품 이름을 Property 이름 참고용",
                f"본문 + 청구항 / 수치 사양 {n_num}건")
    return (f"{maker} 명의 구조 특허 / 부품 배치·형상이 주 내용 / 수치 사양 없음 / "
            f"제어 언급 {n_ctrl}건도 '설계 조건에 따라 조절' 수준 / 부품 이름만 참고 가능",
            "본문 전체 (수치 없음)")



def main():
    sys.stdout.reconfigure(encoding="utf-8")
    os.makedirs(OUT, exist_ok=True)
    zeus = {e["no"]: e for e in json.load(open(os.path.join(R, "_zeus_mapping.json"), encoding="utf-8"))}
    itube = {e["no"]: e for e in json.load(open(os.path.join(R, "_itube_mapping.json"), encoding="utf-8"))}
    resv = {e["폴더"]: e["활용번호"] for e in
            json.load(open(os.path.join(R, "_zeus_resv_mapping.json"), encoding="utf-8"))}

    index, cap_todo = [], []
    pslots = paper_slots_by_equip()
    patslots = patent_slots_by_pub()
    patquote = patent_slots_by_pub(with_quote=True)
    zeus_tbl = keti_zeus_slots.load()
    patquote = patent_slots_by_pub(with_quote=True)
    zeus_tbl = keti_zeus_slots.load()

    for grp, aas, folders, keys in B.AAS:
        dst = os.path.join(OUT, grp, B.safe_aas(aas))
        os.makedirs(dst, exist_ok=True)
        rows = []

        # ── 본설비 (KETI 보유) ─────────────────────────────────────
        for fr in folders:
            nm = os.path.basename(fr)
            no = int(nm[:2])
            d = os.path.join(R, fr)
            ze, te = zeus.get(no, {}), itube.get(no, {})

            links = []
            if ze.get("zeus_id"):
                links.append(("ZEUS 등록장비 상세", ZEUS_READ + ze["zeus_id"]))
            # 등록장비가 없는데 예약 활용번호만 있는 경우는 옛 탐색 흔적이다.
            # 44_PEALD 가 그렇다 — 그 번호는 KETI 성남의 다른 장비(아이작리서치 iMV-DX4)를 가리킨다.
            if resv.get(nm):
                links.append(("ZEUS 장비예약 상세", ZEUS_RESV + resv[nm]))
            if te.get("itube_epn"):
                links.append(("i-Tube 상세", ITUBE_VIEW + te["itube_epn"]))

            # 매뉴얼 PDF 는 원본을 그대로 옮기고 내려받은 주소를 함께 남긴다
            pdfs = []
            att = te.get("매뉴얼_첨부", [])
            for p in sorted(glob.glob(os.path.join(d, "*.pdf"))):
                fn = os.path.basename(p)
                shutil.copy2(p, os.path.join(dst, fn))
                url = ITUBE_FILE + att[0]["파일번호"] if att else ""
                pdfs.append((fn, url))

            cap = f"캡처_본설비_{safe(nm)}.jpg"
            # 이미 찍어 둔 캡처가 있으면 다시 찍지 않는다
            if links and not os.path.exists(os.path.join(dst, cap)):
                pick = next((u for lb, u in links if "장비예약" in lb), links[0][1])
                cap_todo.append({"AAS": aas, "장비명": nm, "url": pick,
                                 "저장경로": os.path.abspath(os.path.join(dst, cap))})

            base = {"구분": "Fab장비", "AAS": aas, "분류": grp, "장비명": nm,
                    "제작사(모델)": ze.get("zeus_제작사") or te.get("itube_정보", {}).get("제작사", ""),
                    "보유기관": "한국전자기술연구원 (KETI 전주)", "등급": "Fab장비"}
            if not links and not pdfs:
                rows.append({**base, "자료": "(포털 미등록)", "링크": "", "파일": ""})
            for label, url in links:
                extra = {}
                if "i-Tube" in label:
                    extra["위치"] = ("대장정보(제작사·설치기관·담당자·내용연수) / "
                                     "매뉴얼·카탈로그 첨부 / 사양 기재 없음")
                rows.append({**base, **extra, "자료": label, "링크": url,
                             "파일": cap if os.path.exists(os.path.join(dst, cap)) else ""})
            for fn, url in pdfs:
                rows.append({**base, "자료": "설비 매뉴얼 PDF", "링크": url, "파일": fn})

        # ── 유사장비 (타 기관) ────────────────────────────────────
        for fr in folders:
            for f in sorted(glob.glob(os.path.join(R, fr, "유사장비", "*.txt"))):
                b = os.path.basename(f).replace(".txt", "")
                if keys != ["*"] and not any(k in b for k in keys):
                    continue
                t = open(f, encoding="utf-8").read()
                mk, org, nline = meta_of(t)
                g, why = verdict_of(aas, b)
                m = re.search(r"시설장비활용번호\s*\n\s*([A-Za-z0-9\-]+)", t)
                url = ZEUS_RESV + m.group(1) if m else ""
                cap = f"캡처_{safe(b)}.jpg"
                # 이전 단계에서 찍어 둔 캡처가 있으면 옮겨 온다
                old = os.path.join(BASE, r"docs\원본자료\keti-fab\유사장비후보", grp, B.safe_aas(aas), cap)
                if os.path.exists(old) and not os.path.exists(os.path.join(dst, cap)):
                    shutil.copy2(old, os.path.join(dst, cap))
                if url and not os.path.exists(os.path.join(dst, cap)):
                    cap_todo.append({"AAS": aas, "장비명": b, "url": url,
                                     "저장경로": os.path.abspath(os.path.join(dst, cap))})
                rows.append({"구분": "유사장비", "AAS": aas, "분류": grp, "장비명": b,
                             "제작사(모델)": mk, "보유기관": org, "등급": g,
                             "자료": "ZEUS 장비예약 상세" if url else "(예약 페이지 없음)",
                             "링크": url,
                             "파일": cap if os.path.exists(os.path.join(dst, cap)) else "",
                             "사유": why, "특성 줄수": nline})

        # ── KOSMO 참고자료 (사용자 제공 zip 에서 꺼내 넣는다) ──────────
        for member, label in KOSMO_FILES.get(aas, []):
            fn = safe(os.path.basename(member), 80)
            tgt = os.path.join(dst, fn)
            if not os.path.exists(tgt):
                with zipfile.ZipFile(KOSMO_ZIP) as z:
                    if member in z.namelist():
                        with z.open(member) as src_f, open(tgt, "wb") as out_f:
                            shutil.copyfileobj(src_f, out_f)
            if os.path.exists(tgt):
                rows.append({"구분": "KOSMO 참고", "AAS": aas, "분류": grp, "장비명": label,
                             "제작사(모델)": "", "보유기관": "KOSMO", "등급": "참고자료",
                             "자료": os.path.splitext(fn)[1].lstrip(".").upper() + " 파일",
                             "링크": "", "파일": fn, "사유": KOSMO_NOTE})

        # ── 논문 (그 설비로 돌린 공정 조건) ─────────────────────────
        eq_names = {os.path.basename(x) for x in folders}
        if os.path.isdir(PAPER_DIR):
            for f in sorted(os.listdir(PAPER_DIR)):
                if not f.endswith(".txt"):
                    continue
                eq = f.split("__")[0]
                if eq not in eq_names:
                    continue
                head = open(os.path.join(PAPER_DIR, f), encoding="utf-8").read().split(chr(10))
                title = head[1] if len(head) > 1 else f
                url = head[2] if len(head) > 2 else ""
                mode = next((l for l in head[:8] if l.startswith("[추출모드")), "")
                only_abs = "초록" in "".join(head[:8])
                doi_key = f.split("__")[-1].replace(".txt", "").replace("_", "/", 1)
                mine = sorted({n for e, d in pslots.items() if e.startswith(eq)
                               for n, srcs in d.items()
                               for doi, _ in srcs if doi_key.split("/")[-1] in doi})
                rows.append({"구분": "논문", "AAS": aas, "분류": grp, "장비명": title[:90],
                             "Property후보": ", ".join(mine[:10]),
                             "제작사(모델)": "", "보유기관": "", "등급": "배경자료",
                             "자료": "논문 본문" if not only_abs else "논문 초록",
                             "링크": url, "파일": f, "위치": paper_location(os.path.join(PAPER_DIR, f)),
                             "사유": ("뽑은 슬롯 후보: " + ", ".join(mine)) if mine
                                     else "이 논문에서는 조정 파라미터를 뽑지 못했다"})

        # ── 특허 (제조사 명의) ─────────────────────────────────────
        if os.path.isdir(PATENT_DIR):
            for f in sorted(os.listdir(PATENT_DIR)):
                if not f.endswith(".json"):
                    continue
                maker = f.split("__")[0]
                if aas not in MAKER_AAS.get(maker, []):
                    continue
                d = json.load(open(os.path.join(PATENT_DIR, f), encoding="utf-8"))
                rows.append({"구분": "특허", "AAS": aas, "분류": grp,
                             "장비명": (d.get("제목") or d.get("KIPRIS_제목", ""))[:90],
                             "제작사(모델)": maker, "보유기관": "", "등급": "배경자료",
                             "자료": "특허 명세서·청구항", "링크": d.get("url", ""), "파일": f,
                             "유사성직접": patent_verdict(
                                 d.get("명세서", "") + d.get("청구항", ""), maker)[0],
                             # 위치는 '본문 어디' 가 아니라 Ctrl+F 로 칠 검색어를 준다
                             "위치": search_keywords(
                                 patquote.get(d.get("공개번호", ""), [])) or "본문 전체 (Property 미추출)",
                             "Property후보": ", ".join(
                                 sorted(set(patslots.get(d.get("공개번호", ""), [])))[:10]),
                             "사유": ""})

        # ── 이미 확보해 둔 제조사 특허 (포털수집/<설비>/특허/) ─────────
        for x in folders:
            pdir = os.path.join(R, x, "특허")
            if not os.path.isdir(pdir):
                continue
            for f in sorted(os.listdir(pdir)):
                if not f.endswith(".txt") or f.startswith("목록"):
                    continue
                head = open(os.path.join(pdir, f), encoding="utf-8").read().split(chr(10))[:5]
                title = head[0] if head else f
                meta = head[1] if len(head) > 1 else ""
                url = next((l for l in head if l.startswith("http")), "")
                maker = meta.split("|")[1].strip() if "|" in meta else ""
                body = open(os.path.join(pdir, f), encoding="utf-8").read()
                has_claim = "청구항" in body
                p_sim, p_loc = patent_verdict(body, maker or "제조사")
                mpub = re.search(r"_(KR[0-9A-Z]+)_", f)          # 01_KR101592435B1_제목.txt
                pub = mpub.group(1) if mpub else ""
                rows.append({"구분": "특허", "AAS": aas, "분류": grp, "장비명": title[:90],
                             "제작사(모델)": maker, "보유기관": "", "등급": "배경자료",
                             "자료": "특허 명세서" + ("·청구항" if has_claim else ""),
                             "링크": url, "파일": f,
                             "위치": search_keywords(patquote.get(pub, []))
                                     or "본문 전체 (Property 미추출)",
                             "사유": p_sim, "유사성직접": p_sim,
                             "Property후보": ", ".join(sorted(set(patslots.get(pub, [])))[:12])})

        # ── 폴더별 자료목록 엑셀 ──────────────────────────────────
        wb = Workbook()
        cols = ["구분", "장비명", "제작사(모델)", "보유기관", "등급", "자료", "링크",
                "KETI FAB 설비와의 유사성", "확인할 위치", "Property 후보"]

        # Property 후보가 있는 자료와 없는 자료를 시트로 가른다.
        # 없는 쪽도 조사 흔적이라 버리지 않고 '기타' 로 남긴다.
        main_rows, other_rows = [], []
        for r in rows:
            if r.get("등급") == "제외":
                continue
            if not r.get("Property후보"):
                got = keti_zeus_slots.slots_for(zeus_tbl, aas, r["구분"],
                                                r["장비명"], r.get("자료", ""))
                if got:
                    r["Property후보"] = ", ".join(got[:14])
            sim, loc = similarity_of(aas, r["장비명"])
            if r.get("위치"):
                loc = r["위치"]
            if not sim and r["구분"] == "논문":
                sim = paper_similarity(os.path.join(PAPER_DIR, r["파일"]), aas)
            if r.get("유사성직접"):
                sim = r["유사성직접"]
            r["유사성"], r["파라미터위치"] = sim, loc
            # Fab장비·유사장비·KOSMO 는 설비 자체의 출처라 Property 가 없어도 목록에 남긴다.
            # 기타로 빼는 것은 Property 를 주지 못한 배경자료(논문·특허)다.
            keep = (r["구분"] not in ("논문", "특허")) or (r.get("Property후보") or "").strip()
            (main_rows if keep else other_rows).append(r)

        def fill(ws, src):
            ws.append(cols)
            for c in range(1, len(cols) + 1):
                ws.cell(1, c).fill = HEAD
                ws.cell(1, c).font = Font(bold=True)
            for r in src:
                is_doc = r["구분"] in ("논문", "특허")
                ws.append([r["구분"], r["장비명"], r["제작사(모델)"], r["보유기관"],
                           grade_of(r["구분"], r.get("등급", ""), r["장비명"]),
                           "" if is_doc else r["자료"], "",
                           bullets(r.get("유사성") or ""),
                           bullets_loc(r.get("파라미터위치") or ""),
                           bullets_list(r.get("Property후보") or "")])
                i = ws.max_row
                if r["링크"]:
                    c = ws.cell(i, 7)
                    c.value = r["링크"].split("//")[-1].split("/")[0] + " 열기"
                    c.hyperlink = r["링크"]
                    c.font = LINK
            for i, w in enumerate([10, 40, 22, 20, 10, 20, 20, 62, 40, 34], 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            for row in ws.iter_rows(min_row=2):
                for c in row:
                    c.alignment = Alignment(vertical="top", wrap_text=True)
            ws.freeze_panes = "A2"
            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions

        ws = wb.active
        ws.title = "자료목록"
        fill(ws, main_rows)
        fill(wb.create_sheet("기타"), other_rows)

        wb.save(os.path.join(dst, "자료목록.xlsx"))

        index += rows
        n_main = sum(1 for r in rows if r["구분"] == "본설비")
        print(f"{grp:5} {aas:30} 본설비 {n_main:>2}행 / 유사장비 {len(rows)-n_main:>2}행")

    # ── 전체 요약 엑셀: 어느 자료에서 파라미터를 얻을 수 있는지 한눈에 ──
    wb = Workbook()
    ws = wb.active
    ws.title = "자료 총괄"
    cols = ["분류", "AAS", "구분", "장비·문서명", "제작사/출처", "등급", "자료", "링크",
            "KETI FAB 설비와의 유사성", "확인할 위치", "Property 후보", "파라미터 확보"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(1, c).fill = HEAD
        ws.cell(1, c).font = Font(bold=True)

    def yields_param(r):
        """이 자료에서 파라미터(슬롯)를 실제로 얻을 수 있는가."""
        sim = r.get("유사성", "") or ""
        if r["구분"] in ("Fab장비", "유사장비", "KOSMO 참고"):
            return "없음" if "자료 없음" in (r.get("파라미터위치") or "") else "확보"
        if r["구분"] == "특허":
            return "확보" if "Property 후보 확보" in sim else (
                "이름만" if "부품 이름" in sim else "없음")
        if r["구분"] == "논문":
            return "확보" if "실제로 돌린 논문" in sim else "참고"
        return ""

    order = {"증착": 0, "포토": 1, "프린터": 2, "참고": 3}
    AAS_ORDER = ['박막증착장비-PECVD', '박막증착장비-Sputter', '박막증착장비-DryEtcher', '박막증착장비-ThermalEvaporator', '유기증착기-PlasmaChamber', '유기증착기-OrganicChamber', '유기증착기-MetalChamber', 'PEALD', '현상장비', '마스크 얼라이너', '식각/스트립', '프린팅', 'CBD']
    for r in sorted(index, key=lambda x: (AAS_ORDER.index(x["AAS"])
                                     if x["AAS"] in AAS_ORDER else 99, x["구분"])):
        if r.get("등급") == "제외":
            continue
        is_doc = r["구분"] in ("논문", "특허")
        ws.append([r["분류"], r["AAS"], r["구분"], r["장비명"], r["제작사(모델)"] or r["보유기관"],
                   grade_of(r["구분"], r.get("등급", ""), r["장비명"]),
                   "" if is_doc else r["자료"], "", bullets(r.get("유사성") or ""),
                   bullets_loc(r.get("파라미터위치") or ""),
                   bullets_list(r.get("Property후보") or ""), yields_param(r)])
        i = ws.max_row
        if r.get("링크"):
            c = ws.cell(i, 8)
            c.value = r["링크"].split("//")[-1].split("/")[0] + " 열기"
            c.hyperlink = r["링크"]
            c.font = LINK
    for i, w in enumerate([8, 24, 10, 40, 20, 10, 18, 20, 56, 40, 34, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    # ── 요약 시트: AAS 별로 자료가 어디서 얼마나 왔는지 ──────────
    ws2 = wb.create_sheet("AAS 요약", 0)
    ws2.append(["분류", "AAS", "본설비", "유사장비", "KOSMO", "논문", "특허",
                "자료 계", "뽑은 슬롯 후보", "가장 두꺼운 근거"])
    for c in range(1, 11):
        ws2.cell(1, c).fill = HEAD
        ws2.cell(1, c).font = Font(bold=True)

    # 슬롯 후보 집계 — 논문 TSV 는 '설비'(24_유기증착기), 특허 TSV 는 'AAS' 를 키로 쓴다.
    # 설비 → AAS 는 B.AAS 의 폴더 목록으로 되돌린다.
    eq2aas = {}
    for _g, _a, _folders, _k in B.AAS:
        for _f in _folders:
            eq2aas.setdefault(os.path.basename(_f), []).append(_a)

    slot_n = {}

    def _bump(a, n=1):
        slot_n[a] = slot_n.get(a, 0) + n

    if os.path.exists(PAPER_SLOT):
        txt = open(PAPER_SLOT, encoding="utf-8-sig").read().replace(chr(13) + chr(10), chr(10))
        for r in csv.DictReader(txt.split(chr(10)), delimiter="	"):
            eq = (r.get("설비") or "").strip()
            for cand, aases in eq2aas.items():
                if eq.startswith(cand):
                    for a in aases:
                        _bump(a)
                    break

    if os.path.exists(PATENT_SLOT):
        txt = open(PATENT_SLOT, encoding="utf-8-sig").read().replace(chr(13) + chr(10), chr(10))
        for r in csv.DictReader(txt.split(chr(10)), delimiter="	"):
            for a in (r.get("AAS") or "").split(" · "):
                a = a.strip()
                if a:
                    _bump(a)

    seen_aas = []
    for r in index:
        if (r["분류"], r["AAS"]) not in seen_aas:
            seen_aas.append((r["분류"], r["AAS"]))
    for grp, aas in sorted(seen_aas, key=lambda x: (AAS_ORDER.index(x[1])
                                               if x[1] in AAS_ORDER else 99)):
        mine = [r for r in index if r["AAS"] == aas]
        cnt = {k: sum(1 for r in mine if r["구분"] == k and yields_param(r) == "확보")
               for k in ("Fab장비", "유사장비", "KOSMO 참고", "논문", "특허")}
        # 슬롯 후보는 설비명(논문)·AAS명(특허) 두 키로 세어 둔 것을 합친다
        n_slot = slot_n.get(aas, 0)
        best = max(cnt.items(), key=lambda x: x[1])
        ws2.append([grp, aas, cnt["Fab장비"], cnt["유사장비"], cnt["KOSMO 참고"],
                    cnt["논문"], cnt["특허"], sum(cnt.values()), n_slot,
                    f"{best[0]} {best[1]}건" if best[1] else "-"])
    for i, w in enumerate([8, 28, 8, 10, 8, 8, 8, 8, 14, 18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "C2"

    wb.save(os.path.join(OUT, "_자료_총괄.xlsx"))

    with open(os.path.join(OUT, "_자료목록_전체.json"), "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)
    with open(os.path.join(OUT, "_캡처작업목록.json"), "w", encoding="utf-8") as f:
        json.dump(cap_todo, f, ensure_ascii=False, indent=2)
    print(f"\n총 {len(index)}행 / 캡처 필요 {len(cap_todo)}건 → {OUT}")


if __name__ == "__main__":
    main()
