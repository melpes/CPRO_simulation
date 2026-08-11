"""FAB_시뮬레이션데이터통합.xlsx 의 공정 코드와 MBOM 을 채운다.

사용자가 재구성한 시트 구조를 그대로 두고 필요한 칸만 갱신한다.
공정 코드는 "PI-1 (SiO₂/Si₃H₄ Dep.)" 처럼 코드와 공정명을 함께 적는다.
"""
import openpyxl, re, shutil, datetime

# ── 하드코딩 ────────────────────────────────────────────────────────────
XLSX = r"C:/Users/KangTaehui/KG/keti/CPRO_조립공정/시뮬레이션/Package/docs/FAB/FAB_시뮬레이션데이터통합.xlsx"
RUNSHEET = r"C:/Users/KangTaehui/KG/keti/keti-fab/[KETI]전북본부 FAB 장비 및 공정 정보/3. TFT backplane runsheet_스마트.xlsx"

HEAD = {"MBOM": 8, "공정 정보": 9}          # 헤더 마지막 행 (데이터는 그 다음 행부터)
COL_MB = {"제품": 2, "재료": 3, "공정": 4}
COL_PF = {"제품": 2, "그룹": 3, "공정명": 4, "코드": 5, "선행": 6}

# 제품별 공정 코드 접두어 — 공정흐름도 1~5번
PREFIX = {
    "PI": "PI",
    "Bottom emission OLED 소재 평가용 TEG-cell": "BE",
    "Top emission OLED 소재 평가용 TEG-cell": "TE",
    "전극 스크린 프린팅": "SP",
    "잉크젯 프린팅": "IJ",
}

# 공정흐름도 2~5번 박스별 투입 재료 (흐름도에 화살표로 붙은 것)
FLOW_MAT = {
    ("Bottom emission OLED 소재 평가용 TEG-cell", "Anode ITO 증착"): ["Indium Tin Oxide", "Ar", "O2"],
    ("Bottom emission OLED 소재 평가용 TEG-cell", "포토레지스트 패터닝 공정"): ["Photoresist", "TMAH 2.38% 수용액"],
    ("Bottom emission OLED 소재 평가용 TEG-cell", "습식 식각"): ["ITO etchant"],
    ("Bottom emission OLED 소재 평가용 TEG-cell", "포토레지스트 제거"): ["Organic strip chemical"],
    ("Bottom emission OLED 소재 평가용 TEG-cell", "유기절연막 패터닝 공정"): ["유기절연막", "TMAH 2.38% 수용액"],
    ("Top emission OLED 소재 평가용 TEG-cell", "cathode 증착"): ["ITO/Ag/ITO 다층박막", "Ar"],
    ("Top emission OLED 소재 평가용 TEG-cell", "포토레지스트 패터닝 공정"): ["Photoresist", "TMAH 2.38% 수용액"],
    ("Top emission OLED 소재 평가용 TEG-cell", "습식 식각"): ["금속 etchant"],
    ("Top emission OLED 소재 평가용 TEG-cell", "포토레지스트 제거"): ["Organic strip chemical"],
    ("Top emission OLED 소재 평가용 TEG-cell", "유기절연막 패터닝 공정"): ["유기절연막", "TMAH 2.38% 수용액"],
    ("전극 스크린 프린팅", "제판 설계 및 제작"): ["제판"],
    ("전극 스크린 프린팅", "Screen Printing"): ["Glass", "Flim", "Paste"],
    ("잉크젯 프린팅", "Ink 및 기판 준비"): ["Ink", "Patterned glass or Flat glass"],
}
GASES = ["Ar", "N₂", "SiH4", "N2O", "NH₃", "O2", "CF4", "Cl2"]


# ── runsheet 에서 공정별 재료 읽기 ──────────────────────────────────────
def read_runsheet():
    ws = openpyxl.load_workbook(RUNSHEET, data_only=True)["PI Run sheet"]
    rows = [r + (None,) * 6 for r in ws.iter_rows(values_only=True)]
    step = None
    recs, cur = [], None
    for r in rows:
        p, v = r[2], r[3]
        if p and v is None:
            step = str(p).strip()
            continue
        if p == "Process parameter":
            cur = {"step": step, "mats": [], "gases": []}
            recs.append(cur)
            continue
        if cur is None or not p:
            continue
        k = str(p).strip()
        if k in ("Material", "Strip chemical"):
            cur["mats"].append(str(v).strip())
        elif "gas" in k.lower():
            for nm in re.findall(r"[A-Za-z][A-Za-z0-9₂₃₄]*", k.split("(")[0]):
                if nm in GASES and nm not in cur["gases"]:
                    cur["gases"].append(nm)
    return recs


runsheet_recs = read_runsheet()

# ── 백업 후 열기 ────────────────────────────────────────────────────────
stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
shutil.copyfile(XLSX, XLSX.replace(".xlsx", f".bak-{stamp}.xlsx"))
wb = openpyxl.load_workbook(XLSX)

# ── 공정 정보 : 코드에 공정명을 붙이고 2~5번에도 코드 부여 ──────────────
ws = wb["공정 정보"]
hr = HEAD["공정 정보"]
seq = {}                                   # 제품별 일련번호
labels = []                                # (행, 제품, 표기)
for r in range(hr + 1, ws.max_row + 1):
    prod = ws.cell(r, COL_PF["제품"]).value
    name = ws.cell(r, COL_PF["공정명"]).value
    if not prod or not name:
        continue
    pre = PREFIX.get(str(prod).strip(), str(prod).strip()[:2].upper())
    seq[pre] = seq.get(pre, 0) + 1
    label = f"{pre}-{seq[pre]} ({name})"
    ws.cell(r, COL_PF["코드"]).value = label
    labels.append((r, str(prod).strip(), str(name).strip(), label))

# 선행 공정 = 같은 제품의 직전 공정
prev_by_prod = {}
for r, prod, name, label in labels:
    ws.cell(r, COL_PF["선행"]).value = prev_by_prod.get(prod)
    prev_by_prod[prod] = label

label_of = {(prod, name): label for _, prod, name, label in labels}

# ── MBOM : 공정 × 투입 재료로 전개 ──────────────────────────────────────
ws = wb["MBOM"]
hr = HEAD["MBOM"]
if ws.max_row > hr:
    ws.delete_rows(hr + 1, ws.max_row - hr)

rows_out = []
for i, c in enumerate(runsheet_recs, 1):                    # ① TFT backplane
    label = label_of.get(("PI", c["step"]))
    for nm in c["mats"] + c["gases"]:
        rows_out.append(("PI", nm, label))
for (prod, step), mats in FLOW_MAT.items():                 # ②~⑤ 흐름도
    label = label_of.get((prod, step))
    for nm in mats:
        rows_out.append((prod, nm, label))

for prod, nm, label in rows_out:
    ws.append([None, prod, nm, label])

wb.save(XLSX)

from collections import Counter
print(f"공정 정보 : 코드 {len(labels)}건 부여")
print("  제품별:", dict(Counter(p for _, p, _, _ in labels)))
print(f"MBOM     : {len(rows_out)}행")
print("  제품별:", dict(Counter(p for p, _, _ in rows_out)))
print("  미매칭 :", sum(1 for _, _, l in rows_out if not l), "건")
