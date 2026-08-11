"""KETI_FAB_AAS_파라미터슬롯.xlsx 에서 ZEUS·i-Tube·매뉴얼 유래 슬롯 이름을 꺼낸다.

특허·논문은 명세서를 읽어 슬롯을 뽑았지만 ZEUS 사양은 그 작업을 따로 하지 않았다.
대신 그 결과가 이미 파라미터슬롯 엑셀에 들어 있으므로 여기서 되가져온다.

출처 문자열이 곧 키다.
  '유사장비: 05_표면처리기(전남대…)'  → 유사장비 행
  'ZEUS 등록장비 상세'                 → 본설비의 등록장비 화면
  '43_스핀 트랙 시스템 · 설비 매뉴얼 PDF (p80)' → 본설비 매뉴얼
"""

import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import bundle_keti_aas_rawtext as B

BASE = r"C:\Users\KangTaehui\KG\keti\CPRO_조립공정\시뮬레이션\Package"
XLSX = os.path.join(BASE, r"docs\원본자료\keti-fab\KETI_FAB_AAS_파라미터슬롯.xlsx")


def _norm(s):
    s = re.sub(r"^\d{2}_", "", (s or "").strip())
    return re.sub(r"[^0-9a-z가-힣]", "", s.lower())


def load():
    """(AAS, 출처원문) → 슬롯 이름 목록. 병합 셀 때문에 빈 출처는 위 값을 잇는다."""
    if not os.path.exists(XLSX):
        return {}
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    out = {}
    for name in wb.sheetnames:
        if name == "요약":
            continue
        aas = B.from_safe(name)      # 시트명은 / 를 못 써서 _ 로 저장돼 있다
        cur = ""
        for r in wb[name].iter_rows(min_row=2, values_only=True):
            src = (r[6] or "").strip()
            if src:
                cur = src
            slot = (r[3] or "").strip()
            if not slot or not cur:
                continue
            out.setdefault((aas, cur), [])
            if slot not in out[(aas, cur)]:
                out[(aas, cur)].append(slot)
    wb.close()
    return out


def slots_for(table, aas, 구분, 장비명, 자료):
    """자료목록 한 행에 붙일 슬롯 이름을 찾는다.

    출처 표기가 여러 형태다.
      '43_스핀 트랙 시스템 · 본설비 / … · 본설비(예약페이지)'
      '43_스핀 트랙 시스템 · 설비 매뉴얼 PDF (p80)'
      'ZEUS 등록장비 상세 / ZEUS 장비예약 상세'
      '유사장비: 포토 트랙'
    그래서 설비명이나 화면 종류가 출처 어딘가에 들어 있으면 잡는다.
    """
    hit = []
    eq = _norm(장비명)[:10]

    if 구분 == "유사장비":
        for (a, src), names in table.items():
            if a != aas or "유사장비" not in src:
                continue
            if eq and eq in _norm(src):
                hit += names
        return hit

    if 구분 == "Fab장비":
        is_manual = "매뉴얼" in (자료 or "")
        is_resv = "예약" in (자료 or "")
        is_itube = "i-Tube" in (자료 or "")
        for (a, src), names in table.items():
            if a != aas or src.startswith("유사장비") or src.startswith("KOSMO"):
                continue
            n = _norm(src)
            same_eq = (eq and eq in n)
            if is_manual:
                if same_eq and "매뉴얼" in src:
                    hit += names
            elif is_resv:
                if ("예약" in src) and (same_eq or "zeus" in n):
                    hit += names
            elif is_itube:
                continue          # i-Tube 는 ZEUS 와 같은 사양이라 중복을 피한다
            else:                 # ZEUS 등록장비 상세
                if "매뉴얼" in src:
                    continue
                if same_eq or "등록장비" in src:
                    hit += names
        return hit

    if 구분 == "KOSMO 참고":
        for (a, src), names in table.items():
            if a == aas and src.startswith("KOSMO"):
                hit += names
        return hit
    return []
