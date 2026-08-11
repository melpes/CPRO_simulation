"""FAB 요청 자료 엑셀 생성 — 작업자·BOM자재·설비공정 3시트, 미확보분 노란색."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 하드코딩 ────────────────────────────────────────────────────────────
OUT = r"C:/Users/KangTaehui/KG/keti/CPRO_조립공정/시뮬레이션/Package/docs/FAB/FAB_요청자료.xlsx"

HEAD_FILL = PatternFill("solid", fgColor="D9E2F3")
NEED_FILL = PatternFill("solid", fgColor="FFF2CC")      # 요청 필요 = 노란색
HAVE_FILL = PatternFill("solid", fgColor="E2EFDA")      # 확보
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = ["구분", "AAS 항목", "저장 위치 (AAS 경로)", "현재 상태", "FAB 대응 자료", "요청 내용"]
WIDTH = [16, 24, 44, 10, 34, 52]

# 상태: O 확보 / X 요청필요
SHEETS = {
    "① 작업자": [
        ("인원", "worker_count", "AssemblyByWorker > WorkstationInformation > 워크스테이션 > WorkstationConfigurationRecords",
         "X", "없음", "공정별 또는 설비별 작업 인원 수. 레코드 1건 = 1명 구조"),
        ("담당", "AssignedProcessGroups", "AssemblyByWorker > WorkstationInformation > 워크스테이션 > AssignedProcessGroups",
         "X", "없음", "각 작업자가 어느 공정(설비)을 담당하는지. 설비별 전담인지 공정별 순회인지"),
        ("동시 담당", "UnitsPerWorker", "AssemblyByWorker > WorkstationInformation > 워크스테이션 > UnitsPerWorker",
         "X", "없음", "한 사람이 동시에 보는 설비 대수"),
        ("숙련도", "SkillLevel", "AssemblyByWorker > … > WorkstationConfigurationRecords > 레코드 > SkillLevel",
         "X", "없음", "작업자 숙련 구분이 있는지 (없으면 단일 등급으로 처리)"),
        ("근무 시작", "WorkStartTime", "AssemblyByWorker > WorkstationWorkerMatchingData > WorkStartTime",
         "X", "로그상 08~09시 시작 추정", "공식 근무 시작 시각"),
        ("근무 종료", "WorkEndTime", "AssemblyByWorker > WorkstationWorkerMatchingData > WorkEndTime",
         "X", "로그상 16~17시 종료 추정", "공식 근무 종료 시각"),
        ("휴게", "BreakDurationMin (min/max)", "AssemblyByWorker > WorkstationWorkerMatchingData > BreakDurationMin",
         "X", "로그상 12시대 공백", "점심 등 휴게 시작·종료 시각"),
        ("운반", "(신규)", "—",
         "X", "없음", "설비 간 기판을 누가 옮기는지, 1회 몇 매씩 옮기는지 (카세트 25매 단위인지)"),
        ("무인 운전", "(신규)", "—",
         "X", "로그상 야간·주말 실행 0건", "야간·주말 무인 운전이 가능한지, 안 하는 것인지 못 하는 것인지"),
        ("작업 지시", "(신규)", "—",
         "X", "스케줄러 로그에 수동 조작 1,450건", "다음 작업을 누가 어떤 기준으로 정하는지 (작업지시서·일정표 유무)"),
    ],
    "② BOM·자재": [
        ("제품 정의", "제품 Entity", "MODEL_N > HierarchicalStructures > Entity",
         "X", "TFT backplane (PI기판 / 유리기판)", "시뮬 대상 제품을 무엇으로 잡을지. 기판 종류별인지, 의뢰 건별인지"),
        ("BOM 구조", "부품 Entity + Quantity", "MODEL_N > HierarchicalStructures > Entity 제품 > Entity 부품",
         "X", "기판 1장 투입 → 1장 산출", "BOM을 뽑을 수 있는 구조인지. 판(기판) 1장 기준으로 투입 자재 목록을 정리할 수 있는지"),
        ("자재 분류", "BOMCategory", "MODEL_N > HierarchicalStructures > BOMCategory",
         "X", "기판·타깃·가스·감광재·약액 5분류 도출", "FAB에서 쓰는 자재 분류 체계가 따로 있는지"),
        ("공정별 투입", "InputBOM", "MODEL_N > ManufacturingProcess > 공정그룹 > 공정코드 > InputBOM",
         "△", "runsheet Material 열 (품목만)", "공정별 투입 자재의 소요량. 가스는 로그로 산출 가능하므로 제외"),
        ("타깃 소모", "InputBOM 수량", "동일",
         "X", "기판 370×470 확보로 부착량 계산 가능", "타깃 교체 주기 (몇 매 처리 후 또는 몇 시간 사용 후). Mo·ITO·IGZO·Ag"),
        ("감광재 소모", "InputBOM 수량", "동일",
         "X", "도포 회전수·시간만 있음", "PR 1회 도포 토출량 (ZPP1700PG-30)"),
        ("약액 소모", "InputBOM 수량", "동일",
         "X", "온도·침지 시간만 있음", "약액 배스 용량과 교체 전 처리 매수 (TMAH·ITO etchant·Organic strip·NMP)"),
        ("재고 하한", "MinStock", "MODEL_N > HierarchicalStructures > BOMCategory > 카테고리 > MinStock",
         "X", "없음", "자재별 발주점 (재고가 얼마 이하면 주문하는지)"),
        ("재고 상한", "MaxStock", "동일",
         "X", "없음", "자재별 최대 보유량"),
        ("발주 비율", "OrderRatio", "동일",
         "X", "없음", "1회 발주량 기준"),
        ("조달 리드타임", "ReplenishLeadDay", "PSM > DefaultParameters > ReplenishLeadDay",
         "X", "없음", "자재 주문 후 입고까지 걸리는 일수. 품목별로 다르면 품목별로"),
        ("용기 단위", "ContainerCapacity", "PSM > DefaultParameters > (SolderCreamParam 대응)",
         "X", "카세트 25매만 확인", "가스 봄베 용량·약액 통 용량·타깃 1장 규격"),
        ("자재 수명", "ShelfLife", "동일",
         "X", "없음", "개봉·교체 후 사용 가능 기간 (약액 배스, 타깃)"),
    ],
    "③ 설비·공정": [
        ("공정 목록", "공정그룹 > 공정코드", "MODEL_N > ManufacturingProcess",
         "O", "runsheet 31공정 (PI) / 30공정 (유리)", "—"),
        ("선후 관계", "DepPrev", "MODEL_N > ManufacturingProcess > 공정그룹 > 공정코드 > DepPrev",
         "O", "runsheet 공정 순서", "—"),
        ("의존 유형", "DepType", "동일",
         "△", "전부 SEQUENCE로 추정", "동시 진행하거나 합류하는 공정이 있는지 (JOIN/FORK)"),
        ("설비 매핑", "AssignedProcessGroups", "PSM > KnowledgeGraph > Action > AssignedProcessGroups",
         "O", "runsheet 사용장비 열", "—"),
        ("처리 시간", "CycleTimeSec", "MODEL_N > ManufacturingProcess > 공정그룹 > 공정코드 > CycleTimeSec",
         "△", "박막증착장비 chA·chC 실측 / 나머지는 runsheet 공정시간만",
         "설비별 1회 처리 시간. 기판을 넣고 빼는 시간까지 포함한 설비 점유 시간"),
        ("처리 단위", "(신규)", "—",
         "X", "일부 확인 (카세트 25매, 700장/월)", "설비별 처리 방식 — 낱장인지, 배치(몇 매)인지, 컨베이어 연속인지"),
        ("설비 대수", "(신규)", "—",
         "X", "없음", "같은 공정을 처리하는 설비가 몇 대인지"),
        ("공정 간 대기", "(신규)", "—",
         "X", "없음 (장비 밖은 로그가 안 봄)", "앞 공정이 끝나고 다음 공정을 시작할 때까지 보통 얼마나 걸리는지"),
        ("공정 간 이동", "(신규)", "—",
         "X", "박막증착장비 내부만 확보 (13~19초)", "설비에서 설비로 기판을 옮기는 데 걸리는 시간"),
        ("전체 리드타임", "(신규)", "—",
         "X", "runsheet 일정란 공백", "기판 1장이 전 공정을 마치는 데 며칠 걸리는지. 하루에 보통 몇 공정 진행하는지"),
        ("불량률", "DefectRate", "MODEL_N > ManufacturingProcess > 공정그룹 > 공정코드 > DefectRate",
         "X", "runsheet에 rework 기입 1건", "불량률과 그 판정 단위 (공정별인지 층별인지 기판 전체인지)"),
        ("재작업", "(신규)", "—",
         "X", "없음", "재작업이 가능한 공정 범위와 발생 빈도"),
        ("설비 전력", "RatedPowerKw", "MODEL_N > ManufacturingProcess > 공정그룹 > 공정코드 > RatedPowerKw",
         "X", "44종 중 유틸리티 사양 1건뿐", "설비별 정격전력 또는 분전반 계측값"),
        ("공장 기저전력", "DefaultProcessConsumedPowerKw", "PSM > DefaultParameters > DefaultProcessConsumedPowerKw",
         "X", "없음", "공정과 무관하게 상시 소비되는 전력 (공조·펌프·칠러 등)"),
        ("후처리 대기", "CuringTimeSec", "MODEL_N > ManufacturingProcess > 공정그룹 > 공정코드 > CuringTimeSec",
         "O", "Annealing UV 2h + 380℃ 3h", "—"),
        ("주문 수량", "PurchaseOrder", "PSM > PurchaseOrder > 모델_N",
         "X", "제작일정 수량 7→6→3매", "제품·주문·납기 개념이 성립하는지. 성립한다면 1회 의뢰 단위와 수량"),
        ("주문일", "RegisteredDay", "PSM > PurchaseOrder > 모델_N > Qualifier RegisteredDay",
         "X", "runsheet 시작예정일 열 공백", "의뢰를 받은 날짜 기준"),
        ("납기일", "DueDay", "PSM > PurchaseOrder > 모델_N > Qualifier DueDay",
         "X", "runsheet 종료예정일 열 공백", "언제까지 납품해야 하는지. 생산계획이 문서로 있는지"),
        ("설비 정지", "(신규)", "—",
         "X", "없음", "타깃·약액 교체나 예방정비로 설비가 서는 시간과 주기"),
    ],
}


def make_sheet(wb, name, rows, first):
    ws = wb.active if first else wb.create_sheet()
    ws.title = name
    ws.append(COLS)
    for i, c in enumerate(ws[1], 1):
        c.font = Font(bold=True)
        c.fill = HEAD_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[c.column_letter].width = WIDTH[i - 1]
    ws.freeze_panes = "A2"

    for gubun, ids, path, st, cur, req in rows:
        ws.append([gubun, ids, path, {"O": "확보", "△": "일부", "X": "요청"}[st], cur, req])
        r = ws.max_row
        for c in ws[r]:
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
        if st == "O":
            ws.cell(r, 4).fill = HAVE_FILL
        else:
            for col in (2, 4, 6):                 # 항목·상태·요청내용에 노란색
                ws.cell(r, col).fill = NEED_FILL
    return ws


wb = openpyxl.Workbook()
for i, (name, rows) in enumerate(SHEETS.items()):
    make_sheet(wb, name, rows, i == 0)
wb.save(OUT)

for name, rows in SHEETS.items():
    o = sum(1 for r in rows if r[3] == "O")
    t = sum(1 for r in rows if r[3] == "△")
    x = sum(1 for r in rows if r[3] == "X")
    print(f"{name}: 총 {len(rows)}  확보 {o} · 일부 {t} · 요청 {x}")
print("→", OUT)
