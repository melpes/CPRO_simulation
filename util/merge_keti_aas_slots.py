"""에이전트가 뽑은 AAS별 파라미터 TSV 를 하나의 엑셀로 합친다.

파싱으로는 계층 구조를 못 살려서, 원문을 에이전트에게 넘겨 정리한 결과를 쓴다.
자체 자료(ZEUS·i-Tube·매뉴얼·유사장비)와 KOSMO 참고 슬롯을 열로 구분한다.
"""

import csv
import glob
import io
import bundle_keti_aas_rawtext as B
import os
import re
import sys
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = r"C:\Users\KangTaehui\KG\keti\CPRO_조립공정\시뮬레이션\Package"
SRC = os.path.join(BASE, r"docs\원본자료\keti-fab\AAS원문묶음")
OUT = os.path.join(BASE, r"docs\원본자료\keti-fab\KETI_FAB_AAS_파라미터슬롯.xlsx")
KOSMO_ZIP = os.path.join(BASE, r"docs\원본자료\keti-fab\kasmo_후보데이터.zip")

KOSMO_NOTE = "사용자 제공 (kasmo_후보데이터.zip) — 공개 URL 없음"

# 도메인 검토에서 제외하기로 한 유사장비 (AAS, 출처에 들어 있는 식별 문자열)
# 03_마그네트론 스퍼터: 값이 전부 수량(1set·3set)이라 공정 파라미터가 없다.
#   같은 슬롯을 01·02 가 값과 함께 갖고 있어 더할 것이 없다.
DROP_SOURCE = {
    "박막증착장비-Sputter": ["03_마그네트론 스퍼터"],
    # 플라즈마 애싱은 진공 플라즈마 PR 제거라 박막증착장비의 Dry Etcher 챔버와 계통이 겹친다.
    # 식각·스트립은 습식(비진공) 라인이므로 여기서는 뺀다.
    "식각/스트립": ["플라즈마 애싱"],
}

GROUP = {
    "박막증착장비-PECVD": "증착", "박막증착장비-Sputter": "증착",
    "박막증착장비-DryEtcher": "증착", "박막증착장비-ThermalEvaporator": "증착",
    "유기증착기-PlasmaChamber": "증착", "유기증착기-OrganicChamber": "증착",
    "유기증착기-MetalChamber": "증착", "PEALD": "증착",
    "현상장비": "포토", "마스크 얼라이너": "포토", "식각/스트립": "포토",
    "프린팅": "프린터", "CBD": "참고",
}

# KOSMO AAS 슬롯을 붙일 곳 — 기능이 실제로 같은 AAS 에만
KOSMO_MAP = {
    "유기증착기-PlasmaChamber": [("표면처리기.aasx", "OperationData", "KOSMO 표면처리기 · OperationData")],
    "박막증착장비-DryEtcher": [
        ("[일.6]SemiconductorCircuitEtchingEquipment.aasx", "TechnicalData", "KOSMO 반도체회로 에칭장비 · TechnicalData"),
        ("[일.6]SemiconductorCircuitEtchingEquipment.aasx", "OperationalData", "KOSMO 반도체회로 에칭장비 · OperationalData")],
    "CBD": [
        ("[신규.2]ElectroplatingBath.aasx", "TechnicalData", "KOSMO 전해도금조 · TechnicalData"),
        ("[신규.2]ElectroplatingBath.aasx", "OperationalData", "KOSMO 전해도금조 · OperationalData")],
}


# KOSMO 슬롯 중 그 AAS 에 실제로 맞는 것만 남기는 필터.
# 사용자 도메인 판단 반영:
#   Dry Etcher — 가이드 2-(나) '반도체 에칭 장비 개요' 의 플라즈마 건식식각 관리 데이터만.
#                생산관리·제조사정보 계통은 다른 유사장비로 대체 가능하다.
#   CBD        — 전해도금조와 용액 침지까지는 같으나 코팅 방식이 다르다(전기도금 vs 화학반응).
#                탱크·순환·히터·필터·배기 같은 설비 공통만 쓰고 전기도금 공정제어는 뺀다.
KOSMO_KEEP = {
    "박막증착장비-DryEtcher": (
        re.compile(r"EtchingChamber|ESC|Radiofrequency|RF|GasInjectionShowerhead|MFC|Flowrate"
                   r"|GasMixing|VacuumPumpMotor|GateValve|Scrubber|Temperaturecontroller|Coolant"
                   r"|Pressure|VacuumLevel|Powersupply|Etching|Chamber|Plasma|Selfbias|Leakage"
                   r"|CleaningCycle|Humidity|Temperature|Rectifier|VacuumPump|ModelNumber"
                   r"|MaterialName|BaseUnitSetting|Unit$", re.I),
        re.compile(r"WorkOrder|ProductionPlan|ProductProductionLog|ProductionStatus|EnergyData"
                   r"|Transferrobot|Batch(ID|Data)|Product(ID|Type|Code|ion)|Allocation|Defective"
                   r"|RecipeID|WorkerID|LineID|Rawmaterial|Throughput|ChamberOpenCount|StartDate"
                   r"|EndDate|StartTime|EndTime|Timestamp|RetentionTime|ProductSize|Remaining"
                   r"|GeneralInformation|ProductClassifications|FurtherInformation", re.I)),
    "CBD": (
        re.compile(r"Tank|Bath|Temperature|Flowrate|Pressure|Filter|Pump|Circulat|Heater|Exhaust"
                   r"|Level|Width|Depth|Height|Availablecapacity|MaterialName|Unit$"
                   r"|Differentialpressure|Inlet|Outlet", re.I),
        re.compile(r"Currentdensity|DCVoltage|DCCurrent|DCElectricpower|Rectifier|Anode|Cathode"
                   r"|Electrode|Electricpower|WorkOrder|ProductionPlan|ProductProductionLog"
                   r"|ProductionStatus|EnergyData|Batch(ID)|Product(ID|Type|Code|ion)|Allocation"
                   r"|Defective|RecipeID|WorkerID|LineID|Rawmaterial|Timestamp|StartDate|EndDate"
                   r"|StartTime|EndTime|GeneralInformation|ProductClassifications", re.I)),
}


def kosmo_filter(aas, slots):
    """그 AAS 에 맞는 KOSMO 슬롯만 남긴다. 규칙이 없으면 전부 통과."""
    rule = KOSMO_KEEP.get(aas)
    if not rule:
        return slots
    keep, drop = rule
    return [s for s in slots if keep.search(s) and not drop.search(s)]


def kosmo_slots(member, submodel):
    """KOSMO AASX 에서 슬롯 이름을 계층 경로로 뽑는다."""
    import xml.etree.ElementTree as ET
    NS = "{https://admin-shell.io/aas/3/0}"
    if not os.path.exists(KOSMO_ZIP):
        return []
    z = zipfile.ZipFile(KOSMO_ZIP)
    if member not in z.namelist():
        return []
    inner = zipfile.ZipFile(io.BytesIO(z.read(member)))
    xn = [n for n in inner.namelist() if n.endswith(".aas.xml")][0]
    root = ET.fromstring(inner.read(xn))

    def ids(e):
        c = e.find(NS + "idShort")
        return c.text if c is not None else ""

    def walk(e, prefix, out):
        kids = e.find(NS + "submodelElements")
        if kids is None:
            kids = e.find(NS + "value")
        for se in (kids if kids is not None else []):
            n = ids(se)
            tag = se.tag.split("}")[-1]
            path = f"{prefix}.{n}" if prefix else n
            if tag in ("submodelElementCollection", "submodelElementList"):
                walk(se, path, out)
            elif n:
                out.append(path)
        return out

    for sm in root.iter():
        if sm.tag.split("}")[-1] != "submodel":
            continue
        if ids(sm) != submodel:
            continue
        return walk(sm, "", [])
    return []


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    wb = Workbook()
    HF = PatternFill("solid", fgColor="DDEBF7")
    LINK = Font(color="0563C1", underline="single")
    summary, first = [], True

    def src_key(v):
        """같은 유사장비끼리 묶기 위한 정렬 키. '유사장비: 이름(기관)' 의 이름 부분."""
        v = (v or "").strip()
        m = re.match(r"유사장비:\s*([^(/]+)", v)
        return ("2_" + m.group(1).strip()) if m else ("1_" + v[:40])

    for f in sorted(glob.glob(os.path.join(SRC, "out_*.tsv"))):
        name = B.from_safe(os.path.basename(f).replace(".tsv", "").split("_", 2)[2])
        grp = GROUP.get(name, "-")
        rows = list(csv.DictReader(open(f, encoding="utf-8"), delimiter="\t"))

        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = B.safe_aas(name)[:31]
        ws.append(["분류", "AAS", "자료구분", "파라미터 슬롯", "확보된 값", "단위", "출처", "근거 링크"])
        for c in range(1, 9):
            ws.cell(1, c).fill = HF
            ws.cell(1, c).font = Font(bold=True)

        # 같은 유사장비·같은 링크가 이어지도록 출처 기준으로 정렬한다 (병합 전제)
        rows.sort(key=lambda r: src_key(r.get("출처")))

        drops = DROP_SOURCE.get(name, [])
        n_own = 0
        for r in rows:
            nm = (r.get("파라미터명") or "").strip()
            if not nm:
                continue
            if any(d in (r.get("출처") or "") for d in drops):
                continue
            # URL 이 ' / ' 로 이어져 있으나 http 앞뒤가 붙어 오는 경우가 있어 정규식으로 뽑는다
            urls = re.findall(r"https?://[^\s/]+[^\s]*", r.get("근거URL") or "")
            ws.append([grp, name, "자체 자료", nm,
                       (r.get("값") or "")[:250], (r.get("단위") or "")[:20],
                       (r.get("출처") or "")[:200], ""])
            cell = ws.cell(ws.max_row, 8)
            if urls:
                # 표시 텍스트에 장비 식별자를 넣는다. 전부 'zeus.go.kr 열기' 로 같으면
                # 링크가 다른데도 같은 칸처럼 보여 구분이 안 된다.
                tail = urls[0].rstrip("/").split("/")[-1]
                cell.value = f"ZEUS {tail}"
                cell.hyperlink = urls[0]
                cell.font = LINK
            n_own += 1

        n_ref = 0
        for member, sub, label in KOSMO_MAP.get(name, []):
            for slot in kosmo_filter(name, kosmo_slots(member, sub)):
                ws.append([grp, name, "KOSMO 참고", slot, "", "", label, KOSMO_NOTE])
                n_ref += 1

        # 같은 유사장비(출처)·같은 링크끼리 세로 병합.
        # 링크 열은 표시 텍스트가 모두 'www.zeus.go.kr 열기' 로 같으므로
        # 실제 하이퍼링크 주소를 기준으로 삼는다. 안 그러면 다른 장비가 한 칸으로 뭉친다.
        def key_of(r, col):
            c = ws.cell(r, col)
            if col == 8:
                return c.hyperlink.target if c.hyperlink else ("_" + str(c.value))
            return c.value

        for col in (1, 2, 3, 7, 8):
            start, prev = 2, None
            for r in range(2, ws.max_row + 2):
                cur = key_of(r, col) if r <= ws.max_row else object()
                if r == 2:
                    prev, start = cur, 2
                    continue
                if cur != prev:
                    if r - 1 > start:
                        ws.merge_cells(start_row=start, start_column=col,
                                       end_row=r - 1, end_column=col)
                    start, prev = r, cur
        for i, w in enumerate([8, 26, 12, 40, 46, 10, 40, 26], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        summary.append((grp, name, n_own, n_ref, n_own + n_ref, ws.title))
        print(f"{grp:5} {name:30} 자체 {n_own:>4} + KOSMO {n_ref:>4} = {n_own + n_ref:>4}")

    ws = wb.create_sheet("요약", 0)
    ws.append(["분류", "AAS", "자체 자료", "KOSMO 참고", "합계", "시트"])
    for c in range(1, 7):
        ws.cell(1, c).fill = HF
        ws.cell(1, c).font = Font(bold=True)
    order = {"증착": 0, "포토": 1, "프린터": 2, "참고": 3}
    for g, n, a, b, t, sh in sorted(summary, key=lambda x: (order.get(x[0], 9), -x[4])):
        ws.append([g, n, a, b, t, sh])
    ws.append([])
    ws.append(["합계", "", sum(s[2] for s in summary), sum(s[3] for s in summary),
               sum(s[4] for s in summary), ""])
    for i, w in enumerate([10, 32, 12, 12, 10, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(OUT)
    print(f"\n자체 {sum(s[2] for s in summary)} + KOSMO {sum(s[3] for s in summary)} "
          f"= {sum(s[4] for s in summary)}행 → {OUT}")


if __name__ == "__main__":
    main()
