"""유사장비 후보를 분류/AAS 별 폴더로 재구성하고 폴더마다 링크 리스트 엑셀을 만든다.

기존 폴더(포털수집/<분류>/<설비>/유사장비/)는 '설비' 기준이라 AAS 와 어긋난다.
예를 들어 26_박막증착장비 하나에 Sputter·PECVD·DryEtcher·ThermalEvaporator 용이 섞여 있다.
여기서는 AAS 기준으로 다시 묶어, 그 AAS 를 만들 때 볼 자료만 한 폴더에 모은다.
"""

import glob
import json
import os
import re
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import bundle_keti_aas_rawtext as B
from keti_similar_verdict import verdict_of, OK

BASE = r"C:\Users\KangTaehui\KG\keti\CPRO_조립공정\시뮬레이션\Package"
OUT = os.path.join(BASE, r"docs\원본자료\keti-fab\유사장비후보")
ZEUS_RESV = "https://www.zeus.go.kr/resv/equip/read/"
ZEUS_READ = "https://www.zeus.go.kr/search/equip/read/"


def meta_of(text):
    """저장된 ZEUS 페이지 텍스트에서 제작사·기관·활용번호·특성 줄수를 읽는다."""
    def one(pat):
        m = re.search(pat + r"\s*\n\s*([^\n]+)", text)
        return m.group(1).strip() if m else ""
    i, j = text.find("특성"), text.find("용도설명")
    lines = [l.strip() for l in text[i + 2:j].split("\n") if l.strip()] if 0 < i < j else []
    return {
        "제작사(모델)": one(r"제작사명\(모델명\)") or one(r"제작사명 \| 모델명"),
        "보유기관": one(r"보유기관명"),
        "활용번호": one(r"시설장비활용번호"),
        "특성줄수": len(lines),
        "특성": "\n".join(lines),
    }


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    # 폴더를 통째로 지우면 탐색기·엑셀이 열어둔 경우 실패한다.
    # 캡처 이미지도 살려야 하므로 지우지 않고 덮어쓴다.
    os.makedirs(OUT, exist_ok=True)
    index = []

    for grp, name, folders, keys in B.AAS:
        rows = []
        dst = os.path.join(OUT, grp, name)
        for fr in folders:
            src_dir = os.path.join(B.R, fr, "유사장비")
            for f in sorted(glob.glob(os.path.join(src_dir, "*.txt"))):
                b = os.path.basename(f).replace(".txt", "")
                if keys != ["*"] and not any(k in b for k in keys):
                    continue
                t = open(f, encoding="utf-8").read()
                m = meta_of(t)
                v, why = verdict_of(name, b)
                os.makedirs(dst, exist_ok=True)
                # 원문 txt·html 을 AAS 폴더로 복사 (원본은 설비 폴더에 그대로 둔다)
                for ext in (".txt", ".html"):
                    s = f.replace(".txt", ext)
                    if os.path.exists(s):
                        shutil.copy2(s, os.path.join(dst, os.path.basename(s)))
                url = (ZEUS_RESV + m["활용번호"]) if m["활용번호"] else ""
                rows.append({
                    "분류": grp, "AAS": name, "장비명": b,
                    "제작사(모델)": m["제작사(모델)"], "보유기관": m["보유기관"],
                    "판정": v, "판정 사유": why, "특성 줄수": m["특성줄수"],
                    "ZEUS 링크": url, "파일": os.path.basename(f),
                })
        if not rows:
            continue

        # 폴더별 링크 리스트 엑셀
        wb = Workbook()
        ws = wb.active
        ws.title = "유사장비 목록"
        cols = ["분류", "AAS", "장비명", "제작사(모델)", "보유기관", "판정", "판정 사유",
                "특성 줄수", "ZEUS 링크", "캡처 이미지", "파일"]
        ws.append(cols)
        HF = PatternFill("solid", fgColor="DDEBF7")
        for c in range(1, len(cols) + 1):
            ws.cell(1, c).fill = HF
            ws.cell(1, c).font = Font(bold=True)
        for r in rows:
            ws.append([r["분류"], r["AAS"], r["장비명"], r["제작사(모델)"], r["보유기관"],
                       r["판정"], r["판정 사유"], r["특성 줄수"], "", "", r["파일"]])
            i = ws.max_row
            if r["ZEUS 링크"]:
                cell = ws.cell(i, 9)
                cell.value = "ZEUS 열기"
                cell.hyperlink = r["ZEUS 링크"]
                cell.font = Font(color="0563C1", underline="single")
        for i, w in enumerate([8, 24, 44, 26, 22, 10, 52, 10, 14, 14, 30], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(os.path.join(dst, "유사장비_목록.xlsx"))

        index += rows
        n_ok = sum(1 for r in rows if r["판정"] in OK)
        print(f"{grp:5} {name:30} {len(rows):>2}건 (유효 {n_ok}) → {grp}/{name}/")

    with open(os.path.join(OUT, "_유사장비_전체.json"), "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)
    print(f"\n총 {len(index)}건 → {OUT}")


if __name__ == "__main__":
    main()
