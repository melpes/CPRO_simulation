"""논문·특허에서 뽑은 슬롯 후보를 'AAS 별 슬롯 목록' 으로 정리한다.

TSV 는 근거 한 줄에 한 행이라 같은 슬롯이 여러 번 나온다.
사람이 볼 때 필요한 건 '이 AAS 에 무슨 슬롯이 있고, 근거가 몇 건이며, 값이 어떻게 생겼나' 다.
그래서 슬롯 이름으로 묶고 대표 원문과 링크를 붙인다.
"""

import csv
import os
import re
import sys
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = r"C:\Users\KangTaehui\KG\keti\CPRO_조립공정\시뮬레이션\Package"
KETI = os.path.join(BASE, r"docs\원본자료\keti-fab")
PATENT_TSV = os.path.join(KETI, r"특허수집\_특허_슬롯후보.tsv")
PAPER_TSV = os.path.join(KETI, r"논문수집\_논문_슬롯후보.tsv")
OUT = os.path.join(KETI, r"AAS자료모음\_슬롯_목록.xlsx")

GP = "https://patents.google.com/patent/{}/ko"
HEAD = PatternFill("solid", fgColor="DDEBF7")
LINK = Font(color="0563C1", underline="single")

# 설비 폴더명 → AAS (논문 TSV 는 설비 기준이라 되돌린다)
EQ2AAS = {
    "24_유기증착기": "유기증착기-OrganicChamber",
    "26_박막증착장비": "박막증착장비-Sputter",
    "27_화학 습식 증착(CBD)": "CBD",
    "44_PEALD": "PEALD",
    "43_스핀 트랙 시스템": "현상장비",
    "18_현상장비": "현상장비",
    "34_스핀디벨로퍼": "현상장비",
    "19_엣쳐_스트리퍼": "식각/스트립",
    "35_유기스트리퍼": "식각/스트립",
    "20_스크린프린터": "프린팅",
    "23_잉크젯 프린터(lab)": "프린팅",
    "30_잉크젯프린터(lab #2)": "프린팅",
    "31_리버스 옵셋 프린터": "프린팅",
    "09_잉크젯 프린터 for PLED #1": "프린팅",
    "13_잉크젯 프린터 for PLED #2": "프린팅",
    "21_마스크얼라이너": "마스크 얼라이너",
    "42_마스크 얼라이너(8인치)": "마스크 얼라이너",
}

ORDER = ['박막증착장비-PECVD', '박막증착장비-Sputter', '박막증착장비-DryEtcher', '박막증착장비-ThermalEvaporator', '유기증착기-PlasmaChamber', '유기증착기-OrganicChamber', '유기증착기-MetalChamber', 'PEALD', '현상장비', '마스크 얼라이너', '식각/스트립', '프린팅', 'CBD']


def read_tsv(path, delim="\t"):
    if not os.path.exists(path):
        return []
    txt = open(path, encoding="utf-8-sig").read().replace("\r\n", "\n")
    return list(csv.DictReader(txt.split("\n"), delimiter=delim))


def norm(s):
    """표기가 조금 다른 같은 슬롯을 묶기 위한 키."""
    s = re.sub(r"[()（）\[\]]", " ", (s or "").lower())
    return re.sub(r"[^0-9a-z가-힣]", "", s)


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    # (AAS, 슬롯키) → 정보
    slots = OrderedDict()

    for r in read_tsv(PATENT_TSV):
        aas_raw = (r.get("AAS") or "").strip()
        nm = (r.get("슬롯후보") or "").strip()
        if not aas_raw or not nm:
            continue
        for aas in [a.strip() for a in aas_raw.split(" · ") if a.strip()]:
            k = (aas, norm(nm))
            e = slots.setdefault(k, {"AAS": aas, "이름": nm, "출처": set(), "근거": [],
                                     "표기": set()})
            e["표기"].add(nm)
            e["출처"].add("특허")
            e["근거"].append(("특허", r.get("제조사", ""), r.get("공개번호", ""),
                              (r.get("원문표현") or "").strip(), (r.get("위치") or "").strip()))

    for r in read_tsv(PAPER_TSV):
        eq = (r.get("설비") or "").strip()
        nm = (r.get("슬롯후보") or "").strip()
        if not nm:
            continue
        aas = next((v for kk, v in EQ2AAS.items() if eq.startswith(kk)), None)
        if not aas:
            continue
        k = (aas, norm(nm))
        e = slots.setdefault(k, {"AAS": aas, "이름": nm, "출처": set(), "근거": [],
                                 "표기": set()})
        e["표기"].add(nm)
        e["출처"].add("논문")
        e["근거"].append(("논문", "", (r.get("DOI") or "").strip(),
                          (r.get("원문표현") or "").strip(), (r.get("절위치") or "").strip()))

    wb = Workbook()
    ws = wb.active
    ws.title = "AAS별 슬롯"
    cols = ["AAS", "슬롯 이름", "출처", "근거 건수", "값이 어떻게 나오나 (대표 원문)",
            "다른 표기", "근거 링크1", "근거 링크2", "근거 링크3"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(1, c).fill = HEAD
        ws.cell(1, c).font = Font(bold=True)

    def sort_key(item):
        (aas, _), e = item
        return (ORDER.index(aas) if aas in ORDER else 99, -len(e["근거"]), e["이름"])

    for (aas, _), e in sorted(slots.items(), key=sort_key):
        # 대표 원문 — 숫자가 들어간 것을 우선한다 (값이 어떻게 생겼는지 보여야 한다)
        best = sorted(e["근거"], key=lambda x: (0 if re.search(r"\d", x[3]) else 1, -len(x[3])))
        rep = best[0][3][:160] if best else ""
        others = " / ".join(sorted(t for t in e["표기"] if t != e["이름"]))[:60]
        ws.append([aas, e["이름"], "·".join(sorted(e["출처"])), len(e["근거"]),
                   rep, others, "", "", ""])
        i = ws.max_row
        for n, (kind, maker, ident, _t, _loc) in enumerate(best[:3]):
            cell = ws.cell(i, 7 + n)
            if kind == "특허" and ident:
                cell.value = f"{maker} {ident}"
                cell.hyperlink = GP.format(ident)
                cell.font = LINK
            elif kind == "논문" and ident:
                cell.value = ident.replace("https://doi.org/", "DOI ")[:28]
                cell.hyperlink = ident if ident.startswith("http") else "https://doi.org/" + ident
                cell.font = LINK

    for i, w in enumerate([26, 30, 10, 10, 70, 26, 24, 24, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions

    # 요약 시트
    ws2 = wb.create_sheet("요약", 0)
    ws2.append(["AAS", "슬롯 수", "특허만", "논문만", "둘 다"])
    for c in range(1, 6):
        ws2.cell(1, c).fill = HEAD
        ws2.cell(1, c).font = Font(bold=True)
    per = {}
    for (aas, _), e in slots.items():
        d = per.setdefault(aas, [0, 0, 0, 0])
        d[0] += 1
        if e["출처"] == {"특허"}:
            d[1] += 1
        elif e["출처"] == {"논문"}:
            d[2] += 1
        else:
            d[3] += 1
    for aas in sorted(per, key=lambda a: ORDER.index(a) if a in ORDER else 99):
        ws2.append([aas] + per[aas])
    ws2.append([])
    ws2.append(["합계", sum(v[0] for v in per.values()), sum(v[1] for v in per.values()),
                sum(v[2] for v in per.values()), sum(v[3] for v in per.values())])
    for i, w in enumerate([28, 10, 10, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ── 특허 색인 시트: 명세서를 안 읽어도 무슨 장비 얘기인지 알게 ──────
    import glob
    import json
    ws3 = wb.create_sheet("특허 색인")
    ws3.append(["제조사", "제목", "무슨 발명인가 (본문 첫 문장)", "이 특허가 준 슬롯",
                "본문 길이", "링크"])
    for c in range(1, 7):
        ws3.cell(1, c).fill = HEAD
        ws3.cell(1, c).font = Font(bold=True)

    by_pub = {}
    for (_a, _k), e in slots.items():
        for kind, _m, ident, _t, _l in e["근거"]:
            if kind == "특허" and ident:
                by_pub.setdefault(ident, set()).add(e["이름"])

    rows3 = []
    for f in glob.glob(os.path.join(KETI, r"특허수집\명세서\*.json")):
        d = json.load(open(f, encoding="utf-8"))
        body = d.get("명세서", "")
        pub = d.get("공개번호", "")
        # 한국 특허는 첫 단락이 '본 발명은 ~에 관한 것이다' 로 무엇인지 밝힌다
        m = re.search(r"본\s*발명은[^.]{5,180}\.", body)
        gist = " ".join(m.group(0).split()) if m else " ".join(body[:150].split())
        rows3.append((d.get("제조사", ""), d.get("제목", "")[:60], gist[:180],
                      ", ".join(sorted(by_pub.get(pub, []))[:8]), len(body), pub))
    for maker, title, gist, mine, ln, pub in sorted(rows3, key=lambda x: (-len(x[3]), x[0])):
        ws3.append([maker, title, gist, mine or "(슬롯 없음)", ln, ""])
        if pub:
            c = ws3.cell(ws3.max_row, 6)
            c.value = pub
            c.hyperlink = GP.format(pub)
            c.font = LINK
    for i, w in enumerate([12, 40, 76, 44, 10, 20], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for row in ws3.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws3.freeze_panes = "C2"
    ws3.auto_filter.ref = ws3.dimensions

    wb.save(OUT)
    print(f"슬롯 {len(slots)}종 / 특허 색인 {len(rows3)}건 → {OUT}")
    for aas in sorted(per, key=lambda a: ORDER.index(a) if a in ORDER else 99):
        print(f"   {aas:28} {per[aas][0]:>4}종")


if __name__ == "__main__":
    main()
