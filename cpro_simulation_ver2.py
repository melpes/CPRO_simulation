import simpy
import random
import os
import time
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.nn.functional as F
from collections import defaultdict
from dataclasses import dataclass
from enum import Enum
from path_extractor import (
    load_aas, AASModel,
    ProcessNode, SkillLevel,
)

# ████████████████████████████████████████████████████████████████████
# §A. DATA LOADERS
# ████████████████████████████████████████████████████████████████████

try:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    BASE_DIR = os.getcwd()

POLICY_PATH  = os.path.join(BASE_DIR, 'ppo_policy.pt')
RESULT_PATH  = os.path.join(BASE_DIR, 'simulation_results.xlsx')
AAS_JSON_PATHS = {
    'MODEL_A': os.path.join(BASE_DIR, 'MODEL_A.json'),
    'MODEL_B': os.path.join(BASE_DIR, 'MODEL_B.json'),
    'MODEL_C': os.path.join(BASE_DIR, 'MODEL_C.json'),
}

RANDOM_SEED = 42
DAY_SEC     = 24 * 3600
MAX_DAYS    = 365
_active_schedule: dict = {}

def _apply_schedule(schedule_dict: dict):

    _active_schedule.update(schedule_dict)
    required = ('work_start_sec', 'work_end_sec',
                'lunch_start_sec', 'lunch_end_sec', 'break_duration_sec')
    missing = [k for k in required if k not in _active_schedule]
    if missing:
        raise RuntimeError(
            f'AAS 가 근무 스케줄 미제공: {missing}. '
            f'WorkstationWorkerMatchingData 의 WorkStartTime / WorkEndTime / '
            f'BreakDurationMin(Range) 을 확인하세요.')

PCB_PER_UNIT       = 1
MAG_SIZE           = 15
TRUCK_SIZE         = 30

SOLDER_G           = 500
SOLDER_VALID_SEC   = 24 * 3600
SOLDER_USE_G       = 0.07

CT_STD_RATIO       = 0.10
DEFECT_FLOOR       = 0.00001

SKILL_CT_FACTOR = {1: 1.20, 2: 1.00, 3: 0.80}
SKILL_DR_FACTOR = {1: 1.50, 2: 1.00, 3: 0.60}

MIN_STOCK                  = 100
CRITICAL_STOCK             = 5
REPLENISH_LEAD_DAY         = 1
REPLENISH_QTY_MULT         = 10
WIP_CAP_RATIO              = 1.5

OQC_RATE     = 0.05
OQC_TIME_SEC = 600
AOI_DEFECT_ACTION = 'repair'

RMA_REPAIR_TIME_MEAN_SEC = 300
RMA_REPAIR_TIME_STD_SEC  = 60
RMA_REPAIR_TIME_MIN_SEC  = 60

SMT_BREAKDOWN_PROB  = 0.000005
SMT_MTTR_DEFAULT_HR = 0.5
WORKER_ABSENT_PROB  = 0.0005
THT_DELAY_PROB      = 0.02
THT_DELAY_MIN_SEC   = 1  * 3600
THT_DELAY_MAX_SEC   = 24 * 3600
THT_OUTSOURCE_SEC   = 12 * 3600

TRAIN_MONITOR_INTERVAL = DAY_SEC
INFER_MONITOR_STEP_HR  = 0.1
INFER_MONITOR_INTERVAL = int(INFER_MONITOR_STEP_HR * 3600)
MONITOR_MIN_WALL_SEC   = 0.05


PCB_MAP = {
    'MODEL_A': '03203204',
    'MODEL_B': '03203145',
    'MODEL_C': '03203315',
}
THT_PCB_BY_MODEL = {
    'MODEL_A': ['03902715', '03903424'],
    'MODEL_B': ['03902608', '03902730'],
    'MODEL_C': ['03903388', '03903391'],
}
THT_PCB = {c for codes in THT_PCB_BY_MODEL.values() for c in codes}
THT_RAW_SUFFIX = '_RAW'
def tht_raw_code(pcb_code):
    return f'{pcb_code}{THT_RAW_SUFFIX}'

SMT_LINE_IDS = ['L1', 'L2']
PCB_INITIAL_RATIO          = 0.8
BOM_INITIAL_RATIO          = 0.6
BOM_LOT_RATIO              = 0.5
WAREHOUSE_BOM_INIT_FLOOR   = 50
WAREHOUSE_BOM_LOT_FLOOR    = 50
WAREHOUSE_NONBOM_INIT_MULT = 10
RATED_POWER_KW = {
    'SMT_LOADER_L1':    0.66,
    'SMT_PRINTER_L1':   0.84,
    'SMT_SPI_L1':       2.20,
    'SMT_MOUNTER_H_L1': 19.93,
    'SMT_MOUNTER_M_L1': 4.64,
    'SMT_REFLOW_L1':    63.26,
    'SMT_UNLOADER_L1':  0.33,
    'SMT_LOADER_L2':    0.66,
    'SMT_PRINTER_L2':   1.72,
    'SMT_SPI_L2':       1.29,
    'SMT_MOUNTER_H_L2': 10.13,
    'SMT_MOUNTER_M_L2': 4.64,
    'SMT_REFLOW_L2':    48.03,
    'SMT_UNLOADER_L2':  0.33,
    'SMT_AOI':          0.29,
    'MODULE_FW':        0.0,
    'NVD_40_FOCUS':     0.36,
    'MODULE':          23.38,
    'SEMI':            25.50,
    'SET':             33.67,
    'INSP':             1.22,
    'AGING':            1.84,
    'OQC':              0.44,
    'PACK':             8.31,
    'PACK_LABEL':       0.37,
    'RMA':              0.50,
}

def get_rated_power_kw(process_code: str, process_group: str = '',
                       capacity: int = 1) -> float:
    base = RATED_POWER_KW.get(str(process_code))
    if base is None:
        base = RATED_POWER_KW.get(str(process_group), 0.0)
    return base / max(int(capacity), 1)
WORKER_DEFAULT_CAP = {
    'WORKER_RMA': 6,
}

SET_INSP_HEADCOUNT = 3
PROCESS_GROUP_TO_WORKER_GROUP = {
    'MODULE':       None,
    'MODULE_FW':    'WORKER_FW',
    'NVD_40_FOCUS': 'WORKER_SENSOR_FOCUS',
    'SEMI':         'WORKER_SEMI',
    'SET':          'WORKER_SET',
    'INSP':         'WORKER_AGING',
    'AGING':        'WORKER_AGING',
    'OQC':          'WORKER_OQC',
    'PACK':         'WORKER_PACK',
    'RMA':          'WORKER_RMA',
}


def _find_pack_entry(data, model_id):
    try:
        procs = data.get_model_procs(model_id)
    except Exception:
        return None
    grp_of = {str(r['process_code']): str(r.get('process_group', '') or '')
              for _, r in procs.iterrows()}
    for _, r in procs.iterrows():
        if str(r.get('process_group', '') or '') != 'PACK':
            continue
        prevs = [p.strip() for p in
                 str(r.get('dep_prev_codes', '') or '').split(';') if p.strip()]
        for p in prevs:
            if grp_of.get(p) == 'INSP':
                return str(r['process_code'])
    return None

WORKER_GROUPS = {
    'WORKER_FW', 'WORKER_SENSOR_FOCUS', 'WORKER_LENS_HOLDER',
    'WORKER_SEMI', 'WORKER_SET', 'WORKER_SET_INSP',
    'WORKER_AGING', 'WORKER_OQC', 'WORKER_PACK', 'WORKER_RMA',
}

LOCATION_LABEL = {
    'WORKER_FW'          : 'F/W 입력',
    'WORKER_LENS_HOLDER' : 'LENS HOLDER 조립',
    'WORKER_SENSOR_FOCUS': 'FOCUS',
    'WORKER_SET'         : 'SET 조립',
    'WORKER_SET_INSP'    : 'SET 조립 (INSP)',
    'WORKER_SEMI'        : '반 조립 라인',
    'WORKER_RMA'         : 'RMA',
    'WORKER_OQC'         : 'OQC',
    'WORKER_AGING'       : 'Aging test',
    'WORKER_PACK'        : '포장',
}
LOCATION_ORDER = [
    'WORKER_FW', 'WORKER_LENS_HOLDER', 'WORKER_SENSOR_FOCUS',
    'WORKER_SET', 'WORKER_SET_INSP', 'WORKER_SEMI',
    'WORKER_RMA', 'WORKER_OQC', 'WORKER_AGING', 'WORKER_PACK',
]


# ══════════════════════════════════════════════════════════
# M02. 정적 데이터 로더 (FallbackDataLoader)
# ══════════════════════════════════════════════════════════

class FallbackDataLoader:

    _PF_COLS = ('process_code', 'process_group', 'dep_type', 'dep_prev_codes',
                'worker_group', 'worker_count', 'cycle_time_sec', 'defect_rate',
                'transfer_qty', 'transport_mode', 'transfer_time_sec')

    _PF_ALL_ROWS = [
        ('RMA_REPAIR',       'RMA', 'SEQUENCE', '',                 'WORKER_RMA', 6, 600.0, 0.0,   0, '',         0.0),
        ('SMT_LOADER_L1',    'SMT', 'SEQUENCE', '',                 'OPERATOR',   2,   2.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_LOADER_L2',    'SMT', 'SEQUENCE', '',                 'OPERATOR',   2,   2.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_PRINTER_L1',   'SMT', 'SEQUENCE', 'SMT_LOADER_L1',    'OPERATOR',   2,  43.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_PRINTER_L2',   'SMT', 'SEQUENCE', 'SMT_LOADER_L2',    'OPERATOR',   2,  43.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_SPI_L1',       'SMT', 'SEQUENCE', 'SMT_PRINTER_L1',   'OPERATOR',   2,  15.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_SPI_L2',       'SMT', 'SEQUENCE', 'SMT_PRINTER_L2',   'OPERATOR',   2,  15.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_MOUNTER_H_L1', 'SMT', 'SEQUENCE', 'SMT_SPI_L1',       'OPERATOR',   2,  80.0, 0.003, 1, 'CONVEYOR', 5.0),
        ('SMT_MOUNTER_H_L2', 'SMT', 'SEQUENCE', 'SMT_SPI_L2',       'OPERATOR',   2,  80.0, 0.003, 1, 'CONVEYOR', 5.0),
        ('SMT_MOUNTER_M_L1', 'SMT', 'SEQUENCE', 'SMT_MOUNTER_H_L1', 'OPERATOR',   2,  45.0, 0.005, 1, 'CONVEYOR', 5.0),
        ('SMT_MOUNTER_M_L2', 'SMT', 'SEQUENCE', 'SMT_MOUNTER_H_L2', 'OPERATOR',   2,  45.0, 0.005, 1, 'CONVEYOR', 5.0),
        ('SMT_REFLOW_L1',    'SMT', 'SEQUENCE', 'SMT_MOUNTER_M_L1', 'OPERATOR',   2, 410.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_REFLOW_L2',    'SMT', 'SEQUENCE', 'SMT_MOUNTER_M_L2', 'OPERATOR',   2, 410.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_UNLOADER_L1',  'SMT', 'SEQUENCE', 'SMT_REFLOW_L1',    'OPERATOR',   2,   5.0, 0.001, 0, 'CART',     7.0),
        ('SMT_UNLOADER_L2',  'SMT', 'SEQUENCE', 'SMT_REFLOW_L2',    'OPERATOR',   2,   5.0, 0.001, 0, 'CART',     7.0),
        ('SMT_AOI', 'SMT', 'JOIN', 'SMT_UNLOADER_L1;SMT_UNLOADER_L2', 'OPERATOR', 2,  30.0, 0.003, 1, '',         0.0),
    ]

    _RESOURCE_MTTR_HR = {
        'SMT_AOI':          3.3,
        'SMT_MOUNTER_H_L1': 1.5,
        'SMT_MOUNTER_H_L2': 1.4,
        'SMT_MOUNTER_M_L1': 1.5,
        'SMT_MOUNTER_M_L2': 1.5,
        'SMT_PRINTER_L1':   4.0,
        'SMT_PRINTER_L2':   4.5,
        'SMT_REFLOW_L1':    8.0,
        'SMT_REFLOW_L2':    8.0,
        'SMT_SPI_L2':       2.3,
    }

    def __init__(self):
        rows = []
        for tup in self._PF_ALL_ROWS:
            rec = dict(zip(self._PF_COLS, tup))
            rec['model_id']    = 'ALL'
            rec['ref_no']      = ''
            rec['process_name']= ''
            rec['dep_wait_hr'] = 0.0
            rows.append(rec)
        self.pf = pd.DataFrame(rows)
        self.pf['defect_rate'] = self.pf['defect_rate'].apply(
            lambda x: x if x > 0 else DEFECT_FLOOR)

        self.bom = pd.DataFrame(columns=[
            'item_code', 'item_name', 'smt_side', 'min_stock_qty',
            'lot_size', 'critical_stock_qty', 'defect_rate'])
        self.bom_struct = pd.DataFrame(columns=[
            'model_id', 'parent_type', 'parent_code', 'item_code',
            'item_name', 'qty_per_parent'])

        self.workers   = {}
        self.resources = []

        self._pc_map = {str(r['process_code']): r for _, r in self.pf.iterrows()}
        self._grp_kw = defaultdict(float)
        self._mttr   = {pc: hr * 3600 for pc, hr in self._RESOURCE_MTTR_HR.items()}
        self._bom_idx = defaultdict(list)
        self._min_stock_cache      = {}
        self._lot_size_cache       = {}
        self._item_name_cache      = {}
        self._smt_side_cache       = {}
        self._critical_stock_cache = {}
        self._bom_dup_merge_log    = []

    def get_proc(self, pc):
        return self._pc_map.get(str(pc))

    def get_kw(self, process_code, process_group, capacity=1):
        kw = get_rated_power_kw(process_code, process_group, capacity)
        if kw > 0:
            return kw
        return {'SMT': 5.0, 'MODULE': 1.0, 'SEMI': 2.0, 'SET': 2.0,
                'INSP': 0.5, 'OQC': 0.2, 'PACK': 1.0, 'RMA': 0.1
                }.get(process_group, 0.5) / max(capacity, 1)

    def get_mttr(self, process_code):
        return self._mttr.get(str(process_code), SMT_MTTR_DEFAULT_HR * 3600)

    def get_min_stock(self, item_code):
        return self._min_stock_cache.get(str(item_code), float(MIN_STOCK))

    def get_critical_stock(self, item_code):
        """페널티(violation) 임계. min_stock(reorder point) 와 분리."""
        return self._critical_stock_cache.get(str(item_code), float(CRITICAL_STOCK))

    def get_lot_size(self, item_code):
        c = str(item_code)
        lot = self._lot_size_cache.get(c, -1)
        if lot > 0:
            return lot
        return int(self.get_min_stock(c) * REPLENISH_QTY_MULT)

    def get_bom_parts(self, model_id, parent_code):
        return self._bom_idx.get((model_id, str(parent_code)), [])

    def get_pcb_parts(self, pcb_code):
        """PCB 코드 → BOM 부품 목록. fallback 은 PCB BOM 을 보유하지 않으므로 빈 리스트.

        실제 PCB 부품 데이터는 AAS HierarchicalStructures (PCB_xxxxx Entity 의
        HasPart_yyyy RelationshipElement) 가 단일 출처이며, CombinedDataLoader.
        get_pcb_parts 에서 _hs_bom_idx 를 통해 조회한다.
        """
        return []

    def get_item_name(self, item_code):
        return self._item_name_cache.get(str(item_code), '')

    def smt_side(self, item_code):
        return 'double'

    def iter_all_bom_items(self):
        """fallback 자체는 BOM item 미보유 (AAS 가 단일 출처)."""
        return set()

    def get_all_bom_codes(self):
        if not hasattr(self, '_all_bom_codes_cache'):
            self._all_bom_codes_cache = self.iter_all_bom_items()
        return self._all_bom_codes_cache

    def get_model_procs(self, model_id):
        df = self.pf
        return df[df['model_id'].isin([model_id, 'ALL'])
                  & ~df['process_group'].isin(
                      ['SMT', 'LOGISTICS', 'SMT_SHARED', 'RMA'])].copy()

# ══════════════════════════════════════════════════════════
# M02b. AAS 데이터 통합 로더
# ══════════════════════════════════════════════════════════

class CombinedDataLoader:
    """path_extractor.load_aas() 결과(AASModel)를 소비하는 통합 인터페이스.

    데이터 흐름:
      path_extractor.load_aas(model_id, path) → AASModel
        .ManufacturingProcess  {ProcessCode: ProcessNode}
        .WorkstationWorkerMatchingData {WorkstationId: WorkstationData}
        .SkillLevelType        {name: SkillLevel}
        .HierarchicalStructures HierarchicalStructuresData
        .schedule              {WorkStartTime, WorkEndTime, ...}
        .group_to_workstation  {GroupIdShort: WorkstationId}

    FallbackDataLoader (SMT/RMA 정적 데이터) 와 AASModel 을 통합하여
    시뮬레이션 전체에서 단일 data 객체로 접근 가능하게 한다.
    """

    def __init__(self, static_loader: 'FallbackDataLoader', aas_models: dict):
        self.static   = static_loader
        self.aas_map  = aas_models

        self._pc_map = dict(static_loader._pc_map)
        for model_id, aas in aas_models.items():
            for ProcessCode, node in aas.ManufacturingProcess.items():
                self._pc_map[ProcessCode] = {
                    'process_code'     : node.ProcessCode,
                    'model_id'         : model_id,
                    'process_group'    : node.ProcessGroup,
                    'worker_group'     : self._resolve_worker(node, aas),
                    'cycle_time_sec'   : float(node.CycleTimeSec),
                    'defect_rate'      : max(float(node.DefectRate), DEFECT_FLOOR),
                    'dep_type'         : node.DepType,
                    'dep_prev_codes'   : ';'.join(node.DepPrev),
                    'dep_wait_hr'      : 0.0,
                    'transfer_qty'     : 1,
                    'transfer_time_sec': 0.0,
                    'transport_mode'   : '',
                }

        self._bom_idx = dict(static_loader._bom_idx)
        for model_id, aas in aas_models.items():
            for ProcessCode, node in aas.ManufacturingProcess.items():
                key = (model_id, ProcessCode)
                entries = [(item.item_code, item.Quantity) for item in node.InputBOM]
                if entries:
                    self._bom_idx[key] = entries

        self._hs_bom_idx = {}
        for model_id, aas in aas_models.items():
            for pcb_id, pcb in aas.HierarchicalStructures.pcb_entries.items():
                key = (model_id, pcb_id)
                self._hs_bom_idx[key] = [
                    (comp.item_code, comp.Quantity, comp.Category)
                    for comp in pcb.components
                ]
            for part_id, part in aas.HierarchicalStructures.assembly_parts.items():
                key = (model_id, '__UNIT__')
                self._hs_bom_idx.setdefault(key, [])
                self._hs_bom_idx[key].append((part_id, part.Quantity, part.Category))

        self.workers = {}
        for aas in aas_models.values():
            for ws_id, ws in aas.WorkstationWorkerMatchingData.items():
                wgrp = self._ws_to_worker(ws_id)
                if wgrp:
                    self.workers[wgrp] = ws.WorkstationConfigurationRecords
        for wgrp, cap in WORKER_DEFAULT_CAP.items():
            self.workers.setdefault(wgrp, cap)
        set_total = self.workers.get('WORKER_SET', 0)
        set_insp  = max(int(set_total * 0.2), SET_INSP_HEADCOUNT) if set_total else SET_INSP_HEADCOUNT
        self.workers.setdefault('WORKER_SET_INSP', set_insp)

        self.worker_skill = {}
        self.skill_ct     = {}
        self.skill_dr     = {}
        for aas in aas_models.values():
            for ws_id, ws in aas.WorkstationWorkerMatchingData.items():
                wgrp = self._ws_to_worker(ws_id)
                if wgrp:
                    self.worker_skill[wgrp] = ws.SkillLevel
            for name, sl in aas.SkillLevelType.items():
                self.skill_ct[sl.rank] = sl.ct_factor
                self.skill_dr[sl.rank] = sl.dr_factor
            break
        if not self.skill_ct:
            self.skill_ct = dict(SKILL_CT_FACTOR)
            self.skill_dr = dict(SKILL_DR_FACTOR)

        self.worker_groups  = set(self.workers.keys())
        self.location_order = [ws for ws in LOCATION_ORDER if ws in self.worker_groups]
        self.location_label = dict(LOCATION_LABEL)

        self.pcb_map          = {}
        self.tht_pcb_by_model = {}
        for model_id, aas in aas_models.items():
            main_pcbs = [
                pcb_id[4:] for pcb_id, pcb in aas.HierarchicalStructures.pcb_entries.items()
                if pcb.components
            ]
            tht_pcbs = [
                pcb_id[4:] for pcb_id, pcb in aas.HierarchicalStructures.pcb_entries.items()
                if not pcb.components
            ]
            if main_pcbs:
                self.pcb_map[model_id] = main_pcbs[0]
            if tht_pcbs:
                self.tht_pcb_by_model[model_id] = tht_pcbs
        if not self.pcb_map:
            self.pcb_map = dict(PCB_MAP)
        if not self.tht_pcb_by_model:
            self.tht_pcb_by_model = dict(THT_PCB_BY_MODEL)

        self.schedule = {}
        for aas in aas_models.values():
            s = aas.schedule
            if s:
                self.schedule = {
                    'work_start_sec'    : s.get('WorkStartTime', 32400),
                    'work_end_sec'      : s.get('WorkEndTime', 64800),
                    'lunch_start_sec'   : s.get('BreakDurationMin_min', 43200),
                    'lunch_end_sec'     : s.get('BreakDurationMin_max', 46800),
                    'break_duration_sec': s.get('BreakDurationMin_max', 46800)
                                        - s.get('BreakDurationMin_min', 43200),
                }
                break

        self.bom        = static_loader.bom
        self.bom_struct = static_loader.bom_struct
        self.resources  = static_loader.resources
        self._pf_combined = self._build_combined_pf()

    _WWM_LINE_TO_WORKER = {
        'WWM_FwInputLine'      : 'WORKER_FW',
        'WWM_LensHolderLine'   : 'WORKER_LENS_HOLDER',
        'WWM_FocusLine'        : 'WORKER_SENSOR_FOCUS',
        'WWM_SemiAssemblyLine' : 'WORKER_SEMI',
        'WWM_SetAssemblyLine'  : 'WORKER_SET',
        'WWM_AgingLine'        : 'WORKER_AGING',
        'WWM_OqcLine'          : 'WORKER_OQC',
        'WWM_RMALine'          : 'WORKER_RMA',
        'WWM_PackagingLine'    : 'WORKER_PACK',
    }

    def _ws_to_worker(self, ws_id: str) -> str:
        return self._WWM_LINE_TO_WORKER.get(ws_id, '')

    def _resolve_worker(self, node: ProcessNode, aas: AASModel) -> str:
        ws_id = aas.group_to_workstation.get(node.GroupIdShort, '')
        wgrp  = self._ws_to_worker(ws_id)
        if not wgrp:
            pg = node.ProcessGroup
            wgrp = PROCESS_GROUP_TO_WORKER_GROUP.get(pg, 'WORKER_SEMI') or 'WORKER_SEMI'
        return wgrp

    def _build_combined_pf(self) -> pd.DataFrame:
        aas_rows = list(self._pc_map.values())
        static_rows = self.static.pf.to_dict('records')
        df = pd.DataFrame(static_rows + [r for r in aas_rows if isinstance(r, dict)])
        for col in ['cycle_time_sec','defect_rate','transfer_qty','transfer_time_sec','dep_wait_hr']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        if 'defect_rate' in df.columns:
            df['defect_rate'] = df['defect_rate'].apply(lambda x: x if x > 0 else DEFECT_FLOOR)
        return df

    def get_proc(self, pc):
        return self._pc_map.get(str(pc))

    def get_kw(self, process_code, process_group, capacity=None):
        return self.static.get_kw(process_code, process_group, capacity or 1)

    def get_mttr(self, process_code):
        return self.static.get_mttr(process_code)

    def get_min_stock(self, item_code):
        return self.static.get_min_stock(item_code)

    def get_critical_stock(self, item_code):
        return self.static.get_critical_stock(item_code)

    def get_lot_size(self, item_code):
        return self.static.get_lot_size(item_code)

    def get_item_name(self, item_code):
        return self.static.get_item_name(item_code)

    def get_bom_parts(self, model_id, parent_code):
        return self._bom_idx.get((model_id, str(parent_code)), [])

    def get_pcb_parts(self, pcb_code):
        for model_id in self.aas_map:
            result = self._hs_bom_idx.get((model_id, f'PCB_{pcb_code}'), [])
            if result:
                return [(ic, qty) for ic, qty, _ in result]
        return []

    def smt_side(self, item_code):
        c = str(item_code)
        for aas in self.aas_map.values():
            pcb = aas.HierarchicalStructures.pcb_entries.get(f'PCB_{c}')
            if pcb:
                return pcb.SMT_Side
        return 'double'

    def get_all_bom_codes(self):
        codes = set()
        for (mid, pc), items in self._bom_idx.items():
            for ic, _ in items:
                codes.add(str(ic))
        return codes

    def iter_all_bom_items(self):
        return self.get_all_bom_codes()

    def get_model_procs(self, model_id):
        df = self._pf_combined
        return df[
            df['model_id'].isin([model_id, 'ALL'])
            & ~df['process_group'].isin(['SMT','LOGISTICS','SMT_SHARED','RMA'])
        ].copy()

# ══════════════════════════════════════════════════════════
# M03. 공정 지식 그래프
# ══════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════
# M03. 공정 지식 그래프 & Ready 게이트
# ══════════════════════════════════════════════════════════

class ReadyStatus(Enum):
    READY       = 'ready'
    WAIT_PRED   = 'wait_pred'
    WAIT_STOCK  = 'wait_stock'
    WAIT_WORKER = 'wait_worker'
    UNKNOWN_PC  = 'unknown_pc'


@dataclass
class ReadyContext:
    """ready 판정에 필요한 런타임 핸들 묶음."""
    kg       : 'ProcessKnowledgeGraph'
    done_set : set
    wh       : 'Warehouse'
    wres     : dict
    data     : object
    model_id : str


def resolve_worker_group(pc: str, node: dict) -> str:
    """SET 공정 중 process_code 접미사가 INSP 인 경우 WORKER_SET_INSP 로 매핑."""
    wgrp = str(node.get('worker_group', '') or '')
    grp  = str(node.get('process_group', '') or '')
    if wgrp == 'WORKER_SET' and grp == 'SET':
        if str(pc).rsplit('_', 1)[-1].upper() == 'INSP':
            wgrp = 'WORKER_SET_INSP'
    return wgrp


def _bom_satisfied(pc: str, ctx: ReadyContext) -> bool:
    for code, qty in ctx.data.get_bom_parts(ctx.model_id, pc):
        if ctx.wh.stock[str(code)] < qty:
            return False
    return True


def is_process_ready(pc: str, ctx: ReadyContext) -> ReadyStatus:
    pc = str(pc)
    if pc not in ctx.kg.nodes:
        return ReadyStatus.UNKNOWN_PC

    node  = ctx.kg.nodes[pc]
    preds = [p for (p, t, _) in ctx.kg.edges if t == pc]

    if preds:
        if not all(p in ctx.done_set for p in preds):
            return ReadyStatus.WAIT_PRED
        if not _bom_satisfied(pc, ctx):
            return ReadyStatus.WAIT_STOCK
    else:
        if not _bom_satisfied(pc, ctx):
            return ReadyStatus.WAIT_STOCK

    wgrp = resolve_worker_group(pc, node)
    if wgrp:
        res = ctx.wres.get(wgrp)
        if res is None or res.count >= res.capacity:
            return ReadyStatus.WAIT_WORKER

    return ReadyStatus.READY


class ProcessKnowledgeGraph:
    """공정 DAG.

    노드 특징 벡터 (6차원):
      [cycle_time 정규화, defect_rate×1000 클리핑,
       worker_count/20, rated_kw/100, is_fork, is_join]
    """

    def __init__(self, data, model_id: str):
        self.model_id = model_id
        self._data    = data
        self.nodes    = {}
        self.edges    = []
        df     = data.get_model_procs(model_id)
        max_ct = max(df['cycle_time_sec'].max(), 1)
        for _, r in df.iterrows():
            pc   = str(r['process_code'])
            wgrp = str(r.get('worker_group', '') or '')
            kw   = data.get_kw(pc, str(r.get('process_group', '') or ''))
            dt   = str(r.get('dep_type', 'SEQUENCE') or 'SEQUENCE').upper()
            self.nodes[pc] = {
                'process_code'  : pc,
                'process_group' : str(r.get('process_group', '') or ''),
                'cycle_time_sec': float(r['cycle_time_sec'] or 0),
                'defect_rate'   : float(r['defect_rate'] or DEFECT_FLOOR),
                'dep_wait_hr'   : float(r['dep_wait_hr'] or 0),
                'worker_group'  : wgrp,
                'worker_count'  : data.workers.get(wgrp, 1),
                'rated_kw'      : kw,
                'transfer_time' : float(r['transfer_time_sec'] or 0),
                'dep_type'      : dt,
                'feat': np.array([
                    float(r['cycle_time_sec'] or 0) / max_ct,
                    min(float(r['defect_rate'] or DEFECT_FLOOR) * 1000, 1.0),
                    data.workers.get(wgrp, 1) / 20,
                    min(kw / 100, 1.0),
                    1.0 if dt == 'FORK' else 0.0,
                    1.0 if dt == 'JOIN'  else 0.0,
                ], dtype=np.float32)
            }
            for prev in [p.strip() for p in
                         str(r.get('dep_prev_codes', '') or '').split(';') if p.strip()]:
                if prev != pc:
                    self.edges.append((prev, pc, dt))

    def get_feat_matrix(self):
        pcs = list(self.nodes.keys())
        return pcs, np.stack([self.nodes[p]['feat'] for p in pcs])

    def get_adj(self):
        pcs = list(self.nodes.keys())
        idx = {p: i for i, p in enumerate(pcs)}
        N   = len(pcs)
        adj = np.zeros((N, N), dtype=np.float32)
        for f, t, _ in self.edges:
            if f in idx and t in idx:
                adj[idx[f]][idx[t]] = 1.0
        return adj

    def ready_processes(self, ctx: ReadyContext) -> list:
        """is_process_ready 를 통과한 공정 코드 목록을 반환.

        에이전트는 이 목록 중에서 최적 공정을 선택한다.
        """
        return [pc for pc in self.nodes
                if pc not in ctx.done_set
                and is_process_ready(pc, ctx) == ReadyStatus.READY]


# ══════════════════════════════════════════════════════════
# M04. 창고 / WIPTracker
# ══════════════════════════════════════════════════════════

class Warehouse:
    def __init__(self, data, order: dict):
        self.data    = data
        self.order   = order
        total_qty    = sum(order.values())

        bom_codes             = data.get_all_bom_codes()
        self._bom_codes       = bom_codes
        self._bom_init_stock  = WAREHOUSE_BOM_INIT_FLOOR
        self._init_stock      = float(total_qty * WAREHOUSE_NONBOM_INIT_MULT)

        demand = defaultdict(float)
        for model_id, qty in order.items():
            try:
                procs = data.get_model_procs(model_id)
            except Exception:
                continue
            for _, row in procs.iterrows():
                pc = str(row['process_code'])
                for ic, q in data.get_bom_parts(model_id, pc):
                    demand[str(ic)] += float(q) * int(qty)

        self.stock = defaultdict(lambda: self._init_stock)
        self._demand = dict(demand)
        self._initial_stocks = {}
        for code in bom_codes:
            d   = demand.get(str(code), 0.0)
            init = max(WAREHOUSE_BOM_INIT_FLOOR, int(d * BOM_INITIAL_RATIO))
            self.stock[str(code)] = float(init)
            self._initial_stocks[str(code)] = init

        for model_id, qty in order.items():
            main_pcb = self.data.pcb_map.get(model_id)
            if main_pcb is not None:
                self.stock[str(main_pcb)] = float(int(qty * PCB_INITIAL_RATIO))
            for tht_pcb in self.data.tht_pcb_by_model.get(model_id, []):
                self.stock[str(tht_pcb)] = float(int(qty * PCB_INITIAL_RATIO))

        for pcb_code in {c for codes in self.data.tht_pcb_by_model.values() for c in codes}:
            self.stock[tht_raw_code(pcb_code)] = 0.0

        self.consumed        = defaultdict(int)
        self.violations      = defaultdict(int)
        self.history         = defaultdict(list)
        self._pending_orders = set()
        self._wait_events    = defaultdict(list)
        self.snapshots       = defaultdict(list)
        self.reorder_log     = []
        self.reorder_count   = defaultdict(int)

        self._pcb_codes = set(self.data.pcb_map.values()) | {c for codes in self.data.tht_pcb_by_model.values() for c in codes}
        self.outsource_log       = []
        self.unit_completions    = {}
        self.smt_per_model       = defaultdict(int)
        self.pcb_flow            = defaultdict(lambda: defaultdict(int))
        self.skipped_pcs         = defaultdict(int)
        self.smt_model_choices   = []
        self.kg_incomplete_log   = []
        self.smt_single_side_log = []
        self.stuck_wait_log      = []

    def consume(self, item_code, qty, sim_time=0):
        c = str(item_code)
        self.stock[c] = max(0, self.stock[c] - int(qty))
        self.consumed[c] += int(qty)
        if not c.endswith(THT_RAW_SUFFIX):
            is_pcb = c in self._pcb_codes
            if (not is_pcb
                    and self.stock[c] < self.data.get_min_stock(c)
                    and c not in self._pending_orders):
                self._pending_orders.add(c)
                self.snapshots[c].append((sim_time, self.stock[c]))
            if self.stock[c] < self.data.get_critical_stock(c):
                self.violations[c] += 1
        self.history[c].append((sim_time, self.stock[c]))

    def restore(self, item_code, qty, sim_time=0):
        c = str(item_code)
        self.stock[c] += int(qty)
        self.history[c].append((sim_time, self.stock[c]))
        if c.endswith(THT_RAW_SUFFIX):
            base = c[:-len(THT_RAW_SUFFIX)]
            self.pcb_flow[base]['outsource_in'] += int(qty)
        elif c in self._pcb_codes:
            self.pcb_flow[c]['restore_from_smt_or_outsource'] += int(qty)
        self._notify_waiters(c)

    def wait_stock(self, env, item_code, qty, max_wait_sec=None):
        """item_code 재고가 qty 이상이 될 때까지 대기 후 consume.

        max_wait_sec 초 이내에 재고가 확보되지 않으면 데드락 탈출을 위해
        clamp consume 으로 진행한다 (stuck_wait_log 에 기록).
        """
        if max_wait_sec is None:
            max_wait_sec = MAX_DAYS * (
                _active_schedule['work_end_sec']
                - _active_schedule['work_start_sec']
                - _active_schedule['break_duration_sec']
            )
        c = str(item_code)
        blocked = False
        start_t = float(env.now)
        while self.stock[c] < qty:
            if not blocked and not c.endswith(THT_RAW_SUFFIX):
                self.violations[c] += 1
                blocked = True
            ev = env.event()
            self._wait_events[c].append(ev)
            remaining = max_wait_sec - (env.now - start_t)
            if remaining <= 0:
                break
            timeout_ev = env.timeout(remaining)
            result = yield env.any_of([ev, timeout_ev])
            if timeout_ev in result and ev not in result:
                self.stuck_wait_log.append({
                    'item_code'    : c,
                    'qty'          : int(qty),
                    'wait_start_h' : start_t / 3600,
                    'wait_end_h'   : float(env.now) / 3600,
                    'stock_at_end' : float(self.stock[c]),
                })
                self.consume(c, qty, env.now)
                return
        self.consume(c, qty, env.now)

    def _notify_waiters(self, item_code):
        c = str(item_code)
        for ev in self._wait_events.pop(c, []):
            if not ev.triggered:
                ev.succeed()

    def stock_penalty(self):
        return sum(self.violations.values())

    def _lot_for(self, item_code):
        """부품별 발주 lot_size — 예상 demand × BOM_LOT_RATIO.

        demand 정보 없으면 (BOM 외 부품 등) data.get_lot_size 로 fallback.
        floor 는 WAREHOUSE_BOM_LOT_FLOOR (소량 부품도 너무 잦은 발주 방지).
        """
        c = str(item_code)
        d = self._demand.get(c, 0.0)
        if d > 0:
            return max(WAREHOUSE_BOM_LOT_FLOOR, int(d * BOM_LOT_RATIO))
        return int(self.data.get_lot_size(c))

    def replenish(self, item_code, sim_time=0, order_time=None, stock_at_order=None):
        c = str(item_code)
        if c.endswith(THT_RAW_SUFFIX):
            return
        if c in self._pcb_codes:
            self._pending_orders.discard(c)
            return
        lot      = self._lot_for(c)
        min_s    = self.data.get_min_stock(c)
        incoming = max(lot, int(min_s) - self.stock[c] + lot)
        self.stock[c] += incoming
        self._pending_orders.discard(c)
        self.history[c].append((sim_time, self.stock[c]))
        is_pcb = c in self._pcb_codes
        self.reorder_log.append({
            'item_code'     : c,
            'order_time'    : float(order_time) if order_time is not None else float(sim_time),
            'arrive_time'   : float(sim_time),
            'lot_size'      : int(lot),
            'incoming'      : int(incoming),
            'stock_at_order': (int(stock_at_order) if stock_at_order is not None
                               else int(self.stock[c] - incoming)),
            'is_pcb'        : is_pcb,
        })
        self.reorder_count[c] += 1
        if is_pcb:
            self.pcb_flow[c]['external_replenish_arrived'] += 1
        self._notify_waiters(c)

    def snapshot_loop(self, env, interval=3600):
        """1시간마다 추적 대상 부품의 재고를 스냅샷으로 찍는 SimPy 프로세스.
        추적 대상: (a) 현재 stock dict 에 등장한 부품 + (b) BOM 마스터 전체 item_code.
        소비·입고가 한 번도 없었던 부품도 초기재고로 행에 나와야 하므로 BOM 마스터 포함.
        """
        tracked = set()
        try:
            for c in self.data.iter_all_bom_items():
                tracked.add(c)
        except AttributeError:
            pass
        while True:
            yield env.timeout(interval)
            tracked.update(self.stock.keys())
            now = env.now
            for c in tracked:
                q = self.stock[c]
                self.snapshots[c].append((now, int(q)))



class WIPTracker:
    """
    재고품 수량 상한 = 총 주문 x WIP_CAP_RATIO x 3 (임의, ConWIP 기반)
    Spearman et al. (1990) ConWIP; Ekerete et al. (2026 IRE Journals)
    """
    def __init__(self, order: dict):
        self.wip  = defaultdict(int)
        self.cap  = defaultdict(int)
        self.viol = defaultdict(int)
        self.history = defaultdict(list)
        self.snapshots = defaultdict(list)
        total = sum(order.values())
        for grp in ['MODULE','SEMI','SET','INSP','PACK','SMT']:
            self.cap[grp] = max(int(total * WIP_CAP_RATIO * 3), 10)

    def enter(self, grp):
        self.wip[grp] += 1
        if self.wip[grp] > self.cap.get(grp, 9999):
            self.viol[grp] += 1

    def leave(self, grp):
        self.wip[grp] = max(0, self.wip[grp] - 1)

    def violations(self):
        return sum(self.viol.values())

    def snapshot_loop(self, env, interval=3600):
        """1시간마다 그룹별 WIP 스냅샷 기록 (WIP_Timeseries 시트용)."""
        tracked_grps = ['SMT', 'MODULE', 'SEMI', 'SET', 'INSP', 'PACK', 'RMA']
        while True:
            yield env.timeout(interval)
            now = env.now
            for grp in tracked_grps:
                self.snapshots[grp].append((now, int(self.wip.get(grp, 0))))


# ══════════════════════════════════════════════════════════
# M05. 전력 로거
# ══════════════════════════════════════════════════════════

class EnergyLogger:
    def __init__(self, data):
        self.data   = data
        self.by_pc  = defaultdict(float)
        self.by_grp = defaultdict(float)
        self.total  = 0.0
        self.history = []

    def record(self, pc, grp, ct, sim_time=0, capacity=None):
        if capacity is None:
            wgrp = PROCESS_GROUP_TO_WORKER_GROUP.get(str(grp))
            if wgrp:
                capacity = int(self.data.workers.get(wgrp, 1))
            else:
                capacity = 1
        kw  = self.data.get_kw(pc, grp, capacity)
        kwh = kw * float(ct) / 3600
        self.by_pc[pc]   += kwh
        self.by_grp[grp] += kwh
        self.total       += kwh
        self.history.append((sim_time, self.total))
        return kwh

    def report(self):
        print('\n[공정그룹별 전력 소비 (kWh)]')
        print(f'  {"그룹":12s} {"kWh":>10s}')
        for g in sorted(self.by_grp):
            print(f'  {g:12s} {self.by_grp[g]:>10.4f}')
        print(f'  {"합계":12s} {self.total:>10.4f}')


# ══════════════════════════════════════════════════════════
# M06. 유휴·숙련도 추적기
# ══════════════════════════════════════════════════════════

class IdleTracker:
    """capacity 기반 per-person idle 적분 추적기.

    의도: "각 워커 개개인이, 근무시간에, 자기 그룹에 할 일이 남았는데 앞 공정
    병목으로 일을 못 하고 쉬는 시간" 의 person·sec 합계.

    동작:
      - 그룹 g 의 capacity = N. 매 시점 t 에 점유 중인 워커 수 = active(t).
        idle 워커 수 = max(N - active(t), 0).
      - acquire/release 이벤트 사이의 (cap - active) × Δt(근무시간) 를 적분.
      - 그룹별 자기 할당량 (set_target 으로 등록된 수) 처리 완료 시점
        (_completed_at[g]) 이후는 idle 카운트 X — "내 할 일 끝나면 idle 아님".

    호환:
      - 기존 mark_busy(env, name) API 는 SMT pc/AOI 등 cap=1 binary 추적용으로 유지.
        WORKER_GROUPS 대상은 acquire/release 로 교체됐고, mark_busy 결과는
        worker_idle_penalty 에 포함되지 않음 (그룹 capacity 등록된 항목만 합산).
    """
    def __init__(self):
        self._capacity            = {}
        self._active              = defaultdict(int)
        self._last_t              = defaultdict(float)
        self.total_idle           = defaultdict(float)
        self.absent_groups        = set()
        self._completed_at        = {}
        self._completion_target   = {}
        self._completion_counter  = defaultdict(int)
        self._worker_groups       = set()

    def configure(self, capacity_map: dict, worker_groups: set = None):
        for g, cap in capacity_map.items():
            self._capacity[g] = int(cap)
            self._active.setdefault(g, 0)
            self._last_t.setdefault(g, 0.0)
        if worker_groups:
            self._worker_groups = set(worker_groups)

    def set_target(self, g: str, target: int):
        """그룹별 처리해야 할 work item 총수 등록 (자기 할당량 완료 판정용)."""
        if target > 0:
            self._completion_target[g] = int(target)

    def _flush(self, env, g):
        """그룹 g 의 (last_t, env.now) 사이 idle 워커 person·sec 누적."""
        now = float(env.now)
        last = self._last_t.get(g, 0.0)
        if g not in self._capacity:
            self._last_t[g] = now
            return
        if now <= last:
            return
        cap    = self._capacity[g]
        active = self._active.get(g, 0)
        idle_workers = max(cap - active, 0)
        completed = self._completed_at.get(g)
        if completed is not None and last >= completed:
            self._last_t[g] = now
            return
        eff_end = now if completed is None else min(now, completed)
        if eff_end > last and idle_workers > 0:
            self.total_idle[g] += idle_workers * _work_seconds_between(last, eff_end)
        self._last_t[g] = now

    def flush_all(self, env):
        """모든 등록 그룹의 마지막 구간까지 idle 반영. report/reward 직전 호출."""
        for g in list(self._capacity.keys()):
            self._flush(env, g)

    def acquire(self, env, g):
        """그룹 g 의 워커 1 명이 점유 시작 — yield res.request() 직후 호출."""
        if g not in self._capacity:
            self._capacity[g] = 1
            self._last_t.setdefault(g, float(env.now))
        self._flush(env, g)
        self._active[g] = self._active.get(g, 0) + 1

    def release(self, env, g):
        """그룹 g 의 워커 1 명이 점유 종료 — res.release() 직전 호출.

        그룹별 처리량 카운터 +=1. 등록된 target 도달 시 mark_completed 자동 발화 →
        그룹 g 의 자기 할당량 완료 시각이 정확히 잡힘 (전체 PACK 완성 시점이 아님).
        """
        if g not in self._capacity:
            return
        self._flush(env, g)
        self._active[g] = max(self._active.get(g, 0) - 1, 0)
        self._completion_counter[g] += 1
        target = self._completion_target.get(g)
        if (target is not None
                and self._completion_counter[g] >= target
                and g not in self._completed_at):
            self._completed_at[g] = float(env.now)


    def mark_completed(self, wgrp: str, sim_time: float):
        existing = self._completed_at.get(wgrp)
        if existing is None or sim_time > existing:
            self._completed_at[wgrp] = float(sim_time)

    def worker_idle_penalty(self, threshold=300):
        """그룹 capacity 등록된 (= acquire/release 추적된) 그룹의 누적 idle 만 합산."""
        total = 0.0
        for name in self._capacity:
            if name not in self._worker_groups:
                continue
            v = self.total_idle.get(name, 0.0)
            if v > threshold:
                total += v
        return total

    def report(self):
        print('\n[작업자 유휴 시간 상위 10 (그룹 person·hh:mm:ss)]')
        workers = {k: v for k, v in self.total_idle.items() if k in self._worker_groups}
        for n, v in sorted(workers.items(), key=lambda x: -x[1])[:10]:
            cap  = self._capacity.get(n, 1)
            flag = ' <- 임계초과' if v > 1800 * cap else ''
            sec = int(max(v, 0))
            h, rem = divmod(sec, 3600)
            mm, ss = divmod(rem, 60)
            print(f'  {n:25s}(cap={cap:2d}): {h:02d}:{mm:02d}:{ss:02d}{flag}')


# ══════════════════════════════════════════════════════════
# M07. SMT 라인
# ══════════════════════════════════════════════════════════

class SolderCream:
    def __init__(self, env, name):
        self.env       = env
        self.name      = name
        self.stock_g   = SOLDER_G
        self.open_time = None

    def use(self):
        if self.open_time is None:
            self.open_time = self.env.now
        if self.env.now - self.open_time >= SOLDER_VALID_SEC:
            self.stock_g   = SOLDER_G
            self.open_time = self.env.now
        self.stock_g -= SOLDER_USE_G
        if self.stock_g <= 0:
            self.stock_g   = SOLDER_G
            self.open_time = self.env.now


class OutsourceTruckPool:
    """모든 SMT 라인이 공유하는 외주 트럭 풀.

    THT 보드를 종류 무관하게 동일 트럭에 적재 (옵션 b).
    (2026-05-06 변경) 트럭 적재량 제한 제거 — 트럭은 충분히 크다고 가정하고
    SMT 라인 종료 시점에 누적된 모든 보드를 한 번에 convoy 로 출발.
    TRUCK_SIZE 는 모니터링/엑셀 통계 표시 시 "트럭 1대당 환산 보드 수" 로만 사용.
    THT_DELAY_PROB 는 convoy 단위로 적용 (한 번 늦으면 모두 같이 늦음).
    """

    def __init__(self, env, wh, stats, smt_lines_ref):
        self.env = env
        self.wh = wh
        self.stats = stats
        self.smt_lines = smt_lines_ref
        self.truck = []
        self.dispatched_count = 0
        self.in_transit = []
        self.truck_log = []

    def add_board(self, line_sid, pcb_code, model_id, board_id):
        """SMT 라인이 THT 보드 1장 외주 발사. 트럭에 적재. (자동 출발 안 함)"""
        raw_code = tht_raw_code(pcb_code)
        self.wh.restore(raw_code, 1, self.env.now)
        self.stats['tht_out'] = self.stats.get('tht_out', 0) + 1
        ev_idx = len(self.wh.outsource_log)
        self.wh.outsource_log.append({
            'pcb_code'    : pcb_code,
            'model_id'    : model_id,
            'board_id'    : board_id,
            'send_time'   : float(self.env.now),
            'return_time' : None,
            'delay_sec'   : 0.0,
            'status'      : 'queued',
        })
        self.truck.append({
            'pcb_code': pcb_code,
            'model_id': model_id,
            'board_id': board_id,
            'line_sid': line_sid,
            'ev_idx'  : ev_idx,
        })

    def flush_now(self):
        """SMT 라인 모두 종료 시 누적된 보드 한 번에 출발 (convoy)."""
        if self.truck:
            self._dispatch()

    def _dispatch(self):
        truck = self.truck
        self.truck = []
        delay = 0.0
        if random.random() < THT_DELAY_PROB:
            delay = random.uniform(THT_DELAY_MIN_SEC, THT_DELAY_MAX_SEC)
        send_t = float(self.env.now)
        for entry in truck:
            log = self.wh.outsource_log[entry['ev_idx']]
            log['send_time'] = send_t
            log['delay_sec'] = float(delay)
            log['status']    = 'in_flight'
        eta = send_t + THT_OUTSOURCE_SEC + delay
        self.dispatched_count += 1
        breakdown = defaultdict(int)
        for entry in truck:
            breakdown[entry['pcb_code']] += 1
        truck_log_entry = {
            'dispatch_id'  : self.dispatched_count,
            'send_t'       : send_t,
            'eta'          : eta,
            'delay_sec'    : float(delay),
            'size'         : len(truck),
            'breakdown'    : dict(breakdown),
            'truck_count'  : max(1, (len(truck) + TRUCK_SIZE - 1) // TRUCK_SIZE),
            'return_t'     : None,
        }
        self.truck_log.append(truck_log_entry)
        transit_meta = {'size': len(truck), 'send_t': send_t, 'eta': eta,
                        'log_entry': truck_log_entry}
        self.in_transit.append(transit_meta)
        self.env.process(self._truck_arrive(truck, delay, transit_meta))

    def _truck_arrive(self, truck, delay, transit_meta):
        yield self.env.timeout(THT_OUTSOURCE_SEC + delay)
        try:
            self.in_transit.remove(transit_meta)
        except ValueError:
            pass
        log_entry = transit_meta.get('log_entry')
        if log_entry is not None:
            log_entry['return_t'] = float(self.env.now)
        for entry in truck:
            pcb_code = entry['pcb_code']
            line_sid = entry['line_sid']
            line = self.smt_lines.get(line_sid)
            raw_code = tht_raw_code(pcb_code)
            self.wh.consume(raw_code, 1, self.env.now)
            log = self.wh.outsource_log[entry['ev_idx']]
            log['return_time'] = float(self.env.now)
            log['status']      = 'returned'
            self.wh.pcb_flow[pcb_code]['outsource_returned'] += 1
            if line is None:
                self.wh.restore(pcb_code, 1, self.env.now)
                continue
            line.mag_buf[pcb_code] += 1
            in_flight = (self.wh.pcb_flow[pcb_code].get('outsource_in', 0)
                         - self.wh.pcb_flow[pcb_code].get('outsource_returned', 0))
            if line.mag_buf[pcb_code] >= MAG_SIZE:
                flush = MAG_SIZE
            elif in_flight == 0 and line.mag_buf[pcb_code] > 0:
                flush = line.mag_buf[pcb_code]
            else:
                flush = 0
            if flush > 0:
                line.mag_buf[pcb_code] -= flush
                self.wh.restore(pcb_code, flush, self.env.now)
                line.pcb_count[pcb_code] += flush
                self.stats['pcb_done'] = self.stats.get('pcb_done', 0) + flush
                self.wh.smt_per_model[(line_sid, entry['model_id'], pcb_code)] += flush


class SMTLine:
    def __init__(self, env, suffix, data, wh,
                 aoi_res, rma_store, energy, idle, wip, stats, broken_flag,
                 outsource_pool=None):
        self.env          = env
        self.sfx          = suffix
        self.data         = data
        self.wh           = wh
        self.aoi_res      = aoi_res
        self.rma          = rma_store
        self.outsource_pool = outsource_pool
        self.energy       = energy
        self.idle         = idle
        self.wip          = wip
        self.stats        = stats
        self.broken_flag  = broken_flag
        self.solder       = SolderCream(env, f'SMT_{suffix}')
        self.mag_buf      = defaultdict(int)
        self.assigned_model = None
        self._res = {pc: simpy.Resource(env, capacity=1) for pc in [
            f'SMT_LOADER_{suffix}',    f'SMT_PRINTER_{suffix}',
            f'SMT_SPI_{suffix}',       f'SMT_MOUNTER_H_{suffix}',
            f'SMT_MOUNTER_M_{suffix}', f'SMT_REFLOW_{suffix}',
            f'SMT_UNLOADER_{suffix}']}
        self.pcb_count = defaultdict(int)
        self.stage_active = {}

    def process_board(self, pcb_code, board_id, model_id, is_second=False):
        """SMT 보드 1장 처리.

        결함 처리 (2026-05-06): 조립 공정의 INSP 검출 로직과 동일하게,
        LOADER~UNLOADER stage 중 어느 곳에서든 결함이 발생하면 board_has_defect
        플래그만 표시하고 보드는 계속 진행한다. 마지막 SMT_AOI 가 누적 결함을
        자체 dr 확률로 검출하면 RMA 큐 (src='AOI') 로 투입한다 — 이 경로에서
        AOI_DEFECT_ACTION 으로 수리/폐기 토글 가능.
        """
        seq = [f'SMT_LOADER_{self.sfx}',    f'SMT_PRINTER_{self.sfx}',
               f'SMT_SPI_{self.sfx}',       f'SMT_MOUNTER_H_{self.sfx}',
               f'SMT_MOUNTER_M_{self.sfx}', f'SMT_REFLOW_{self.sfx}',
               f'SMT_UNLOADER_{self.sfx}']

        self.wip.enter('SMT')
        board_has_defect = False
        for pc in seq:
            pr = self.data.get_proc(pc)
            if pr is None:
                yield self.env.timeout(0.001)
                continue
            ct = float(pr['cycle_time_sec'] or 0.001)
            tt = float(pr['transfer_time_sec'] or 0)
            dr = float(pr['defect_rate'] or DEFECT_FLOOR)

            if not _is_work_time(self.env.now):
                yield self.env.timeout(_next_work_start(self.env.now) - self.env.now)

            while self.broken_flag.get(pc, False):
                yield self.env.timeout(60)

            with self._res[pc].request() as req:
                yield req
                _start_t = float(self.env.now)
                self.stage_active[pc] = (pcb_code, board_id, is_second)
                try:
                    act = max(random.normalvariate(ct, ct * CT_STD_RATIO), ct * 0.5)
                    yield self.env.timeout(act)
                finally:
                    self.stage_active.pop(pc, None)
            self.energy.record(pc, 'SMT', act, self.env.now)
            if 'PRINTER' in pc:
                self.solder.use()
            if tt > 0:
                yield self.env.timeout(tt)
            if random.random() < dr:
                self.stats['smt_defect'] += 1
                board_has_defect = True

        aoi = self.data.get_proc('SMT_AOI')
        aoi_ct = float(aoi['cycle_time_sec'] or 30)
        aoi_dr = float(aoi['defect_rate'] or DEFECT_FLOOR)

        if not _is_work_time(self.env.now):
            yield self.env.timeout(_next_work_start(self.env.now) - self.env.now)

        with self.aoi_res.request() as req:
            yield req
            _aoi_start_t = float(self.env.now)
            self.stage_active['SMT_AOI'] = (pcb_code, board_id, is_second)
            try:
                yield self.env.timeout(aoi_ct)
            finally:
                self.stage_active.pop('SMT_AOI', None)
        self.energy.record('SMT_AOI', 'SMT_SHARED', aoi_ct, self.env.now)
        detected = board_has_defect and (random.random() < aoi_dr)
        if detected:
            self.stats['aoi_defect'] += 1
            self.wip.leave('SMT')
            yield self.rma.put({'src':'AOI','board':board_id,
                                'pcb':pcb_code,'grp':'SMT_SHARED','model':model_id})
            return

        self.wip.leave('SMT')

        side = self.data.smt_side(pcb_code)
        if not is_second and side == 'double':
            self.wh.pcb_flow[pcb_code]['double_first_pass'] += 1
            yield self.env.process(
                self.process_board(pcb_code, board_id, model_id,
                                    is_second=True))
            return

        if side == 'double' and not is_second:
            self.wh.smt_single_side_log.append({
                'pcb_code': pcb_code, 'model_id': model_id,
                'board_id': board_id, 'time_h': self.env.now/3600,
                'reason': 'double_pcb_flushed_without_second_pass',
            })

        if pcb_code in {c for codes in data.tht_pcb_by_model.values() for c in codes}:
            if self.outsource_pool is not None:
                self.outsource_pool.add_board(self.sfx, pcb_code, model_id, board_id)
            else:
                self.wh.restore(pcb_code, 1, self.env.now)
            return

        self.mag_buf[pcb_code] += 1
        if self.mag_buf[pcb_code] >= MAG_SIZE:
            self.mag_buf[pcb_code] -= MAG_SIZE
            self.wh.restore(pcb_code, MAG_SIZE, self.env.now)
            self.pcb_count[pcb_code] += MAG_SIZE
            self.stats['pcb_done'] += MAG_SIZE
            self.wh.smt_per_model[(self.sfx, model_id, pcb_code)] += MAG_SIZE


# ████████████████████████████████████████████████████████████████████
# §C. SIMULATION
# ████████████████████████████████████████████████████████████████████


# ══════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════

def _is_work_time(sim_now_sec):
    _s = _active_schedule
    t  = sim_now_sec % DAY_SEC
    if not (_s['work_start_sec'] <= t < _s['work_end_sec']):
        return False
    if _s['lunch_start_sec'] <= t < _s['lunch_end_sec']:
        return False
    return True


def _next_work_start(sim_now_sec):
    _s      = _active_schedule
    day_num = int(sim_now_sec // DAY_SEC)
    t       = sim_now_sec % DAY_SEC
    if _s['lunch_start_sec'] <= t < _s['lunch_end_sec']:
        return day_num * DAY_SEC + _s['lunch_end_sec']
    if t < _s['work_start_sec']:
        return day_num * DAY_SEC + _s['work_start_sec']
    return (day_num + 1) * DAY_SEC + _s['work_start_sec']


def work_timeout(env, duration):
    """근무시간 내에서만 경과. 점심·퇴근 boundary 만나면 재개 시점까지 pause."""
    _s        = _active_schedule
    remaining = float(max(duration, 0))
    while remaining > 1e-6:
        if not _is_work_time(env.now):
            yield env.timeout(_next_work_start(env.now) - env.now)
            continue
        t = env.now % DAY_SEC
        if t < _s['lunch_start_sec']:
            next_boundary = env.now - t + _s['lunch_start_sec']
        else:
            next_boundary = env.now - t + _s['work_end_sec']
        chunk = min(remaining, max(next_boundary - env.now, 0))
        if chunk <= 0:
            yield env.timeout(_next_work_start(env.now) - env.now)
            continue
        yield env.timeout(chunk)
        remaining -= chunk


def _work_seconds_between(start_sec, end_sec):
    """두 시뮬레이션 시각 사이의 실 근무시간(초)을 반환.
    야간·점심을 제외한 순수 근무시간만 계산. IdleTracker.mark_busy 에서 호출.
    """
    _s = _active_schedule
    if end_sec <= start_sec:
        return 0.0
    total = 0.0
    t     = float(start_sec)
    end   = float(end_sec)
    while t < end:
        if not _is_work_time(t):
            nxt = _next_work_start(t)
            if nxt >= end:
                break
            t = nxt
            continue
        day_t = t % DAY_SEC
        if day_t < _s['lunch_start_sec']:
            boundary = t - day_t + _s['lunch_start_sec']
        else:
            boundary = t - day_t + _s['work_end_sec']
        chunk = min(end, boundary) - t
        if chunk > 0:
            total += chunk
        t = min(end, boundary)
    return total


# ══════════════════════════════════════════════════════════
# M08. 단일 공정 실행
# ══════════════════════════════════════════════════════════

def run_process(env, prow, done_ev, wres, wh, rma, energy,
                idle, wip, stats, mid, data, uid=0,
                unit_defect_flag=None, progress=None):
    """단일 공정 실행.

    unit_defect_flag : {uid: bool} 공유 딕셔너리.
        비INSP 공정에서 불량 발생 시 True 로 표시하고 공정은 계속 진행.
        INSP 공정에서 플래그가 있으면 defect_rate 확률로 검출 -> RMA 투입.
    progress : ManufacturingEnv.progress 공유 딕셔너리. RMA 투입 시 전달.
    """
    pc    = str(prow['process_code'])
    grp   = str(prow.get('process_group','') or '')
    ct    = float(prow['cycle_time_sec'] or 0)
    dr    = float(prow['defect_rate'] or DEFECT_FLOOR)
    tt    = float(prow['transfer_time_sec'] or 0)
    whr   = float(prow['dep_wait_hr'] or 0)
    wgrp  = str(prow.get('worker_group','') or '')
    dtype = str(prow.get('dep_type','SEQUENCE') or 'SEQUENCE').upper()
    prevs = [p.strip() for p in
             str(prow.get('dep_prev_codes','') or '').split(';') if p.strip()]

    if wgrp == 'WORKER_SET' and grp == 'SET':
        if pc.rsplit('_', 1)[-1].upper() == 'INSP':
            wgrp = 'WORKER_SET_INSP'

    skill_base   = data.worker_skill.get(wgrp, 2)
    skill_actual = max(1, skill_base - (1 if wgrp in idle.absent_groups else 0))
    ct = ct * data.skill_ct.get(skill_actual, 1.0)
    dr = dr * data.skill_dr.get(skill_actual, 1.0)

    if prevs:
        wait_evs = [done_ev[p] for p in prevs if p in done_ev]
        if dtype == 'JOIN':
            if wait_evs:
                yield simpy.AllOf(env, wait_evs)
        elif wait_evs:
            yield wait_evs[0]

    if whr > 0:
        yield env.timeout(whr * 3600)
    if tt > 0:
        yield env.timeout(tt)

    if not _is_work_time(env.now):
        yield env.timeout(_next_work_start(env.now) - env.now)

    wip.enter(grp)
    for item_code, qty in data.get_bom_parts(mid, pc):
        yield from wh.wait_stock(env, item_code, qty)

    res = wres.get(wgrp)
    req = res.request() if res else None
    acquired = False
    if req:
        yield req
        idle.acquire(env, wgrp)
        acquired = True

    try:
        act = max(random.normalvariate(ct, ct * CT_STD_RATIO), ct * 0.5) if ct > 0 else 0.001
        yield from work_timeout(env, act)
        energy.record(pc, grp, act, env.now)
        _is_insp = (grp == 'SET' and pc.rsplit('_', 1)[-1].upper() == 'INSP')

        if _is_insp:
            has_prior_defect = (unit_defect_flag is not None
                                and unit_defect_flag.get(uid, False))
            detected = has_prior_defect and (random.random() < dr)
            if detected:
                stats['assy_defect'] = stats.get('assy_defect', 0) + 1
                if unit_defect_flag is not None:
                    unit_defect_flag[uid] = False
                    unit_defect_flag['_routed_to_rma'] = True
                wip.leave(grp)
                yield rma.put({'src': pc, 'grp': grp, 'model': mid,
                               'uid': uid,
                               'progress': progress})
                if req and res:
                    if acquired:
                        idle.release(env, wgrp)
                        acquired = False
                    res.release(req)
                ev = done_ev.get(pc)
                if ev and not ev.triggered:
                    ev.succeed()
                return
        else:
            if random.random() < dr:
                stats['assy_defect'] = stats.get('assy_defect', 0) + 1
                if unit_defect_flag is not None:
                    unit_defect_flag[uid] = True
    finally:
        if req and res:
            try:
                if acquired:
                    idle.release(env, wgrp)
                    acquired = False
                res.release(req)
            except Exception:
                pass

    wip.leave(grp)
    ev = done_ev.get(pc)
    if ev and not ev.triggered:
        ev.succeed()


# ══════════════════════════════════════════════════════════
# M09. RMA 수리 및 재투입
# ══════════════════════════════════════════════════════════

def run_rma(env, rma, wres, wh, energy, idle, wip, stats, data,
            progress=None):
    """RMA 큐 디스패처.

    AOI 보드 불량 처리 — AOI_DEFECT_ACTION 으로 분기:
      'repair' (기본): RMA 수리 후 wh.restore() 로 PCB 인벤토리 직접 재투입.
      'scrap'        : 폐기. SMT scheduler 가 추가 생산해 보전.
    (2026-05-06) SMT stage 자체 결함은 더 이상 즉시 폐기되지 않음 — process_board
    가 board_has_defect 플래그만 표시하고 AOI 가 검출하므로 모든 SMT 결함 경로는
    src='AOI' 로 통합됨.
    INSP 공정에서 검출된 유닛 불량은 _rma_repair_and_reinsert 로 수리 후 PACK 재투입.
    """
    while True:
        item = yield rma.get()
        src = str(item.get('src', ''))
        if src == 'AOI':
            if AOI_DEFECT_ACTION == 'repair':
                env.process(_rma_repair_aoi_board(
                    env, item, wres, wh, energy, idle, wip, stats, data))
                continue
            else:
                stats['smt_rma_scrap'] = stats.get('smt_rma_scrap', 0) + 1
                continue
        env.process(_rma_repair_and_reinsert(
            env, item, wres, wh, energy, idle, wip, stats, data,
            progress=progress))


def _rma_repair_aoi_board(env, item, wres, wh, energy, idle, wip, stats, data):
    """AOI 검출 불량 SMT 보드 수리 → PCB 인벤토리 직접 재투입 (요청사항 0506 ④).

    수리 시 부품 소모 (2026-05-06): PCB 분해·재납땜 작업이므로 그 PCB 의
    BOM 부품 (HierarchicalStructures PCB-typed) 한 세트를 다시 소모한다.
    부품 부족하면 wait_stock 으로 대기 (RMA 수리 비동기 대기).
    """
    pcb_code = str(item.get('pcb', ''))
    model    = str(item.get('model', ''))
    res = wres.get('WORKER_RMA')
    req = res.request() if res else None
    acquired = False
    try:
        if req:
            yield req
            idle.acquire(env, 'WORKER_RMA')
            acquired = True
        rt = max(random.normalvariate(RMA_REPAIR_TIME_MEAN_SEC,
                                       RMA_REPAIR_TIME_STD_SEC),
                 RMA_REPAIR_TIME_MIN_SEC)
        yield from work_timeout(env, rt)
        energy.record('RMA_REPAIR', 'RMA', rt, env.now)
        stats['aoi_repaired'] = stats.get('aoi_repaired', 0) + 1

        if pcb_code:
            for part_code, part_qty in data.get_pcb_parts(pcb_code):
                yield from wh.wait_stock(env, part_code, part_qty)
            wh.restore(pcb_code, 1, env.now)
    finally:
        if req:
            if acquired:
                idle.release(env, 'WORKER_RMA')
            res.release(req)


def _sample_defective_predecessor(data, model_id, src_pc):
    """src_pc(INSP) 의 트랜지티브 선행 공정 중 F/W~SET (MODULE/SEMI/SET 그룹,
    INSP suffix 제외) 후보를 모아 defect_rate 가중치로 1개 sampling.

    "INSP 가 잡아낸 결함이 어느 조립 공정에서 발생했는지 모르므로 (기록 없음)
    그 INSP 앞 단계의 비-INSP 공정 중 불량률에 비례해 1개 추정" 의도.
    가중치 합이 0 이거나 후보가 없으면 None.
    """
    try:
        procs = data.get_model_procs(model_id)
    except Exception:
        return None
    prow_map = {str(r['process_code']): r for _, r in procs.iterrows()}

    visited = set()
    queue = [str(src_pc)]
    candidates = []
    while queue:
        cur = queue.pop(0)
        prow = prow_map.get(cur)
        if prow is None:
            continue
        prevs = [p.strip() for p in
                 str(prow.get('dep_prev_codes', '') or '').split(';') if p.strip()]
        for prev in prevs:
            if prev in visited:
                continue
            visited.add(prev)
            queue.append(prev)
            prev_row = prow_map.get(prev)
            if prev_row is None:
                continue
            grp  = str(prev_row.get('process_group', '') or '')
            wgrp = str(prev_row.get('worker_group', '') or '')
            if grp not in ('MODULE', 'SEMI', 'SET'):
                continue
            if prev.rsplit('_', 1)[-1].upper() == 'INSP':
                continue
            if wgrp == 'WORKER_SET_INSP':
                continue
            candidates.append(prev_row)

    if not candidates:
        return None
    weights = [max(float(c.get('defect_rate', 0) or 0), 0.0) for c in candidates]
    if sum(weights) <= 0:
        return None
    chosen = random.choices(candidates, weights=weights, k=1)[0]
    return str(chosen['process_code'])


def _rma_repair_and_reinsert(env, item, wres, wh, energy, idle, wip, stats, data,
                             progress=None):
    """RMA 수리 후 PACK 첫 공정으로 재투입.

    흐름:
      SET/INSP 공정 불량 발생
        -> run_rma 큐에서 꺼냄
        -> WORKER_RMA 자원 확보
        -> 수리 (임의 평균 300초)
        -> 해당 모델의 PACK 첫 공정(_find_pack_entry 자동 추출) 실행
        -> 완성 카운터 증가
    """
    if not _is_work_time(env.now):
        yield env.timeout(_next_work_start(env.now) - env.now)

    res = wres.get('WORKER_RMA')
    req = res.request() if res else None
    acquired = False
    if req:
        yield req
        idle.acquire(env, 'WORKER_RMA')
        acquired = True

    model    = item.get('model', '')
    src_pc   = item.get('src', '')
    rma_uid  = int(item.get('uid', 0))
    progress = item.get('progress') or progress or {}
    pc_rma   = 'RMA_REPAIR'

    try:
        rt = max(random.normalvariate(RMA_REPAIR_TIME_MEAN_SEC,
                                       RMA_REPAIR_TIME_STD_SEC),
                 RMA_REPAIR_TIME_MIN_SEC)
        yield from work_timeout(env, rt)
        energy.record('RMA_REPAIR', 'RMA', rt, env.now)
        stats['rma_repaired'] = stats.get('rma_repaired', 0) + 1

        # MODULE/SEMI/SET 그룹의 비-INSP 공정을 defect_rate 가중치로 1개 sampling
        if src_pc:
            sampled_pc = _sample_defective_predecessor(data, model, src_pc)
            if sampled_pc:
                for part_code, part_qty in data.get_bom_parts(model, sampled_pc):
                    yield from wh.wait_stock(env, part_code, part_qty)

    finally:
        if req and res:
            try:
                if acquired:
                    idle.release(env, 'WORKER_RMA')
                    acquired = False
                res.release(req)
            except Exception:
                pass

    pack_pc = _find_pack_entry(data, model)
    if pack_pc is None:
        _do_complete(stats, progress, model, env.now, wh=wh, src_pc=src_pc)
        return

    prow = data.get_proc(pack_pc)
    if prow is None:
        _do_complete(stats, progress, model, env.now, wh=wh, src_pc=src_pc)
        return

    done_ev_rma = {}
    pack_pcs = _get_pack_sequence(data, model, pack_pc)
    for p_pc in pack_pcs:
        p_prow = data.get_proc(p_pc)
        if p_prow is None:
            continue
        p_prow_copy = p_prow.copy()
        p_prow_copy['dep_prev_codes'] = ''
        p_prow_copy['dep_wait_hr']    = 0
        done_ev_rma[p_pc] = env.event()
        yield env.process(
            run_process(env, p_prow_copy, done_ev_rma, wres, wh,
                        simpy.Store(env), energy, idle, wip, stats, model, data,
                        unit_defect_flag=None,
                        progress=progress))

    _do_complete(stats, progress, model, env.now, wh=wh, src_pc=src_pc)


def _do_complete(stats, progress, model, now, wh=None, src_pc=None):
    if not model or model not in progress:
        return
    done, total = progress[model]
    if done >= total:
        if wh is not None:
            wh.unit_completions[('RMA', f'{model}_{src_pc}_{int(now)}')] = {
                'path': 'rma_blocked_by_quota',
                'end_time': float(now),
                'done_n': -1,
                'total_n': -1,
                'rma_count': 1,
            }
        return
    stats[f'{model}_done'] = stats.get(f'{model}_done', 0) + 1
    progress[model] = (done + 1, total)
    pct = (done + 1) / max(total, 1) * 100
    if wh is not None:
        rma_idx = sum(1 for k in wh.unit_completions
                      if isinstance(k, tuple) and k[0] == model and
                      isinstance(k[1], str) and k[1].startswith('rma_'))
        wh.unit_completions[(model, f'rma_{rma_idx}')] = {
            'path': 'rma',
            'end_time': float(now),
            'done_n': -1,
            'total_n': -1,
            'rma_count': 1,
            'src_pc': str(src_pc) if src_pc else '',
        }


def _get_pack_sequence(data, model, first_pack_pc):
    """PACK 공정 연속 순서 반환. first_pack_pc 부터 시작해 SEQUENCE로 연결된 공정들.

    CombinedDataLoader / FallbackDataLoader 양쪽을 지원.
    get_model_procs() 가 있으면 해당 메서드 사용, 없으면 빈 리스트.
    """
    result = [first_pack_pc]
    visited = {first_pack_pc}
    try:
        procs = data.get_model_procs(model)
    except Exception:
        return result

    next_map = {}
    for _, row in procs.iterrows():
        pc   = str(row['process_code'])
        prev = str(row.get('dep_prev_codes', '') or '')
        grp  = str(row.get('process_group', '') or '')
        if grp != 'PACK':
            continue
        for p in [x.strip() for x in prev.split(';') if x.strip()]:
            next_map[p] = pc
    cur = first_pack_pc
    while cur in next_map:
        nxt = next_map[cur]
        if nxt in visited:
            break
        result.append(nxt)
        visited.add(nxt)
        cur = nxt
    return result


# ══════════════════════════════════════════════════════════
# M10. 단일 제품 생산
# ══════════════════════════════════════════════════════════

def produce_unit(env, model_id, unit_id, data, kg,
                 wres, wh, rma, energy, idle, wip, stats, progress, menv=None):
    done_ev = {pc: env.event() for pc in kg.nodes}

    pcs     = list(kg.nodes.keys())
    idx_map = {pc: i for i, pc in enumerate(pcs)}
    if menv is not None and hasattr(menv, '_H_cache'):
        H   = menv._H_cache[model_id]
        adj = menv._adj_cache[model_id]
    else:
        _, H_np = kg.get_feat_matrix()
        H   = torch.tensor(H_np,  dtype=torch.float32)
        adj = torch.tensor(kg.get_adj(), dtype=torch.float32)
    done_set = {'SMT_COMPLETE', 'SMT_THT'}
    kg_done  = set()
    kg_total = set(kg.nodes.keys())
    unit_defect_flag = {unit_id: False}

    unit_key = (model_id, unit_id)
    us = None
    if menv is not None and hasattr(menv, 'unit_states'):
        us = menv.unit_states
        us[unit_key] = {'state': 'SMT_WAIT', 'pc': '-', 'done_n': 0,
                        'total_n': len(kg.nodes), 'ready': []}

    ready_ctx = ReadyContext(
        kg=kg, done_set=done_set, wh=wh, wres=wres,
        data=data, model_id=model_id)
    while kg_done != kg_total:
        ready_pcs = kg.ready_processes(ready_ctx)

        if us is not None:
            us[unit_key].update({
                'done_n': len(kg_done),
                'ready': list(ready_pcs)[:5],
                'state': 'READY_WAIT' if not ready_pcs else 'RUNNING',
            })

        if not ready_pcs:
            yield env.timeout(1)
            continue

        if menv is not None and menv.agent is not None:
            s = menv.get_state()
            ready_mask = torch.zeros(len(pcs), dtype=torch.bool)
            for pc in ready_pcs:
                if pc in idx_map:
                    ready_mask[idx_map[pc]] = True
            a, lp, v, emb, mask_bytes = menv.agent.act(s, H, adj, ready_mask)
            next_pc = pcs[a] if (a < len(pcs) and ready_mask[a]) else ready_pcs[0]
            r = menv.reward()
            menv.agent.store(s, emb, a, r, lp, v, mask=mask_bytes, model_id=model_id)
        else:
            next_pc = ready_pcs[0]

        if us is not None:
            us[unit_key]['pc'] = next_pc

        prow = data.get_proc(next_pc)
        if prow is None:
            if menv is not None and hasattr(menv, 'wh'):
                menv.wh.skipped_pcs[(model_id, str(next_pc))] += 1
            done_set.add(next_pc)
            kg_done.add(next_pc)
            continue

        yield env.process(
            run_process(env, prow, done_ev, wres, wh, rma,
                        energy, idle, wip, stats, model_id, data,
                        unit_defect_flag=unit_defect_flag,
                        progress=progress))

        done_set.add(next_pc)
        kg_done.add(next_pc)

        if unit_defect_flag.get('_routed_to_rma'):
            if us is not None:
                us[unit_key].update({'state': 'ROUTED_TO_RMA', 'pc': '-'})
            if menv is not None and hasattr(menv, 'wh'):
                menv.wh.unit_completions[(model_id, unit_id)] = {
                    'path': 'routed_to_rma',
                    'end_time': float(env.now),
                    'done_n': len(kg_done),
                    'total_n': len(kg_total),
                    'rma_count': 0,
                }
            return

    missing = kg_total - kg_done
    if missing:
        stats['kg_incomplete_units'] = stats.get('kg_incomplete_units', 0) + 1
        if menv is not None and hasattr(menv, 'wh'):
            menv.wh.kg_incomplete_log.append({
                'model_id': model_id, 'unit_id': unit_id,
                'time_h': float(env.now)/3600,
                'missing_pcs': sorted(missing),
            })

    if us is not None:
        us[unit_key].update({'state': 'DONE', 'pc': '-',
                             'done_n': len(kg_done)})

    if random.random() < OQC_RATE:
        if not _is_work_time(env.now):
            yield env.timeout(_next_work_start(env.now) - env.now)
        res = wres.get('WORKER_OQC')
        req = res.request() if res else None
        acquired = False
        if req:
            yield req
            idle.acquire(env, 'WORKER_OQC')
            acquired = True
        try:
            yield from work_timeout(env, OQC_TIME_SEC)
            energy.record('OQC', 'INSP', OQC_TIME_SEC, env.now)
        finally:
            if req and res:
                if acquired:
                    idle.release(env, 'WORKER_OQC')
                    acquired = False
                res.release(req)
        stats['oqc_inspected'] = stats.get('oqc_inspected', 0) + 1

    done, total = progress[model_id]
    if done >= total:
        if menv is not None and hasattr(menv, 'wh'):
            menv.wh.unit_completions[(model_id, unit_id)] = {
                'path': 'normal_blocked_by_quota',
                'end_time': float(env.now),
                'done_n': len(kg_done),
                'total_n': len(kg_total),
                'rma_count': 0,
            }
        return
    stats[f'{model_id}_done'] = stats.get(f'{model_id}_done', 0) + 1
    progress[model_id] = (done + 1, total)
    pct = (done + 1) / total * 100
    if done + 1 >= total and menv is not None:
        for _pc, node in kg.nodes.items():
            wgrp = node.get('worker_group', '')
            if wgrp and wgrp not in menv.idle._completed_at:
                menv.idle.mark_completed(wgrp, float(env.now))
    if menv is not None and hasattr(menv, 'wh'):
        menv.wh.unit_completions[(model_id, unit_id)] = {
            'path': 'normal',
            'end_time': float(env.now),
            'done_n': len(done_set),
            'total_n': len(kg.nodes),
            'rma_count': 0,
        }


# ══════════════════════════════════════════════════════════
# M11. 콘솔 모니터
# ══════════════════════════════════════════════════════════

    """
    상태 : 시뮬레이션 스칼라 벡터 + GNN 그래프 임베딩
    행동 : 조립 공정 내 실행 가능 공정 중 우선순위 선택 (GNN 노드 스코어 기반)
          경험 (s, emb, a, r, lp, v, mask)은 produce_unit에서 store()로 직접 수집
    보상 : w1*(이번스텝시간감소) + w2*(-전력증가) - w3*WIP초과
           - w4*재고부족(critical_stock 기준) + w5*납기 + w6*(-작업자유휴)
    수렴 : 최근 CONV_WINDOW 에피소드 평균 보상 변화 < CONV_THRESHOLD
    """
    def __init__(self, data, order: dict):
        """
        data  : CombinedDataLoader 인스턴스 (FallbackDataLoader 도 호환)
        order : {model_id: qty} 주문 딕셔너리
        """
        self.data  = data
        self.order = order

    def _state_dim(self):
        n_models  = len(self.order)
        n_workers = len(self.data.workers)
        return n_models + n_workers + 6

    def _build_tensors(self, kg: ProcessKnowledgeGraph):
        _, H_np = kg.get_feat_matrix()
        adj_np  = kg.get_adj()
        return (torch.tensor(H_np, dtype=torch.float32),
                torch.tensor(adj_np, dtype=torch.float32))

    def run_ppo_training(self, max_episodes=5000):
        print(f'\n[PPO+GNN 학습] 상한 {max_episodes} 에피소드, 수렴 시 조기 종료')
        s_dim = self._state_dim()
        gnn = ProcessGNN(in_dim=6, hidden=32, out_dim=16)
        agent = PPOAgent(s_dim, gnn)

        if os.path.exists(POLICY_PATH):
            agent.load()
            start_ep = len(agent.ep_rewards)
            print(f'  이전 학습 이어서 진행 (이미 {start_ep} 에피소드 완료)')
        else:
            start_ep = 0

        CKPT_EVERY = 5
        train_t0 = time.time()
        ep = 0
        try:
            for ep in range(1, max_episodes + 1):
                ep_t0 = time.time()
                menv = ManufacturingEnv(self.data, self.order)
                menv._init_sim()
                menv.agent = agent

                menv.run(training=True)

                graphs_cache = {
                    m: self._build_tensors(kg)
                    for m, kg in menv.graphs.items()
                }
                ep_r = agent.update(graphs_cache=graphs_cache)

                decomp = list(getattr(menv, '_reward_decomp_sum', [0.0]*6))
                agent.ep_rewards_decomp.append(decomp)

                ep_dur   = time.time() - ep_t0
                elapsed  = time.time() - train_t0
                avg_dur  = elapsed / ep
                remain   = max_episodes - ep
                eta_min  = avg_dur * remain / 60
                avg_r    = float(np.mean(agent.ep_rewards[-min(100, len(agent.ep_rewards)):])) \
                           if agent.ep_rewards else 0.0
                done_now = sum(menv.stats.get(f'{m}_done', 0) for m in menv.order)
                tot_need = sum(menv.order.values())
                ms_h     = menv.env.now / 3600
                print(f'  ep {start_ep+ep:4d}/{start_ep+max_episodes:<4d} '
                      f'| sim={ms_h:6.2f}h done={done_now}/{tot_need} '
                      f'| R={ep_r:+8.3f} avg={avg_r:+8.3f} '
                      f'| {ep_dur:5.1f}s | ETA {eta_min:5.1f}m',
                      flush=True)

                if ep % CKPT_EVERY == 0:
                    agent.save(verbose=False)

                if agent.is_converged():
                    print(f'  수렴 (ep{start_ep + ep}). 학습 종료.')
                    break
        except KeyboardInterrupt:
            print(f'\n  [학습 중단됨] ep{start_ep + ep} 까지 완료. 현재 상태 저장.')

        agent.save()
        self._last_agent = agent
        total_min = (time.time() - train_t0) / 60
        print(f'\n[학습 완료] 총 {start_ep + ep} 에피소드 ({total_min:.1f}분)')
        return agent

    def run_inference(self, agent=None):
        """추론 실행.

        agent: run_ppo_training()이 반환한 PPOAgent 객체.
               None이면 ppo_policy.pt 파일에서 로드 시도.
               파일도 없으면 greedy(ready_pcs[0]) 모드로 실행.
        """
        print('\n[추론 실행]')
        menv = ManufacturingEnv(self.data, self.order)

        if agent is not None:
            agent.eval()
            menv.agent = agent
            self._last_agent = agent
            print('  학습된 정책 적용.')
        elif os.path.exists(POLICY_PATH):
            try:
                gnn   = ProcessGNN(in_dim=6, hidden=32, out_dim=16)
                s_dim = len(menv.get_state())
                loaded = PPOAgent(s_dim, gnn)
                loaded.load()
                loaded.eval()
                menv.agent = loaded
                self._last_agent = loaded
                print(f'  저장된 정책 로드: {POLICY_PATH}')
            except Exception as e:
                print(f'  [경고] 정책 로드 실패 -> greedy 모드: {e}')
                menv.agent = None
        else:
            print('  저장된 정책 없음 -> greedy 모드로 진행.')
            menv.agent = None

        summary = menv.run()
        self._last_menv = menv
        return summary

    def save_results(self, inference_summary=None, path=RESULT_PATH):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb       = openpyxl.Workbook()
        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_fill = PatternFill('solid', fgColor='2E4053')

        def _hdr(ws, headers):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font  = hdr_font
                c.fill  = hdr_fill
                c.alignment = Alignment(horizontal='center')

        def _aw(ws):
            for col in ws.columns:
                ml = max((len(str(c.value)) for c in col if c.value), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+2,30)

        ws = wb.active
        ws.title = 'Inference'
        if inference_summary:
            _hdr(ws, ['항목','값'])
            items = [
                ('makespan_hr',  inference_summary.get('makespan_hr', 0)),
                ('total_kwh',    inference_summary.get('total_kwh', 0)),
                ('재고품초과',   inference_summary.get('재고품초과', 0)),
                ('stock_pen',    inference_summary.get('stock_pen', 0)),
                ('total_defect', inference_summary.get('total_defect', 0)),
                ('total_done',   inference_summary.get('total_done', 0)),
                ('total_order',  inference_summary.get('total_order', 0)),
            ]
            for r, (k, v) in enumerate(items, 2):
                ws.cell(r, 1, k)
                ws.cell(r, 2, round(float(v),6) if isinstance(v, float) else v)
            row_i = len(items) + 3
            ws.cell(row_i, 1, '공정그룹')
            ws.cell(row_i, 2, 'kWh')
            for grp, kwh in inference_summary.get('by_grp_kwh', {}).items():
                row_i += 1
                ws.cell(row_i, 1, grp)
                ws.cell(row_i, 2, round(kwh, 6))
        _aw(ws)

        agent = getattr(self, '_last_agent', None)
        if agent and agent.ep_rewards:
            ws2 = wb.create_sheet('Training_Rewards')
            _hdr(ws2, ['episode','reward','rolling_avg_100'])
            for i, r in enumerate(agent.ep_rewards, 1):
                ws2.cell(i+1, 1, i)
                ws2.cell(i+1, 2, round(float(r), 6))
                if i >= 100:
                    ws2.cell(i+1, 3, round(float(np.mean(agent.ep_rewards[i-100:i])),6))
            _aw(ws2)




        # ══════════════════════════════════════════════════════════
        # ══════════════════════════════════════════════════════════
        menv = getattr(self, '_last_menv', None)
        if menv is not None:
            wh = menv.wh
            makespan_s = int(menv.env.now)
            max_hour = max(1, makespan_s // 3600 + 1)
            if max_hour > 720:
                print(f'[경고] makespan {max_hour}h > 720h - Stock_Timeseries 열 수가 많아 '
                      f'openpyxl 쓰기가 오래 걸릴 수 있습니다.')

            all_items = set()
            try:
                all_items = wh.data.iter_all_bom_items()
            except AttributeError:
                pass
            all_items.update(wh.stock.keys())
            all_items.update(wh.snapshots.keys())
            all_items = sorted(all_items)

            ws_s = wb.create_sheet('Stock_Summary')
            _hdr(ws_s, ['item_code', 'item_name', 'initial_stock',
                        'total_consumed', 'final_stock', 'min_stock_qty',
                        'lot_size', 'violations_count', 'reorder_count'])
            r = 2
            for code in all_items:
                try:
                    name = wh.data.get_item_name(code)
                except AttributeError:
                    name = ''
                ws_s.cell(r, 1, code)
                ws_s.cell(r, 2, name)
                init_val = wh._initial_stocks.get(code,
                    wh._bom_init_stock if code in wh._bom_codes else wh._init_stock)
                ws_s.cell(r, 3, int(init_val))
                ws_s.cell(r, 4, int(wh.consumed.get(code, 0)))
                ws_s.cell(r, 5, int(wh.stock.get(code, init_val)))
                try:
                    ws_s.cell(r, 6, float(wh.data.get_min_stock(code)))
                except Exception:
                    ws_s.cell(r, 6, MIN_STOCK)
                try:
                    ws_s.cell(r, 7, int(wh._lot_for(code)))
                except Exception:
                    ws_s.cell(r, 7, 0)
                ws_s.cell(r, 8, int(wh.violations.get(code, 0)))
                ws_s.cell(r, 9, int(wh.reorder_count.get(code, 0)))
                r += 1
            _aw(ws_s)

            ws_t = wb.create_sheet('Stock_Timeseries')
            header = ['item_code', 'item_name'] + [f't={h}h' for h in range(max_hour + 1)]
            _hdr(ws_t, header)
            r = 2
            for code in all_items:
                try:
                    name = wh.data.get_item_name(code)
                except AttributeError:
                    name = ''
                ws_t.cell(r, 1, code)
                ws_t.cell(r, 2, name)
                by_hour = {}
                for t_sec, q in wh.snapshots.get(code, []):
                    by_hour[int(t_sec // 3600)] = int(q)
                prev = int(wh._initial_stocks.get(code,
                    wh._bom_init_stock if code in wh._bom_codes else wh._init_stock))
                for h in range(max_hour + 1):
                    if h in by_hour:
                        prev = by_hour[h]
                    ws_t.cell(r, 3 + h, prev)
                r += 1

            ws_e = wb.create_sheet('Stock_Events')
            _hdr(ws_e, ['time_sec', 'time_hr', 'item_code', 'stock_after'])
            r = 2
            EVENT_CAP = 60000
            count = 0
            stop = False
            for code in all_items:
                if stop:
                    break
                for t_sec, q in wh.history.get(code, []):
                    ws_e.cell(r, 1, float(t_sec))
                    ws_e.cell(r, 2, round(float(t_sec) / 3600, 3))
                    ws_e.cell(r, 3, code)
                    ws_e.cell(r, 4, int(q))
                    r += 1
                    count += 1
                    if count >= EVENT_CAP:
                        ws_e.cell(r, 1, '[TRUNCATED]')
                        ws_e.cell(r, 3, f'events > {EVENT_CAP}')
                        stop = True
                        break
            _aw(ws_e)

            ws_p = wb.create_sheet('Process_Log')
            pheader = ['process_code'] + [f't={h}h' for h in range(max_hour + 1)]
            _hdr(ws_p, pheader)
            r = 2
            _aw(ws_p)

            ws_r = wb.create_sheet('Reorder_Log')
            _hdr(ws_r, ['item_code', 'order_time_hr', 'arrive_time_hr',
                        'lead_hr', 'lot_size', 'incoming', 'stock_at_order'])
            r = 2
            for entry in wh.reorder_log:
                ot = float(entry.get('order_time', 0)) / 3600
                at = float(entry.get('arrive_time', 0)) / 3600
                ws_r.cell(r, 1, entry.get('item_code', ''))
                ws_r.cell(r, 2, round(ot, 3))
                ws_r.cell(r, 3, round(at, 3))
                ws_r.cell(r, 4, round(at - ot, 3))
                ws_r.cell(r, 5, int(entry.get('lot_size', 0)))
                ws_r.cell(r, 6, int(entry.get('incoming', 0)))
                ws_r.cell(r, 7, int(entry.get('stock_at_order', 0)))
                r += 1
            _aw(ws_r)

            print(f'  재고·공정 시트 5개 추가 (items={len(all_items)}, '
                  f'events={count}, reorders={len(wh.reorder_log)})')

            # ══════════════════════════════════════════════════════
            # ══════════════════════════════════════════════════════

            ws_dms = wb.create_sheet('Debug_Model_Stats')
            _hdr(ws_dms, ['model_id', 'order_qty', 'stats_done',
                          'normal_completions', 'rma_completions',
                          'blocked_by_quota', 'first_event_h',
                          'last_event_h', 'first_after_24h',
                          'first_after_48h'])
            comps = wh.unit_completions
            r = 2
            for model in menv.order:
                qty = menv.order[model]
                stats_done = int(menv.stats.get(f'{model}_done', 0))
                model_comps = [v for k, v in comps.items()
                               if isinstance(k, tuple) and k[0] == model]
                n_normal = sum(1 for c in model_comps if c['path'] == 'normal')
                n_rma    = sum(1 for c in model_comps if c['path'] == 'rma')
                n_blocked = sum(1 for c in model_comps
                                if 'blocked_by_quota' in c['path'])
                m_evs = [e for e in events_all if e.get('mid') == model]
                first_h = round(min((e['start'] for e in m_evs), default=0)/3600, 2)
                last_h  = round(max((e['end']   for e in m_evs), default=0)/3600, 2)
                after_24 = round(min((e['start'] for e in m_evs
                                      if e['start'] >= 24*3600), default=0)/3600, 2)
                after_48 = round(min((e['start'] for e in m_evs
                                      if e['start'] >= 48*3600), default=0)/3600, 2)
                ws_dms.cell(r, 1, model)
                ws_dms.cell(r, 2, qty)
                ws_dms.cell(r, 3, stats_done)
                ws_dms.cell(r, 4, n_normal)
                ws_dms.cell(r, 5, n_rma)
                ws_dms.cell(r, 6, n_blocked)
                ws_dms.cell(r, 7, first_h)
                ws_dms.cell(r, 8, last_h)
                ws_dms.cell(r, 9, after_24)
                ws_dms.cell(r, 10, after_48)
                r += 1
            _aw(ws_dms)

            ws_dus = wb.create_sheet('Debug_Unit_Status')
            _hdr(ws_dus, ['model_id', 'unit_key', 'completion_path',
                          'end_time_h', 'done_n', 'total_n',
                          'live_state', 'live_pc', 'live_done_n',
                          'live_total_n'])
            r = 2
            live_us = getattr(menv, 'unit_states', {}) or {}
            seen_keys = set()
            for k, c in sorted(comps.items(),
                                key=lambda kv: (str(kv[0][0]), str(kv[0][1]))):
                model = k[0]
                uid_or_tag = k[1]
                ws_dus.cell(r, 1, str(model))
                ws_dus.cell(r, 2, str(uid_or_tag))
                ws_dus.cell(r, 3, c.get('path', ''))
                ws_dus.cell(r, 4, round(c.get('end_time', 0)/3600, 3))
                ws_dus.cell(r, 5, int(c.get('done_n', -1)))
                ws_dus.cell(r, 6, int(c.get('total_n', -1)))
                if isinstance(uid_or_tag, int):
                    ls = live_us.get((model, uid_or_tag), {})
                    ws_dus.cell(r, 7, str(ls.get('state', '')))
                    ws_dus.cell(r, 8, str(ls.get('pc', '')))
                    ws_dus.cell(r, 9, int(ls.get('done_n', 0)))
                    ws_dus.cell(r, 10, int(ls.get('total_n', 0)))
                    seen_keys.add((model, uid_or_tag))
                r += 1
            for (model, uid), ls in sorted(live_us.items(),
                                           key=lambda kv: (str(kv[0][0]), kv[0][1])):
                if (model, uid) in seen_keys:
                    continue
                ws_dus.cell(r, 1, str(model))
                ws_dus.cell(r, 2, str(uid))
                ws_dus.cell(r, 3, 'cutoff')
                ws_dus.cell(r, 4, round(menv.env.now/3600, 3))
                ws_dus.cell(r, 5, int(ls.get('done_n', 0)))
                ws_dus.cell(r, 6, int(ls.get('total_n', 0)))
                ws_dus.cell(r, 7, str(ls.get('state', '')))
                ws_dus.cell(r, 8, str(ls.get('pc', '')))
                ws_dus.cell(r, 9, int(ls.get('done_n', 0)))
                ws_dus.cell(r, 10, int(ls.get('total_n', 0)))
                r += 1
            _aw(ws_dus)

            ws_dpc = wb.create_sheet('Debug_Process_Coverage')
            _hdr(ws_dpc, ['model_id', 'process_code', 'process_group',
                          'worker_group', 'expected_qty', 'actual_count',
                          'first_h', 'last_h'])
            r = 2
            for model in menv.order:
                qty = menv.order[model]
                kg_nodes = menv.graphs[model].nodes if model in menv.graphs else {}
                for pc in sorted(kg_nodes.keys()):
                    node = kg_nodes[pc]
                    pc_evs = [e for e in events_all
                              if e.get('mid') == model and e.get('pc') == pc]
                    cnt = len(pc_evs)
                    f_h = round(min((e['start'] for e in pc_evs),
                                    default=0)/3600, 2) if pc_evs else ''
                    l_h = round(max((e['end']   for e in pc_evs),
                                    default=0)/3600, 2) if pc_evs else ''
                    ws_dpc.cell(r, 1, model)
                    ws_dpc.cell(r, 2, pc)
                    ws_dpc.cell(r, 3, node.get('process_group', ''))
                    ws_dpc.cell(r, 4, node.get('worker_group', ''))
                    ws_dpc.cell(r, 5, qty)
                    ws_dpc.cell(r, 6, cnt)
                    ws_dpc.cell(r, 7, f_h)
                    ws_dpc.cell(r, 8, l_h)
                    r += 1
            _aw(ws_dpc)

            ws_dpf = wb.create_sheet('Debug_PCB_Flow')
            _hdr(ws_dpf, ['pcb_code', 'role', 'model_hint',
                          'initial_stock', 'final_stock', 'total_consumed',
                          'smt_or_outsource_restore',
                          'outsource_in', 'outsource_returned',
                          'external_order_trigger', 'external_replenish_arrived',
                          'is_bug_candidate'])
            r = 2
            model_for_pcb = {}
            for m, pc_main in self.data.pcb_map.items():
                model_for_pcb[pc_main] = (m, 'main')
            for m, ths in self.data.tht_pcb_by_model.items():
                for pc_t in ths:
                    model_for_pcb[pc_t] = (m, 'tht')
            for code in sorted(wh._pcb_codes):
                role = model_for_pcb.get(code, ('-', '-'))
                flow = wh.pcb_flow.get(code, {})
                ext_trig = int(flow.get('external_order_trigger', 0))
                ext_arr  = int(flow.get('external_replenish_arrived', 0))
                bug = ext_trig > 0 or ext_arr > 0
                ws_dpf.cell(r, 1, code)
                ws_dpf.cell(r, 2, role[1])
                ws_dpf.cell(r, 3, role[0])
                ws_dpf.cell(r, 4, int(wh._initial_stocks.get(code,
                    wh._bom_init_stock if code in wh._bom_codes else wh._init_stock)))
                ws_dpf.cell(r, 5, int(wh.stock.get(code, 0)))
                ws_dpf.cell(r, 6, int(wh.consumed.get(code, 0)))
                ws_dpf.cell(r, 7, int(flow.get('restore_from_smt_or_outsource', 0)))
                ws_dpf.cell(r, 8, int(flow.get('outsource_in', 0)))
                ws_dpf.cell(r, 9, int(flow.get('outsource_returned', 0)))
                ws_dpf.cell(r, 10, ext_trig)
                ws_dpf.cell(r, 11, ext_arr)
                ws_dpf.cell(r, 12, 'BUG' if bug else '')
                r += 1
            r += 1
            ws_dpf.cell(r, 1, '== SMT 처리량 (line × model × pcb) ==')
            r += 1
            ws_dpf.cell(r, 1, 'line')
            ws_dpf.cell(r, 2, 'model')
            ws_dpf.cell(r, 3, 'pcb_code')
            ws_dpf.cell(r, 4, 'restored_qty')
            r += 1
            for (sid, m, pc_code), q in sorted(wh.smt_per_model.items()):
                ws_dpf.cell(r, 1, sid)
                ws_dpf.cell(r, 2, m)
                ws_dpf.cell(r, 3, pc_code)
                ws_dpf.cell(r, 4, int(q))
                r += 1
            _aw(ws_dpf)

            ws_dol = wb.create_sheet('Debug_Outsource_Log')
            _hdr(ws_dol, ['pcb_code', 'model_id', 'board_id',
                          'send_time_h', 'return_time_h', 'transit_h',
                          'delay_h', 'status'])
            r = 2
            for entry in wh.outsource_log:
                ret = entry.get('return_time')
                send_h = round(entry.get('send_time', 0)/3600, 3)
                ret_h  = round(ret/3600, 3) if ret is not None else ''
                transit = round((ret - entry.get('send_time', 0))/3600, 3) \
                          if ret is not None else ''
                ws_dol.cell(r, 1, entry.get('pcb_code', ''))
                ws_dol.cell(r, 2, entry.get('model_id', ''))
                ws_dol.cell(r, 3, entry.get('board_id', ''))
                ws_dol.cell(r, 4, send_h)
                ws_dol.cell(r, 5, ret_h)
                ws_dol.cell(r, 6, transit)
                ws_dol.cell(r, 7, round(entry.get('delay_sec', 0)/3600, 3))
                ws_dol.cell(r, 8, entry.get('status', ''))
                r += 1
            _aw(ws_dol)

            ws_dpr = wb.create_sheet('Debug_PCB_Reorders')
            _hdr(ws_dpr, ['item_code', 'order_h', 'arrive_h',
                          'lot_size', 'incoming', 'stock_at_order'])
            r = 2
            for entry in wh.reorder_log:
                if not entry.get('is_pcb'):
                    continue
                ws_dpr.cell(r, 1, entry.get('item_code', ''))
                ws_dpr.cell(r, 2, round(entry.get('order_time', 0)/3600, 3))
                ws_dpr.cell(r, 3, round(entry.get('arrive_time', 0)/3600, 3))
                ws_dpr.cell(r, 4, int(entry.get('lot_size', 0)))
                ws_dpr.cell(r, 5, int(entry.get('incoming', 0)))
                ws_dpr.cell(r, 6, int(entry.get('stock_at_order', 0)))
                r += 1
            _aw(ws_dpr)

            ws_dsk = wb.create_sheet('Debug_Skipped_PCs')
            _hdr(ws_dsk, ['model_id', 'process_code', 'skip_count',
                          'in_excel_pf', 'in_kg_nodes'])
            r = 2
            for (m, pc), c in sorted(wh.skipped_pcs.items()):
                in_pf = wh.data._pc_map.get(pc) is not None
                in_kg = (m in menv.graphs and
                         pc in menv.graphs[m].nodes)
                ws_dsk.cell(r, 1, m)
                ws_dsk.cell(r, 2, pc)
                ws_dsk.cell(r, 3, int(c))
                ws_dsk.cell(r, 4, 'Y' if in_pf else 'N')
                ws_dsk.cell(r, 5, 'Y' if in_kg else 'N')
                r += 1
            _aw(ws_dsk)

            ws_dsm = wb.create_sheet('Debug_SMT_Choices')
            _hdr(ws_dsm, ['time_h', 'chosen_model'])
            r = 2
            for (t, m) in wh.smt_model_choices:
                ws_dsm.cell(r, 1, round(t/3600, 3))
                ws_dsm.cell(r, 2, m)
                r += 1
            _aw(ws_dsm)

            ws_dsa = wb.create_sheet('Debug_Safety_Alarms')
            _hdr(ws_dsa, ['alarm_type', 'detail_1', 'detail_2',
                          'detail_3', 'time_h', 'extra'])
            r = 2
            for entry in wh.kg_incomplete_log:
                ws_dsa.cell(r, 1, 'B1_kg_incomplete')
                ws_dsa.cell(r, 2, entry['model_id'])
                ws_dsa.cell(r, 3, str(entry['unit_id']))
                ws_dsa.cell(r, 4, '')
                ws_dsa.cell(r, 5, round(entry['time_h'], 3))
                ws_dsa.cell(r, 6, ','.join(entry['missing_pcs']))
                r += 1
            for entry in wh.smt_single_side_log:
                ws_dsa.cell(r, 1, 'B5_single_side_double_pcb')
                ws_dsa.cell(r, 2, entry['pcb_code'])
                ws_dsa.cell(r, 3, entry['model_id'])
                ws_dsa.cell(r, 4, str(entry['board_id']))
                ws_dsa.cell(r, 5, round(entry['time_h'], 3))
                ws_dsa.cell(r, 6, entry.get('reason', ''))
                r += 1
            dup_log = getattr(wh.data, '_bom_dup_merge_log', [])
            for msg in dup_log:
                ws_dsa.cell(r, 1, 'B4_bom_smt_side_conflict')
                ws_dsa.cell(r, 2, str(msg))
                r += 1
            for entry in wh.stuck_wait_log:
                ws_dsa.cell(r, 1, 'wait_stock_timeout_fallback')
                ws_dsa.cell(r, 2, entry['item_code'])
                ws_dsa.cell(r, 3, str(entry['qty']))
                ws_dsa.cell(r, 4, f'stock={entry["stock_at_end"]}')
                ws_dsa.cell(r, 5, round(entry['wait_end_h'], 3))
                ws_dsa.cell(r, 6, f'wait_for_{entry["wait_end_h"]-entry["wait_start_h"]:.2f}h')
                r += 1
            _aw(ws_dsa)

            ws_dpe = wb.create_sheet('Debug_Plogger_Events')
            _hdr(ws_dpe, ['pc', 'mid', 'uid', 'grp', 'wgrp', 'slot',
                          'start_h', 'end_h', 'dur_s', 'work_timed'])
            r = 2
            for e in evs:
                if e.get('grp') != 'PACK':
                    continue
                ws_dpe.cell(r, 1, str(e.get('pc', '')))
                ws_dpe.cell(r, 2, str(e.get('mid', '')))
                ws_dpe.cell(r, 3, int(e.get('uid', 0)))
                ws_dpe.cell(r, 4, str(e.get('grp', '')))
                ws_dpe.cell(r, 5, str(e.get('wgrp', '')))
                ws_dpe.cell(r, 6, int(e.get('slot', -1)))
                ws_dpe.cell(r, 7, round(float(e.get('start', 0))/3600, 3))
                ws_dpe.cell(r, 8, round(float(e.get('end', 0))/3600, 3))
                ws_dpe.cell(r, 9, round(float(e.get('end', 0))-float(e.get('start', 0)), 2))
                ws_dpe.cell(r, 10, str(e.get('work_timed', False)))
                r += 1
            _aw(ws_dpe)

            pack_no_wgrp = sum(1 for e in evs
                               if e.get('grp') == 'PACK' and not e.get('wgrp'))
            pack_bad_slot = sum(1 for e in evs
                                if e.get('grp') == 'PACK' and e.get('slot', -1) < 0)
            from collections import Counter
            pack_per_model = Counter(e.get('mid') for e in evs
                                     if e.get('grp') == 'PACK')

            print(f'  디버그 시트 10개 추가 (unit_comps={len(comps)}, '
                  f'outsource_evs={len(wh.outsource_log)}, '
                  f'skipped={sum(wh.skipped_pcs.values())}, '
                  f'smt_choices={len(wh.smt_model_choices)}, '
                  f'kg_incomplete={len(wh.kg_incomplete_log)}, '
                  f'single_side_alarm={len(wh.smt_single_side_log)})')
            print(f'  [PACK 이벤트 진단] no_wgrp={pack_no_wgrp}, bad_slot={pack_bad_slot}, '
                  f'per_model={dict(pack_per_model)}')

            wip = menv.wip
            ws_wt = wb.create_sheet('WIP_Timeseries')
            grps = ['SMT', 'MODULE', 'SEMI', 'SET', 'INSP', 'PACK', 'RMA']
            _hdr(ws_wt, ['hour'] + grps)
            hours = sorted({int(t // 3600) for g in grps
                            for t, _ in wip.snapshots.get(g, [])})
            lookup = {g: {int(t // 3600): n for t, n in wip.snapshots.get(g, [])}
                      for g in grps}
            r = 2
            for h in hours:
                ws_wt.cell(r, 1, h)
                for ci, g in enumerate(grps, 2):
                    ws_wt.cell(r, ci, lookup[g].get(h, 0))
                r += 1
            _aw(ws_wt)

            pool = getattr(menv, 'outsource_pool', None)
            ws_tl = wb.create_sheet('Truck_Log')
            _hdr(ws_tl, ['dispatch_id', 'send_h', 'eta_h', 'return_h',
                         'transit_h', 'delay_h', 'boards', 'truck_count_eq',
                         'pcb_breakdown'])
            r = 2
            for entry in (pool.truck_log if pool else []):
                ret_h = (entry['return_t'] / 3600) if entry.get('return_t') else None
                send_h = entry['send_t'] / 3600
                eta_h  = entry['eta'] / 3600
                bk = ', '.join(f"{c[-4:]}×{n}"
                               for c, n in entry['breakdown'].items())
                ws_tl.cell(r, 1, entry['dispatch_id'])
                ws_tl.cell(r, 2, round(send_h, 3))
                ws_tl.cell(r, 3, round(eta_h, 3))
                ws_tl.cell(r, 4, round(ret_h, 3) if ret_h else '')
                ws_tl.cell(r, 5, round((ret_h - send_h), 3) if ret_h else '')
                ws_tl.cell(r, 6, round(entry['delay_sec'] / 3600, 3))
                ws_tl.cell(r, 7, entry['size'])
                ws_tl.cell(r, 8, entry['truck_count'])
                ws_tl.cell(r, 9, bk)
                r += 1
            _aw(ws_tl)


        wb.save(path)
        print(f'결과 저장: {path}')

def main():
    if os.name == 'nt':
        os.system('')

    random.seed(RANDOM_SEED)
    np.random.seed(RANDOM_SEED)
    torch.manual_seed(RANDOM_SEED)

    print('=== CPRO 제조 공정 시뮬레이션 (재고·공정 추적판) ===')

    static_data = FallbackDataLoader()
    print(f'  [Fallback] SMT/RMA PROCESS_FLOW:{len(static_data.pf)}공정 | '
          f'MTTR 설비:{len(static_data._mttr)}개')

    aas_models = {}
    for model_id, json_path in AAS_JSON_PATHS.items():
        if not os.path.exists(json_path):
            print(f'  [AAS]  {model_id}: JSON 없음 — 시뮬에서 제외')
            continue
        aas = load_aas(model_id, json_path)
        if aas.ManufacturingProcess:
            aas_models[model_id] = aas
            print(f'  [AAS]  {model_id}: {len(aas.ManufacturingProcess)}공정 | '
                  f'InputBOM:{sum(len(n.InputBOM) for n in aas.ManufacturingProcess.values())}건')
        else:
            print(f'  [AAS]  {model_id}: ManufacturingProcess 미파싱 — 시뮬에서 제외')

    data = CombinedDataLoader(static_data, aas_models)
    if not data.schedule:
        raise RuntimeError(
            'AAS 가 근무 스케줄을 제공하지 않음. '
            '적어도 1개 모델의 AAS JSON 이 WorkstationWorkerMatchingData 를 포함해야 함.')
    _apply_schedule(data.schedule)
    print(f'  [통합] 총 process_code:{len(data._pc_map)}개 | '
          f'workers:{len(data.workers)}그룹 | '
          f'schedule: {data.schedule.get("work_start_sec",0)//3600:02d}h~'
          f'{data.schedule.get("work_end_sec",0)//3600:02d}h '
          f'(break {data.schedule.get("break_duration_sec",0)//60}min)')

    order = {}
    if not aas_loaders:
        print('  [경고] AAS 로드된 모델이 없어 주문 입력을 건너뜁니다.')
    for m in aas_loaders:
        while True:
            try:
                qty = int(input(f'{m} 주문 수량: '))
                if qty > 0:
                    order[m] = qty
                    break
                print('  1 이상의 정수를 입력하세요.')
            except ValueError:
                print('  숫자를 입력하세요.')

    max_eps = 5000
    while True:
        s = input('학습 에피소드 최대 개수 (엔터=5000, 0=학습건너뛰기): ').strip()
        if not s:
            max_eps = 5000
            break
        try:
            v = int(s)
            if v >= 0:
                max_eps = v
                break
            print('  0 이상의 정수를 입력하세요.')
        except ValueError:
            print('  숫자를 입력하세요.')

    runner = ExperimentRunner(data, order)
    if max_eps == 0:
        print('\n[학습 생략] 기존 ppo_policy.pt 가 있으면 그대로 추론에 사용합니다.')
        agent = None
    else:
        agent = runner.run_ppo_training(max_episodes=max_eps)

    if agent is not None:
        print(f'\n[학습 완료 요약]')
        print(f'  총 에피소드     : {len(agent.ep_rewards)}')
        if agent.ep_rewards:
            print(f'  최종 보상 평균  : {float(np.mean(agent.ep_rewards[-100:])):.4f}')
            print(f'  최고 에피소드 보상: {max(agent.ep_rewards):.4f}')

    s = runner.run_inference(agent=agent)
    runner.save_results(inference_summary=s)


if __name__ == '__main__':
    main()