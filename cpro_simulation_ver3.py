# cpro_simulation_ver2.py
#
# [모듈 구성 — 4 섹터 + sub-section]
#
# 향후 패키지 분할 시 각 섹터가 cpro/{data_loaders, domain, simulation, rl}.py
# 또는 그 하위 패키지로 그대로 떨어지도록 정렬했다. 단일 파일 유지는 의도.
#
#   §A. DATA LOADERS  — 정적/AAS 데이터 → CombinedDataLoader
#       M01 config         : 전역 상수 및 경로
#       M02 fallback       : FallbackDataLoader (SMT + RMA + MTTR 정적 표)
#       M02b aas           : AASJsonLoader (모델별 조립~포장 공정)
#       M02c combined      : CombinedDataLoader (두 로더 통합 인터페이스)
#
#   §B. DOMAIN        — 시뮬 상태 컨테이너 (도메인 객체)
#       M03 graph          : ProcessKnowledgeGraph (공정 DAG, GNN 입력)
#       M03b ready_gate    : is_process_ready (단일 게이트 — PR1 행동보존)
#       M04 warehouse·wip  : Warehouse + WIPTracker
#       M05 energy         : EnergyLogger
#       M06 idle           : IdleTracker (유휴·숙련도)
#       M07 smt            : SolderCream + OutsourceTruckPool + SMTLine
#
#   §C. SIMULATION    — SimPy 프로세스·이벤트 루프
#       schedule helpers   : _is_work_time / work_timeout / _work_seconds_between
#       M08 run_process    : 단일 공정 실행
#       M09 run_rma        : RMA 수리·재투입
#       M10 produce_unit   : 단일 제품 오케스트레이션 (PPO 액션 진입점)
#       M11 monitor        : ProcessActivityLogger + 콘솔 모니터
#
#   §D. RL            — GNN + PPO + 환경 + 실험 드라이버
#       M12 gnn            : ProcessGNN
#       M13 ppo            : PPOAgent
#       M14 env            : ManufacturingEnv (RL 환경)
#       M15 runner         : ExperimentRunner
#       M16 main           : 진입점 (main)
#
# [데이터 소스 분리]
#   SMT 라인 + RMA  : 정적 하드코딩 (FallbackDataLoader)
#   조립~포장 공정  : MODEL_X.json (AAS) — AASJsonLoader
#   통합 인터페이스 : CombinedDataLoader (두 로더를 합산)
#
# [강화학습 구조]
#   상태  : 완성률, 작업자가동률, kWh율, 경과시간비율,
#           재고부족율, 재고품초과율, 작업자유휴율,
#           SMT설비고장율
#   행동  : 조립 공정 내 실행 가능 공정 우선순위 (GNN 노드 스코어 기반)
#           ready 조건 = 선행공정 완료 AND 필요 부품 재고 확보
#   보상  : w1*(이번스텝시간감소) + w2*(-전력증가) - w3*재고품초과
#           - w4*재고부족(CRITICAL_STOCK 하한 기준) + w5*납기 + w6*(-작업자유휴)
#   수렴  : 최근 100 에피소드 평균 보상 변화 < CONV_THRESHOLD
#
# [근로 시간]
#   AAS JSON WorkstationWorkerMatchingData 에서 로드.
#   WorkStartTime / WorkEndTime / BreakDurationMin 기준.
#   점심 시작 시각은 JSON 에 없으므로 fallback 12시00분 유지,
#   lunch_end = lunch_start + break_duration 으로 역산.
#
# [동적 이벤트]
#   - SMT 설비 고장 (포아송 확률)
#   - 작업자 결근 (일별 확률)
#   - 외주 납기 지연 (THT 외주 시)
#
# [참고 문헌]
#   Schulman et al. (2017) PPO
#   Luo et al. (2023 NeurIPS) GNN+스케줄링
#   Sun et al. (2025 Engineering Vol.46) PPO_S 수렴 ~4000 에피소드
#   Wang & Liao (2024 J.Intell.Manuf.) SimPy+PPO
#   Ekerete et al. (2026 IRE Journals) WIP 재고 관리

import simpy
import random
import re
import os
import sys
import time
import json
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.nn.functional as F
from collections import defaultdict
from dataclasses import dataclass
from enum import Enum
# 엑셀 시트 의존 제거 (2026-05-06): FallbackDataLoader 가 정적 데이터를 보유.
# openpyxl 자체는 결과 저장(save_results)에서만 import 해서 사용한다.


# ████████████████████████████████████████████████████████████████████
# §A. DATA LOADERS
# ████████████████████████████████████████████████████████████████████
# 책임: 정적 표(SMT/RMA/MTTR) + AAS JSON(조립~포장) → 통합 인터페이스 제공.
# 외부 export: CombinedDataLoader, FallbackDataLoader, AASJsonLoader,
#              _apply_schedule, _active_schedule, AAS_JSON_PATHS, RANDOM_SEED.
# 향후 파일 분할 위치: cpro/data/{config, fallback, aas, combined}.py


# ══════════════════════════════════════════════════════════
# M01. 전역 상수 / 정책값 / AAS 미보유 데이터
# ══════════════════════════════════════════════════════════
# 본 섹션 항목은 모두 "AAS JSON 이 직접 제공하지 않는" 데이터.
# 출처별 그룹:
#   §0  시스템 / 환경
#   §1  근무 스케줄        — AAS 가 work_start/end/break_duration 제공.
#                            점심 시작 시각만 AAS 미보유 → 12시 hardcoded.
#   §2  PCB·모델 메타정보  — AAS HS 에 PCB Entity 들 있지만 main vs THT 구분 X.
#   §3  설비 / 시설 메타  — RATED_POWER_KW, SMT_LINE_IDS.
#   §4  물류·물성 단위    — PCB_PER_UNIT, MAG_SIZE, TRUCK_SIZE, SOLDER_*.
#   §5  시뮬 정책값       — 발주·재고·OQC·AOI·RMA. 사용자 튜닝 가능.
#   §6  통계·이벤트 추정  — 고장·결근·외주지연 확률.
#   §7  Worker fallback   — AAS 8그룹 제공. RMA 만 미보유 + SET_INSP_HEADCOUNT.
#   §8  매핑·UI            — process_group→worker, LOCATION_*, 모니터.
#   §9  파생값·헬퍼함수    — 위 상수에서 파생.

# ── §0 시스템 / 환경 ─────────────────────────────────────
try:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    BASE_DIR = os.getcwd()

POLICY_PATH  = os.path.join(BASE_DIR, 'ppo_policy.pt')
RESULT_PATH  = os.path.join(BASE_DIR, 'simulation_results.xlsx')
# (2026-05-06) KETI_PATH 제거 — 엑셀 시트 의존을 FallbackDataLoader 로 흡수.

# AAS JSON 경로 — 모델별 조립~포장 공정 데이터
AAS_JSON_PATHS = {
    'MODEL_A': os.path.join(BASE_DIR, 'MODEL_A.json'),
    'MODEL_B': os.path.join(BASE_DIR, 'MODEL_B.json'),
    'MODEL_C': os.path.join(BASE_DIR, 'MODEL_C.json'),
}

RANDOM_SEED = 42
DAY_SEC     = 24 * 3600
MAX_DAYS    = 365   # sim 강제 종료 상한 (§5 정책값보다 위 — DAY_SEC 동급)

# ── §1 근무 스케줄 ─────────────────────────────────────
# AAS WorkstationWorkerMatchingData 가 work_start/end/break_duration 단일 출처.
# ⚠ break(점심)의 *시작 시각* 만 AAS 미보유 → 12:00 hardcoded.
# main() 의 AAS 로드 후 _apply_schedule() 호출 필수. AAS 미제공 시 RuntimeError.
_active_schedule = {
    'lunch_start_sec': 12 * 3600,   # ⚠ AAS 미보유 — hardcoded
    # work_start_sec / work_end_sec / break_duration_sec / lunch_end_sec 는
    # _apply_schedule() 호출 시 채워짐 (AAS 가 단일 출처).
}

def _apply_schedule(schedule_dict: dict):
    """AAS schedule 적용 + lunch_end_sec 역산.
    AAS 가 work_start/end/break_duration 미제공 시 RuntimeError.
    """
    _active_schedule.update(schedule_dict)
    required = ('work_start_sec', 'work_end_sec', 'break_duration_sec')
    missing = [k for k in required if k not in _active_schedule]
    if missing:
        raise RuntimeError(
            f'AAS 가 근무 스케줄 미제공: {missing}. '
            f'적어도 1개 모델 AAS JSON 이 WorkstationWorkerMatchingData '
            f'(WorkStartTime/WorkEndTime/BreakDurationMin) 를 제공해야 함.')
    _active_schedule['lunch_end_sec'] = (
        _active_schedule['lunch_start_sec']
        + _active_schedule['break_duration_sec']
    )

# ── §4 물류·물성 단위 (AAS 미보유) ────────────────────
PCB_PER_UNIT       = 1            # 유닛 1개당 PCB 1장
MAG_SIZE           = 15           # 비-THT PCB 매거진 batching 단위
TRUCK_SIZE         = 30           # THT 외주 트럭 적재 환산 (모니터링용)

SOLDER_G           = 500          # 솔더크림 1통 g
SOLDER_VALID_SEC   = 24 * 3600    # 솔더크림 사용 유효시간
SOLDER_USE_G       = 0.07         # 1 cycle 당 사용량 (도메인 추정)

CT_STD_RATIO       = 0.10         # cycle_time 표준편차 비율 (정규분포 σ/μ)
DEFECT_FLOOR       = 0.00001      # 미정의 공정 최소 불량률

# ── §5 시뮬 정책값 (AAS 미보유 — 사용자 튜닝 가능) ────
# 발주·재고
MIN_STOCK                  = 100  # 발주 트리거 임계
CRITICAL_STOCK             = 5    # 페널티 임계 (min_stock 과 분리)
REPLENISH_LEAD_DAY         = 1    # 발주 후 입고 소요 일수
REPLENISH_QTY_MULT         = 10   # (deprecated) 균일 lot 배수
WIP_CAP_RATIO              = 1.5

# OQC 표본 추출 정책: §G PROCESS_GROUP_SPECS['OQC'] 가 단일 출처
# (sampling_rate / sample_time_sec). 호환용 OQC_RATE / OQC_TIME_SEC 는
# §G-derive 에서 자동 생성.

# AOI 결함 처리: 'repair' (RMA 후 PCB 재투입) / 'scrap' (폐기)
# AOI 는 SMT 라인 stage 라 process_group 단위 아님 — 별도 정책 상수.
AOI_DEFECT_ACTION = 'repair'

# RMA 수리 시간 (현장 추정 — 정규분포)
RMA_REPAIR_TIME_MEAN_SEC = 300
RMA_REPAIR_TIME_STD_SEC  = 60
RMA_REPAIR_TIME_MIN_SEC  = 60

# ── §6 통계·이벤트 추정 (AAS 미보유) ────────────────────
SMT_BREAKDOWN_PROB  = 0.000005    # SMT 설비 1대당 10분 check 시 고장 확률
SMT_MTTR_DEFAULT_HR = 0.5         # MTTR 기본 (RESOURCE 미제공 시)
WORKER_ABSENT_PROB  = 0.0005      # 그룹별 매 근무일 결근 발생 확률
THT_DELAY_PROB      = 0.02        # 외주 트럭 납기 지연 확률
THT_DELAY_MIN_SEC   = 1  * 3600
THT_DELAY_MAX_SEC   = 24 * 3600

# ── 모니터 출력 간격 ──────────────────────────────────
TRAIN_MONITOR_INTERVAL = DAY_SEC                            # 학습 중 1일 단위
INFER_MONITOR_STEP_HR  = 0.1                                # 추론 1 frame 진행 시뮬 시간 (h)
INFER_MONITOR_INTERVAL = int(INFER_MONITOR_STEP_HR * 3600)  # 초 환산
MONITOR_MIN_WALL_SEC   = 0.05                                # frame 당 최소 wall 시간 (sim 너무 빠를 때)

# 이벤트 버퍼 (monitor 의 [최근 이벤트] 표시용)
_EVENT_BUF = []
_EVENT_BUF_MAX = 100

def _log_event(sim_now_sec, msg):
    _EVENT_BUF.append((float(sim_now_sec), str(msg)))
    if len(_EVENT_BUF) > _EVENT_BUF_MAX:
        del _EVENT_BUF[:len(_EVENT_BUF) - _EVENT_BUF_MAX]

# ── §2 PCB·모델 메타정보 (AAS 미보유) ─────────────────
# AAS HS 에 PCB Entity 들 있지만 어느 게 메인이고 어느 게 THT 외주 대상인지
# 구분 정보 없음. 도메인 지식으로 직접 명시.
PCB_MAP = {                     # 모델별 메인 PCB
    'MODEL_A': '03203204',
    'MODEL_B': '03203145',
    'MODEL_C': '03203315',
}
THT_PCB_BY_MODEL = {            # 모델별 THT (외주) PCB 목록
    'MODEL_A': ['03902715', '03903424'],
    'MODEL_B': ['03902608', '03902730'],
    'MODEL_C': ['03903388', '03903391'],
}
THT_PCB = {c for codes in THT_PCB_BY_MODEL.values() for c in codes}

# THT 외주 raw/완료 재고 분리: raw 풀 +1 (발사) → -1 (도착) / pcb +1.
THT_RAW_SUFFIX = '_RAW'
def tht_raw_code(pcb_code):
    return f'{pcb_code}{THT_RAW_SUFFIX}'

# (참고) PCB 양면/단면 정보는 AAS HS 의 PCB Entity SMT_Side qualifier 가
# 단일 출처. 별도 fallback dict 없음.

# ── §3 설비 / 시설 메타 (AAS 미보유) ──────────────────
SMT_LINE_IDS = ['L1', 'L2']

# THT 외주 운송 시간 (도메인 상수, AAS 미보유)
THT_OUTSOURCE_SEC = 12 * 3600

# ── §5b PCB·Warehouse 정책 (AAS 미보유) ───────────────
# 만들어야 할 PCB 수량의 PCB_INITIAL_RATIO 만큼이 초기 재고로 존재.
# SMT 라인은 (1 - PCB_INITIAL_RATIO) × 수요만 추가 생산.
PCB_INITIAL_RATIO          = 0.8
# BOM 부품: 부품별 예상 demand × ratio 로 초기 재고/발주량 결정 (demand-aware).
# 너무 작은 부품(1~2 ea/order) 은 *_FLOOR 보장.
BOM_INITIAL_RATIO          = 0.6   # 부품 초기 재고 / demand
BOM_LOT_RATIO              = 0.5   # 발주 lot_size / demand
WAREHOUSE_BOM_INIT_FLOOR   = 50    # BOM 부품 초기 재고 최소값
WAREHOUSE_BOM_LOT_FLOOR    = 50    # BOM 부품 1회 발주량 최소값
WAREHOUSE_NONBOM_INIT_MULT = 10    # 비-BOM 부품 / total_qty

# ── §3b 설비 전력 RATED_POWER_KW (AAS 미보유, resource.xlsx 출처) ─
# SMT 라인 stage 는 process_code 단위 (라인별로 정격 다름). 한 표에 모은다.
# 조립~포장 그룹 전력은 §G ProcessGroupSpec 의 rated_kw 에서 derive.
RATED_POWER_KW_SMT = {
    'SMT_LOADER_L1':    0.66,
    'SMT_PRINTER_L1':   0.84,
    'SMT_SPI_L1':       2.20,
    'SMT_MOUNTER_H_L1': 19.93,  # NXT-II 2대
    'SMT_MOUNTER_M_L1': 4.64,
    'SMT_REFLOW_L1':    63.26,
    'SMT_UNLOADER_L1':  0.33,
    'SMT_LOADER_L2':    0.66,
    'SMT_PRINTER_L2':   1.72,
    'SMT_SPI_L2':       1.29,
    'SMT_MOUNTER_H_L2': 10.13,  # NXT-II 1대
    'SMT_MOUNTER_M_L2': 4.64,
    'SMT_REFLOW_L2':    48.03,
    'SMT_UNLOADER_L2':  0.33,
    'SMT_AOI':          0.29,
}

# ── §7 Worker fallback / SET sub-group ─────────────────
# AAS WorkstationConfigurationRecords 가 FW/LENS_HOLDER/SENSOR_FOCUS/SEMI/SET/
# AGING/OQC/PACK 8개 그룹 단일 출처. RMA 만 AAS 에 없음 — §G WorkerGroupSpec
# 의 fallback_cap 으로 한 자리에 등록.
#
# SET 내 INSP 인원 (sub-group). AAS 가 그룹 매핑은 하지만 인원 별도 정보 X.
# WORKER_SET 풀 중 INSP 전담 3명 — simpy.Resource 분리:
#   WORKER_SET      capacity = data.workers['WORKER_SET'] − SET_INSP_HEADCOUNT
#   WORKER_SET_INSP capacity = SET_INSP_HEADCOUNT
# 전력은 EnergyLogger 가 풀 cap (data.workers['WORKER_SET']) 그대로 사용.
SET_INSP_HEADCOUNT = 3


# ══════════════════════════════════════════════════════════
# §G. 그룹 메타 단일 출처 (Process / Worker SPECS)
# ══════════════════════════════════════════════════════════
# 새 그룹·공정 추가 시 *spec 표에 행 1개만* 등록하면 RATED_POWER / OQC 정책 /
# PACK 진입점 / 워커 매핑 / UI 라벨 / idle penalty 대상이 한 번에 적용된다.
# 하단의 derived dict (RATED_POWER_KW, OQC_RATE, ... LOCATION_ORDER) 는
# spec 에서 자동 생성 — 호출처 코드는 변경 없이 그대로 동작 (행동보존).

def _find_pack_entry(data, model_id):
    """모델의 첫 PACK 공정 process_code 를 AAS process flow 에서 자동 추출.

    RMA 수리 후 *조립부터 다시* 가 아니라 PACK 단계로 점프할 때 진입점이 필요.
    조건: process_group='PACK' AND dep_prev 에 INSP 그룹 공정 포함.
    매칭 실패 시 None — _do_complete 가 카운터만 올리고 종료.

    PROCESS_GROUP_SPECS['PACK'].rma_jump_entry 로도 호출 가능.
    """
    try:
        procs = data.get_model_procs(model_id)
    except Exception:
        return None
    grp_of = {str(r['process_code']): str(r.get('process_group', '') or '')
              for _, r in procs.iterrows()}
    for _, r in procs.iterrows():
        if str(r.get('process_group', '') or '') != 'PACK':
            continue
        prevs = [p.strip() for p in
                 str(r.get('dep_prev_codes', '') or '').split(';') if p.strip()]
        for p in prevs:
            if grp_of.get(p) == 'INSP':
                return str(r['process_code'])
    return None


@dataclass(frozen=True)
class ProcessGroupSpec:
    """공정 그룹별 단일 메타. AAS 미보유 항목만 여기에 등록.

    필드:
      name              : process_group 키 (AAS row 의 process_group 과 매칭)
      rated_kw          : 그룹 가동 시 시간당 평균 소비 (kW). EnergyLogger 가 사용.
      primary_worker    : 대표 worker_group (없으면 None — 호출자가 직접 지정).
                          AAS row 의 worker_group 이 우선이고 이건 fallback.
      on_defect         : 결함 발생 시 분기 라벨 (informational, 호출 분기는 호출처).
                          'route_to_rma' (INSP) / 'flag_only' (조립 일반) / None (RMA 자체)
      sampling_rate     : 표본 검사 확률 (OQC=0.05, 그 외=0).
      sample_time_sec   : 표본 검사 시간 (OQC=600s, 그 외=0).
      rma_jump_entry    : RMA 수리 후 진입점 lookup. 현재 PACK 만 _find_pack_entry.
      requires_outsource: THT 외주 대상 (현재 SMT 단의 PCB 코드 기준이라 사용 X — 표시용).
    """
    name              : str
    rated_kw          : float
    primary_worker    : str | None     = None
    on_defect         : str | None     = None
    sampling_rate     : float          = 0.0
    sample_time_sec   : int            = 0
    rma_jump_entry    : object         = None
    requires_outsource: bool           = False


@dataclass(frozen=True)
class WorkerGroupSpec:
    """작업자 그룹별 단일 메타.

    필드:
      name        : worker_group 키 (AAS row 의 worker_group 과 매칭)
      label_ko    : 간트차트·UI 라벨
      sort_order  : LOCATION_ORDER 정렬용 (낮을수록 위)
      fallback_cap: AAS WorkstationConfig 미제공 시 fallback (현재 RMA 만)
      track_idle  : idle penalty 계산 대상 여부 (보상 r6 분모)
    """
    name        : str
    label_ko    : str
    sort_order  : int
    fallback_cap: int | None = None
    track_idle  : bool       = True


PROCESS_GROUP_SPECS: dict = {
    'MODULE'      : ProcessGroupSpec('MODULE',       23.38, primary_worker=None,
                                     on_defect='flag_only'),
    'MODULE_FW'   : ProcessGroupSpec('MODULE_FW',     0.0,  primary_worker='WORKER_FW',
                                     on_defect='flag_only'),
    'NVD_40_FOCUS': ProcessGroupSpec('NVD_40_FOCUS',  0.36, primary_worker='WORKER_SENSOR_FOCUS',
                                     on_defect='flag_only'),
    'SEMI'        : ProcessGroupSpec('SEMI',         25.50, primary_worker='WORKER_SEMI',
                                     on_defect='flag_only'),
    'SET'         : ProcessGroupSpec('SET',          33.67, primary_worker='WORKER_SET',
                                     on_defect='flag_only'),
    'INSP'        : ProcessGroupSpec('INSP',          1.22, primary_worker='WORKER_AGING',
                                     on_defect='route_to_rma'),
    'AGING'       : ProcessGroupSpec('AGING',         1.84, primary_worker='WORKER_AGING',
                                     on_defect='flag_only'),
    'OQC'         : ProcessGroupSpec('OQC',           0.44, primary_worker='WORKER_OQC',
                                     on_defect='flag_only',
                                     sampling_rate=0.05, sample_time_sec=600),
    'PACK'        : ProcessGroupSpec('PACK',          8.31, primary_worker='WORKER_PACK',
                                     on_defect='flag_only',
                                     rma_jump_entry=_find_pack_entry),
    'PACK_LABEL'  : ProcessGroupSpec('PACK_LABEL',    0.37, primary_worker='WORKER_PACK',
                                     on_defect='flag_only'),
    'RMA'         : ProcessGroupSpec('RMA',           0.50, primary_worker='WORKER_RMA',
                                     on_defect=None),
}

WORKER_SPECS: dict = {
    'WORKER_FW'          : WorkerGroupSpec('WORKER_FW',           'F/W 입력',          0),
    'WORKER_LENS_HOLDER' : WorkerGroupSpec('WORKER_LENS_HOLDER',  'LENS HOLDER 조립',  1),
    'WORKER_SENSOR_FOCUS': WorkerGroupSpec('WORKER_SENSOR_FOCUS', 'FOCUS',             2),
    'WORKER_SET'         : WorkerGroupSpec('WORKER_SET',          'SET 조립',          3),
    'WORKER_SET_INSP'    : WorkerGroupSpec('WORKER_SET_INSP',     'SET 조립 (INSP)',   4),
    'WORKER_SEMI'        : WorkerGroupSpec('WORKER_SEMI',         '반 조립 라인',      5),
    'WORKER_RMA'         : WorkerGroupSpec('WORKER_RMA',          'RMA',               6, fallback_cap=6),
    'WORKER_OQC'         : WorkerGroupSpec('WORKER_OQC',          'OQC',               7),
    'WORKER_AGING'       : WorkerGroupSpec('WORKER_AGING',        'Aging test',        8),
    'WORKER_PACK'        : WorkerGroupSpec('WORKER_PACK',         '포장',              9),
}


# ── §G-derive. 기존 dict/상수 — spec 에서 자동 생성 (호출처 변경 0) ──

# ProcessGroupSpec → derived
RATED_POWER_KW = {**RATED_POWER_KW_SMT,
                  **{n: s.rated_kw for n, s in PROCESS_GROUP_SPECS.items()}}
PROCESS_GROUP_TO_WORKER_GROUP = {
    n: s.primary_worker for n, s in PROCESS_GROUP_SPECS.items()
}
OQC_RATE     = PROCESS_GROUP_SPECS['OQC'].sampling_rate   # 5% 표본 추출
OQC_TIME_SEC = PROCESS_GROUP_SPECS['OQC'].sample_time_sec  # 표본 검사 시간 (s)
# AOI_DEFECT_ACTION 은 §5 에서 정의 — SMT 라인 stage 라 그룹 spec 단위 아님.

# WorkerGroupSpec → derived
WORKER_GROUPS = {n for n, s in WORKER_SPECS.items() if s.track_idle}
WORKER_DEFAULT_CAP = {n: s.fallback_cap for n, s in WORKER_SPECS.items()
                      if s.fallback_cap is not None}
LOCATION_LABEL = {n: s.label_ko for n, s in WORKER_SPECS.items()}
LOCATION_ORDER = sorted(WORKER_SPECS, key=lambda n: WORKER_SPECS[n].sort_order)


def get_rated_power_kw(process_code: str, process_group: str = '',
                       capacity: int = 1) -> float:
    """단위 공정당 단위시간 평균 소비 (kW). capacity 가 N 이면 record 당 kw/N."""
    base = RATED_POWER_KW.get(str(process_code))
    if base is None:
        base = RATED_POWER_KW.get(str(process_group), 0.0)
    return base / max(int(capacity), 1)


# ══════════════════════════════════════════════════════════
# M02. 정적 데이터 로더 (FallbackDataLoader)
# ══════════════════════════════════════════════════════════
# 엑셀 시트 (KETI.시뮬레이션데이터통합.xlsx) 의존을 제거하고, AAS JSON 이
# 제공하지 않는 데이터만 하드코딩한다.
#
# AAS 제공 (FallbackDataLoader 미보유):
#   - 모델별 ManufacturingProcess (조립~포장 공정 흐름)
#   - HierarchicalStructures (PCB→부품, 단품 등록)
#   - WorkstationWorkerMatchingData (작업자 수, 숙련도, 근무 스케줄)
#
# Fallback 보유 (AAS 미제공):
#   - SMT 라인 stage 행 (LOADER~AOI for L1/L2) — 두 모델 공통 SMT 설비
#   - RMA_REPAIR 행
#   - RESOURCE.mttr_hr (설비 고장 평균 수리시간)
#
# 모델 A/C 가 AAS 에 있으면 모델 B 도 동일하게 AAS 만 신뢰. AAS JSON 이
# 없는 모델은 ProcessKnowledgeGraph 가 빈 그래프로 즉시 완료된다.

class FallbackDataLoader:
    """SMT 라인 + RMA + 설비 MTTR 만 하드코딩한 정적 데이터 로더.

    원본 ExcelDataLoader 와 동일한 인터페이스 (pf / bom / bom_struct / workers /
    resources / get_proc / get_kw / get_mttr / get_min_stock / get_critical_stock /
    get_lot_size / get_bom_parts / get_pcb_parts / get_item_name / smt_side /
    iter_all_bom_items / get_all_bom_codes / get_repair_parts / get_model_procs)
    를 제공해 CombinedDataLoader 에서 그대로 사용 가능.

    모델별 process flow / BOM 은 보유하지 않음 — AAS 가 단일 출처.
    """

    # PROCESS_FLOW row 컬럼 정의 (tuple 위치)
    _PF_COLS = ('process_code', 'process_group', 'dep_type', 'dep_prev_codes',
                'worker_group', 'worker_count', 'cycle_time_sec', 'defect_rate',
                'transfer_qty', 'transport_mode', 'transfer_time_sec')

    # ── SMT 라인 + RMA (model_id='ALL', 공용 설비) ────────────
    # 출처: KETI.시뮬레이션데이터통합.xlsx PROCESS_FLOW 의 ALL 행 (snapshot).
    _PF_ALL_ROWS = [
        ('RMA_REPAIR',       'RMA', 'SEQUENCE', '',                 'WORKER_RMA', 6, 600.0, 0.0,   0, '',         0.0),
        ('SMT_LOADER_L1',    'SMT', 'SEQUENCE', '',                 'OPERATOR',   2,   2.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_LOADER_L2',    'SMT', 'SEQUENCE', '',                 'OPERATOR',   2,   2.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_PRINTER_L1',   'SMT', 'SEQUENCE', 'SMT_LOADER_L1',    'OPERATOR',   2,  43.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_PRINTER_L2',   'SMT', 'SEQUENCE', 'SMT_LOADER_L2',    'OPERATOR',   2,  43.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_SPI_L1',       'SMT', 'SEQUENCE', 'SMT_PRINTER_L1',   'OPERATOR',   2,  15.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_SPI_L2',       'SMT', 'SEQUENCE', 'SMT_PRINTER_L2',   'OPERATOR',   2,  15.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_MOUNTER_H_L1', 'SMT', 'SEQUENCE', 'SMT_SPI_L1',       'OPERATOR',   2,  80.0, 0.003, 1, 'CONVEYOR', 5.0),
        ('SMT_MOUNTER_H_L2', 'SMT', 'SEQUENCE', 'SMT_SPI_L2',       'OPERATOR',   2,  80.0, 0.003, 1, 'CONVEYOR', 5.0),
        ('SMT_MOUNTER_M_L1', 'SMT', 'SEQUENCE', 'SMT_MOUNTER_H_L1', 'OPERATOR',   2,  45.0, 0.005, 1, 'CONVEYOR', 5.0),
        ('SMT_MOUNTER_M_L2', 'SMT', 'SEQUENCE', 'SMT_MOUNTER_H_L2', 'OPERATOR',   2,  45.0, 0.005, 1, 'CONVEYOR', 5.0),
        ('SMT_REFLOW_L1',    'SMT', 'SEQUENCE', 'SMT_MOUNTER_M_L1', 'OPERATOR',   2, 410.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_REFLOW_L2',    'SMT', 'SEQUENCE', 'SMT_MOUNTER_M_L2', 'OPERATOR',   2, 410.0, 0.001, 1, 'CONVEYOR', 5.0),
        ('SMT_UNLOADER_L1',  'SMT', 'SEQUENCE', 'SMT_REFLOW_L1',    'OPERATOR',   2,   5.0, 0.001, 0, 'CART',     7.0),
        ('SMT_UNLOADER_L2',  'SMT', 'SEQUENCE', 'SMT_REFLOW_L2',    'OPERATOR',   2,   5.0, 0.001, 0, 'CART',     7.0),
        ('SMT_AOI', 'SMT', 'JOIN', 'SMT_UNLOADER_L1;SMT_UNLOADER_L2', 'OPERATOR', 2,  30.0, 0.003, 1, '',         0.0),
    ]

    # ── 설비 MTTR (RESOURCE 시트 mttr_hr 정보 보유 행만) ──────
    # 같은 process_code 에 복수 설비면 max (원본 _build_mttr 동작 동일).
    _RESOURCE_MTTR_HR = {
        'SMT_AOI':          3.3,
        'SMT_MOUNTER_H_L1': 1.5,
        'SMT_MOUNTER_H_L2': 1.4,
        'SMT_MOUNTER_M_L1': 1.5,
        'SMT_MOUNTER_M_L2': 1.5,
        'SMT_PRINTER_L1':   4.0,
        'SMT_PRINTER_L2':  44.5,
        'SMT_REFLOW_L1':    8.0,
        'SMT_REFLOW_L2':    8.0,
        'SMT_SPI_L2':       2.3,
    }

    def __init__(self):
        # ── pf (DataFrame): SMT/RMA 행을 model_id='ALL' 로 등록 ──
        rows = []
        for tup in self._PF_ALL_ROWS:
            rec = dict(zip(self._PF_COLS, tup))
            rec['model_id']    = 'ALL'
            rec['ref_no']      = ''
            rec['process_name']= ''
            rec['dep_wait_hr'] = 0.0
            rows.append(rec)
        self.pf = pd.DataFrame(rows)
        # defect_rate 0 → DEFECT_FLOOR (원본 _load_pf 동작과 동일)
        self.pf['defect_rate'] = self.pf['defect_rate'].apply(
            lambda x: x if x > 0 else DEFECT_FLOOR)

        # ── bom / bom_struct (DataFrame): fallback 미보유 → 빈 ──
        # AAS HierarchicalStructures + InputBOM 이 모델별 BOM 단일 출처.
        self.bom = pd.DataFrame(columns=[
            'item_code', 'item_name', 'smt_side', 'min_stock_qty',
            'lot_size', 'critical_stock_qty', 'defect_rate'])
        self.bom_struct = pd.DataFrame(columns=[
            'model_id', 'parent_type', 'parent_code', 'item_code',
            'item_name', 'qty_per_parent'])

        # ── workers stub: CombinedDataLoader 가 AAS + WORKER_DEFAULT_CAP 으로 결정 ──
        self.workers   = {}
        self.resources = []  # rated_power_kw 는 RATED_POWER_KW (M01) 사용

        # ── 인덱스 / 캐시 ────────────────────────────────────
        self._pc_map = {str(r['process_code']): r for _, r in self.pf.iterrows()}
        # _grp_kw 은 RATED_POWER_KW 로 fallback 되므로 빈 dict.
        self._grp_kw = defaultdict(float)
        self._mttr   = {pc: hr * 3600 for pc, hr in self._RESOURCE_MTTR_HR.items()}
        # _bom_idx 비어있음 — get_bom_parts 는 항상 빈 리스트 반환.
        # CombinedDataLoader 에서 AAS bom_records 로 충전.
        self._bom_idx = defaultdict(list)
        # 캐시 빈 상태로 초기화 (모든 lookup 이 M01 기본값 fallback).
        self._min_stock_cache      = {}
        self._lot_size_cache       = {}
        self._item_name_cache      = {}
        self._smt_side_cache       = {}
        self._critical_stock_cache = {}
        self._bom_dup_merge_log    = []

    # ── 공개 인터페이스 (ExcelDataLoader 와 동일) ─────────────
    def get_proc(self, pc):
        return self._pc_map.get(str(pc))

    def get_kw(self, process_code, process_group, capacity=1):
        # 요청사항 0506 ②: M01 RATED_POWER_KW (resource.xlsx 기준) 우선 사용.
        kw = get_rated_power_kw(process_code, process_group, capacity)
        if kw > 0:
            return kw
        # 그룹 단위 최후 fallback (RATED_POWER_KW 매칭 실패 시).
        return {'SMT': 5.0, 'MODULE': 1.0, 'SEMI': 2.0, 'SET': 2.0,
                'INSP': 0.5, 'OQC': 0.2, 'PACK': 1.0, 'RMA': 0.1
                }.get(process_group, 0.5) / max(capacity, 1)

    def get_mttr(self, process_code):
        return self._mttr.get(str(process_code), SMT_MTTR_DEFAULT_HR * 3600)

    def get_min_stock(self, item_code):
        return self._min_stock_cache.get(str(item_code), float(MIN_STOCK))

    def get_critical_stock(self, item_code):
        """페널티(violation) 임계. min_stock(reorder point) 와 분리."""
        return self._critical_stock_cache.get(str(item_code), float(CRITICAL_STOCK))

    def get_lot_size(self, item_code):
        c = str(item_code)
        lot = self._lot_size_cache.get(c, -1)
        if lot > 0:
            return lot
        return int(self.get_min_stock(c) * REPLENISH_QTY_MULT)

    def get_bom_parts(self, model_id, parent_code):
        return self._bom_idx.get((model_id, str(parent_code)), [])

    def get_pcb_parts(self, pcb_code):
        """PCB 코드 → BOM 부품 목록. fallback 은 PCB BOM 을 보유하지 않으므로 빈 리스트.

        실제 PCB 부품 데이터는 AAS HierarchicalStructures (PCB_xxxxx Entity 의
        HasPart_yyyy RelationshipElement) 가 단일 출처이며, CombinedDataLoader.
        get_pcb_parts 에서 _hs_bom_idx 를 통해 조회한다.
        """
        return []

    def get_item_name(self, item_code):
        return self._item_name_cache.get(str(item_code), '')

    def smt_side(self, item_code):
        # AAS HierarchicalStructures SMT_Side qualifier 가 단일 출처.
        # CombinedDataLoader.smt_side 에서 AAS 우선 조회. 여기까지 오면 AAS 정보
        # 없는 경우이므로 'double' 기본 (THT 든 메인이든 도메인상 모두 양면).
        return 'double'

    def iter_all_bom_items(self):
        """fallback 자체는 BOM item 미보유 (AAS 가 단일 출처)."""
        return set()

    def get_all_bom_codes(self):
        if not hasattr(self, '_all_bom_codes_cache'):
            self._all_bom_codes_cache = self.iter_all_bom_items()
        return self._all_bom_codes_cache

    def get_repair_parts(self, src_pc, model_id):
        """[deprecated] 신 RMA 로직은 _rma_repair_aoi_board / _rma_repair_and_reinsert
        에서 직접 get_pcb_parts / get_bom_parts 를 호출한다. 하위 호환용 빈 반환."""
        return []

    def get_model_procs(self, model_id):
        df = self.pf
        # RMA 그룹은 정규 그래프에서 제외 (불량 품목은 run_rma 경로로만 처리).
        return df[df['model_id'].isin([model_id, 'ALL'])
                  & ~df['process_group'].isin(
                      ['SMT', 'LOGISTICS', 'SMT_SHARED', 'RMA'])].copy()

# ══════════════════════════════════════════════════════════
# M02b. AAS JSON 로더  (조립~포장 공정 전용)
# ══════════════════════════════════════════════════════════

def _aas_get_prop(elements, id_short):
    """AAS SubmodelElementCollection 내에서 idShort 로 Property 값을 조회."""
    for el in elements:
        if el.get('idShort') == id_short:
            return el.get('value')
    return None

def _aas_get_qualifier(obj, q_type):
    """AAS 요소의 qualifiers 에서 type 으로 값을 조회."""
    for q in obj.get('qualifiers', []):
        if q.get('type') == q_type:
            return q.get('value')
    return None

def _aas_extract_item_code(reference_elem):
    """InputBOM ReferenceElement 에서 item_code 추출.

    두 가지 참조 패턴 처리:
      1) ExternalReference keys[0].value = '.../cd/ITEM_CODE/1/0'
         -> ITEM_CODE 부분 추출 (PCB_ 접두사 유지)
      2) ModelReference keys → Entity idShort
    PCB_ 접두사를 가진 경우: PCB_XXXXX → 숫자 코드 XXXXX 만 반환.
    """
    val = reference_elem.get('value', {})
    keys = val.get('keys', [])
    for k in keys:
        raw = str(k.get('value', ''))
        if '/cd/' in raw:
            part = raw.rstrip('/').split('/')
            # '/1/0' suffix 제거 후 마지막 의미 토큰
            token = next((p for p in reversed(part)
                          if p and p not in ('1', '0')), '')
            if token.startswith('PCB_'):
                return token[4:]    # PCB_ 접두사 제거 → 숫자 코드
            if token:
                return token
    return ''

def _aas_worker_group_from_supplemental(group_el):
    """Group 요소의 supplementalSemanticIds 에서 WORKER 그룹 ID 추출.

    패턴: '.../cd/WORKER_XXX/1/0' 형태 URL 에서 WORKER_XXX 부분.
    """
    for sup in group_el.get('supplementalSemanticIds', []):
        for k in sup.get('keys', []):
            raw = str(k.get('value', ''))
            if '/cd/WORKER_' in raw:
                part = raw.rstrip('/').split('/')
                token = next((p for p in reversed(part)
                              if p.startswith('WORKER_')), '')
                if token:
                    return token
    return ''


class AASJsonLoader:
    """AAS JSON 파일에서 조립~포장 공정 데이터를 로드.

    지원 서브모델:
      - ManufacturingProcess : 공정 노드·의존성·BOM 파트 목록
      - WorkstationWorkerMatchingData : 작업장별 작업자 수 (미기재 시 WORKER_DEFAULT_CAP 사용)

    로드 결과:
      self.pf_records  : list[dict]  — 공정 흐름 레코드 (FallbackDataLoader.pf 행과 동일 스키마)
      self.bom_records : list[dict]  — (model_id, parent_code, item_code, qty) 튜플 목록
      self.workers     : dict[str,int] — worker_group -> 작업자 수
    """

    # supplementalSemanticIds URL → WORKER_GROUP 매핑 (AAS JSON 에 작업자 그룹 명시가 없을 때 fallback)
    _WORKER_FALLBACK = {
        'WWM_FwInputLine'       : 'WORKER_FW',
        'WWM_LensHolderLine'    : 'WORKER_LENS_HOLDER',
        'WWM_FocusLine'         : 'WORKER_SENSOR_FOCUS',
        'WWM_SemiAssemblyLine'  : 'WORKER_SEMI',
        'WWM_SetAssemblyLine'   : 'WORKER_SET',
        'WWM_AgingLine'         : 'WORKER_AGING',
        'WWM_OqcLine'           : 'WORKER_OQC',
        'WWM_PackagingLine'     : 'WORKER_PACK',
    }

    # ProcessGroup → 기본 worker_group 매핑 (supplementalSemanticIds 가 없는 그룹 fallback)
    _PG_WORKER_FALLBACK = {
        'MODULE' : 'WORKER_SEMI',      # 센서·렌즈 조립 등 반조립 계열 (임의)
        'SET'    : 'WORKER_SET',
        'SEMI'   : 'WORKER_SEMI',
        'INSP'   : 'WORKER_AGING',     # Aging Test 검사
        'PACK'   : 'WORKER_PACK',
    }

    # 그룹 idShort 키워드 → worker_group 오버라이드 매핑
    _GROUP_KW_WORKER = {
        'FwInput'       : 'WORKER_FW',
        'Focus'         : 'WORKER_SENSOR_FOCUS',
        'Sensor'        : 'WORKER_LENS_HOLDER',
        'LensHolder'    : 'WORKER_LENS_HOLDER',
        'Gimbal'        : 'WORKER_SET',
        'Bottom'        : 'WORKER_SEMI',
        'Combined'      : 'WORKER_SET',
        # 요청사항 0506 ③: 자원·idle 은 WORKER_SET_INSP (cap=SET_INSP_HEADCOUNT) 로 분리,
        # 전력은 grp='SET' 으로 통합 (run_process 에서 wgrp 만 재매핑).
        'Inspection'    : 'WORKER_SET_INSP',
        'UpperHousing'  : 'WORKER_SEMI',
        'SetAssembly'   : 'WORKER_SET',
        'AgingTest'     : 'WORKER_AGING',
        'Packaging'     : 'WORKER_PACK',
    }

    def __init__(self, model_id: str, json_path: str):
        self.model_id    = model_id
        self.json_path   = json_path
        self.pf_records     = []   # [{process_code, model_id, process_group, ...}, ...]
        self.bom_records    = []   # [(model_id, parent_code, item_code, qty), ...]
        self.hs_bom_records = []   # HierarchicalStructures 파싱 결과
                                   # (model_id, parent_code, item_code, qty, category)
                                   #   parent='PCB_XXXXX' → PCB 실장 부품 목록
                                   #   parent='__UNIT__'  → 모델 단위 조립 단품 전체
        # PCB Entity 의 SMT_Side qualifier (AAS 명시 정보 우선 활용).
        self.pcb_sides      = {}   # pcb_code (접두사 'PCB_' 제거) -> 'double'/'single'
        self.workers        = {}   # worker_group -> count
        self.worker_skill   = {}   # worker_group -> rank int (1/2/3)
        self.skill_ct       = {}   # rank int -> CT factor
        self.skill_dr       = {}   # rank int -> DR factor
        self.schedule       = {}   # work_start_sec, work_end_sec, break_duration_sec

        if not os.path.exists(json_path):
            print(f'[AASJsonLoader] 경고: {json_path} 없음 — {model_id} AAS 데이터 미로드')
            return

        with open(json_path, 'r', encoding='utf-8') as f:
            raw = json.load(f)

        submodels = raw.get('submodels', [])
        self._parse_skill_levels(submodels)                       # CT/DR factor 먼저
        self._parse_manufacturing_process(submodels, model_id)
        self._parse_worker_matching(submodels)
        self._parse_hierarchical_structures(submodels, model_id)  # BOM 구조

    # ── 내부 파싱 메서드 ──────────────────────────────────

    def _parse_skill_levels(self, submodels):
        """SkillLevelType SMC 에서 rank 숫자 · CT factor · DR factor 파싱.

        JSON 구조:
          WorkstationWorkerMatchingData > SkillLevelType [SMC]
            LOW      value=1  description: "...CT factor of 1.20 ... DR factor of 1.50..."
            STANDARD value=2  description: "...CT factor of 1.00 ... DR factor of 1.00..."
            HIGH     value=3  description: "...CT factor of 0.80 ... DR factor of 0.60..."

        결과: self.skill_ct = {1: 1.20, 2: 1.00, 3: 0.80}
              self.skill_dr = {1: 1.50, 2: 1.00, 3: 0.60}
        """
        ww = next(
            (sm for sm in submodels
             if sm.get('idShort') == 'WorkstationWorkerMatchingData'),
            None)
        if ww is None:
            return
        for el in ww.get('submodelElements', []):
            if el.get('idShort') != 'SkillLevelType':
                continue
            for prop in (el.get('value') or []):
                try:
                    rank = int(prop.get('value', 0))
                except (ValueError, TypeError):
                    continue
                desc = next(
                    (x.get('text', '') for x in (prop.get('description') or [])
                     if x.get('language') == 'en'),
                    '')
                ct_m = re.search(r'cycle time correction factor of ([0-9.]+)', desc)
                dr_m = re.search(r'defect rate correction factor of ([0-9.]+)', desc)
                if ct_m:
                    self.skill_ct[rank] = float(ct_m.group(1))
                if dr_m:
                    self.skill_dr[rank] = float(dr_m.group(1))

    def _build_skill_name_map(self, submodels) -> dict:
        """SkillLevelType Property idShort(이름) → value(rank 정수) 역방향 맵.

        예: {'LOW': 1, 'STANDARD': 2, 'HIGH': 3}
        """
        result = {}
        ww = next(
            (sm for sm in submodels
             if sm.get('idShort') == 'WorkstationWorkerMatchingData'),
            None)
        if ww is None:
            return result
        for el in ww.get('submodelElements', []):
            if el.get('idShort') != 'SkillLevelType':
                continue
            for prop in (el.get('value') or []):
                name = str(prop.get('idShort', '')).upper()
                try:
                    rank = int(prop.get('value', 0))
                except (ValueError, TypeError):
                    continue
                if name and rank:
                    result[name] = rank
        return result

    def _parse_worker_matching(self, submodels):
        """WorkstationWorkerMatchingData 파싱.

        각 workstation 에서 추출:
          - worker_count   : WorkstationConfigurationRecords 항목 수
          - dominant_skill : 레코드별 SkillLevel 이름 → rank 정수(최빈값)
          - schedule       : WorkStartTime / WorkEndTime / BreakDurationMin
                             (모든 workstation 동일 → 첫 번째만 self.schedule 에 저장)
        결과: self.workers[wgrp] = worker_count
              self.worker_skill[wgrp] = dominant_rank
              self.schedule = {work_start_sec, work_end_sec, break_duration_sec}
        """
        ww = next(
            (sm for sm in submodels
             if sm.get('idShort') == 'WorkstationWorkerMatchingData'),
            None)
        if ww is None:
            return

        skill_name_to_rank = self._build_skill_name_map(submodels)
        schedule_set = False

        for top_el in ww.get('submodelElements', []):
            if top_el.get('idShort') != 'GeneralWorkstationData':
                continue
            for sub_el in (top_el.get('value') or []):
                if sub_el.get('idShort') != 'WorkstationInformation':
                    continue
                for ws_el in (sub_el.get('value') or []):
                    ws_id = ws_el.get('idShort', '')
                    wgrp  = self._WORKER_FALLBACK.get(ws_id)
                    if not wgrp:
                        continue

                    props = ws_el.get('value') or []

                    # ── 근무 스케줄 (전 workstation 동일 → 첫 번째만) ──
                    if not schedule_set:
                        def _getv(id_s):
                            for p in props:
                                if isinstance(p, dict) and p.get('idShort') == id_s:
                                    return p.get('value')
                            return None
                        ws_t = str(_getv('WorkStartTime') or '9:00')
                        we_t = str(_getv('WorkEndTime')   or '18:00')
                        brk  = int(_getv('BreakDurationMin') or 60)
                        h_s, m_s = (int(x) for x in ws_t.split(':'))
                        h_e, m_e = (int(x) for x in we_t.split(':'))
                        self.schedule = {
                            'work_start_sec'    : h_s * 3600 + m_s * 60,
                            'work_end_sec'      : h_e * 3600 + m_e * 60,
                            'break_duration_sec': brk * 60,
                        }
                        schedule_set = True

                    # ── WorkstationConfigurationRecords 탐색 ──────────
                    records = next(
                        (p.get('value') for p in props
                         if isinstance(p, dict)
                         and p.get('idShort') == 'WorkstationConfigurationRecords'
                         and isinstance(p.get('value'), list)),
                        None)

                    if isinstance(records, list):
                        worker_count = len(records)
                        skill_votes  = []
                        for rec in records:
                            if not isinstance(rec, dict):
                                continue
                            for inner in (rec.get('value') or []):
                                if (isinstance(inner, dict)
                                        and inner.get('idShort') == 'SkillLevel'):
                                    name = str(inner.get('value', '')).upper()
                                    rank = skill_name_to_rank.get(name)
                                    if rank is not None:
                                        skill_votes.append(rank)
                        dominant = (
                            max(set(skill_votes), key=skill_votes.count)
                            if skill_votes else 2
                        )
                    else:
                        worker_count = 1
                        dominant     = 2

                    self.workers[wgrp]      = worker_count
                    self.worker_skill[wgrp] = dominant

    def _normalize_dep_prev(self, dep_prev_raw: str, known_pcs: set) -> str:
        """AAS DepPrev 값의 suffix 를 제거해 실제 process_code 로 정규화.

        AAS 패턴: '{process_code}_{SUFFIX}' (예: VD7_10_FW → VD7_10)
        멀티 선행: 세미콜론(;) 구분. 각 항목을 개별 정규화.

        정규화 규칙:
          - known_pcs 에 그대로 있으면 그대로 반환 (SMT_COMPLETE 등 가상 키 포함)
          - 없으면 마지막 '_' 기준으로 suffix 를 하나씩 제거하여 known_pcs 와 매칭
        """
        if not dep_prev_raw:
            return ''
        parts = [p.strip() for p in dep_prev_raw.split(';') if p.strip()]
        normalized = []
        for raw in parts:
            if raw in known_pcs:
                normalized.append(raw)
                continue
            # suffix 순차 제거
            candidate = raw
            found = False
            while '_' in candidate:
                candidate = candidate.rsplit('_', 1)[0]
                if candidate in known_pcs:
                    normalized.append(candidate)
                    found = True
                    break
            if not found:
                # 매칭 실패 — 원본 그대로 (그래프 엣지가 끊기지 않도록)
                normalized.append(raw)
        return ';'.join(normalized)

    def _parse_manufacturing_process(self, submodels, model_id):
        """ManufacturingProcess 서브모델 파싱 → pf_records, bom_records."""
        mp = next(
            (sm for sm in submodels if sm.get('idShort') == 'ManufacturingProcess'),
            None)
        if mp is None:
            return

        # 1단계: 모든 process_code 수집 (DepPrev 정규화에 필요)
        known_pcs = {'SMT_COMPLETE', 'SMT_THT'}
        for group_el in mp.get('submodelElements', []):
            if group_el.get('modelType') != 'SubmodelElementCollection':
                continue
            if group_el.get('idShort') == 'ProcessType':
                continue
            for proc_el in (group_el.get('value') or []):
                if proc_el.get('modelType') == 'SubmodelElementCollection':
                    known_pcs.add(str(proc_el.get('idShort', '')))

        # 2단계: 실제 파싱
        for group_el in mp.get('submodelElements', []):
            if group_el.get('modelType') != 'SubmodelElementCollection':
                continue
            if group_el.get('idShort') == 'ProcessType':
                continue  # 타입 정의 컬렉션, 공정 아님

            # 공정 그룹 정보
            process_group = _aas_get_qualifier(group_el, 'ProcessGroup') or 'MODULE'
            worker_group  = _aas_worker_group_from_supplemental(group_el)

            # supplementalSemanticIds 에 WORKER 참조가 없는 그룹:
            # 그룹 idShort 키워드 → worker_group 오버라이드 우선,
            # 없으면 ProcessGroup 기반 기본값 사용.
            if not worker_group:
                group_id = group_el.get('idShort', '')
                for kw, wgrp in self._GROUP_KW_WORKER.items():
                    if kw.lower() in group_id.lower():
                        worker_group = wgrp
                        break
                if not worker_group:
                    worker_group = self._PG_WORKER_FALLBACK.get(process_group, 'WORKER_SEMI')  # (임의)

            for proc_el in (group_el.get('value') or []):
                if proc_el.get('modelType') != 'SubmodelElementCollection':
                    continue

                pc_raw  = str(proc_el.get('idShort', ''))
                elems   = proc_el.get('value') or []

                dep_type     = _aas_get_prop(elems, 'DepType') or 'SEQUENCE'
                dep_prev_raw = _aas_get_prop(elems, 'DepPrev') or ''
                # DepPrev suffix 정규화: VD7_10_FW → VD7_10
                dep_prev     = self._normalize_dep_prev(dep_prev_raw, known_pcs)
                ct_raw       = _aas_get_prop(elems, 'CycleTimeSec')
                dr_raw       = _aas_get_prop(elems, 'DefectRate')

                cycle_time  = float(ct_raw) if ct_raw is not None else 0.0
                defect_rate = float(dr_raw) if dr_raw is not None else DEFECT_FLOOR

                # INSP 접미사 공정 → worker_group 런타임 라우팅
                # (run_process 내 SET+INSP 분기와 동일 로직)
                wgrp = worker_group

                pc = pc_raw

                record = {
                    'process_code'    : pc,
                    'model_id'        : model_id,
                    'process_group'   : process_group,
                    'worker_group'    : wgrp,
                    'cycle_time_sec'  : cycle_time,
                    'defect_rate'     : defect_rate if defect_rate > 0 else DEFECT_FLOOR,
                    'dep_type'        : dep_type.upper(),
                    'dep_prev_codes'  : dep_prev,
                    'dep_wait_hr'     : 0.0,
                    'transfer_qty'    : 1,
                    'transfer_time_sec': 0.0,
                    'transport_mode'  : '',
                }
                self.pf_records.append(record)

                # BOM 파트 파싱
                input_bom_el = next(
                    (e for e in elems if e.get('idShort') == 'InputBOM'), None)
                if input_bom_el is not None:
                    for bom_item in (input_bom_el.get('value') or []):
                        item_code = _aas_extract_item_code(bom_item)
                        qty_raw   = _aas_get_qualifier(bom_item, 'Quantity')
                        qty       = float(qty_raw) if qty_raw is not None else 1.0
                        if item_code:
                            self.bom_records.append(
                                (model_id, pc, item_code, qty))

    # ── 공개 인터페이스 ────────────────────────────────────

    def get_all_item_codes(self) -> set:
        """ManufacturingProcess InputBOM + HierarchicalStructures 두 소스의
        item_code 합집합 반환.
        Warehouse 초기 재고 설정 및 iter_all_bom_items 에서 활용.
        """
        codes = {ic for _, _, ic, _ in self.bom_records}
        codes |= {ic for _, _, ic, _, _ in self.hs_bom_records}
        return codes

    def get_bom_parts(self, model_id: str, parent_code: str) -> list:
        """(model_id, parent_code) 에 해당하는 [(item_code, qty), ...] 반환."""
        return [
            (ic, qty)
            for (mid, pc, ic, qty) in self.bom_records
            if mid == model_id and pc == parent_code
        ]

    def _parse_hierarchical_structures(self, submodels, model_id):
        """HierarchicalStructures 서브모델 파싱 → hs_bom_records.

        JSON 구조:
          MODEL_X_VD7 [Entity]
            ├─ PCB_XXXXX [Entity]  (PCB 복합 부품: SMT 실장 부품 목록)
            │     └─ HasPart_YYYYYYY [RelationshipElement]
            │           qualifier Quantity → qty
            │           qualifier Category → category (Capacitor, Resistor 등)
            └─ PXXXXXXXX [Entity]  (조립 공정 투입 단품: parts=0)
                  qualifier Quantity → qty
                  qualifier Category → 부품 카테고리 (LENS, CABLE 등)

        결과:
          PCB 계층: (model_id, 'PCB_XXXXX', 'YYYYYYY', qty, category)
            → PCB 에 실장되는 전자부품. item_code = HasPart_ 접두사 제거 후 숫자코드.
          단품 등록: (model_id, '__UNIT__', 'PXXXXXXX', qty, category)
            → parent='__UNIT__' 은 모델 전체 단위 단품 목록을 나타내는 가상 키.
              Warehouse 초기 재고 보완 및 iter_all_bom_items 확장에 활용.
        """
        hs = next(
            (sm for sm in submodels if sm.get('idShort') == 'HierarchicalStructures'),
            None)
        if hs is None:
            return

        model_entity = next(
            (e for e in hs.get('submodelElements', [])
             if e.get('modelType') == 'Entity'),
            None)
        if model_entity is None:
            return

        def _get_q(obj, q_type):
            for q in (obj.get('qualifiers') or []):
                if q.get('type') == q_type:
                    return q.get('value')
            return None

        for st in (model_entity.get('statements') or []):
            parent_id   = str(st.get('idShort', ''))
            inner_parts = st.get('statements') or []

            # PCB Entity 의 SMT_Side qualifier 추출 (parent_id='PCB_XXXXX' 형태).
            if parent_id.startswith('PCB_'):
                side_v = _get_q(st, 'SMT_Side')
                if side_v:
                    pcb_code = parent_id[len('PCB_'):]
                    self.pcb_sides[pcb_code] = str(side_v).lower()

            if inner_parts:
                # PCB 복합 부품: HasPart_YYYYYYY RelationshipElement 목록
                for rel in inner_parts:
                    if rel.get('modelType') != 'RelationshipElement':
                        continue
                    has_part_id = str(rel.get('idShort', ''))
                    if not has_part_id.startswith('HasPart_'):
                        continue
                    item_code = has_part_id[len('HasPart_'):]  # 'HasPart_10200032' → '10200032'
                    qty_raw   = _get_q(rel, 'Quantity')
                    category  = _get_q(rel, 'Category') or ''
                    try:
                        qty = float(qty_raw) if qty_raw is not None else 1.0
                    except (ValueError, TypeError):
                        qty = 1.0
                    self.hs_bom_records.append(
                        (model_id, parent_id, item_code, qty, category))
            else:
                # 단품 등록 (P-코드): 하위 없음, qualifier 에 Quantity/Category
                qty_raw  = _get_q(st, 'Quantity')
                category = _get_q(st, 'Category') or ''
                try:
                    qty = float(qty_raw) if qty_raw is not None else 1.0
                except (ValueError, TypeError):
                    qty = 1.0
                self.hs_bom_records.append(
                    (model_id, '__UNIT__', parent_id, qty, category))


class CombinedDataLoader:
    """FallbackDataLoader (SMT/RMA/MTTR) + AASJsonLoader (모델별 조립~포장) 통합 인터페이스.

    ProcessKnowledgeGraph, run_process, Warehouse 등 기존 코드가
    data.get_proc(), data.get_bom_parts(), data.workers 등을 호출할 때
    이 클래스 하나로 두 소스의 데이터를 투명하게 제공한다.

    우선순위: AAS JSON > Fallback (동일 process_code 충돌 시 AAS 우선).
    Fallback 은 SMT/RMA/MTTR 만 보유하므로 충돌은 발생하지 않음.
    """

    def __init__(self, static_loader: 'FallbackDataLoader',
                 aas_loaders: dict):
        """
        static_loader : FallbackDataLoader 인스턴스 (SMT 라인 + RMA + MTTR)
        aas_loaders   : {model_id: AASJsonLoader} 딕셔너리
        """
        self.static  = static_loader
        self.aas_map = aas_loaders   # {model_id: AASJsonLoader}

        # ── 통합 process_code 맵 구축 ─────────────────────
        # 1) Fallback 공정 등재 (SMT 라인 + RMA)
        self._pc_map = dict(static_loader._pc_map)
        # 2) AAS 공정 추가 (조립~포장; SMT 와 충돌 없음)
        for aas in aas_loaders.values():
            for rec in aas.pf_records:
                self._pc_map[rec['process_code']] = rec

        # ── 통합 BOM 인덱스 구축 ──────────────────────────
        # Fallback _bom_idx 는 항상 비어있음 — AAS 가 단일 출처.
        self._bom_idx = dict(static_loader._bom_idx)
        for aas in aas_loaders.values():
            for (mid, pc, ic, qty) in aas.bom_records:
                key = (mid, pc)
                if key not in self._bom_idx:
                    self._bom_idx[key] = []
                if not any(e[0] == ic for e in self._bom_idx[key]):
                    self._bom_idx[key].append((ic, qty))

        # ── HierarchicalStructures BOM 인덱스 구축 ────────
        # (model_id, parent_code) → [(item_code, qty, category), ...]
        # parent_code 예시:
        #   'PCB_03203204' → PCB 에 실장되는 전자부품 목록 (SMT 소모 부품)
        #   '__UNIT__'     → 모델 단위 조립 단품 전체 목록 (조립 공정 소모 부품)
        self._hs_bom_idx = {}
        for aas in aas_loaders.values():
            for (mid, parent, ic, qty, cat) in getattr(aas, 'hs_bom_records', []):
                key = (mid, parent)
                if key not in self._hs_bom_idx:
                    self._hs_bom_idx[key] = []
                if not any(e[0] == ic for e in self._hs_bom_idx[key]):
                    self._hs_bom_idx[key].append((ic, qty, cat))

        # ── 통합 workers ──────────────────────────────────
        # 요청사항 0506 ②: 우선순위 = AAS WorkstationConfigurationRecords →
        # WORKER_DEFAULT_CAP (M01 상단). FallbackDataLoader.workers 는 빈 dict.
        self.workers = {}
        for aas in aas_loaders.values():
            for wgrp, count in aas.workers.items():
                if count:                        # AAS 에서 실제 수가 파싱된 경우만 기록
                    self.workers[wgrp] = count
        # AAS 누락분은 M01 fallback 적용
        for wgrp, default_cap in WORKER_DEFAULT_CAP.items():
            self.workers.setdefault(wgrp, default_cap)
        # 요청사항 0506 ③: WORKER_SET_INSP 자원·idle 만 분리 (cap=SET_INSP_HEADCOUNT).
        # 전력은 run_process 에서 grp='SET' 유지로 SET 풀 통합 처리.
        self.workers.setdefault('WORKER_SET_INSP', SET_INSP_HEADCOUNT)

        # ── AAS 근무 스케줄 통합 ─────────────────────────
        # 모든 workstation 이 동일 스케줄을 가지므로 첫 번째 유효한 것을 사용
        self.schedule = {}
        for aas in aas_loaders.values():
            if getattr(aas, 'schedule', {}):
                self.schedule = aas.schedule
                break

        # ── AAS 숙련도 · CT/DR factor 통합 ───────────────
        self.worker_skill = {}
        self.skill_ct     = {}
        self.skill_dr     = {}
        for aas in aas_loaders.values():
            self.worker_skill.update(getattr(aas, 'worker_skill', {}))
            if not self.skill_ct:
                self.skill_ct = getattr(aas, 'skill_ct', {})
            if not self.skill_dr:
                self.skill_dr = getattr(aas, 'skill_dr', {})

        # ── 위임 속성: Fallback 로더 메서드를 그대로 노출 ───
        self.bom        = static_loader.bom
        self.bom_struct = static_loader.bom_struct
        self.resources  = static_loader.resources

        # ── 통합 공정 DataFrame (ProcessKnowledgeGraph 용) ─
        self._pf_combined = self._build_combined_pf()

    def _build_combined_pf(self) -> pd.DataFrame:
        """Fallback PF (SMT/RMA) + AAS PF (모델별) 를 합쳐 하나의 DataFrame.

        Fallback 은 model_id='ALL' 로만 등록되어 있으므로 모델별 충돌 없음.
        AAS 로드된 모델은 그 모델의 AAS 행이 단일 출처.
        AAS 없는 모델은 PF 행이 없으므로 ProcessKnowledgeGraph 빈 그래프 → 즉시 완료.
        """
        aas_rows = []
        for aas in self.aas_map.values():
            aas_rows.extend(aas.pf_records)

        static_rows = self.static.pf.to_dict('records')

        combined = static_rows + aas_rows
        df = pd.DataFrame(combined)
        for col in ['cycle_time_sec', 'defect_rate', 'transfer_qty',
                    'transfer_time_sec', 'dep_wait_hr']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        if 'defect_rate' in df.columns:
            df['defect_rate'] = df['defect_rate'].apply(
                lambda x: x if x > 0 else DEFECT_FLOOR)
        return df

    # ── FallbackDataLoader 위임 메서드 ────────────────────

    def get_proc(self, pc):
        return self._pc_map.get(str(pc))

    def get_kw(self, process_code, process_group, capacity=None):
        # capacity 미지정 시 1 (SMT 라인 stage 는 capacity=1).
        if capacity is None:
            capacity = 1
        return self.static.get_kw(process_code, process_group, capacity)

    def get_mttr(self, process_code):
        return self.static.get_mttr(process_code)

    def get_min_stock(self, item_code):
        return self.static.get_min_stock(item_code)

    def get_critical_stock(self, item_code):
        return self.static.get_critical_stock(item_code)

    def get_lot_size(self, item_code):
        return self.static.get_lot_size(item_code)

    def get_item_name(self, item_code):
        return self.static.get_item_name(item_code)

    def get_repair_parts(self, src_pc, model_id):
        """[deprecated] 신 RMA 로직은 _rma_repair_aoi_board / _rma_repair_and_reinsert
        에서 직접 get_pcb_parts / get_bom_parts 를 호출. 빈 리스트만 반환."""
        return []

    def smt_side(self, item_code):
        # AAS HierarchicalStructures 의 PCB Entity SMT_Side qualifier 가 단일 출처.
        c = str(item_code)
        for aas in self.aas_map.values():
            v = aas.pcb_sides.get(c)
            if v:
                return v
        # AAS 정보 없음 → fallback ('double' 기본 — 도메인상 메인·THT 모두 양면)
        return self.static.smt_side(item_code)

    def iter_all_bom_items(self):
        """AAS InputBOM + HierarchicalStructures 두 소스의 item_code 합집합."""
        items = self.static.iter_all_bom_items()  # 비어있음
        for aas in self.aas_map.values():
            items |= aas.get_all_item_codes()
        for (_mid, _parent), parts in self._hs_bom_idx.items():
            for ic, _qty, _cat in parts:
                items.add(ic)
        return items

    def get_all_bom_codes(self) -> set:
        if not hasattr(self, '_all_bom_codes_cache'):
            self._all_bom_codes_cache = self.iter_all_bom_items()
        return self._all_bom_codes_cache

    def get_bom_parts(self, model_id: str, parent_code: str) -> list:
        """공정 BOM 조회. AAS InputBOM 이 단일 출처."""
        key = (model_id, str(parent_code))
        return self._bom_idx.get(key, [])

    def get_pcb_parts(self, pcb_code: str) -> list:
        """PCB 코드 → BOM 부품 목록 [(item_code, qty), ...] 반환.

        AOI 보드 수리 시 PCB 분해·재조립으로 SMT 부품을 한 세트 다시 소모.
        AAS HierarchicalStructures 의 'PCB_<code>' Entity 를 단일 출처로 사용.
        같은 PCB 가 여러 모델에 등재돼도 BOM 동일하다고 가정하고 첫 매칭 사용.
        """
        c = str(pcb_code)
        candidates = (c, f'PCB_{c}')
        for (_mid, parent), parts in self._hs_bom_idx.items():
            if parent in candidates:
                return [(ic, qty) for ic, qty, _cat in parts]
        return []

    def get_hs_bom_parts(self, model_id: str, parent_code: str) -> list:
        """HierarchicalStructures 기반 BOM 조회.

        반환: [(item_code, qty, category), ...]

        사용처:
          parent_code = 'PCB_XXXXX' : 해당 PCB 실장 전자부품 전체
          parent_code = '__UNIT__'  : 해당 모델의 조립 단품 전체 등록 목록
        """
        return self._hs_bom_idx.get((model_id, parent_code), [])

    def get_all_hs_item_codes(self, model_id: str) -> set:
        """model_id 에 속한 HierarchicalStructures 전체 item_code 집합."""
        return {
            ic
            for (mid, _parent), parts in self._hs_bom_idx.items()
            if mid == model_id
            for ic, _qty, _cat in parts
        }

    def get_model_procs(self, model_id: str) -> pd.DataFrame:
        """모델별 공정 DataFrame 반환 (SMT/LOGISTICS/SMT_SHARED/RMA 그룹 제외)."""
        exclude = {'SMT', 'LOGISTICS', 'SMT_SHARED', 'RMA'}
        df = self._pf_combined
        return df[
            (df['model_id'] == model_id) &
            (~df['process_group'].isin(exclude))
        ].copy()

# ████████████████████████████████████████████████████████████████████
# §B. DOMAIN
# ████████████████████████████████████████████████████████████████████
# 책임: 시뮬레이션의 *상태*를 보유하는 도메인 객체. SimPy 의 이벤트 루프와는
# 분리 — 상태 변경 메서드만 노출하고, 시간 진행은 §C 가 담당.
# 외부 export: ProcessKnowledgeGraph, is_process_ready / ReadyContext / ReadyStatus,
#              Warehouse, WIPTracker, EnergyLogger, IdleTracker,
#              SolderCream, OutsourceTruckPool, SMTLine.
# 향후 파일 분할 위치: cpro/domain/{kg, ready_gate, warehouse, wip, energy,
#                                   idle, smt, outsource}.py


# ══════════════════════════════════════════════════════════
# M03. 공정 지식 그래프
# ══════════════════════════════════════════════════════════

class ProcessKnowledgeGraph:
    """
    공정 DAG. 노드 특징 6차원:
    [cycle_time정규화, defect_rate*1000(0~1 클리핑),
     worker_count/20, rated_kw/100, is_fork, is_join]
    """
    def __init__(self, data, model_id: str):
        self.model_id = model_id
        self._data    = data   # BOM 재고 ready 체크용
        self.nodes    = {}
        self.edges    = []
        df     = data.get_model_procs(model_id)
        max_ct = max(df['cycle_time_sec'].max(), 1)
        for _, r in df.iterrows():
            pc   = str(r['process_code'])
            wgrp = str(r.get('worker_group','') or '')
            kw   = data.get_kw(pc, str(r.get('process_group','') or ''))
            dt   = str(r.get('dep_type','SEQUENCE') or 'SEQUENCE').upper()
            self.nodes[pc] = {
                'process_code' : pc,
                'process_group': str(r.get('process_group','') or ''),
                'cycle_time_sec': float(r['cycle_time_sec'] or 0),
                'defect_rate'  : float(r['defect_rate'] or DEFECT_FLOOR),
                'dep_wait_hr'  : float(r['dep_wait_hr'] or 0),
                'worker_group' : wgrp,
                'worker_count' : data.workers.get(wgrp, 1),
                'rated_kw'     : kw,
                'transfer_time': float(r['transfer_time_sec'] or 0),
                'dep_type'     : dt,
                'feat': np.array([
                    float(r['cycle_time_sec'] or 0) / max_ct,
                    min(float(r['defect_rate'] or DEFECT_FLOOR) * 1000, 1.0),
                    data.workers.get(wgrp, 1) / 20,
                    min(kw / 100, 1.0),
                    1.0 if dt == 'FORK' else 0.0,
                    1.0 if dt == 'JOIN'  else 0.0,
                ], dtype=np.float32)
            }
            for prev in [p.strip() for p in
                         str(r.get('dep_prev_codes','') or '').split(';') if p.strip()]:
                if prev != pc:
                    self.edges.append((prev, pc, dt))

    def get_feat_matrix(self):
        pcs = list(self.nodes.keys())
        return pcs, np.stack([self.nodes[p]['feat'] for p in pcs])

    def get_adj(self):
        pcs = list(self.nodes.keys())
        idx = {p: i for i, p in enumerate(pcs)}
        N   = len(pcs)
        adj = np.zeros((N, N), dtype=np.float32)
        for f, t, _ in self.edges:
            if f in idx and t in idx:
                adj[idx[f]][idx[t]] = 1.0
        return adj

    def ready_processes(self, ctx) -> list:
        """게이트(M03b is_process_ready)로 즉시 실행 가능한 공정만 반환.

        조건 ①선행공정 완료 ②부품재고 충족 ③작업자 idle — 셋 다 통과한 pc 만.
        도식(rl_action) 의 ready 의도와 코드가 일치하는 단일 진실원이 됐다.

        ctx : ReadyContext (kg / done_set / wh / wres / data / model_id 보유)

        ⚠ run_process 내 wait_stock·wres.request 는 *유지*한다 (backup 안전망).
        ready 평가와 실제 자원 점유 사이에 다른 process 가 먼저 가져갈 수 있어
        assertion 으로 강등하면 race 시 crash. 게이트는 _최선 추정_, wait 은 안전망.
        """
        return [pc for pc in self.nodes
                if pc not in ctx.done_set
                and is_process_ready(pc, ctx) == ReadyStatus.READY]


# ══════════════════════════════════════════════════════════
# M03b. 공정 Ready 게이트 (요청사항_0506 — ready_rule 통합)
# ══════════════════════════════════════════════════════════
# 한 공정이 *진짜로* 실행 가능한지 한 함수에서 판정한다.
# 게이트:
#   ① 선행공정 완료    : kg.edges 의 prev 들 ⊆ done_set
#   ② 부품 재고 확보   : data.get_bom_parts(model, pc) 의 (item, qty) 모두
#                        wh.stock[item] ≥ qty
#   ③ 작업자 idle      : pc 의 worker_group (SET→SET_INSP override 포함) 의
#                        simpy.Resource 에 빈 자리 있음
#
# 현 단계 (PR1, 행동보존):
#   - kg.ready_processes / run_process 의 wait_stock·wres.request 는 그대로.
#   - 이 함수는 진단·단위테스트·디버그 용도로만 사용.
#   - 즉 이 섹션 추가만으로는 시뮬레이션 결과가 1bit 도 바뀌지 않는다.
#
# 다음 단계 (PR3, 행동변경):
#   - kg.ready_processes 가 is_process_ready 호출.
#   - run_process 내 wait_stock·wres.request 는 assertion 으로 강등.
#   - PPO 의 action mask 가 좁아져 정책 분포가 바뀜 → 재학습 필요.


class ReadyStatus(Enum):
    READY       = 'ready'
    WAIT_PRED   = 'wait_pred'
    WAIT_STOCK  = 'wait_stock'
    WAIT_WORKER = 'wait_worker'
    UNKNOWN_PC  = 'unknown_pc'   # kg.nodes 에 없음 (가상 노드 등)


@dataclass
class ReadyContext:
    """ready 판정에 필요한 모든 런타임 핸들. menv 가 이 모두를 보유하므로
    호출처에서 menv 로부터 한 번에 만들 수 있다."""
    kg       : 'ProcessKnowledgeGraph'
    done_set : set
    wh       : 'Warehouse'
    wres     : dict          # worker_group → simpy.Resource
    data     : object        # CombinedDataLoader (forward ref)
    model_id : str


def resolve_worker_group(pc: str, node: dict) -> str:
    """run_process 의 SET→SET_INSP override 와 동일 규칙. 여러 곳에서 같은
    분기를 반복 안 하도록 한 함수에 모은다.

    규칙 (run_process L2236-2241 참조):
      worker_group == 'WORKER_SET' AND process_group == 'SET' AND
      process_code 의 마지막 토큰이 'INSP' 이면 → 'WORKER_SET_INSP' 로 매핑.
    """
    wgrp = str(node.get('worker_group', '') or '')
    grp  = str(node.get('process_group', '') or '')
    if wgrp == 'WORKER_SET' and grp == 'SET':
        if str(pc).rsplit('_', 1)[-1].upper() == 'INSP':
            wgrp = 'WORKER_SET_INSP'
    return wgrp


def is_process_ready(pc: str, ctx: ReadyContext) -> ReadyStatus:
    """공정 한 개가 지금 즉시 실행 가능한지 판정. 단일 의사결정점.

    행동보존 모드(PR1)에서는 호출 사이드 이펙트가 없도록 RNG 미사용.
    """
    pc = str(pc)
    if pc not in ctx.kg.nodes:
        return ReadyStatus.UNKNOWN_PC
    node = ctx.kg.nodes[pc]

    # ① 선행공정 완료
    preds = [p for (p, t, _) in ctx.kg.edges if t == pc]
    if not all(p in ctx.done_set for p in preds):
        return ReadyStatus.WAIT_PRED

    # ② 부품 재고 — wh.stock 은 defaultdict 라 None 직접 비교 불가, 0 fallback
    for code, qty in ctx.data.get_bom_parts(ctx.model_id, pc):
        if ctx.wh.stock[str(code)] < qty:
            return ReadyStatus.WAIT_STOCK

    # ③ 작업자
    wgrp = resolve_worker_group(pc, node)
    if wgrp:
        res = ctx.wres.get(wgrp)
        if res is None or res.count >= res.capacity:
            return ReadyStatus.WAIT_WORKER

    return ReadyStatus.READY


def ready_processes_with_status(ctx: ReadyContext) -> dict:
    """진단용. {ReadyStatus: [pc, ...]} 형태로 분류 반환.
    어느 공정이 어느 게이트에서 막혀 있는지 한눈에 보기 위해."""
    out = {st: [] for st in ReadyStatus}
    for pc in ctx.kg.nodes:
        if pc in ctx.done_set:
            continue
        out[is_process_ready(pc, ctx)].append(pc)
    return out


# ══════════════════════════════════════════════════════════
# M04. 창고 / WIPTracker
# ══════════════════════════════════════════════════════════

class Warehouse:
    def __init__(self, data, order: dict):
        self.data    = data
        self.order   = order
        total_qty    = sum(order.values())

        # BOM 부품 초기 재고: demand-aware (2026-05-06).
        # 부품별 예상 총 소비량 (sum of qty × order across processes consuming it)
        # 을 계산하고 그 값 × BOM_INITIAL_RATIO 만큼 초기 보유. 단 항상
        # WAREHOUSE_BOM_INIT_FLOOR 이상.
        # PCB(메인·THT)는 별도로 PCB_INITIAL_RATIO (=0.8) 적용.
        bom_codes             = data.get_all_bom_codes()
        self._bom_codes       = bom_codes
        self._bom_init_stock  = WAREHOUSE_BOM_INIT_FLOOR  # (Stock_Summary 표시용)
        self._init_stock      = float(total_qty * WAREHOUSE_NONBOM_INIT_MULT)

        # 부품별 demand 계산
        demand = defaultdict(float)
        for model_id, qty in order.items():
            try:
                procs = data.get_model_procs(model_id)
            except Exception:
                continue
            for _, row in procs.iterrows():
                pc = str(row['process_code'])
                for ic, q in data.get_bom_parts(model_id, pc):
                    demand[str(ic)] += float(q) * int(qty)

        self.stock = defaultdict(lambda: self._init_stock)
        # 부품별 예상 demand 보존 — 초기 재고 / 발주 lot_size 계산에 사용.
        self._demand = dict(demand)
        # Stock_Summary 시트 등에서 부품별 초기 재고를 정확히 보여주기 위해 보존.
        self._initial_stocks = {}
        for code in bom_codes:
            d   = demand.get(str(code), 0.0)
            init = max(WAREHOUSE_BOM_INIT_FLOOR, int(d * BOM_INITIAL_RATIO))
            self.stock[str(code)] = float(init)
            self._initial_stocks[str(code)] = init

        # PCB(메인·THT) 초기 재고 = 모델별 주문수량 × PCB_INITIAL_RATIO
        for model_id, qty in order.items():
            main_pcb = PCB_MAP.get(model_id)
            if main_pcb is not None:
                self.stock[str(main_pcb)] = float(int(qty * PCB_INITIAL_RATIO))
            for tht_pcb in THT_PCB_BY_MODEL.get(model_id, []):
                self.stock[str(tht_pcb)] = float(int(qty * PCB_INITIAL_RATIO))

        # THT raw 풀: 외주 시작 시 +1, 완료 시 -1. 시작 시점 0.
        for pcb_code in THT_PCB:
            self.stock[tht_raw_code(pcb_code)] = 0.0

        self.consumed        = defaultdict(int)
        self.violations      = defaultdict(int)
        self.history         = defaultdict(list)
        self._pending_orders = set()
        self._wait_events    = defaultdict(list)
        self.snapshots       = defaultdict(list)
        self.reorder_log     = []
        self.reorder_count   = defaultdict(int)

        # ── 디버그 추적용 속성 ──────────────────────────────────────
        # PCB_MAP(메인) + THT_PCB(수삽): 외부 발주 대상 제외 코드 집합
        self._pcb_codes          = set(PCB_MAP.values()) | set(THT_PCB)
        # 외주 이벤트 로그: (start_time, return_time, pcb_code, model, ...)
        self.outsource_log       = []
        # 유닛별 완성 경로: (model, uid) -> {path, end_time, done_n, total_n, ...}
        self.unit_completions    = {}
        # SMT 라인 처리량: (line_id, model_id, pcb_code) -> 완성 board 수
        self.smt_per_model       = defaultdict(int)
        # PCB 코드별 흐름 카운터: smt restore / outsource in·out / external order
        self.pcb_flow            = defaultdict(lambda: defaultdict(int))
        # prow=None 으로 스킵된 (model_id, process_code) → count
        self.skipped_pcs         = defaultdict(int)
        # SMT 라인 모델 선택 로그: (sim_time_sec, selected_model)
        self.smt_model_choices   = []
        # produce_unit 종료 시점에 미처리 공정이 있는 유닛 기록
        self.kg_incomplete_log   = []
        # 양면 PCB 에서 한 면만 SMT 처리된 board 기록
        self.smt_single_side_log = []
        # wait_stock timeout fallback 이 발생한 케이스 기록
        self.stuck_wait_log      = []

    def consume(self, item_code, qty, sim_time=0):
        c = str(item_code)
        self.stock[c] = max(0, self.stock[c] - int(qty))
        self.consumed[c] += int(qty)
        # PCB(메인·THT) 와 raw 코드는 외부 발주 대상이 아님.
        # PCB 는 SMT 라인 또는 외주 경로로만 보충된다.
        if not c.endswith(THT_RAW_SUFFIX):
            is_pcb = c in self._pcb_codes
            # 발주 트리거: PCB 제외, min_stock 미만이고 발주가 진행 중이 아닐 때.
            # 발주 자체는 정상 재주문 행위이므로 음의 보상 신호가 아님.
            if (not is_pcb
                    and self.stock[c] < self.data.get_min_stock(c)
                    and c not in self._pending_orders):
                self._pending_orders.add(c)
                self.snapshots[c].append((sim_time, self.stock[c]))
            # critical_stock 미만일 때만 페널티 카운트 (min_stock 진입은 포함 안 함).
            if self.stock[c] < self.data.get_critical_stock(c):
                self.violations[c] += 1
        self.history[c].append((sim_time, self.stock[c]))

    def restore(self, item_code, qty, sim_time=0):
        c = str(item_code)
        self.stock[c] += int(qty)
        self.history[c].append((sim_time, self.stock[c]))
        # 디버그: PCB 가 SMT 또는 외주 경로로 복귀할 때 카운트. raw 코드는 외주 시작.
        if c.endswith(THT_RAW_SUFFIX):
            base = c[:-len(THT_RAW_SUFFIX)]
            self.pcb_flow[base]['outsource_in'] += int(qty)
        elif c in self._pcb_codes:
            self.pcb_flow[c]['restore_from_smt_or_outsource'] += int(qty)
        self._notify_waiters(c)

    def wait_stock(self, env, item_code, qty, max_wait_sec=None):
        """item_code 재고가 qty 이상이 될 때까지 대기 후 consume.

        max_wait_sec 초 이내에 재고가 확보되지 않으면 데드락 탈출을 위해
        clamp consume 으로 진행한다 (stuck_wait_log 에 기록).
        """
        if max_wait_sec is None:
            max_wait_sec = MAX_DAYS * (
                _active_schedule['work_end_sec']
                - _active_schedule['work_start_sec']
                - _active_schedule['break_duration_sec']
            )
        c = str(item_code)
        blocked = False
        start_t = float(env.now)
        while self.stock[c] < qty:
            if not blocked and not c.endswith(THT_RAW_SUFFIX):
                self.violations[c] += 1
                blocked = True
            ev = env.event()
            self._wait_events[c].append(ev)
            remaining = max_wait_sec - (env.now - start_t)
            if remaining <= 0:
                break
            timeout_ev = env.timeout(remaining)
            result = yield env.any_of([ev, timeout_ev])
            if timeout_ev in result and ev not in result:
                # timeout 도달: 데드락 탈출. clamp consume 으로 진행 보장.
                self.stuck_wait_log.append({
                    'item_code'    : c,
                    'qty'          : int(qty),
                    'wait_start_h' : start_t / 3600,
                    'wait_end_h'   : float(env.now) / 3600,
                    'stock_at_end' : float(self.stock[c]),
                })
                # consume clamp (재고 부족해도 0으로 클램프 후 진행)
                self.consume(c, qty, env.now)
                return
        self.consume(c, qty, env.now)

    def _notify_waiters(self, item_code):
        # restore/replenish 호출 시 대기 중인 이벤트를 깨움.
        c = str(item_code)
        for ev in self._wait_events.pop(c, []):
            if not ev.triggered:
                ev.succeed()

    def stock_penalty(self):
        return sum(self.violations.values())

    def _lot_for(self, item_code):
        """부품별 발주 lot_size — 예상 demand × BOM_LOT_RATIO.

        demand 정보 없으면 (BOM 외 부품 등) data.get_lot_size 로 fallback.
        floor 는 WAREHOUSE_BOM_LOT_FLOOR (소량 부품도 너무 잦은 발주 방지).
        """
        c = str(item_code)
        d = self._demand.get(c, 0.0)
        if d > 0:
            return max(WAREHOUSE_BOM_LOT_FLOOR, int(d * BOM_LOT_RATIO))
        return int(self.data.get_lot_size(c))

    def replenish(self, item_code, sim_time=0, order_time=None, stock_at_order=None):
        c = str(item_code)
        if c.endswith(THT_RAW_SUFFIX):
            return
        # PCB 는 외부 발주 대상이 아님.
        if c in self._pcb_codes:
            self._pending_orders.discard(c)
            return
        lot      = self._lot_for(c)
        min_s    = self.data.get_min_stock(c)
        # 입고 후에도 min_stock_qty 미만이면 그 차이만큼 추가 입고
        incoming = max(lot, int(min_s) - self.stock[c] + lot)
        self.stock[c] += incoming
        self._pending_orders.discard(c)
        self.history[c].append((sim_time, self.stock[c]))
        # reorder_log: order_time 이 넘어오지 않으면 sim_time 을 fallback 으로 사용
        is_pcb = c in self._pcb_codes
        self.reorder_log.append({
            'item_code'     : c,
            'order_time'    : float(order_time) if order_time is not None else float(sim_time),
            'arrive_time'   : float(sim_time),
            'lot_size'      : int(lot),
            'incoming'      : int(incoming),
            'stock_at_order': (int(stock_at_order) if stock_at_order is not None
                               else int(self.stock[c] - incoming)),
            'is_pcb'        : is_pcb,   # True 면 PCB 가 외부 발주됨(버그 후보).
        })
        self.reorder_count[c] += 1
        if is_pcb:
            self.pcb_flow[c]['external_replenish_arrived'] += 1
        self._notify_waiters(c)

    def snapshot_loop(self, env, interval=3600):
        """1시간마다 추적 대상 부품의 재고를 스냅샷으로 찍는 SimPy 프로세스.
        추적 대상: (a) 현재 stock dict 에 등장한 부품 + (b) BOM 마스터 전체 item_code.
        소비·입고가 한 번도 없었던 부품도 초기재고로 행에 나와야 하므로 BOM 마스터 포함.
        """
        tracked = set()
        try:
            for c in self.data.iter_all_bom_items():
                tracked.add(c)
        except AttributeError:
            pass  # helper 미존재 시 stock dict 만 사용
        while True:
            yield env.timeout(interval)
            # stock dict 에 새로 등장한 code 도 편입
            tracked.update(self.stock.keys())
            now = env.now
            for c in tracked:
                # defaultdict lambda 트리거로 초기재고 반환
                q = self.stock[c]
                self.snapshots[c].append((now, int(q)))



class WIPTracker:
    """
    재고품 수량 상한 = 총 주문 x WIP_CAP_RATIO x 3 (임의, ConWIP 기반)
    Spearman et al. (1990) ConWIP; Ekerete et al. (2026 IRE Journals)
    """
    def __init__(self, order: dict):
        self.wip  = defaultdict(int)
        self.cap  = defaultdict(int)
        self.viol = defaultdict(int)
        self.history = defaultdict(list)
        # 1시간 간격 WIP 스냅샷: snapshots[grp] = [(t_sec, count), ...]
        self.snapshots = defaultdict(list)
        total = sum(order.values())
        for grp in ['MODULE','SEMI','SET','INSP','PACK','SMT']:
            self.cap[grp] = max(int(total * WIP_CAP_RATIO * 3), 10)  # (임의)

    def enter(self, grp):
        self.wip[grp] += 1
        if self.wip[grp] > self.cap.get(grp, 9999):
            self.viol[grp] += 1

    def leave(self, grp):
        self.wip[grp] = max(0, self.wip[grp] - 1)

    def violations(self):
        return sum(self.viol.values())

    def snapshot_loop(self, env, interval=3600):
        """1시간마다 그룹별 WIP 스냅샷 기록 (WIP_Timeseries 시트용)."""
        tracked_grps = ['SMT', 'MODULE', 'SEMI', 'SET', 'INSP', 'PACK', 'RMA']
        while True:
            yield env.timeout(interval)
            now = env.now
            for grp in tracked_grps:
                self.snapshots[grp].append((now, int(self.wip.get(grp, 0))))


# ══════════════════════════════════════════════════════════
# M05. 전력 로거
# ══════════════════════════════════════════════════════════

class EnergyLogger:
    def __init__(self, data):
        self.data   = data
        self.by_pc  = defaultdict(float)
        self.by_grp = defaultdict(float)
        self.total  = 0.0
        self.history = []

    def record(self, pc, grp, ct, sim_time=0, capacity=None):
        # 요청사항 0506 ②: 단위 공정 단위시간 소비 = 그룹 평균 / 동시 가동수.
        # capacity 미지정 시 PROCESS_GROUP_TO_WORKER_GROUP 매핑으로 self.data.workers 에서 추론.
        # SMT 라인 stage (SMT_LOADER_L1 등) 은 RATED_POWER_KW 가 정확 매칭 + cap=1 fallback.
        if capacity is None:
            wgrp = PROCESS_GROUP_TO_WORKER_GROUP.get(str(grp))
            if wgrp:
                capacity = int(self.data.workers.get(wgrp, 1))
            else:
                capacity = 1
        kw  = self.data.get_kw(pc, grp, capacity)
        kwh = kw * float(ct) / 3600
        self.by_pc[pc]   += kwh
        self.by_grp[grp] += kwh
        self.total       += kwh
        self.history.append((sim_time, self.total))
        return kwh

    def report(self):
        print('\n[공정그룹별 전력 소비 (kWh)]')
        print(f'  {"그룹":12s} {"kWh":>10s}')
        for g in sorted(self.by_grp):
            print(f'  {g:12s} {self.by_grp[g]:>10.4f}')
        print(f'  {"합계":12s} {self.total:>10.4f}')


# ══════════════════════════════════════════════════════════
# M06. 유휴·숙련도 추적기
# ══════════════════════════════════════════════════════════

class IdleTracker:
    """capacity 기반 per-person idle 적분 추적기.

    의도: "각 워커 개개인이, 근무시간에, 자기 그룹에 할 일이 남았는데 앞 공정
    병목으로 일을 못 하고 쉬는 시간" 의 person·sec 합계.

    동작:
      - 그룹 g 의 capacity = N. 매 시점 t 에 점유 중인 워커 수 = active(t).
        idle 워커 수 = max(N - active(t), 0).
      - acquire/release 이벤트 사이의 (cap - active) × Δt(근무시간) 를 적분.
      - 그룹별 자기 할당량 (set_target 으로 등록된 수) 처리 완료 시점
        (_completed_at[g]) 이후는 idle 카운트 X — "내 할 일 끝나면 idle 아님".

    호환:
      - 기존 mark_busy(env, name) API 는 SMT pc/AOI 등 cap=1 binary 추적용으로 유지.
        WORKER_GROUPS 대상은 acquire/release 로 교체됐고, mark_busy 결과는
        worker_idle_penalty 에 포함되지 않음 (그룹 capacity 등록된 항목만 합산).
    """
    def __init__(self):
        self._capacity     = {}                    # group → cap
        self._active       = defaultdict(int)      # group → 현재 점유 워커 수
        self._last_t       = defaultdict(float)    # group → 마지막 적분 시각
        self.total_idle    = defaultdict(float)    # group → 누적 person·sec
        self.absent_groups = set()
        self._completed_at = {}                    # group → 자기 할당량 완료 시각
        self._completion_target  = {}              # group → 처리해야 할 work item 총수
        self._completion_counter = defaultdict(int)
        self._last         = {}                    # 호환: mark_busy binary 추적

    def configure(self, capacity_map: dict):
        """그룹 capacity 등록. ManufacturingEnv.__init__ 에서 한 번 호출."""
        for g, cap in capacity_map.items():
            self._capacity[g] = int(cap)
            self._active.setdefault(g, 0)
            self._last_t.setdefault(g, 0.0)

    def set_target(self, g: str, target: int):
        """그룹별 처리해야 할 work item 총수 등록 (자기 할당량 완료 판정용)."""
        if target > 0:
            self._completion_target[g] = int(target)

    def _flush(self, env, g):
        """그룹 g 의 (last_t, env.now) 사이 idle 워커 person·sec 누적."""
        now = float(env.now)
        last = self._last_t.get(g, 0.0)
        if g not in self._capacity:
            self._last_t[g] = now
            return
        if now <= last:
            return
        cap    = self._capacity[g]
        active = self._active.get(g, 0)
        idle_workers = max(cap - active, 0)
        completed = self._completed_at.get(g)
        # 자기 할당량 완료 이후 구간은 idle 카운트 X
        if completed is not None and last >= completed:
            self._last_t[g] = now
            return
        eff_end = now if completed is None else min(now, completed)
        if eff_end > last and idle_workers > 0:
            self.total_idle[g] += idle_workers * _work_seconds_between(last, eff_end)
        self._last_t[g] = now

    def flush_all(self, env):
        """모든 등록 그룹의 마지막 구간까지 idle 반영. report/reward 직전 호출."""
        for g in list(self._capacity.keys()):
            self._flush(env, g)

    def acquire(self, env, g):
        """그룹 g 의 워커 1 명이 점유 시작 — yield res.request() 직후 호출."""
        if g not in self._capacity:
            self._capacity[g] = 1
            self._last_t.setdefault(g, float(env.now))
        self._flush(env, g)
        self._active[g] = self._active.get(g, 0) + 1

    def release(self, env, g):
        """그룹 g 의 워커 1 명이 점유 종료 — res.release() 직전 호출.

        그룹별 처리량 카운터 +=1. 등록된 target 도달 시 mark_completed 자동 발화 →
        그룹 g 의 자기 할당량 완료 시각이 정확히 잡힘 (전체 PACK 완성 시점이 아님).
        """
        if g not in self._capacity:
            return
        self._flush(env, g)
        self._active[g] = max(self._active.get(g, 0) - 1, 0)
        self._completion_counter[g] += 1
        target = self._completion_target.get(g)
        if (target is not None
                and self._completion_counter[g] >= target
                and g not in self._completed_at):
            self._completed_at[g] = float(env.now)

    def mark_busy(self, env, name):
        """호환용: SMT pc/AOI 등 cap=1 binary 추적. WORKER_GROUPS 는 acquire/release 사용."""
        now = env.now
        if name in self._last:
            self.total_idle[name] += _work_seconds_between(self._last[name], now)
        self._last[name] = now

    def mark_completed(self, wgrp: str, sim_time: float):
        existing = self._completed_at.get(wgrp)
        if existing is None or sim_time > existing:
            self._completed_at[wgrp] = float(sim_time)

    def worker_idle_penalty(self, threshold=300):
        """그룹 capacity 등록된 (= acquire/release 추적된) 그룹의 누적 idle 만 합산."""
        total = 0.0
        for name in self._capacity:
            if name not in WORKER_GROUPS:
                continue
            v = self.total_idle.get(name, 0.0)
            if v > threshold:
                total += v
        return total

    def report(self):
        print('\n[작업자 유휴 시간 상위 10 (그룹 person·hh:mm:ss)]')
        workers = {k: v for k, v in self.total_idle.items() if k in WORKER_GROUPS}
        for n, v in sorted(workers.items(), key=lambda x: -x[1])[:10]:
            cap  = self._capacity.get(n, 1)
            flag = ' <- 임계초과' if v > 1800 * cap else ''
            sec = int(max(v, 0))
            h, rem = divmod(sec, 3600)
            mm, ss = divmod(rem, 60)
            print(f'  {n:25s}(cap={cap:2d}): {h:02d}:{mm:02d}:{ss:02d}{flag}')


# ══════════════════════════════════════════════════════════
# M07. SMT 라인
# ══════════════════════════════════════════════════════════

class SolderCream:
    def __init__(self, env, name):
        self.env       = env
        self.name      = name
        self.stock_g   = SOLDER_G
        self.open_time = None

    def use(self):
        if self.open_time is None:
            self.open_time = self.env.now
        if self.env.now - self.open_time >= SOLDER_VALID_SEC:
            self.stock_g   = SOLDER_G
            self.open_time = self.env.now
        self.stock_g -= SOLDER_USE_G
        if self.stock_g <= 0:
            self.stock_g   = SOLDER_G
            self.open_time = self.env.now


class OutsourceTruckPool:
    """모든 SMT 라인이 공유하는 외주 트럭 풀.

    THT 보드를 종류 무관하게 동일 트럭에 적재 (옵션 b).
    (2026-05-06 변경) 트럭 적재량 제한 제거 — 트럭은 충분히 크다고 가정하고
    SMT 라인 종료 시점에 누적된 모든 보드를 한 번에 convoy 로 출발.
    TRUCK_SIZE 는 모니터링/엑셀 통계 표시 시 "트럭 1대당 환산 보드 수" 로만 사용.
    THT_DELAY_PROB 는 convoy 단위로 적용 (한 번 늦으면 모두 같이 늦음).
    """

    def __init__(self, env, wh, stats, smt_lines_ref):
        self.env = env
        self.wh = wh
        self.stats = stats
        self.smt_lines = smt_lines_ref   # {sid: SMTLine} — mag_buf / pcb_count 갱신용
        # 현재 적재 중인 트럭의 보드들. board: dict(pcb_code, model_id, board_id, line_sid, ev_idx)
        self.truck = []
        # 트럭 출발/도착 추적 (monitor 시각화용)
        self.dispatched_count = 0    # 누적 출발 트럭 수
        self.in_transit = []         # 도착 대기 트럭 list ({size, send_t, eta})
        # 출발한 트럭 단위 로그 (Truck_Log 시트용).
        # entry: {dispatch_id, send_t, eta, return_t, size, board_breakdown(dict)}
        self.truck_log = []

    def add_board(self, line_sid, pcb_code, model_id, board_id):
        """SMT 라인이 THT 보드 1장 외주 발사. 트럭에 적재. (자동 출발 안 함)"""
        raw_code = tht_raw_code(pcb_code)
        # raw 풀에 +1 (외주 진행 중 표식, _smt_schedule 의 in_flight 카운트 소스)
        self.wh.restore(raw_code, 1, self.env.now)
        self.stats['tht_out'] = self.stats.get('tht_out', 0) + 1
        ev_idx = len(self.wh.outsource_log)
        self.wh.outsource_log.append({
            'pcb_code'    : pcb_code,
            'model_id'    : model_id,
            'board_id'    : board_id,
            'send_time'   : float(self.env.now),  # 트럭 출발 시각으로 추후 갱신
            'return_time' : None,
            'delay_sec'   : 0.0,
            'status'      : 'queued',
        })
        self.truck.append({
            'pcb_code': pcb_code,
            'model_id': model_id,
            'board_id': board_id,
            'line_sid': line_sid,
            'ev_idx'  : ev_idx,
        })

    def flush_now(self):
        """SMT 라인 모두 종료 시 누적된 보드 한 번에 출발 (convoy)."""
        if self.truck:
            self._dispatch()

    def _dispatch(self):
        truck = self.truck
        self.truck = []
        # 트럭 단위 지연 (THT_DELAY_PROB 한 번 굴려 안의 모든 보드에 동일 적용)
        delay = 0.0
        if random.random() < THT_DELAY_PROB:
            delay = random.uniform(THT_DELAY_MIN_SEC, THT_DELAY_MAX_SEC)
            _log_event(self.env.now,
                       f'THT 외주 트럭 납기 지연: {len(truck)}보드 +{delay/3600:.1f}h')
        send_t = float(self.env.now)
        for entry in truck:
            log = self.wh.outsource_log[entry['ev_idx']]
            log['send_time'] = send_t
            log['delay_sec'] = float(delay)
            log['status']    = 'in_flight'
        eta = send_t + THT_OUTSOURCE_SEC + delay
        self.dispatched_count += 1
        # 보드 종류 분포
        breakdown = defaultdict(int)
        for entry in truck:
            breakdown[entry['pcb_code']] += 1
        truck_log_entry = {
            'dispatch_id'  : self.dispatched_count,
            'send_t'       : send_t,
            'eta'          : eta,
            'delay_sec'    : float(delay),
            'size'         : len(truck),
            'breakdown'    : dict(breakdown),
            'truck_count'  : max(1, (len(truck) + TRUCK_SIZE - 1) // TRUCK_SIZE),
            'return_t'     : None,
        }
        self.truck_log.append(truck_log_entry)
        transit_meta = {'size': len(truck), 'send_t': send_t, 'eta': eta,
                        'log_entry': truck_log_entry}
        self.in_transit.append(transit_meta)
        _log_event(self.env.now,
                   f'THT 외주 convoy #{self.dispatched_count} 출발: '
                   f'{len(truck)}보드 (트럭 환산 {truck_log_entry["truck_count"]}대, '
                   f'도착 +{(eta - send_t)/3600:.1f}h)')
        self.env.process(self._truck_arrive(truck, delay, transit_meta))

    def _truck_arrive(self, truck, delay, transit_meta):
        yield self.env.timeout(THT_OUTSOURCE_SEC + delay)
        try:
            self.in_transit.remove(transit_meta)
        except ValueError:
            pass
        log_entry = transit_meta.get('log_entry')
        if log_entry is not None:
            log_entry['return_t'] = float(self.env.now)
        for entry in truck:
            pcb_code = entry['pcb_code']
            line_sid = entry['line_sid']
            line = self.smt_lines.get(line_sid)
            raw_code = tht_raw_code(pcb_code)
            self.wh.consume(raw_code, 1, self.env.now)
            log = self.wh.outsource_log[entry['ev_idx']]
            log['return_time'] = float(self.env.now)
            log['status']      = 'returned'
            self.wh.pcb_flow[pcb_code]['outsource_returned'] += 1
            if line is None:
                # 라인 없으면 (안전) 보드 단위 즉시 inventory
                self.wh.restore(pcb_code, 1, self.env.now)
                continue
            line.mag_buf[pcb_code] += 1
            in_flight = (self.wh.pcb_flow[pcb_code].get('outsource_in', 0)
                         - self.wh.pcb_flow[pcb_code].get('outsource_returned', 0))
            if line.mag_buf[pcb_code] >= MAG_SIZE:
                flush = MAG_SIZE
            elif in_flight == 0 and line.mag_buf[pcb_code] > 0:
                flush = line.mag_buf[pcb_code]
            else:
                flush = 0
            if flush > 0:
                line.mag_buf[pcb_code] -= flush
                self.wh.restore(pcb_code, flush, self.env.now)
                line.pcb_count[pcb_code] += flush
                self.stats['pcb_done'] = self.stats.get('pcb_done', 0) + flush
                self.wh.smt_per_model[(line_sid, entry['model_id'], pcb_code)] += flush
        _log_event(self.env.now, f'THT 외주 트럭 도착: {len(truck)}보드')


class SMTLine:
    # 혼류 금지: 한 라인에서 한 모델씩만 처리.
    def __init__(self, env, suffix, data, wh,
                 aoi_res, rma_store, energy, idle, wip, stats, broken_flag,
                 outsource_pool=None):
        self.env          = env
        self.sfx          = suffix
        self.data         = data
        self.wh           = wh
        self.aoi_res      = aoi_res
        self.rma          = rma_store
        self.outsource_pool = outsource_pool   # OutsourceTruckPool (모든 라인 공유)
        self.energy       = energy
        self.idle         = idle
        self.wip          = wip
        self.stats        = stats
        self.broken_flag  = broken_flag   # {pc: bool} 공유 딕셔너리
        self.solder       = SolderCream(env, f'SMT_{suffix}')
        self.mag_buf      = defaultdict(int)
        self.assigned_model = None
        self._res = {pc: simpy.Resource(env, capacity=1) for pc in [
            f'SMT_LOADER_{suffix}',    f'SMT_PRINTER_{suffix}',
            f'SMT_SPI_{suffix}',       f'SMT_MOUNTER_H_{suffix}',
            f'SMT_MOUNTER_M_{suffix}', f'SMT_REFLOW_{suffix}',
            f'SMT_UNLOADER_{suffix}']}
        self.pcb_count = defaultdict(int)
        # monitor 시각화용: 각 stage 자원에서 현재 처리 중인 보드 정보.
        # pc → (pcb_code, board_id, is_second). 자원 release 시 pop.
        self.stage_active = {}
        # SMT_Stage_Activity 시트용 stage 별 events.
        # entry: {pc, pcb_code, board_id, is_second, start, end}
        self.stage_events = []

    def process_board(self, pcb_code, board_id, model_id, is_second=False):
        """SMT 보드 1장 처리.

        결함 처리 (2026-05-06): 조립 공정의 INSP 검출 로직과 동일하게,
        LOADER~UNLOADER stage 중 어느 곳에서든 결함이 발생하면 board_has_defect
        플래그만 표시하고 보드는 계속 진행한다. 마지막 SMT_AOI 가 누적 결함을
        자체 dr 확률로 검출하면 RMA 큐 (src='AOI') 로 투입한다 — 이 경로에서
        AOI_DEFECT_ACTION 으로 수리/폐기 토글 가능.
        """
        seq = [f'SMT_LOADER_{self.sfx}',    f'SMT_PRINTER_{self.sfx}',
               f'SMT_SPI_{self.sfx}',       f'SMT_MOUNTER_H_{self.sfx}',
               f'SMT_MOUNTER_M_{self.sfx}', f'SMT_REFLOW_{self.sfx}',
               f'SMT_UNLOADER_{self.sfx}']

        self.wip.enter('SMT')
        # board_has_defect: stage 중 한번이라도 결함이 발생했는지 (AOI 검출 조건).
        # 조립 공정의 unit_defect_flag 와 동일 의미.
        board_has_defect = False
        for pc in seq:
            pr = self.data.get_proc(pc)
            if pr is None:
                yield self.env.timeout(0.001)
                continue
            ct = float(pr['cycle_time_sec'] or 0.001)
            tt = float(pr['transfer_time_sec'] or 0)
            dr = float(pr['defect_rate'] or DEFECT_FLOOR)

            if not _is_work_time(self.env.now):
                yield self.env.timeout(_next_work_start(self.env.now) - self.env.now)

            # 설비 고장 중이면 복구까지 대기
            while self.broken_flag.get(pc, False):
                yield self.env.timeout(60)

            self.idle.mark_busy(self.env, pc)
            with self._res[pc].request() as req:
                yield req
                _start_t = float(self.env.now)
                self.stage_active[pc] = (pcb_code, board_id, is_second)
                try:
                    act = max(random.normalvariate(ct, ct * CT_STD_RATIO), ct * 0.5)
                    yield self.env.timeout(act)
                finally:
                    self.stage_active.pop(pc, None)
                    self.stage_events.append({
                        'pc': pc, 'pcb_code': pcb_code,
                        'model_id': model_id, 'board_id': board_id,
                        'is_second': is_second,
                        'start': _start_t, 'end': float(self.env.now),
                    })
            self.energy.record(pc, 'SMT', act, self.env.now)
            if 'PRINTER' in pc:
                self.solder.use()
            if tt > 0:
                yield self.env.timeout(tt)
            # stage 결함 발생 시 플래그만 표시, 보드는 계속 진행 (AOI 가 검출).
            if random.random() < dr:
                self.stats['smt_defect'] += 1
                board_has_defect = True

        # SMT_AOI row 는 FallbackDataLoader._PF_ALL_ROWS 가 항상 제공.
        aoi = self.data.get_proc('SMT_AOI')
        aoi_ct = float(aoi['cycle_time_sec'] or 30)
        aoi_dr = float(aoi['defect_rate'] or DEFECT_FLOOR)

        if not _is_work_time(self.env.now):
            yield self.env.timeout(_next_work_start(self.env.now) - self.env.now)

        self.idle.mark_busy(self.env, 'SMT_AOI')
        with self.aoi_res.request() as req:
            yield req
            _aoi_start_t = float(self.env.now)
            self.stage_active['SMT_AOI'] = (pcb_code, board_id, is_second)
            try:
                yield self.env.timeout(aoi_ct)
            finally:
                self.stage_active.pop('SMT_AOI', None)
                self.stage_events.append({
                    'pc': 'SMT_AOI', 'pcb_code': pcb_code,
                    'model_id': model_id, 'board_id': board_id,
                    'is_second': is_second,
                    'start': _aoi_start_t, 'end': float(self.env.now),
                })
        self.energy.record('SMT_AOI', 'SMT_SHARED', aoi_ct, self.env.now)
        # AOI 검출: stage 중 누적 결함이 있을 때만 dr 확률로 검출 → RMA.
        # 결함 없는 보드는 dr 확률과 무관하게 정상 통과.
        # (조립 공정 INSP 의 has_prior_defect 검사 로직과 동일)
        detected = board_has_defect and (random.random() < aoi_dr)
        if detected:
            self.stats['aoi_defect'] += 1
            self.wip.leave('SMT')
            yield self.rma.put({'src':'AOI','board':board_id,
                                'pcb':pcb_code,'grp':'SMT_SHARED','model':model_id})
            return

        self.wip.leave('SMT')

        # 2026-04-29 B5 수정: 양면 PCB 의 두번째 패스를 재귀 호출.
        # 이전 코드는 첫 패스 후 mag_buf 에 적재하고 return 했지만, 두번째 패스
        # 호출 코드가 어디에도 없어 양면 PCB 가 BACK 면 SMT 를 영영 안 하고
        # 끝나는 버그. 또 양면 + THT PCB (03902715, 03902730, 03903424) 는 외주
        # 분기에 도달조차 못해 외주가 0건이었다.
        # 새 로직: 첫 패스 끝에서 두번째 패스를 yield 로 재호출. 두번째 패스
        # 안에서 THT 또는 main 적재 분기가 실행됨. 적재 시점 일관성 유지.
        side = self.data.smt_side(pcb_code)
        if not is_second and side == 'double':
            self.wh.pcb_flow[pcb_code]['double_first_pass'] += 1
            yield self.env.process(
                self.process_board(pcb_code, board_id, model_id,
                                    is_second=True))
            return

        # 안전 로직 (B5): 적재 시점에서 양면 PCB 인데 첫 패스로 도달했다면
        # 두번째 패스 누락. 위 분기에서 재귀 호출되므로 정상 흐름에선 도달
        # 불가능. BOM smt_side 캐시가 잘못됐거나 호출자가 직접 is_second
        # 우회 호출했을 때만 이 로그가 찍힘. 적재는 그대로 진행.
        if side == 'double' and not is_second:
            self.wh.smt_single_side_log.append({
                'pcb_code': pcb_code, 'model_id': model_id,
                'board_id': board_id, 'time_h': self.env.now/3600,
                'reason': 'double_pcb_flushed_without_second_pass',
            })

        if pcb_code in THT_PCB:
            # (2026-05-06) THT 외주는 OutsourceTruckPool 가 관리.
            # 보드를 트럭에 적재만 하고 즉시 return — process_board 종료.
            # 트럭이 TRUCK_SIZE 도달 시 자동 출발, 16h 후 도착 처리는 pool 측에서.
            if self.outsource_pool is not None:
                self.outsource_pool.add_board(self.sfx, pcb_code, model_id, board_id)
            else:
                # outsource_pool 미주입 (이론상 없어야 함) — 안전 fallback
                self.wh.restore(pcb_code, 1, self.env.now)
            return

        self.mag_buf[pcb_code] += 1
        if self.mag_buf[pcb_code] >= MAG_SIZE:
            self.mag_buf[pcb_code] -= MAG_SIZE
            self.wh.restore(pcb_code, MAG_SIZE, self.env.now)
            self.pcb_count[pcb_code] += MAG_SIZE
            self.stats['pcb_done'] += MAG_SIZE
            self.wh.smt_per_model[(self.sfx, model_id, pcb_code)] += MAG_SIZE


# ████████████████████████████████████████████████████████████████████
# §C. SIMULATION
# ████████████████████████████████████████████████████████████████████
# 책임: SimPy 프로세스로 도메인 객체의 상태를 시간축에 따라 진행. 모든
# yield env.timeout / Resource.request 가 이 섹터에 모인다.
# 외부 export: run_process, run_rma, produce_unit, monitor,
#              ProcessActivityLogger, _is_work_time, work_timeout,
#              _next_work_start, _work_seconds_between.
# 향후 파일 분할 위치: cpro/sim/{schedule, process, rma, produce, monitor}.py


# ══════════════════════════════════════════════════════════
# 근무 시간 유틸
# ══════════════════════════════════════════════════════════

def _is_work_time(sim_now_sec):
    _s = _active_schedule
    t  = sim_now_sec % DAY_SEC
    if not (_s['work_start_sec'] <= t < _s['work_end_sec']):
        return False
    if _s['lunch_start_sec'] <= t < _s['lunch_end_sec']:
        return False
    return True


def _next_work_start(sim_now_sec):
    _s      = _active_schedule
    day_num = int(sim_now_sec // DAY_SEC)
    t       = sim_now_sec % DAY_SEC
    if _s['lunch_start_sec'] <= t < _s['lunch_end_sec']:
        return day_num * DAY_SEC + _s['lunch_end_sec']
    if t < _s['work_start_sec']:
        return day_num * DAY_SEC + _s['work_start_sec']
    return (day_num + 1) * DAY_SEC + _s['work_start_sec']


def work_timeout(env, duration):
    """근무시간 내에서만 경과. 점심·퇴근 boundary 만나면 재개 시점까지 pause."""
    _s        = _active_schedule
    remaining = float(max(duration, 0))
    while remaining > 1e-6:
        if not _is_work_time(env.now):
            yield env.timeout(_next_work_start(env.now) - env.now)
            continue
        t = env.now % DAY_SEC
        if t < _s['lunch_start_sec']:
            next_boundary = env.now - t + _s['lunch_start_sec']
        else:  # lunch_end_sec <= t < work_end_sec
            next_boundary = env.now - t + _s['work_end_sec']
        chunk = min(remaining, max(next_boundary - env.now, 0))
        if chunk <= 0:
            yield env.timeout(_next_work_start(env.now) - env.now)
            continue
        yield env.timeout(chunk)
        remaining -= chunk


def _work_seconds_between(start_sec, end_sec):
    """두 시뮬레이션 시각 사이의 실 근무시간(초)을 반환.
    야간·점심을 제외한 순수 근무시간만 계산. IdleTracker.mark_busy 에서 호출.
    """
    _s = _active_schedule
    if end_sec <= start_sec:
        return 0.0
    total = 0.0
    t     = float(start_sec)
    end   = float(end_sec)
    while t < end:
        if not _is_work_time(t):
            nxt = _next_work_start(t)
            if nxt >= end:
                break
            t = nxt
            continue
        day_t = t % DAY_SEC
        if day_t < _s['lunch_start_sec']:
            boundary = t - day_t + _s['lunch_start_sec']
        else:
            boundary = t - day_t + _s['work_end_sec']
        chunk = min(end, boundary) - t
        if chunk > 0:
            total += chunk
        t = min(end, boundary)
    return total


# ══════════════════════════════════════════════════════════
# M08. 단일 공정 실행
# ══════════════════════════════════════════════════════════

def run_process(env, prow, done_ev, wres, wh, rma, energy,
                idle, wip, stats, mid, data, plogger=None, uid=0,
                unit_defect_flag=None, progress=None):
    """단일 공정 실행.

    unit_defect_flag : {uid: bool} 공유 딕셔너리.
        비INSP 공정에서 불량 발생 시 True 로 표시하고 공정은 계속 진행.
        INSP 공정에서 플래그가 있으면 defect_rate 확률로 검출 -> RMA 투입.
    progress : ManufacturingEnv.progress 공유 딕셔너리. RMA 투입 시 전달.
    """
    pc    = str(prow['process_code'])
    grp   = str(prow.get('process_group','') or '')
    ct    = float(prow['cycle_time_sec'] or 0)
    dr    = float(prow['defect_rate'] or DEFECT_FLOOR)
    tt    = float(prow['transfer_time_sec'] or 0)
    whr   = float(prow['dep_wait_hr'] or 0)
    wgrp  = str(prow.get('worker_group','') or '')
    dtype = str(prow.get('dep_type','SEQUENCE') or 'SEQUENCE').upper()
    prevs = [p.strip() for p in
             str(prow.get('dep_prev_codes','') or '').split(';') if p.strip()]

    # 요청사항 0506 ③: SET 의 INSP 공정은 자원·가동률만 WORKER_SET_INSP 로 분리,
    # 전력은 SET 풀(33.67kW / WORKER_SET cap 16)로 통합 처리.
    # → wgrp 만 변경, grp 는 'SET' 유지.
    if wgrp == 'WORKER_SET' and grp == 'SET':
        if pc.rsplit('_', 1)[-1].upper() == 'INSP':
            wgrp = 'WORKER_SET_INSP'

    # 숙련도 보정: data 에서 읽은 AAS SkillLevelType 값 사용
    skill_base   = data.worker_skill.get(wgrp, 2)
    skill_actual = max(1, skill_base - (1 if wgrp in idle.absent_groups else 0))
    ct = ct * data.skill_ct.get(skill_actual, 1.0)
    dr = dr * data.skill_dr.get(skill_actual, 1.0)

    if prevs:
        wait_evs = [done_ev[p] for p in prevs if p in done_ev]
        if dtype == 'JOIN':
            if wait_evs:
                yield simpy.AllOf(env, wait_evs)
        elif wait_evs:
            yield wait_evs[0]

    if whr > 0:
        yield env.timeout(whr * 3600)
    if tt > 0:
        yield env.timeout(tt)

    if not _is_work_time(env.now):
        yield env.timeout(_next_work_start(env.now) - env.now)

    wip.enter(grp)
    # (2026-05-06) BOM 부품을 워커 확보 *전* 에 먼저 wait_stock 으로 받는다.
    # 이전 버그: 워커 점유 후 wait_stock 하면, 부품 도착 대기 동안 워커가
    # 묶여 다른 unit 들이 그 워커 풀을 못 쓰고 라인 전체가 stall. 부품이
    # 들어와야 워커 잡으러 가도록 순서를 뒤집어 라인 throughput 회복.
    for item_code, qty in data.get_bom_parts(mid, pc):
        yield from wh.wait_stock(env, item_code, qty)

    res = wres.get(wgrp)
    req = res.request() if res else None
    acquired = False
    if req:
        yield req
        idle.acquire(env, wgrp)
        acquired = True

    logger_started = False
    ev_id_p = None
    try:
        # 자원 확보 후 실제 가동 구간만 '진행중' 으로 기록
        if plogger is not None:
            _cap = res.capacity if res is not None else 1
            ev_id_p = plogger.mark_start(pc, mid, uid, env.now, grp,
                                         wgrp=wgrp, cap=_cap,
                                         work_timed=True)
            logger_started = True
        act = max(random.normalvariate(ct, ct * CT_STD_RATIO), ct * 0.5) if ct > 0 else 0.001
        # cycle_time 은 근무시간 경계를 준수하도록 work_timeout 으로 실행한다.
        yield from work_timeout(env, act)
        energy.record(pc, grp, act, env.now)
        # ── 불량 처리: 발생과 검출 분리 ──────────────────────────────
        # 비INSP 공정: 불량이 발생해도 공정은 계속 진행하고 unit_defect_flag 표시.
        # INSP 공정:   unit_defect_flag가 있을 때만 defect_rate 확률로 검출 -> RMA.
        _is_insp = (grp == 'SET' and pc.rsplit('_', 1)[-1].upper() == 'INSP')

        if _is_insp:
            # INSP 공정: 이전 공정에서 누적된 불량을 이 확률로 검출
            has_prior_defect = (unit_defect_flag is not None
                                and unit_defect_flag.get(uid, False))
            detected = has_prior_defect and (random.random() < dr)
            if detected:
                stats['assy_defect'] = stats.get('assy_defect', 0) + 1
                if unit_defect_flag is not None:
                    unit_defect_flag[uid] = False
                    # PACK 이중 실행 방지: RMA 경로가 PACK 을 담당하므로 produce_unit 에 종료 신호.
                    unit_defect_flag['_routed_to_rma'] = True
                wip.leave(grp)
                yield rma.put({'src': pc, 'grp': grp, 'model': mid,
                               'uid': uid,
                               'progress': progress})
                if req and res:
                    if acquired:
                        idle.release(env, wgrp)
                        acquired = False
                    res.release(req)
                ev = done_ev.get(pc)
                if ev and not ev.triggered:
                    ev.succeed()
                return
        else:
            # 비INSP 공정: 불량 발생 시 플래그만 표시, 공정 계속 진행
            if random.random() < dr:
                stats['assy_defect'] = stats.get('assy_defect', 0) + 1
                if unit_defect_flag is not None:
                    unit_defect_flag[uid] = True
    finally:
        if plogger is not None and logger_started:
            try:
                plogger.mark_end(pc, env.now, ev_id_p)
            except Exception:
                pass
        if req and res:
            try:
                if acquired:
                    idle.release(env, wgrp)
                    acquired = False
                res.release(req)
            except Exception:
                pass

    wip.leave(grp)
    ev = done_ev.get(pc)
    if ev and not ev.triggered:
        ev.succeed()


# ══════════════════════════════════════════════════════════
# M09. RMA 수리 및 재투입
# ══════════════════════════════════════════════════════════

def run_rma(env, rma, wres, wh, energy, idle, wip, stats, data,
            progress=None, plogger=None):
    """RMA 큐 디스패처.

    AOI 보드 불량 처리 — AOI_DEFECT_ACTION 으로 분기:
      'repair' (기본): RMA 수리 후 wh.restore() 로 PCB 인벤토리 직접 재투입.
      'scrap'        : 폐기. SMT scheduler 가 추가 생산해 보전.
    (2026-05-06) SMT stage 자체 결함은 더 이상 즉시 폐기되지 않음 — process_board
    가 board_has_defect 플래그만 표시하고 AOI 가 검출하므로 모든 SMT 결함 경로는
    src='AOI' 로 통합됨.
    INSP 공정에서 검출된 유닛 불량은 _rma_repair_and_reinsert 로 수리 후 PACK 재투입.
    """
    while True:
        item = yield rma.get()
        src = str(item.get('src', ''))
        if src == 'AOI':
            if AOI_DEFECT_ACTION == 'repair':
                # AOI 불량 보드: RMA 수리 후 PCB 인벤토리 직접 적재 (검사 생략)
                env.process(_rma_repair_aoi_board(
                    env, item, wres, wh, energy, idle, wip, stats, data))
                continue
            else:
                # 'scrap': 폐기. SMT scheduler 가 부족 검출 시 +1 생산.
                stats['smt_rma_scrap'] = stats.get('smt_rma_scrap', 0) + 1
                _log_event(env.now,
                           f'AOI 보드 불량 폐기: pcb={item.get("pcb","")} '
                           f'model={item.get("model","")} '
                           f'board={item.get("board","")}')
                continue
        # INSP 불량(또는 기타 unit-level): 수리 + PACK 재투입
        env.process(_rma_repair_and_reinsert(
            env, item, wres, wh, energy, idle, wip, stats, data,
            progress=progress, plogger=plogger))


def _rma_repair_aoi_board(env, item, wres, wh, energy, idle, wip, stats, data):
    """AOI 검출 불량 SMT 보드 수리 → PCB 인벤토리 직접 재투입 (요청사항 0506 ④).

    수리 시 부품 소모 (2026-05-06): PCB 분해·재납땜 작업이므로 그 PCB 의
    BOM 부품 (HierarchicalStructures PCB-typed) 한 세트를 다시 소모한다.
    부품 부족하면 wait_stock 으로 대기 (RMA 수리 비동기 대기).
    """
    pcb_code = str(item.get('pcb', ''))
    model    = str(item.get('model', ''))
    res = wres.get('WORKER_RMA')
    req = res.request() if res else None
    acquired = False
    try:
        if req:
            yield req
            idle.acquire(env, 'WORKER_RMA')
            acquired = True
        rt = max(random.normalvariate(RMA_REPAIR_TIME_MEAN_SEC,
                                       RMA_REPAIR_TIME_STD_SEC),
                 RMA_REPAIR_TIME_MIN_SEC)
        yield from work_timeout(env, rt)
        energy.record('RMA_REPAIR', 'RMA', rt, env.now)
        stats['aoi_repaired'] = stats.get('aoi_repaired', 0) + 1

        # PCB 부품 한 세트 재소모 (분해 후 모든 SMT 부품 새 걸로 갈아끼움 가정).
        if pcb_code:
            for part_code, part_qty in data.get_pcb_parts(pcb_code):
                yield from wh.wait_stock(env, part_code, part_qty)
            wh.restore(pcb_code, 1, env.now)
        _log_event(env.now,
                   f'AOI 불량 수리 완료 → PCB 재투입: pcb={pcb_code} model={model}')
    finally:
        if req:
            if acquired:
                idle.release(env, 'WORKER_RMA')
            res.release(req)


def _sample_defective_predecessor(data, model_id, src_pc):
    """src_pc(INSP) 의 트랜지티브 선행 공정 중 F/W~SET (MODULE/SEMI/SET 그룹,
    INSP suffix 제외) 후보를 모아 defect_rate 가중치로 1개 sampling.

    "INSP 가 잡아낸 결함이 어느 조립 공정에서 발생했는지 모르므로 (기록 없음)
    그 INSP 앞 단계의 비-INSP 공정 중 불량률에 비례해 1개 추정" 의도.
    가중치 합이 0 이거나 후보가 없으면 None.
    """
    try:
        procs = data.get_model_procs(model_id)
    except Exception:
        return None
    prow_map = {str(r['process_code']): r for _, r in procs.iterrows()}

    # BFS 로 src_pc 의 모든 선행 공정 수집
    visited = set()
    queue = [str(src_pc)]
    candidates = []
    while queue:
        cur = queue.pop(0)
        prow = prow_map.get(cur)
        if prow is None:
            continue
        prevs = [p.strip() for p in
                 str(prow.get('dep_prev_codes', '') or '').split(';') if p.strip()]
        for prev in prevs:
            if prev in visited:
                continue
            visited.add(prev)
            queue.append(prev)
            prev_row = prow_map.get(prev)
            if prev_row is None:
                continue
            grp  = str(prev_row.get('process_group', '') or '')
            wgrp = str(prev_row.get('worker_group', '') or '')
            # F/W ~ SET 범위 = MODULE / SEMI / SET. INSP 성격 공정은 모두 제외:
            #  (1) Excel 식 'XXX_INSP' 접미사
            #  (2) AAS 식 worker_group=WORKER_SET_INSP (VD7Inspection 그룹의 자식)
            if grp not in ('MODULE', 'SEMI', 'SET'):
                continue
            if prev.rsplit('_', 1)[-1].upper() == 'INSP':
                continue
            if wgrp == 'WORKER_SET_INSP':
                continue
            candidates.append(prev_row)

    if not candidates:
        return None
    weights = [max(float(c.get('defect_rate', 0) or 0), 0.0) for c in candidates]
    if sum(weights) <= 0:
        return None
    chosen = random.choices(candidates, weights=weights, k=1)[0]
    return str(chosen['process_code'])


def _rma_repair_and_reinsert(env, item, wres, wh, energy, idle, wip, stats, data,
                             progress=None, plogger=None):
    """RMA 수리 후 PACK 첫 공정으로 재투입.

    흐름:
      SET/INSP 공정 불량 발생
        -> run_rma 큐에서 꺼냄
        -> WORKER_RMA 자원 확보
        -> 수리 (임의 평균 300초)
        -> 해당 모델의 PACK 첫 공정(_find_pack_entry 자동 추출) 실행
        -> 완성 카운터 증가
    """
    if not _is_work_time(env.now):
        yield env.timeout(_next_work_start(env.now) - env.now)

    res = wres.get('WORKER_RMA')
    req = res.request() if res else None
    acquired = False
    if req:
        yield req
        idle.acquire(env, 'WORKER_RMA')
        acquired = True

    model    = item.get('model', '')
    src_pc   = item.get('src', '')
    rma_uid  = int(item.get('uid', 0))
    # item에 담긴 progress 우선, 없으면 인자로 받은 progress 사용
    progress = item.get('progress') or progress or {}
    pc_rma   = 'RMA_REPAIR'
    ev_id_p  = None

    try:
        if plogger is not None:
            _cap = res.capacity if res is not None else 1
            ev_id_p = plogger.mark_start(pc_rma, model or '-', rma_uid,
                                         env.now, 'RMA',
                                         wgrp='WORKER_RMA', cap=_cap,
                                         work_timed=True)
        # 수리 시간: 평균/표준편차/최소값은 M01 상단 상수로 관리.
        rt = max(random.normalvariate(RMA_REPAIR_TIME_MEAN_SEC,
                                       RMA_REPAIR_TIME_STD_SEC),
                 RMA_REPAIR_TIME_MIN_SEC)
        yield from work_timeout(env, rt)
        energy.record('RMA_REPAIR', 'RMA', rt, env.now)
        stats['rma_repaired'] = stats.get('rma_repaired', 0) + 1

        # 수리 시 교체 부품 소모 (2026-05-06 재구현):
        # INSP 는 직전 조립 공정의 결함을 검출. 어느 공정에서 불량이 났는지
        # 추적 정보가 없으므로 src_pc(INSP) 의 트랜지티브 선행 공정 중
        # MODULE/SEMI/SET 그룹의 비-INSP 공정을 defect_rate 가중치로 1개 sampling
        # 한 뒤, 그 공정의 BOM 부품 한 세트를 창고에서 차감한다.
        # 수리는 반조립품 분해·교체 작업이므로 부품이 추가 소모된다.
        # 부품 부족하면 wait_stock 으로 차단 — 부품 없으면 대기.
        if src_pc:
            sampled_pc = _sample_defective_predecessor(data, model, src_pc)
            if sampled_pc:
                for part_code, part_qty in data.get_bom_parts(model, sampled_pc):
                    yield from wh.wait_stock(env, part_code, part_qty)
                _log_event(env.now,
                           f'RMA 추정 결함공정={sampled_pc} → BOM 한 세트 재소모')

        _log_event(env.now, f'RMA 수리 완료: {src_pc} ({model}) -> PACK 진입')
    finally:
        if plogger is not None and ev_id_p is not None:
            try:
                plogger.mark_end(pc_rma, env.now, ev_id_p)
            except Exception:
                pass
        if req and res:
            try:
                if acquired:
                    idle.release(env, 'WORKER_RMA')
                    acquired = False
                res.release(req)
            except Exception:
                pass

    # 수리 완료 후 PACK 첫 공정으로 진입.
    # AAS process flow 에서 자동 추출 (PACK 그룹 중 dep_prev 가 INSP).
    pack_pc = _find_pack_entry(data, model)
    if pack_pc is None:
        # PACK 공정 정보를 찾지 못한 경우 완성 카운터만 증가
        _do_complete(stats, progress, model, env.now, wh=wh, src_pc=src_pc)
        return

    prow = data.get_proc(pack_pc)
    if prow is None:
        _do_complete(stats, progress, model, env.now, wh=wh, src_pc=src_pc)
        return

    # PACK 공정 실행 - 선행 공정 의존성 없이 즉시 시작
    # PACK 내 연속 공정들은 prow의 dep_prev_codes 를 통해 자체적으로 연결됨.
    # RMA 경로에서 진입하므로 첫 PACK 공정의 선행 조건은 이미 충족된 것으로 처리.
    done_ev_rma = {}
    # PACK 연속 공정 전체를 순서대로 실행
    pack_pcs = _get_pack_sequence(data, model, pack_pc)
    for p_pc in pack_pcs:
        p_prow = data.get_proc(p_pc)
        if p_prow is None:
            continue
        p_prow_copy = p_prow.copy()
        p_prow_copy['dep_prev_codes'] = ''  # 선행 의존성 제거
        p_prow_copy['dep_wait_hr']    = 0
        done_ev_rma[p_pc] = env.event()
        yield env.process(
            run_process(env, p_prow_copy, done_ev_rma, wres, wh,
                        simpy.Store(env), energy, idle, wip, stats, model, data,
                        plogger=plogger, uid=rma_uid,
                        unit_defect_flag=None,  # 수리 완료 부품 - 불량 없음
                        progress=progress))

    # PACK 완료 후 완성 카운터 증가
    _do_complete(stats, progress, model, env.now, wh=wh, src_pc=src_pc)


def _do_complete(stats, progress, model, now, wh=None, src_pc=None):
    # RMA 경로 완성 카운터 처리.
    if not model or model not in progress:
        return
    done, total = progress[model]
    # 주문 수량 초과 생산 가드 (2026-04-28). RMA·PACK 재진입 경로에서
    # done >= total 이후에도 호출될 수 있어 카운터 폭주 방지.
    if done >= total:
        # 디버그: RMA 경로가 quota 초과로 차단된 경우.
        if wh is not None:
            wh.unit_completions[('RMA', f'{model}_{src_pc}_{int(now)}')] = {
                'path': 'rma_blocked_by_quota',
                'end_time': float(now),
                'done_n': -1,
                'total_n': -1,
                'rma_count': 1,
            }
        return
    stats[f'{model}_done'] = stats.get(f'{model}_done', 0) + 1
    progress[model] = (done + 1, total)
    pct = (done + 1) / max(total, 1) * 100
    _log_event(now, f'{model} RMA->PACK 완성 ({pct:.0f}%)')
    # 디버그: RMA 경로 완성 기록. 원래 unit_id 를 추적하기 어려워 src_pc 기반 가짜 키.
    if wh is not None:
        # RMA 경로에서 정확한 unit_id 를 확정하기 어려우므로 별도 (model,'rma_N') 키로 누적.
        rma_idx = sum(1 for k in wh.unit_completions
                      if isinstance(k, tuple) and k[0] == model and
                      isinstance(k[1], str) and k[1].startswith('rma_'))
        wh.unit_completions[(model, f'rma_{rma_idx}')] = {
            'path': 'rma',
            'end_time': float(now),
            'done_n': -1,
            'total_n': -1,
            'rma_count': 1,
            'src_pc': str(src_pc) if src_pc else '',
        }


def _get_pack_sequence(data, model, first_pack_pc):
    """PACK 공정 연속 순서 반환. first_pack_pc 부터 시작해 SEQUENCE로 연결된 공정들.

    CombinedDataLoader / FallbackDataLoader 양쪽을 지원.
    get_model_procs() 가 있으면 해당 메서드 사용, 없으면 빈 리스트.
    """
    result = [first_pack_pc]
    visited = {first_pack_pc}
    try:
        procs = data.get_model_procs(model)
    except Exception:
        return result

    # process_code -> dep_prev_codes 역방향 매핑 (다음 공정 찾기용)
    next_map = {}
    for _, row in procs.iterrows():
        pc   = str(row['process_code'])
        prev = str(row.get('dep_prev_codes', '') or '')
        grp  = str(row.get('process_group', '') or '')
        if grp != 'PACK':
            continue
        for p in [x.strip() for x in prev.split(';') if x.strip()]:
            next_map[p] = pc
    cur = first_pack_pc
    while cur in next_map:
        nxt = next_map[cur]
        if nxt in visited:
            break
        result.append(nxt)
        visited.add(nxt)
        cur = nxt
    return result


# ══════════════════════════════════════════════════════════
# M10. 단일 제품 생산
# ══════════════════════════════════════════════════════════

def produce_unit(env, model_id, unit_id, data, kg,
                 wres, wh, rma, energy, idle, wip, stats, progress, menv=None,
                 plogger=None):
    done_ev = {pc: env.event() for pc in kg.nodes}

    pcs     = list(kg.nodes.keys())
    idx_map = {pc: i for i, pc in enumerate(pcs)}
    # H/adj는 에피소드 내 고정. menv 캐시에서 가져와 유닛마다 재계산 방지.
    if menv is not None and hasattr(menv, '_H_cache'):
        H   = menv._H_cache[model_id]
        adj = menv._adj_cache[model_id]
    else:
        _, H_np = kg.get_feat_matrix()
        H   = torch.tensor(H_np,  dtype=torch.float32)
        adj = torch.tensor(kg.get_adj(), dtype=torch.float32)
    # SMT_COMPLETE / SMT_THT 는 가상 노드 (kg.nodes 에 없음, dep_prev 참조 only).
    # (2026-05-06) 사전 PCB 소비 제거. 각 공정이 InputBOM 의 PCB 를 직접 소비
    # 하므로 어느 process 를 ready 로 띄울지가 어떤 PCB 를 먼저 소비할지 결정 →
    # 학습 의미 보존. SMT_COMPLETE/SMT_THT 는 ready 검사 통과용으로 즉시 done.
    done_set = {'SMT_COMPLETE', 'SMT_THT'}
    # kg_done: kg.nodes 기준 완료 추적.
    kg_done  = set()
    kg_total = set(kg.nodes.keys())
    # unit_defect_flag: {unit_id: bool}
    # 비INSP 공정에서 불량 발생 시 True. INSP 공정에서 검출 후 False로 초기화.
    unit_defect_flag = {unit_id: False}

    # 유닛 상태를 menv.unit_states 에 기록해 monitor 에서 관찰.
    unit_key = (model_id, unit_id)
    us = None
    if menv is not None and hasattr(menv, 'unit_states'):
        us = menv.unit_states
        us[unit_key] = {'state': 'SMT_WAIT', 'pc': '-', 'done_n': 0,
                        'total_n': len(kg.nodes), 'ready': []}

    # ReadyContext 는 1회 생성. done_set 은 set 객체 reference 공유라
    # 매 호출마다 ctx 안의 done_set 도 자동으로 최신 상태.
    ready_ctx = ReadyContext(
        kg=kg, done_set=done_set, wh=wh, wres=wres,
        data=data, model_id=model_id)
    while kg_done != kg_total:
        ready_pcs = kg.ready_processes(ready_ctx)

        if us is not None:
            us[unit_key].update({
                'done_n': len(kg_done),
                'ready': list(ready_pcs)[:5],
                'state': 'READY_WAIT' if not ready_pcs else 'RUNNING',
            })

        if not ready_pcs:
            yield env.timeout(1)
            continue

        if menv is not None and menv.agent is not None:
            # 상태 수집
            s = menv.get_state()
            # ready한 공정만 True인 mask 생성 - 학습과 실행이 동일한 mask 사용
            ready_mask = torch.zeros(len(pcs), dtype=torch.bool)
            for pc in ready_pcs:
                if pc in idx_map:
                    ready_mask[idx_map[pc]] = True
            # 에이전트가 공정 선택 + 경험 수집에 필요한 값 반환
            a, lp, v, emb, mask_bytes = menv.agent.act(s, H, adj, ready_mask)
            # 선택한 action이 ready한 노드인지 확인, 아니면 ready_pcs[0] fallback
            next_pc = pcs[a] if (a < len(pcs) and ready_mask[a]) else ready_pcs[0]
            r = menv.reward()
            menv.agent.store(s, emb, a, r, lp, v, mask=mask_bytes, model_id=model_id)
        else:
            next_pc = ready_pcs[0]

        if us is not None:
            us[unit_key]['pc'] = next_pc

        prow = data.get_proc(next_pc)
        if prow is None:
            # 디버그: 어느 (model, pc) 가 prow=None 으로 스킵됐는지 추적.
            if menv is not None and hasattr(menv, 'wh'):
                menv.wh.skipped_pcs[(model_id, str(next_pc))] += 1
            done_set.add(next_pc)
            kg_done.add(next_pc)
            continue

        yield env.process(
            run_process(env, prow, done_ev, wres, wh, rma,
                        energy, idle, wip, stats, model_id, data,
                        plogger=plogger, uid=unit_id,
                        unit_defect_flag=unit_defect_flag,
                        progress=progress))

        done_set.add(next_pc)
        kg_done.add(next_pc)

        # INSP 에서 RMA 경로로 라우팅된 유닛은 후속 공정을 진행하지 않는다.
        # RMA 경로가 동일 유닛에 대해 PACK 체인을 별도로 실행하므로 이중 fire 방지.
        if unit_defect_flag.get('_routed_to_rma'):
            if us is not None:
                us[unit_key].update({'state': 'ROUTED_TO_RMA', 'pc': '-'})
            if menv is not None and hasattr(menv, 'wh'):
                menv.wh.unit_completions[(model_id, unit_id)] = {
                    'path': 'routed_to_rma',
                    'end_time': float(env.now),
                    'done_n': len(kg_done),
                    'total_n': len(kg_total),
                    'rma_count': 0,
                }
            return

    # 안전 로직 (B1 안전망): kg.nodes 전체가 처리됐는지 확인. 누락 발견 시
    # stats 카운터 증가시키고 경고 로그. 학습 신호 오염 방지.
    # 단 _routed_to_rma 인 경우는 위에서 이미 return 했으므로 여기 안 옴.
    missing = kg_total - kg_done
    if missing:
        stats['kg_incomplete_units'] = stats.get('kg_incomplete_units', 0) + 1
        if menv is not None and hasattr(menv, 'wh'):
            menv.wh.kg_incomplete_log.append({
                'model_id': model_id, 'unit_id': unit_id,
                'time_h': float(env.now)/3600,
                'missing_pcs': sorted(missing),
            })
        _log_event(env.now,
                   f'[안전로직] {model_id} #{unit_id+1} 미처리 공정: '
                   f'{sorted(missing)}')

    if us is not None:
        us[unit_key].update({'state': 'DONE', 'pc': '-',
                             'done_n': len(kg_done)})

    # OQC: 전체 생산품의 5% 표본 추출 검사
    if random.random() < OQC_RATE:
        if not _is_work_time(env.now):
            yield env.timeout(_next_work_start(env.now) - env.now)
        res = wres.get('WORKER_OQC')
        req = res.request() if res else None
        acquired = False
        if req:
            yield req
            idle.acquire(env, 'WORKER_OQC')
            acquired = True
        ev_id_oqc = None
        try:
            if plogger is not None:
                _cap = res.capacity if res is not None else 1
                ev_id_oqc = plogger.mark_start(
                    'OQC_SAMPLE', model_id, unit_id, env.now, 'OQC',
                    wgrp='WORKER_OQC', cap=_cap,
                    work_timed=True)
            yield from work_timeout(env, OQC_TIME_SEC)
            energy.record('OQC', 'INSP', OQC_TIME_SEC, env.now)
        finally:
            if plogger is not None and ev_id_oqc is not None:
                try:
                    plogger.mark_end('OQC_SAMPLE', env.now, ev_id_oqc)
                except Exception:
                    pass
            if req and res:
                if acquired:
                    idle.release(env, 'WORKER_OQC')
                    acquired = False
                res.release(req)
        stats['oqc_inspected'] = stats.get('oqc_inspected', 0) + 1

    # 제품 완료 처리 (주문수량 초과 가드)
    done, total = progress[model_id]
    if done >= total:
        # 디버그: 초과 가드로 카운터 증가는 막혔지만 produce_unit 자체는 완주.
        if menv is not None and hasattr(menv, 'wh'):
            menv.wh.unit_completions[(model_id, unit_id)] = {
                'path': 'normal_blocked_by_quota',
                'end_time': float(env.now),
                'done_n': len(kg_done),
                'total_n': len(kg_total),
                'rma_count': 0,
            }
        return
    stats[f'{model_id}_done'] = stats.get(f'{model_id}_done', 0) + 1
    progress[model_id] = (done + 1, total)
    pct = (done + 1) / total * 100
    _log_event(env.now, f'{model_id} #{unit_id+1} 완성 ({pct:.0f}%)')
    # 모든 유닛이 완성되면 관련 worker_group 에 완료 시각을 기록.
    # 이 시각 이후의 유휴는 worker_idle_penalty 계산에서 제외된다.
    if done + 1 >= total and menv is not None:
        for _pc, node in kg.nodes.items():
            wgrp = node.get('worker_group', '')
            if wgrp:
                menv.idle.mark_completed(wgrp, float(env.now))
    # 디버그: 정상 경로 완성 기록.
    if menv is not None and hasattr(menv, 'wh'):
        menv.wh.unit_completions[(model_id, unit_id)] = {
            'path': 'normal',
            'end_time': float(env.now),
            'done_n': len(done_set),
            'total_n': len(kg.nodes),
            'rma_count': 0,
        }


# ══════════════════════════════════════════════════════════
# M11. 콘솔 모니터
# ══════════════════════════════════════════════════════════

class ProcessActivityLogger:
    """공정 위치별 실시간 진행 및 시계열 기록.

    mark_start/mark_end 는 run_process 에서 try/finally 로 감싸 호출.
    log 딕셔너리는 엑셀 Process_Log 시트 소스로 사용.
    current 딕셔너리는 monitor 실시간 표시 소스로 사용.
    """
    def __init__(self):
        self.log = {}          # pc -> {hour_bucket: [labels]}  (Process_Log 시트용)
        self.current = {}      # pc -> (mid, uid, start_sec) - monitor 3-tuple 호환.
                               # 동일 pc 동시 다건이면 "마지막으로 시작한 것" 만 남음.
        self.groups = {}       # pc -> process_group (SEMI/SET/PACK/OQC/MODULE/SMT/RMA)
        # 정밀 이벤트 로그 (2026-4-24). 각 mark_end 에서 append.
        self.events = []
        # slot 할당 (2026-4-24). SimPy Resource 는 pool 이라 "어느 슬롯인지"
        # 정보를 주지 않으므로 독립 관리. wgrp 별로 capacity N 배열을 두고
        # mark_start 시 최저 번호의 빈 슬롯 점유, mark_end 시 해제.
        self.slot_pool = {}    # wgrp -> list[ev_id or None]
        self.max_slot  = {}    # wgrp -> 관찰된 최대 slot index + 1
        # 이벤트 ID 기반 active 트래커 - 같은 pc 동시 다건 처리용.
        self._active    = {}   # ev_id -> meta dict
        self._ev_counter = 0

    def _next_ev_id(self):
        self._ev_counter += 1
        return self._ev_counter

    def mark_start(self, pc, mid, uid, now, grp=None, wgrp=None, cap=None,
                   work_timed=False):
        """이벤트 ID 를 반환. mark_end 호출 시 이 ID 를 함께 넘기면
        동일 pc 가 동시 여러 건 실행돼도 각 이벤트가 정확히 매칭된다.

        work_timed=True 면 timeout 이 work_timeout 으로 wrap 돼 근무시간 경계에서
        실제로 pause 한다는 표식. 플롯 시 boundary 에서 bar 를 분할하는 용도.
        """
        c = str(pc)
        self.current[c] = (str(mid), int(uid), float(now))   # monitor 호환용
        if grp:
            self.groups[c] = str(grp)
        ev_id = self._next_ev_id()
        meta = {
            'pc'        : c,
            'mid'       : str(mid),
            'uid'       : int(uid),
            'start'     : float(now),
            'wgrp'      : '',
            'slot'      : -1,
            'work_timed': bool(work_timed),
        }
        if wgrp:
            wg = str(wgrp)
            pool = self.slot_pool.setdefault(wg, [])
            target_cap = int(cap) if cap and cap > 0 else 1
            while len(pool) < target_cap:
                pool.append(None)
            slot_i = -1
            for i, v in enumerate(pool):
                if v is None:
                    slot_i = i
                    break
            if slot_i < 0:
                slot_i = len(pool)
                pool.append(None)
            pool[slot_i] = ev_id
            meta['wgrp'] = wg
            meta['slot'] = slot_i
            if slot_i + 1 > self.max_slot.get(wg, 0):
                self.max_slot[wg] = slot_i + 1
        self._active[ev_id] = meta
        return ev_id

    def mark_end(self, pc, now, ev_id=None):
        c = str(pc)
        # monitor 표시를 위한 current dict 는 동일 pc 의 "마지막 active" 가 없을 때만 삭제.
        # 간단화를 위해 pop 만 수행 (덮어쓰기로 이미 1건만 남음).
        self.current.pop(c, None)
        meta = None
        if ev_id is not None:
            meta = self._active.pop(ev_id, None)
        if meta is None:
            # ev_id 미전달 또는 무효 -> pc 로 linear search (fallback)
            for eid, m in list(self._active.items()):
                if m['pc'] == c:
                    meta = self._active.pop(eid)
                    ev_id = eid
                    break
        if meta is None:
            return
        t_end = float(now)
        mid, uid, t0 = meta['mid'], meta['uid'], meta['start']
        label = f'{mid}/u{uid+1}'
        h_start = int(t0 // 3600)
        h_end   = int(t_end // 3600)
        if h_end < h_start:
            h_end = h_start
        self.log.setdefault(c, {})
        for h in range(h_start, h_end + 1):
            self.log[c].setdefault(h, []).append(label)
        self.events.append({
            'pc'        : c,
            'mid'       : mid,
            'uid'       : uid,
            'start'     : t0,
            'end'       : t_end,
            'grp'       : self.groups.get(c, ''),
            'wgrp'      : meta['wgrp'],
            'slot'      : meta['slot'],
            'work_timed': meta.get('work_timed', False),
        })
        wg, slot_i = meta['wgrp'], meta['slot']
        if wg:
            pool = self.slot_pool.get(wg)
            if pool and 0 <= slot_i < len(pool) and pool[slot_i] == ev_id:
                pool[slot_i] = None

    def busy_now(self, pc):
        return self.current.get(str(pc))

    def summary(self):
        """Process_Log 시트용 요약: pc -> hour -> label 문자열."""
        out = {}
        for pc, hdict in self.log.items():
            out[pc] = {h: '; '.join(labels) for h, labels in hdict.items()}
        return out


def monitor(env, progress, energy, wh, idle, wip, stats,
            smt_lines, interval=3600, plogger=None, menv=None):
    """실시간 흐름 시각화 모니터.

    부품·PCB·공정 흐름을 한 화면에 시각적으로 표시.
    세부 수치는 simulation_results.xlsx (Stock_Summary, Reorder_Log,
    Debug_Unit_Status, Debug_Plogger_Events 등) 에서 확인.
    """
    _wall_prev = time.time()
    BAR_W = 20

    def _bar(value, max_v, width=BAR_W, fill='█', empty='░'):
        v = max(0, min(int(value), int(max_v)))
        n = int((v / max(int(max_v), 1)) * width)
        return fill * n + empty * (width - n)

    def _pcb_label(model_id, pcb_code):
        """`<모델 약칭> <메인/수삽>-<양면/단면> (코드 끝 4자리)` 형태.

        같은 분류(예: A 수삽-양면) 가 PCB 여러 개일 때 코드 끝 4자리로 식별.
        """
        m_short = (model_id or '?').replace('MODEL_', '')
        is_main = (PCB_MAP.get(model_id) == pcb_code)
        kind = '메인' if is_main else '수삽'
        try:
            side_raw = wh.data.smt_side(pcb_code)
        except Exception:
            side_raw = 'double'   # AAS 정보 없을 시 기본 (도메인상 메인·THT 모두 양면)
        side = '양면' if side_raw == 'double' else '단면'
        tag = (str(pcb_code) or '')[-4:]
        return f'{m_short} {kind}-{side} ({tag})'

    STAGE_KEYS   = ['LOADER', 'PRINTER', 'SPI', 'MOUNTER_H',
                    'MOUNTER_M', 'REFLOW', 'UNLOADER']
    STAGE_LABELS = ['LD', 'PR', 'SP', 'MH', 'MM', 'RF', 'UL']

    while True:
        yield env.timeout(interval)
        # SimPy 가 너무 빨리 돌면 화면이 휙 지나가므로 frame 당 최소 wall 시간 보장.
        _wall_now = time.time()
        _wall_delta = _wall_now - _wall_prev
        if _wall_delta < MONITOR_MIN_WALL_SEC:
            time.sleep(MONITOR_MIN_WALL_SEC - _wall_delta)
            _wall_now = time.time()
            _wall_delta = _wall_now - _wall_prev
        _wall_prev = _wall_now

        try:
            sys.stdout.write('\033[2J\033[H')
            sys.stdout.flush()
        except Exception:
            pass

        day  = int(env.now // DAY_SEC) + 1
        h_in = (env.now % DAY_SEC) / 3600
        pace = f'  {INFER_MONITOR_STEP_HR}h sim ≈ {_wall_delta:.2f}s wall' if _wall_delta > 0 else ''
        header = f' Day{day} {h_in:>4.1f}h │ 누적 {env.now/3600:>6.1f}h │{pace} '
        print('╔' + '═' * (len(header) + 2) + '╗')
        print(f'║ {header} ║')
        print('╚' + '═' * (len(header) + 2) + '╝')

        # ── SMT 라인 stage 진행 ────────────────────────────────
        print('[ SMT 라인 ]')
        print('              ' + '   '.join(f'{s:^3}' for s in STAGE_LABELS))
        for sid, line in smt_lines.items():
            model = line.assigned_model
            row = []
            boards_set = set()
            for stage in STAGE_KEYS:
                pc = f'SMT_{stage}_{sid}'
                if pc in line.stage_active:
                    pcb_code, _bid, _is2 = line.stage_active[pc]
                    row.append(' ●  ')
                    if model:
                        boards_set.add(_pcb_label(model, pcb_code))
                else:
                    row.append(' ─  ')
            head = f'  {sid} ⌊{(model or "-"):^7}⌉'
            print(f'{head} {"".join(row)}')
            if boards_set:
                print(f'         ↳ 처리 중: {", ".join(sorted(boards_set))}')

        # ── 공유 AOI ───────────────────────────────────────
        aoi_label = None
        for sid, line in smt_lines.items():
            if 'SMT_AOI' in line.stage_active:
                pcb_code, _bid, _is2 = line.stage_active['SMT_AOI']
                aoi_label = _pcb_label(line.assigned_model, pcb_code)
                break
        print(f'  AOI (공유):  {"● 진행중 " + aoi_label if aoi_label else "○ 유휴"}')

        # ── THT 외주 ───────────────────────────────────────
        # (도착 후 PCB 는 즉시 인벤으로 가므로 "도착" 별도 표시 안 함 — 인벤토리에서 확인)
        print('\n[ THT 외주 (외주중 = in-flight, mag = 도착 후 미적재) ]')
        pool = getattr(menv, 'outsource_pool', None) if menv is not None else None
        if pool is not None:
            cur_n = len(pool.truck)
            cur_bar = _bar(cur_n, TRUCK_SIZE, width=15)
            in_transit_n = len(pool.in_transit)
            in_transit_boards = sum(t['size'] for t in pool.in_transit)
            print(f'  적재 중 트럭   {cur_bar} {cur_n:>2d}/{TRUCK_SIZE}  '
                  f'│  운송 중: {in_transit_n}대 ({in_transit_boards}보드)  '
                  f'│  누적 출발: {pool.dispatched_count}대')
        # PCB 종류별 진행 — 발사 / 외주중 만 (도착은 PCB 인벤토리 항목으로 흡수됨)
        for model in progress:
            for tht_code in THT_PCB_BY_MODEL.get(model, []):
                flow     = wh.pcb_flow.get(tht_code, {})
                fired    = int(flow.get('outsource_in', 0))
                returned = int(flow.get('outsource_returned', 0))
                in_flight = fired - returned
                mag_remain = sum(line.mag_buf.get(tht_code, 0)
                                 for line in smt_lines.values())
                lbl = _pcb_label(model, tht_code)
                print(f'  {lbl:<13s} 발사 {fired:>3}  외주중 {in_flight:>2}  '
                      f'mag(도착 후 미적재) {mag_remain:>2}')

        # ── PCB 인벤토리 (max = 주문수량) ───────────────────
        print('\n[ PCB 인벤토리 ]   (그래프 max = 주문 수량)')
        for model, (_done, total) in progress.items():
            for pcb in [PCB_MAP.get(model)] + THT_PCB_BY_MODEL.get(model, []):
                if not pcb:
                    continue
                stock = int(wh.stock.get(pcb, 0))
                lbl = _pcb_label(model, pcb)
                bar = _bar(stock, total, width=BAR_W)
                print(f'  {lbl:<13s} {bar} {stock:>3d}/{total}')

        # ── 조립 라인 (가동 워커 수 / 선택 가능 경우 수) ──
        # 분자 = 그 그룹 워커 자원 중 현재 점유된 수 (실제 처리 중 unit 수)
        # 분모 = WIP enter 누적 — wait_stock·worker queue 포함 "선택 가능 경우"
        print('\n[ 조립 라인 (가동 워커 / 선택 가능) ]')
        flow_groups = ['MODULE', 'SEMI', 'SET', 'INSP', 'PACK']
        grp_wgrp = {
            'MODULE': ['WORKER_FW', 'WORKER_LENS_HOLDER', 'WORKER_SENSOR_FOCUS'],
            'SEMI'  : ['WORKER_SEMI'],
            'SET'   : ['WORKER_SET', 'WORKER_SET_INSP'],
            'INSP'  : ['WORKER_AGING'],
            'PACK'  : ['WORKER_PACK'],
        }
        wres = getattr(menv, 'wres', {}) if menv is not None else {}
        cells = []
        for g in flow_groups:
            total  = wip.wip.get(g, 0)
            active = sum(wres[w].count for w in grp_wgrp.get(g, [])
                         if w in wres)
            cells.append(f'{g}({active}/{total})')
        print(f'  {"  ▶  ".join(cells)}')
        print(f'  SMT({wip.wip.get("SMT", 0):>2d})  '
              f'│  RMA({wip.wip.get("RMA", 0):>2d})')

        # ── 완성률 ────────────────────────────────────────
        print('\n[ 완성 ]')
        for m, (done, total) in progress.items():
            bar = _bar(done, total, width=BAR_W)
            pct = done / max(total, 1) * 100
            print(f'  {m}  {bar} {done:>3d}/{total:<3d} ({pct:>3.0f}%)')

        # ── 부품 부족 알림 (top 3) + 발주중 ────────────
        # 외주 raw 풀 (xxxx_RAW) 은 외주 진행 중 가상 stock 이라 부족 임계 의미 없음.
        rows = []
        for c, cur in wh.stock.items():
            if str(c).endswith(THT_RAW_SUFFIX):
                continue
            try:
                ms = wh.data.get_min_stock(c)
            except Exception:
                ms = MIN_STOCK
            if cur < ms:
                rows.append((cur / max(ms, 1), c, int(cur), int(ms)))
        rows.sort()
        if rows:
            top = rows[:3]
            parts = '  '.join(f'{c}={s}/{m}' for _, c, s, m in top)
            pending = len(wh._pending_orders)
            print(f'\n[ 부품 부족 ]  ⚠ {parts}  │  발주중 {pending}건')

        # ── 결근 / 설비 고장 ────────────────────────────
        if idle.absent_groups:
            print(f'\n[ 결근 ] {", ".join(sorted(idle.absent_groups))}')

        # ── 최근 이벤트 5건 ─────────────────────────────
        if _EVENT_BUF:
            print('\n[ 최근 이벤트 ]')
            for t_sec, msg in _EVENT_BUF[-5:]:
                print(f'  {t_sec/3600:>6.2f}h  {msg}')


# ████████████████████████████████████████████████████████████████████
# §D. RL
# ████████████████████████████████████████████████████████████████████
# 책임: 도메인+시뮬을 환경(MDP)으로 추상화하고 정책을 학습/추론. PPO+GNN
# 만 사용 — 다른 RL 알고리즘 도입 시 이 섹터만 갈아끼면 됨.
# 외부 export: ProcessGNN, PPOAgent, ManufacturingEnv, ExperimentRunner, main.
# 향후 파일 분할 위치: cpro/rl/{gnn, ppo, env, runner}.py + cpro_simulation_ver2.py 진입점.


# ══════════════════════════════════════════════════════════
# M12. PyTorch GNN (노드별 스코어 출력)
# ══════════════════════════════════════════════════════════

class ProcessGNN(nn.Module):
    """
    GCN 2층: 각 공정 노드에 대해 '지금 실행할 우선순위 스코어'를 출력.
    입력: 노드 특징 행렬 H (N x in_dim), 인접 행렬 A (N x N)
    출력: 노드별 스코어 (N,) -> PPO 행동 확률 변환에 사용
    Luo et al. (2023 NeurIPS) GNN+스케줄링 참조
    """
    def __init__(self, in_dim=6, hidden=32, out_dim=16):
        super().__init__()
        self.conv1 = nn.Linear(in_dim,  hidden)
        self.conv2 = nn.Linear(hidden,  out_dim)
        self.score = nn.Linear(out_dim, 1)

    def forward(self, H: torch.Tensor, adj: torch.Tensor) -> torch.Tensor:
        # 정규화된 인접 행렬
        deg = adj.sum(dim=1, keepdim=True).clamp(min=1e-6)
        A_n = adj / deg                             # (N, N)
        H1  = F.relu(self.conv1(A_n @ H))          # (N, hidden)
        H2  = F.relu(self.conv2(A_n @ H1))         # (N, out_dim)
        return self.score(H2).squeeze(-1)           # (N,)

    def graph_embed(self, H: torch.Tensor, adj: torch.Tensor) -> torch.Tensor:
        # 그래프 전체 임베딩 (평균 풀링). 상태 벡터에 포함.
        deg = adj.sum(dim=1, keepdim=True).clamp(min=1e-6)
        A_n = adj / deg
        H1  = F.relu(self.conv1(A_n @ H))
        H2  = F.relu(self.conv2(A_n @ H1))
        return H2.mean(dim=0)                       # (out_dim,)


# ══════════════════════════════════════════════════════════
# M13. PPO 에이전트 (PyTorch)
# ══════════════════════════════════════════════════════════

class PPOAgent(nn.Module):
    """
    상태: 시뮬레이션 스칼라 벡터 + GNN 그래프 임베딩
    행동: 조립 공정 내 실행 가능 공정 중 우선순위 선택 (GNN 노드 스코어 기반)
    보상: w1*(이번스텝시간감소) + w2*(-전력증가) - w3*WIP초과
          - w4*재고부족(critical_stock 기준) + w5*납기 + w6*(-작업자유휴)
    가중치: 매 에피소드 Dirichlet([1]*6) 샘플링 (MORL scalarization)
    수렴: 최근 CONV_WINDOW 에피소드 평균 보상 변화 < CONV_THRESHOLD
    Sun et al. (2025 Engineering Vol.46) PPO_S 수렴 ~4000 에피소드
    """
    LR             = 3e-4
    GAMMA          = 0.99
    LAM            = 0.95
    EPS            = 0.2
    EPOCHS         = 4
    CONV_WINDOW    = 100
    CONV_THRESHOLD = 0.01   # (임의)

    def __init__(self, state_dim: int, gnn: ProcessGNN):
        super().__init__()
        self.gnn      = gnn
        embed_dim     = gnn.conv2.out_features   # GNN out_dim
        in_dim        = state_dim + embed_dim

        # 공통 인코더
        self.encoder  = nn.Sequential(
            nn.Linear(in_dim, 128), nn.ReLU(),
            nn.Linear(128,     64), nn.ReLU(),
        )
        # Actor: 스칼라 출력 (노드 스코어와 결합하여 최종 확률 계산)
        self.actor_head  = nn.Linear(64, 1)
        # Critic: 가치 함수
        self.critic_head = nn.Linear(64, 1)

        self.optimizer  = torch.optim.Adam(self.parameters(), lr=self.LR)
        self.buf        = []
        self.ep_rewards = []
        # ep 별 6항 누적 [(w1*r1, w2*r2, ..., w6*r6), ...] — 진단/시각화용.
        self.ep_rewards_decomp = []

    def forward(self, state_vec: torch.Tensor,
                graph_embed: torch.Tensor) -> tuple:
        x   = torch.cat([state_vec, graph_embed], dim=-1)
        enc = self.encoder(x)
        return self.actor_head(enc), self.critic_head(enc)

    def act(self, state_np: np.ndarray,
            H: torch.Tensor, adj: torch.Tensor,
            ready_mask: torch.Tensor) -> tuple:
        """
        state_np : 시뮬레이션 상태 벡터 (numpy)
        H        : 노드 특징 행렬 (N x 6)
        adj      : 인접 행렬 (N x N)
        ready_mask: 실행 가능 노드 마스크 (N,) bool
        반환: (action_idx, log_prob, value, graph_embed, mask_bytes)

        주의 (2026-4-22 수정): 정책은 GNN node_scores 에 ready_mask 를 씌운
        Categorical 분포다. update() 에서 이 분포를 동일하게 재구성해야
        gradient 가 GNN 파라미터까지 흐른다. 샘플링 자체는 no_grad 로 해도
        무방하지만, update() 때 mask 를 재사용해야 하므로 mask_bytes 를
        함께 반환해서 store() 에 남긴다.
        """
        with torch.no_grad():
            s_t  = torch.tensor(state_np, dtype=torch.float32).unsqueeze(0)
            # GNN 중간 결과 H2를 한 번만 계산해 임베딩과 노드 스코어를 모두 추출.
            # graph_embed와 forward가 동일한 conv1/conv2를 공유하므로 중복 제거.
            deg  = adj.sum(dim=1, keepdim=True).clamp(min=1e-6)
            A_n  = adj / deg
            H1   = F.relu(self.gnn.conv1(A_n @ H))
            H2   = F.relu(self.gnn.conv2(A_n @ H1))
            emb  = H2.mean(dim=0)                           # (out_dim,) 그래프 임베딩
            emb_t = emb.unsqueeze(0)
            _, val = self.forward(s_t, emb_t)

            # 노드별 GNN 스코어 -> H2 재사용
            node_scores = self.gnn.score(H2).squeeze(-1)    # (N,)
            node_scores = node_scores.masked_fill(~ready_mask, float('-inf'))
            probs       = torch.softmax(node_scores, dim=0)
            dist        = torch.distributions.Categorical(probs=probs,
                                                          validate_args=False)
            action      = dist.sample()
            log_prob    = dist.log_prob(action)

        mask_bytes = ready_mask.detach().to(torch.bool).numpy().copy()
        return (action.item(), log_prob.item(),
                val.item(), emb.detach().numpy(), mask_bytes)

    def store(self, s, emb, a, r, lp, v, mask=None, model_id=None):
        # 경험 버퍼에 저장. model_id는 update()에서 모델별 H/adj 분리에 사용.
        self.buf.append((s, emb, a, r, lp, v, mask, model_id))

    def update(self, graphs_cache=None):
        """PPO update.

        produce_unit에서 agent.store()로 직접 수집한 경험으로 업데이트.
        graphs_cache: {model_id: (H_tensor, adj_tensor)} - 모델별 그래프 텐서.
        모델별 H/adj를 분리 사용해 다중 모델 혼합 경험을 정확하게 학습.
        """
        if len(self.buf) < 2:
            return 0.0
        rewards   = [b[3] for b in self.buf]
        values    = [b[5] for b in self.buf]
        model_ids = [b[7] for b in self.buf]   # 모델별 H/adj 분리용
        ep_r      = sum(rewards)
        self.ep_rewards.append(ep_r)

        advs, gae = [], 0.0
        for i in reversed(range(len(rewards) - 1)):
            delta = rewards[i] + self.GAMMA * values[i+1] - values[i]
            gae   = delta + self.GAMMA * self.LAM * gae
            advs.insert(0, gae)
        advs_t = torch.tensor(advs, dtype=torch.float32)
        advs_t = (advs_t - advs_t.mean()) / (advs_t.std() + 1e-8)

        for _ in range(self.EPOCHS):
            for i, entry in enumerate(self.buf[:-1]):
                if i >= len(advs):
                    break
                s, emb, a, _, old_lp, _old_v, mask, mid = entry
                s_t   = torch.tensor(s,   dtype=torch.float32).unsqueeze(0)
                emb_t = torch.tensor(emb, dtype=torch.float32).unsqueeze(0)

                # Critic: forward() 는 gradient on 으로 재실행
                _actor_out, critic_out = self.forward(s_t, emb_t)

                # Actor: 이 경험이 속한 모델의 H/adj로 동일 분포 재구성.
                # 모델별로 노드 수가 다를 수 있으므로 model_id로 분기.
                H_m   = graphs_cache.get(mid, (None, None))[0] if graphs_cache else None
                adj_m = graphs_cache.get(mid, (None, None))[1] if graphs_cache else None
                if H_m is not None and adj_m is not None and mask is not None:
                    mask_t = torch.tensor(mask, dtype=torch.bool)
                    node_scores = self.gnn(H_m, adj_m)
                    node_scores = node_scores.masked_fill(~mask_t, float('-inf'))
                    probs       = torch.softmax(node_scores, dim=0)
                    dist        = torch.distributions.Categorical(
                        probs=probs, validate_args=False)
                    new_lp      = dist.log_prob(torch.tensor(int(a)))
                    old_lp_t    = torch.tensor(float(old_lp))
                    ratio       = torch.exp(new_lp - old_lp_t)
                    adv         = advs_t[i]
                    loss_p      = -torch.min(
                        ratio * adv,
                        torch.clamp(ratio, 1-self.EPS, 1+self.EPS) * adv)
                else:
                    loss_p = torch.tensor(0.0)

                loss_v = F.mse_loss(
                    critic_out.squeeze(),
                    torch.tensor(values[i], dtype=torch.float32))
                loss   = loss_p + 0.5 * loss_v

                self.optimizer.zero_grad()
                loss.backward()
                nn.utils.clip_grad_norm_(self.parameters(), 0.5)
                self.optimizer.step()

        self.buf.clear()
        return ep_r

    def is_converged(self):
        if len(self.ep_rewards) < self.CONV_WINDOW * 2:
            return False
        recent = np.mean(self.ep_rewards[-self.CONV_WINDOW:])
        prev   = np.mean(self.ep_rewards[-self.CONV_WINDOW*2:-self.CONV_WINDOW])
        return abs(recent - prev) < self.CONV_THRESHOLD

    def save(self, path=POLICY_PATH, verbose=True):
        torch.save({
            'model_state': self.state_dict(),
            'ep_rewards' : self.ep_rewards,
            'ep_rewards_decomp': self.ep_rewards_decomp,
        }, path)
        if verbose:
            print(f'정책 저장: {path}')

    def load(self, path=POLICY_PATH):
        ckpt = torch.load(path, map_location='cpu')
        self.load_state_dict(ckpt['model_state'])
        self.ep_rewards = ckpt.get('ep_rewards', [])
        self.ep_rewards_decomp = ckpt.get('ep_rewards_decomp', [])
        print(f'정책 불러오기: {path} (학습 에피소드:{len(self.ep_rewards)})')


# ══════════════════════════════════════════════════════════
# M14. 제조 RL 환경
# ══════════════════════════════════════════════════════════

class ManufacturingEnv:
    """
    SimPy 기반 제조 공정 강화학습 환경.

    보상: w1*(시간·전력효율개선) + w2*(-전력증가) - w3*WIP초과
          - w4*재고부족(critical_stock 기준) + w5*납기 + w6*(-작업자유휴)
    가중치: W_DEFAULT 고정 사용.
    에이전트(PPOAgent)는 매 공정 선택 시점에 호출되어 경험을 직접 버퍼에 저장한다.
    """
    W_DEFAULT = (0.30, 0.25, 0.15, 0.10, 0.10, 0.10)

    def __init__(self, data, order, weight_vec=None):
        self.data   = data
        self.order  = order
        self.W      = tuple(weight_vec) if weight_vec else self.W_DEFAULT
        self.graphs = {m: ProcessKnowledgeGraph(data, m) for m in order}
        # 에피소드 내 고정인 H/adj를 미리 tensor로 변환해 캐싱.
        # produce_unit이 유닛마다 재계산하지 않도록 공유.
        self._H_cache = {
            m: torch.tensor(kg.get_feat_matrix()[1], dtype=torch.float32)
            for m, kg in self.graphs.items()
        }
        self._adj_cache = {
            m: torch.tensor(kg.get_adj(), dtype=torch.float32)
            for m, kg in self.graphs.items()
        }
        self._init_sim()

    def _init_sim(self):
        # 새 에피소드 시작 시 이전 이벤트 버퍼 비움 (monitor 진동 방지)
        _EVENT_BUF.clear()
        self.env         = simpy.Environment()
        self.wh          = Warehouse(self.data, self.order)
        self.energy      = EnergyLogger(self.data)
        self.idle        = IdleTracker()
        self.wip         = WIPTracker(self.order)
        self.rma         = simpy.Store(self.env)
        self.stats       = defaultdict(int)
        self.progress    = {m: (0, q) for m, q in self.order.items()}
        # 가동 인원 분리: WORKER_SET 풀(예: 16명) 중 SET_INSP_HEADCOUNT(3명) 은
        # WORKER_SET_INSP 자원으로 빠지므로 일반 SET 가동 인원은 (총원 - INSP).
        # 단, 전력 계산은 EnergyLogger 가 data.workers['WORKER_SET']=16 을 그대로
        # 참조하여 33.67kW/16 으로 분배하므로 영향 없음.
        self.wres = {}
        for _g, _c in self.data.workers.items():
            if _g == 'WORKER_SET':
                _cap = max(int(_c) - SET_INSP_HEADCOUNT, 1)
            else:
                _cap = int(_c)
            self.wres[_g] = simpy.Resource(self.env, capacity=_cap)
        self.aoi_res     = simpy.Resource(self.env, capacity=1)
        self.smt_broken  = defaultdict(bool)   # 설비별 고장 여부
        self.smt_lines   = {}
        # 모든 SMT 라인이 공유하는 외주 트럭 풀. SMT 종류 무관 한 트럭 적재 (옵션 b).
        self.outsource_pool = OutsourceTruckPool(
            self.env, self.wh, self.stats, self.smt_lines)
        for sid in SMT_LINE_IDS:
            line = SMTLine(self.env, sid, self.data, self.wh,
                           self.aoi_res, self.rma, self.energy,
                           self.idle, self.wip, self.stats, self.smt_broken,
                           outsource_pool=self.outsource_pool)
            self.smt_lines[sid] = line
        # 공정 위치별 시계열/실시간 진행 기록기
        self.plogger     = ProcessActivityLogger()
        # 유닛별 진행 상태 진단용
        self.unit_states     = {}
        self.agent           = None
        self._prev_reward_t    = 0.0  # r1 시간차분 기준점
        self._prev_reward_kwh  = 0.0  # r2 전력차분 기준점
        self._prev_wip_viol    = 0    # r3 WIP viol 차분 기준점
        self._prev_stock_pen   = 0    # r4 stock viol 차분 기준점
        self._prev_done        = 0    # r5 완성수 차분 기준점
        self._prev_idle_pen    = 0.0  # r6 idle 차분 기준점

        # IdleTracker capacity 등록 + 그룹별 처리량 target 사전 계산.
        # capacity 는 wres 의 실제 simpy.Resource capacity (WORKER_SET 은 split 후).
        self.idle.configure({g: r.capacity for g, r in self.wres.items()})
        self._compute_idle_targets()

    def _compute_idle_targets(self):
        """그룹별 처리해야 할 work item 총수 = Σ_model (qty × 그 그룹의 process 수).
        run_process 의 INSP suffix override (WORKER_SET → WORKER_SET_INSP) 도 반영.
        RMA 는 불량 발생량에 의존 — target 미설정 (시뮬 종료 시 _check_done 에서 일괄 마킹).
        """
        target = defaultdict(int)
        for mid, qty in self.order.items():
            try:
                procs = self.data.get_model_procs(mid)
            except Exception:
                continue
            for _, r in procs.iterrows():
                wgrp = str(r.get('worker_group', '') or '')
                grp  = str(r.get('process_group', '') or '')
                pc   = str(r['process_code'])
                if wgrp == 'WORKER_SET' and grp == 'SET' \
                        and pc.rsplit('_', 1)[-1].upper() == 'INSP':
                    wgrp = 'WORKER_SET_INSP'
                if wgrp:
                    target[wgrp] += qty
        for g, t in target.items():
            self.idle.set_target(g, t)

    def get_state(self) -> np.ndarray:
        # 상태 벡터 구성:
        #   comp  : 모델별 완성률 (done/total)
        #   wutil : 작업자 그룹별 가동률 (1 = 전원 대기, 0 = 전원 투입 중)
        #   scalar: 단위시간당 전력, 경과시간 비율, 재고부족 페널티,
        #           WIP초과 횟수, 작업자 유휴 페널티, SMT 고장 비율
        comp  = [self.stats.get(f'{m}_done', 0) / max(q, 1)
                 for m, q in self.order.items()]
        wutil = [1 - (r.count / max(r.capacity, 1))
                 for r in self.wres.values()]
        t_max = MAX_DAYS * (
            _active_schedule['work_end_sec']
            - _active_schedule['work_start_sec']
            - _active_schedule['break_duration_sec']
        )
        return np.array(
            comp + wutil + [
                self.energy.total / max(self.env.now + 1, 1),
                self.env.now / t_max,
                self.wh.stock_penalty() / max(sum(self.order.values()), 1),
                self.wip.violations() / max(len(self.wres), 1),
                self.idle.worker_idle_penalty() / max(self.env.now + 1, 1),
                sum(self.smt_broken.values()) / max(len(self.smt_broken) + 1, 1),
            ], dtype=np.float32)

    def reward(self) -> float:
        """보상 함수 — 모든 항이 _직전 호출 이후 발생한 증가분_ (step-incremental).

        PPO 의 시간적 신용 할당 (GAE δ = r + γV' − V) 이 정확히 작동하도록,
        누적값 / 진행률 그대로 부과하지 않고 매 호출마다 차분을 신호로 사용한다.
        에피소드 누적은 자연스럽게 _마지막 누적값/분모_ 또는 unit 진행률 1.0 이 됨.

        w1 r1: 시간차분 -Δt / t_max
        w2 r2: 전력차분 -Δkwh / (kwh_total + 1)
        w3 r3: WIP viol 차분 -Δviol / |wres|
        w4 r4: stock viol 차분 -Δviol / (total × 10)
        w5 r5: 완성 진행 +Δdone / total  (전체 완성 시 마지막 step 에 +1 보너스)
        w6 r6: idle 차분 -Δidle / (Σcap × Δwork_sec)
        """
        total = sum(self.order.values())
        cur   = sum(self.stats.get(f'{m}_done', 0) for m in self.order)
        w1, w2, w3, w4, w5, w6 = self.W

        t_max  = MAX_DAYS * (
            _active_schedule['work_end_sec']
            - _active_schedule['work_start_sec']
            - _active_schedule['break_duration_sec']
        )
        t_now  = float(self.env.now)
        prev_t = getattr(self, '_prev_reward_t', 0.0)
        dt_wall   = t_now - prev_t                              # wall-clock dt (r1)
        dt_work   = _work_seconds_between(prev_t, t_now)         # 근무시간 dt (r6 분모)

        # ── r1: 시간 차분 ─────────────────────────────────────────
        r1 = -dt_wall / max(t_max, 1)

        # ── r2: 전력 차분 ─────────────────────────────────────────
        kwh_now  = self.energy.total
        prev_kwh = getattr(self, '_prev_reward_kwh', 0.0)
        d_kwh    = kwh_now - prev_kwh
        r2       = -d_kwh / max(kwh_now + 1, 1)

        # ── r3: WIP viol 차분 ─────────────────────────────────────
        wip_v_now  = self.wip.violations()
        d_wip      = wip_v_now - getattr(self, '_prev_wip_viol', 0)
        r3         = -d_wip / max(len(self.wres), 1)

        # ── r4: stock viol 차분 ───────────────────────────────────
        stock_now  = self.wh.stock_penalty()
        d_stock    = stock_now - getattr(self, '_prev_stock_pen', 0)
        r4         = -d_stock / max(total * 10, 1)

        # ── r5: 완성 진행 차분 ────────────────────────────────────
        # Δdone/total 누적 = 마지막 step 에서 1.0 (전부 완성 시).
        # 전체 완성 step 에 추가 +1 보너스로 도달 자체에 강한 시그널.
        d_done = cur - getattr(self, '_prev_done', 0)
        r5 = d_done / max(total, 1)
        if cur >= total and getattr(self, '_prev_done', 0) < total:
            r5 += 1.0

        # ── r6: idle 차분 (capacity 기반 per-person·sec) ─────────
        # 분모: 직전 reward() 이후 이번 호출까지 _근무시간 dt × Σ cap_g_.
        #       "그 구간 동안 모든 워커가 내내 idle 일 때" 가 분모.
        self.idle.flush_all(self.env)
        idle_now  = self.idle.worker_idle_penalty()
        d_idle    = idle_now - getattr(self, '_prev_idle_pen', 0.0)
        total_cap = sum(self.idle._capacity.get(g, 0) for g in WORKER_GROUPS)
        r6_denom  = max(total_cap * dt_work, 1.0)
        r6        = -d_idle / r6_denom

        # 다음 호출용 상태 갱신
        self._prev_reward_t   = t_now
        self._prev_reward_kwh = kwh_now
        self._prev_wip_viol   = wip_v_now
        self._prev_stock_pen  = stock_now
        self._prev_done       = cur
        self._prev_idle_pen   = idle_now

        # 항별 분해 누적 (학습 진단용 — ep 끝에 PPOAgent.ep_rewards_decomp 적재).
        contribs = (w1*r1, w2*r2, w3*r3, w4*r4, w5*r5, w6*r6)
        if not hasattr(self, '_reward_decomp_sum'):
            self._reward_decomp_sum = [0.0] * 6
        for i, c in enumerate(contribs):
            self._reward_decomp_sum[i] += c

        return sum(contribs)

    def _event_smt_breakdown(self, env):
        """SMT 설비 고장 이벤트. 10분마다 확률 체크.
        repair_sec: RESOURCE mttr_hr 우선, 없으면 SMT_MTTR_DEFAULT_HR (임의).
        """
        smt_pcs = [pc for sid in self.smt_lines
                   for pc in self.smt_lines[sid]._res]
        while True:
            yield env.timeout(600)
            if not _is_work_time(env.now):
                continue
            for pc in smt_pcs:
                if random.random() < SMT_BREAKDOWN_PROB:
                    self.smt_broken[pc] = True
                    mttr_sec   = self.data.get_mttr(pc)
                    repair_sec = max(random.normalvariate(mttr_sec, mttr_sec * 0.2),
                                     mttr_sec * 0.1)
                    _log_event(env.now,
                               f'SMT 설비 고장: {pc} MTTR={mttr_sec/3600:.1f}h '
                               f'수리예정 {repair_sec/3600:.1f}h')
                    yield env.timeout(repair_sec)
                    self.smt_broken[pc] = False
                    _log_event(env.now, f'SMT 설비 복구: {pc}')

    def _event_worker_absent(self, env):
        """작업자 결근 이벤트. 매 근무일 시작 시 확률 체크."""
        work_day_sec = (
            _active_schedule['work_end_sec']
            - _active_schedule['work_start_sec']
            - _active_schedule['break_duration_sec']
        )
        while True:
            yield env.timeout(_next_work_start(env.now) - env.now)
            for wgrp in self.data.worker_skill:
                if random.random() < WORKER_ABSENT_PROB:
                    self.idle.absent_groups.add(wgrp)
                    cur_skill = self.data.worker_skill.get(wgrp, 2)
                    _log_event(env.now,
                               f'작업자 결근: {wgrp} '
                               f'(숙련도 {cur_skill} -> {max(1, cur_skill - 1)})')
                    yield env.timeout(work_day_sec)
                    self.idle.absent_groups.discard(wgrp)
                    
    def _event_replenishment(self, env):
        """재고 부족 발주 이벤트.
        1시간마다 발주 대기 목록을 확인하고 REPLENISH_LEAD_DAY일 뒤 입고.
        발주 즉시 _pending_orders 에서 제거해 중복 발주를 차단한다.
        도착 후 재고가 다시 부족해지면 consume 이 재등록한다.
        """
        while True:
            yield env.timeout(3600)
            if not _is_work_time(env.now):
                continue
            for item_code in list(self.wh._pending_orders):
                # 발주 즉시 pending 에서 제거 (중복 발주 차단).
                self.wh._pending_orders.discard(item_code)
                # 발주 시각·발주 시점 재고를 입고 시 reorder_log 에 기록하도록 전달
                env.process(self._deliver(env, item_code,
                                           order_time=env.now,
                                           stock_at_order=int(self.wh.stock.get(item_code, 0))))

    def _deliver(self, env, item_code, order_time=None, stock_at_order=None):
        # 발주 후 REPLENISH_LEAD_DAY일 뒤 입고. (임의)
        yield env.timeout(REPLENISH_LEAD_DAY * DAY_SEC)
        self.wh.replenish(item_code, env.now,
                          order_time=order_time, stock_at_order=stock_at_order)

    def _next_model_for_line(self):
        """SMT 라인이 유휴 상태일 때 다음에 처리할 모델을 선택.

        부족도(order_qty - PCB재고합 - 완성품)가 가장 큰 모델을 우선 선택한다.
        동률 시 order dict 삽입 순서를 따른다.
        """
        remaining = [m for m in self.order
                     if self.stats.get(f'smt_done_{m}', 0) < self.order[m]]
        if not remaining:
            return None

        def _shortage(m):
            order_qty = self.order[m]
            codes = [PCB_MAP[m]] + THT_PCB_BY_MODEL.get(m, [])
            stock = sum(self.wh.stock.get(c, 0) for c in codes)
            completed = self.stats.get(f'{m}_done', 0)
            return order_qty - stock - completed

        # 가장 부족한 모델 우선. max() 가 동률일 때 첫번째 발견 항목 반환 →
        # remaining 의 순서가 self.order 삽입 순서를 따라 안정적.
        choice = max(remaining, key=_shortage)
        self.wh.smt_model_choices.append((float(self.env.now), choice))
        return choice

    def _smt_schedule(self):
        # 라인이 유휴 상태가 될 때마다 에이전트가 다음 모델 결정.
        # (2026-05-06) 종료 조건을 누적 공급량 기반으로 교체.
        # 이전 버그: `stock >= target` 으로 비교하면 unit 들이 PCB 를 소비할 때마다
        # stock 이 떨어져 SMT 가 다시 fire → 무한 루프 (10x 과생산, 외주 wave 누적).
        # 수정: target_make = order_qty × (1 - PCB_INITIAL_RATIO). 누적 fire 수가
        # 그 만큼이면 종료. AOI_DEFECT_ACTION='repair' 일 때는 결함 보드도 wh.restore
        # 로 자동 복귀하므로 1회 fire 로 끝남. 'scrap' 일 때는 폐기분만큼 추가 fire.
        def _run_line(env, sid):
            line = self.smt_lines[sid]
            while True:
                model = self._next_model_for_line()
                if model is None:
                    break
                active = {l.assigned_model for l in self.smt_lines.values()
                          if l.assigned_model is not None}
                if model in active:
                    remaining = [m for m in self.order
                                 if self.stats.get(f'smt_done_{m}', 0) < self.order[m]
                                 and m not in active]
                    if not remaining:
                        yield env.timeout(300)  # (임의)
                        continue
                    model = remaining[0]
                line.assigned_model = model
                pcb_codes = ([PCB_MAP[model]] +
                             THT_PCB_BY_MODEL.get(model, []))
                target_total = self.order[model]
                # 초기 재고가 PCB_INITIAL_RATIO 비율로 미리 채워져 있으므로
                # SMT 라인이 만들어야 할 부족분만 발사한다.
                target_make = target_total - int(target_total * PCB_INITIAL_RATIO)
                while True:
                    boards = []
                    for pcb_code in pcb_codes:
                        flow = self.wh.pcb_flow.get(pcb_code, {})
                        # 누적 inventory 도착 (SMT 직접 + 외주 returned + RMA 수리)
                        arrived = int(flow.get('restore_from_smt_or_outsource', 0))
                        # 외주 진행 중 (THT) — 곧 도착 예정
                        in_flight = (int(flow.get('outsource_in', 0))
                                     - int(flow.get('outsource_returned', 0)))
                        # 비-THT mag_buf 에 있는 보드 — 곧 wh.restore 됨
                        in_mag = int(line.mag_buf.get(pcb_code, 0))
                        committed = arrived + in_flight + in_mag
                        if committed >= target_make:
                            continue
                        shortage = target_make - committed
                        for board_id in range(shortage):
                            p = env.process(
                                line.process_board(pcb_code, board_id, model))
                            boards.append(p)
                    if not boards:
                        break
                    yield simpy.AllOf(env, boards)
                    # 발사 분 mag 잔량 회수 (비-THT 만; THT 는 _outsource_return
                    # 가 보드 단위로 wh.restore 직접 호출)
                    for pcb_code in pcb_codes:
                        remainder = line.mag_buf.get(pcb_code, 0)
                        if remainder > 0:
                            line.mag_buf[pcb_code] = 0
                            line.wh.restore(pcb_code, remainder, env.now)
                            line.pcb_count[pcb_code] += remainder
                    # 누적 공급량으로 종료 판정. 외주는 in_flight 로 카운트되므로
                    # 16h 도착 대기 없이 즉시 종료 가능.
                    all_supplied = True
                    for pcb_code in pcb_codes:
                        flow = self.wh.pcb_flow.get(pcb_code, {})
                        arrived = int(flow.get('restore_from_smt_or_outsource', 0))
                        in_flight = (int(flow.get('outsource_in', 0))
                                     - int(flow.get('outsource_returned', 0)))
                        in_mag = int(line.mag_buf.get(pcb_code, 0))
                        if arrived + in_flight + in_mag < target_make:
                            all_supplied = False
                            break
                    if all_supplied:
                        break
                self.stats[f'smt_done_{model}'] = self.order[model]
                line.assigned_model = None

        line_procs = [self.env.process(_run_line(self.env, sid))
                      for sid in SMT_LINE_IDS]

        # 모든 SMT 라인 종료 후 외주 트럭 잔량 flush.
        def _flush_outsource_after_lines_done(env):
            yield simpy.AllOf(env, line_procs)
            self.outsource_pool.flush_now()
        self.env.process(_flush_outsource_after_lines_done(self.env))

    def _report(self, elapsed):
        done = sum(self.stats.get(f'{m}_done', 0) for m in self.order)
        print('\n' + '=' * 60)
        print(f'makespan:{self.env.now/3600:.2f}h | 실행:{elapsed:.2f}s | '
              f'완성:{done}/{sum(self.order.values())}')
        for m, (d, t) in self.progress.items():
            pct = d / max(t, 1) * 100
            bar = '#' * int(pct/5) + '.' * (20-int(pct/5))
            print(f'  {m}: [{bar}] {d}/{t} ({pct:.0f}%)')
        print(f'  불량: SMT={self.stats.get("smt_defect",0)} '
              f'AOI={self.stats.get("aoi_defect",0)} '
              f'조립={self.stats.get("assy_defect",0)} '
              f'수리={self.stats.get("rma_repaired",0)}')
        print(f'  재고부족:{self.wh.stock_penalty()} | 재고품초과:{self.wip.violations()}')
        self.energy.report()
        self.idle.flush_all(self.env)
        self.idle.report()
        print('=' * 60)

    def _summary(self):
        done = sum(self.stats.get(f'{m}_done', 0) for m in self.order)
        return {
            'total_done'  : done,
            'total_order' : sum(self.order.values()),
            'makespan_hr' : self.env.now / 3600,
            'total_kwh'   : self.energy.total,
            '재고품초과'  : self.wip.violations(),
            'stock_pen'   : self.wh.stock_penalty(),
            'total_defect': (self.stats.get('smt_defect', 0) +
                             self.stats.get('aoi_defect', 0) +
                             self.stats.get('assy_defect', 0)),
            'oqc_inspected': self.stats.get('oqc_inspected', 0),
            'by_grp_kwh'  : dict(self.energy.by_grp),
        }
    
    def run(self, training=False):
        t_sec = MAX_DAYS * (
            _active_schedule['work_end_sec']
            - _active_schedule['work_start_sec']
            - _active_schedule['break_duration_sec']
        )
        total_need = sum(self.order.values())
        t0         = time.time()
        stop_event = self.env.event()
        mon_interval = TRAIN_MONITOR_INTERVAL if training else INFER_MONITOR_INTERVAL

        def _check_done(env):
            while True:
                yield env.timeout(30)
                done = sum(self.stats.get(f'{m}_done', 0) for m in self.order)
                if done >= total_need or env.now >= t_sec:
                    if done >= total_need:
                        for wgrp in self.data.workers:
                            self.idle.mark_completed(wgrp, float(env.now))
                    if not stop_event.triggered:
                        stop_event.succeed()
                    return

        self.env.process(run_rma(self.env, self.rma, self.wres, self.wh,
                                 self.energy, self.idle, self.wip,
                                 self.stats, self.data,
                                 progress=self.progress,
                                 plogger=self.plogger))

        if not training:
            self.env.process(monitor(self.env, self.progress, self.energy,
                                     self.wh, self.idle, self.wip, self.stats,
                                     self.smt_lines, interval=mon_interval,
                                     plogger=self.plogger, menv=self))
        self.env.process(_check_done(self.env))
        self.env.process(self._event_smt_breakdown(self.env))
        self.env.process(self._event_worker_absent(self.env))
        self.env.process(self._event_replenishment(self.env))
        # 부품 재고 1h 스냅샷 루프 (Stock_Timeseries 엑셀 시트용)
        self.env.process(self.wh.snapshot_loop(self.env, interval=3600))
        self.env.process(self.wip.snapshot_loop(self.env, interval=3600))
        self._smt_schedule()

        for m in self.order:
            for uid in range(self.order[m]):
                self.env.process(
                    produce_unit(self.env, m, uid, self.data, self.graphs[m],
                                 self.wres, self.wh, self.rma,
                                 self.energy, self.idle, self.wip,
                                 self.stats, self.progress, menv=self,
                                 plogger=self.plogger))

        self.env.run(until=stop_event)
        if not training:
            self._report(time.time() - t0)
        else:
            self._train_elapsed = time.time() - t0
        return self._summary()

# ══════════════════════════════════════════════════════════
# M15. PPO + GNN 그래프 임베딩
# ══════════════════════════════════════════════════════════

class ExperimentRunner:
    """
    상태 : 시뮬레이션 스칼라 벡터 + GNN 그래프 임베딩
    행동 : 조립 공정 내 실행 가능 공정 중 우선순위 선택 (GNN 노드 스코어 기반)
          경험 (s, emb, a, r, lp, v, mask)은 produce_unit에서 store()로 직접 수집
    보상 : w1*(이번스텝시간감소) + w2*(-전력증가) - w3*WIP초과
           - w4*재고부족(critical_stock 기준) + w5*납기 + w6*(-작업자유휴)
    수렴 : 최근 CONV_WINDOW 에피소드 평균 보상 변화 < CONV_THRESHOLD
    """
    def __init__(self, data, order: dict):
        """
        data  : CombinedDataLoader 인스턴스 (FallbackDataLoader 도 호환)
        order : {model_id: qty} 주문 딕셔너리
        """
        self.data  = data
        self.order = order

    def _state_dim(self):
        # get_state() 반환 크기: len(order) + len(wres) + 6 스칼라
        # ManufacturingEnv 전체를 생성하지 않고 직접 계산.
        n_models  = len(self.order)
        n_workers = len(self.data.workers)
        # 스칼라 6개: kWh율, 경과시간, 재고부족, WIP초과, 유휴, SMT고장
        return n_models + n_workers + 6

    def _build_tensors(self, kg: ProcessKnowledgeGraph):
        _, H_np = kg.get_feat_matrix()
        adj_np  = kg.get_adj()
        return (torch.tensor(H_np, dtype=torch.float32),
                torch.tensor(adj_np, dtype=torch.float32))

    def run_ppo_training(self, max_episodes=5000):
        print(f'\n[PPO+GNN 학습] 상한 {max_episodes} 에피소드, 수렴 시 조기 종료')
        s_dim = self._state_dim()
        gnn = ProcessGNN(in_dim=6, hidden=32, out_dim=16)
        agent = PPOAgent(s_dim, gnn)

        if os.path.exists(POLICY_PATH):
            agent.load()
            start_ep = len(agent.ep_rewards)
            print(f'  이전 학습 이어서 진행 (이미 {start_ep} 에피소드 완료)')
        else:
            start_ep = 0

        # 2026-04-29: 학습 중 ep 단위 진행 표시. verbose monitor·_report 는
        # run(training=True) 에서 차단했고 여기서 한 줄 status 만 출력.
        # 2026-05-07: 주기 저장 (CKPT_EVERY ep) + try/finally — Ctrl+C 등 중단 시에도
        # 정책·ep_rewards·ep_rewards_decomp 보존.
        CKPT_EVERY = 5
        train_t0 = time.time()
        ep = 0
        try:
            for ep in range(1, max_episodes + 1):
                ep_t0 = time.time()
                # 고정 가중치(W_DEFAULT) 사용
                menv = ManufacturingEnv(self.data, self.order)
                menv._init_sim()
                # produce_unit이 공정 선택 시 agent.act() + agent.store()를
                # 직접 호출하도록 agent를 환경에 연결
                menv.agent = agent

                # 시뮬레이션 실행 - produce_unit 내부에서 경험 수집이 일어남
                menv.run(training=True)

                # 에피소드 종료 후 PPO 업데이트 - 모델별 H/adj 전달
                graphs_cache = {
                    m: self._build_tensors(kg)
                    for m, kg in menv.graphs.items()
                }
                ep_r = agent.update(graphs_cache=graphs_cache)

                # 항별 분해 적재 (reward() 안에서 menv._reward_decomp_sum 누적)
                decomp = list(getattr(menv, '_reward_decomp_sum', [0.0]*6))
                agent.ep_rewards_decomp.append(decomp)

                # 한 줄 status: ep / makespan / 보상 / 평균 / 벽시계 시간 / ETA
                ep_dur   = time.time() - ep_t0
                elapsed  = time.time() - train_t0
                avg_dur  = elapsed / ep
                remain   = max_episodes - ep
                eta_min  = avg_dur * remain / 60
                avg_r    = float(np.mean(agent.ep_rewards[-min(100, len(agent.ep_rewards)):])) \
                           if agent.ep_rewards else 0.0
                done_now = sum(menv.stats.get(f'{m}_done', 0) for m in menv.order)
                tot_need = sum(menv.order.values())
                ms_h     = menv.env.now / 3600
                print(f'  ep {start_ep+ep:4d}/{start_ep+max_episodes:<4d} '
                      f'| sim={ms_h:6.2f}h done={done_now}/{tot_need} '
                      f'| R={ep_r:+8.3f} avg={avg_r:+8.3f} '
                      f'| {ep_dur:5.1f}s | ETA {eta_min:5.1f}m',
                      flush=True)

                # 주기 저장 (silent — 매 N ep 마다 ppo_policy.pt 갱신)
                if ep % CKPT_EVERY == 0:
                    agent.save(verbose=False)

                if agent.is_converged():
                    print(f'  수렴 (ep{start_ep + ep}). 학습 종료.')
                    break
        except KeyboardInterrupt:
            print(f'\n  [학습 중단됨] ep{start_ep + ep} 까지 완료. 현재 상태 저장.')

        agent.save()
        self._last_agent = agent
        total_min = (time.time() - train_t0) / 60
        print(f'\n[학습 완료] 총 {start_ep + ep} 에피소드 ({total_min:.1f}분)')
        return agent

    def run_inference(self, agent=None):
        """추론 실행.

        agent: run_ppo_training()이 반환한 PPOAgent 객체.
               None이면 ppo_policy.pt 파일에서 로드 시도.
               파일도 없으면 greedy(ready_pcs[0]) 모드로 실행.
        """
        print('\n[추론 실행]')
        menv = ManufacturingEnv(self.data, self.order)

        if agent is not None:
            # 학습 직후 호출된 경우 - 파일 없이 agent 객체 직접 연결
            agent.eval()
            menv.agent = agent
            self._last_agent = agent
            print('  학습된 정책 적용.')
        elif os.path.exists(POLICY_PATH):
            # 이전에 저장된 정책 파일 로드
            try:
                gnn   = ProcessGNN(in_dim=6, hidden=32, out_dim=16)
                s_dim = len(menv.get_state())
                loaded = PPOAgent(s_dim, gnn)
                loaded.load()
                loaded.eval()
                menv.agent = loaded
                self._last_agent = loaded
                print(f'  저장된 정책 로드: {POLICY_PATH}')
            except Exception as e:
                print(f'  [경고] 정책 로드 실패 -> greedy 모드: {e}')
                menv.agent = None
        else:
            print('  저장된 정책 없음 -> greedy 모드로 진행.')
            menv.agent = None

        summary = menv.run()
        self._last_menv = menv
        return summary

    def save_results(self, inference_summary=None, path=RESULT_PATH):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb       = openpyxl.Workbook()
        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_fill = PatternFill('solid', fgColor='2E4053')

        def _hdr(ws, headers):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font  = hdr_font
                c.fill  = hdr_fill
                c.alignment = Alignment(horizontal='center')

        def _aw(ws):
            for col in ws.columns:
                ml = max((len(str(c.value)) for c in col if c.value), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+2,30)

        ws = wb.active
        ws.title = 'Inference'
        if inference_summary:
            _hdr(ws, ['항목','값'])
            items = [
                ('makespan_hr',  inference_summary.get('makespan_hr', 0)),
                ('total_kwh',    inference_summary.get('total_kwh', 0)),
                ('재고품초과',   inference_summary.get('재고품초과', 0)),
                ('stock_pen',    inference_summary.get('stock_pen', 0)),
                ('total_defect', inference_summary.get('total_defect', 0)),
                ('total_done',   inference_summary.get('total_done', 0)),
                ('total_order',  inference_summary.get('total_order', 0)),
            ]
            for r, (k, v) in enumerate(items, 2):
                ws.cell(r, 1, k)
                ws.cell(r, 2, round(float(v),6) if isinstance(v, float) else v)
            row_i = len(items) + 3
            ws.cell(row_i, 1, '공정그룹')
            ws.cell(row_i, 2, 'kWh')
            for grp, kwh in inference_summary.get('by_grp_kwh', {}).items():
                row_i += 1
                ws.cell(row_i, 1, grp)
                ws.cell(row_i, 2, round(kwh, 6))
        _aw(ws)

        agent = getattr(self, '_last_agent', None)
        if agent and agent.ep_rewards:
            ws2 = wb.create_sheet('Training_Rewards')
            _hdr(ws2, ['episode','reward','rolling_avg_100'])
            for i, r in enumerate(agent.ep_rewards, 1):
                ws2.cell(i+1, 1, i)
                ws2.cell(i+1, 2, round(float(r), 6))
                if i >= 100:
                    ws2.cell(i+1, 3, round(float(np.mean(agent.ep_rewards[i-100:i])),6))
            _aw(ws2)

            


        # ══════════════════════════════════════════════════════════
        # 부품 재고 · 공정 추적 시트 (2026-4-22 추가)
        # ══════════════════════════════════════════════════════════
        menv = getattr(self, '_last_menv', None)
        if menv is not None:
            wh = menv.wh
            plogger = menv.plogger
            makespan_s = int(menv.env.now)
            max_hour = max(1, makespan_s // 3600 + 1)
            if max_hour > 720:
                print(f'[경고] makespan {max_hour}h > 720h - Stock_Timeseries 열 수가 많아 '
                      f'openpyxl 쓰기가 오래 걸릴 수 있습니다.')

            # BOM 마스터 전체 item_code (~519개). 한 번도 소비 없던 부품도 포함.
            all_items = set()
            try:
                all_items = wh.data.iter_all_bom_items()
            except AttributeError:
                pass
            all_items.update(wh.stock.keys())
            all_items.update(wh.snapshots.keys())
            all_items = sorted(all_items)

            # ── Stock_Summary ─────────────────────────────────────
            ws_s = wb.create_sheet('Stock_Summary')
            _hdr(ws_s, ['item_code', 'item_name', 'initial_stock',
                        'total_consumed', 'final_stock', 'min_stock_qty',
                        'lot_size', 'violations_count', 'reorder_count'])
            r = 2
            for code in all_items:
                try:
                    name = wh.data.get_item_name(code)
                except AttributeError:
                    name = ''
                ws_s.cell(r, 1, code)
                ws_s.cell(r, 2, name)
                init_val = wh._initial_stocks.get(code,
                    wh._bom_init_stock if code in wh._bom_codes else wh._init_stock)
                ws_s.cell(r, 3, int(init_val))
                ws_s.cell(r, 4, int(wh.consumed.get(code, 0)))
                ws_s.cell(r, 5, int(wh.stock.get(code, init_val)))
                try:
                    ws_s.cell(r, 6, float(wh.data.get_min_stock(code)))
                except Exception:
                    ws_s.cell(r, 6, MIN_STOCK)
                try:
                    ws_s.cell(r, 7, int(wh._lot_for(code)))
                except Exception:
                    ws_s.cell(r, 7, 0)
                ws_s.cell(r, 8, int(wh.violations.get(code, 0)))
                ws_s.cell(r, 9, int(wh.reorder_count.get(code, 0)))
                r += 1
            _aw(ws_s)

            # ── Stock_Timeseries (wide: 행=부품, 열=시간) ─────────
            # 시간 그리드는 1h 간격. snapshots 에서 각 bucket 의 마지막 값 사용,
            # 비어있으면 이전 hour 값 또는 초기재고.
            ws_t = wb.create_sheet('Stock_Timeseries')
            header = ['item_code', 'item_name'] + [f't={h}h' for h in range(max_hour + 1)]
            _hdr(ws_t, header)
            r = 2
            for code in all_items:
                try:
                    name = wh.data.get_item_name(code)
                except AttributeError:
                    name = ''
                ws_t.cell(r, 1, code)
                ws_t.cell(r, 2, name)
                # snapshots[code] = [(t_sec, stock), ...] - hour bucket 으로 재정렬
                by_hour = {}
                for t_sec, q in wh.snapshots.get(code, []):
                    by_hour[int(t_sec // 3600)] = int(q)
                prev = int(wh._initial_stocks.get(code,
                    wh._bom_init_stock if code in wh._bom_codes else wh._init_stock))
                for h in range(max_hour + 1):
                    if h in by_hour:
                        prev = by_hour[h]
                    ws_t.cell(r, 3 + h, prev)
                r += 1

            # ── Stock_Events (long: 디버그용, wh.history 전체 덤프) ─
            ws_e = wb.create_sheet('Stock_Events')
            _hdr(ws_e, ['time_sec', 'time_hr', 'item_code', 'stock_after'])
            r = 2
            # 수십만 row 가능 - 60k 상한으로 truncate (Excel 1M row 한도 내이지만
            # openpyxl 쓰기 속도·파일 크기 고려)
            EVENT_CAP = 60000
            count = 0
            stop = False
            for code in all_items:
                if stop:
                    break
                for t_sec, q in wh.history.get(code, []):
                    ws_e.cell(r, 1, float(t_sec))
                    ws_e.cell(r, 2, round(float(t_sec) / 3600, 3))
                    ws_e.cell(r, 3, code)
                    ws_e.cell(r, 4, int(q))
                    r += 1
                    count += 1
                    if count >= EVENT_CAP:
                        ws_e.cell(r, 1, '[TRUNCATED]')
                        ws_e.cell(r, 3, f'events > {EVENT_CAP}')
                        stop = True
                        break
            _aw(ws_e)

            # ── Process_Log (wide: 행=process_code, 열=시간) ──────
            ws_p = wb.create_sheet('Process_Log')
            pheader = ['process_code'] + [f't={h}h' for h in range(max_hour + 1)]
            _hdr(ws_p, pheader)
            r = 2
            if plogger is not None:
                summary = plogger.summary()
                # 현재 진행중인(mark_end 미호출) 공정도 포함시켜 마지막 시간대 표시
                for pc, (mid, uid, t0) in plogger.current.items():
                    label = f'{mid}/u{uid+1}*'
                    h_start = int(t0 // 3600)
                    h_end = max_hour
                    summary.setdefault(pc, {})
                    for h in range(h_start, h_end + 1):
                        existing = summary[pc].get(h, '')
                        summary[pc][h] = (existing + '; ' + label) if existing else label
                for pc in sorted(summary):
                    ws_p.cell(r, 1, pc)
                    row_map = summary[pc]
                    for h in range(max_hour + 1):
                        val = row_map.get(h, '')
                        if val:
                            ws_p.cell(r, 2 + h, val)
                    r += 1
            _aw(ws_p)

            # ── Reorder_Log (long) ────────────────────────────────
            ws_r = wb.create_sheet('Reorder_Log')
            _hdr(ws_r, ['item_code', 'order_time_hr', 'arrive_time_hr',
                        'lead_hr', 'lot_size', 'incoming', 'stock_at_order'])
            r = 2
            for entry in wh.reorder_log:
                ot = float(entry.get('order_time', 0)) / 3600
                at = float(entry.get('arrive_time', 0)) / 3600
                ws_r.cell(r, 1, entry.get('item_code', ''))
                ws_r.cell(r, 2, round(ot, 3))
                ws_r.cell(r, 3, round(at, 3))
                ws_r.cell(r, 4, round(at - ot, 3))
                ws_r.cell(r, 5, int(entry.get('lot_size', 0)))
                ws_r.cell(r, 6, int(entry.get('incoming', 0)))
                ws_r.cell(r, 7, int(entry.get('stock_at_order', 0)))
                r += 1
            _aw(ws_r)

            print(f'  재고·공정 시트 5개 추가 (items={len(all_items)}, '
                  f'events={count}, reorders={len(wh.reorder_log)})')

            # ══════════════════════════════════════════════════════
            # 디버그 시트 5종 (2026-04-29 추가) — Q1·Q2·Q3 진단용
            # ══════════════════════════════════════════════════════

            # ── Debug_Model_Stats ──────────────────────────────────
            # 모델별 실제 done 카운터 + 정상/RMA 경로 분리 + 1·2일차 첫 공정
            # 시작 시간. "MODEL_B 가 왜 2일차에 시작했나" 직접 답.
            ws_dms = wb.create_sheet('Debug_Model_Stats')
            _hdr(ws_dms, ['model_id', 'order_qty', 'stats_done',
                          'normal_completions', 'rma_completions',
                          'blocked_by_quota', 'first_event_h',
                          'last_event_h', 'first_after_24h',
                          'first_after_48h'])
            comps = wh.unit_completions
            events_all = plogger.events if plogger is not None else []
            r = 2
            for model in menv.order:
                qty = menv.order[model]
                stats_done = int(menv.stats.get(f'{model}_done', 0))
                model_comps = [v for k, v in comps.items()
                               if isinstance(k, tuple) and k[0] == model]
                n_normal = sum(1 for c in model_comps if c['path'] == 'normal')
                n_rma    = sum(1 for c in model_comps if c['path'] == 'rma')
                n_blocked = sum(1 for c in model_comps
                                if 'blocked_by_quota' in c['path'])
                m_evs = [e for e in events_all if e.get('mid') == model]
                first_h = round(min((e['start'] for e in m_evs), default=0)/3600, 2)
                last_h  = round(max((e['end']   for e in m_evs), default=0)/3600, 2)
                after_24 = round(min((e['start'] for e in m_evs
                                      if e['start'] >= 24*3600), default=0)/3600, 2)
                after_48 = round(min((e['start'] for e in m_evs
                                      if e['start'] >= 48*3600), default=0)/3600, 2)
                ws_dms.cell(r, 1, model)
                ws_dms.cell(r, 2, qty)
                ws_dms.cell(r, 3, stats_done)
                ws_dms.cell(r, 4, n_normal)
                ws_dms.cell(r, 5, n_rma)
                ws_dms.cell(r, 6, n_blocked)
                ws_dms.cell(r, 7, first_h)
                ws_dms.cell(r, 8, last_h)
                ws_dms.cell(r, 9, after_24)
                ws_dms.cell(r, 10, after_48)
                r += 1
            _aw(ws_dms)

            # ── Debug_Unit_Status ──────────────────────────────────
            # 모든 (model, uid) 의 최종 상태. 'cutoff' = sim 끝났는데
            # produce_unit while 루프 미완료. BT5_122 미기록 원인 진단.
            ws_dus = wb.create_sheet('Debug_Unit_Status')
            _hdr(ws_dus, ['model_id', 'unit_key', 'completion_path',
                          'end_time_h', 'done_n', 'total_n',
                          'live_state', 'live_pc', 'live_done_n',
                          'live_total_n'])
            r = 2
            live_us = getattr(menv, 'unit_states', {}) or {}
            seen_keys = set()
            for k, c in sorted(comps.items(),
                                key=lambda kv: (str(kv[0][0]), str(kv[0][1]))):
                model = k[0]
                uid_or_tag = k[1]
                ws_dus.cell(r, 1, str(model))
                ws_dus.cell(r, 2, str(uid_or_tag))
                ws_dus.cell(r, 3, c.get('path', ''))
                ws_dus.cell(r, 4, round(c.get('end_time', 0)/3600, 3))
                ws_dus.cell(r, 5, int(c.get('done_n', -1)))
                ws_dus.cell(r, 6, int(c.get('total_n', -1)))
                if isinstance(uid_or_tag, int):
                    ls = live_us.get((model, uid_or_tag), {})
                    ws_dus.cell(r, 7, str(ls.get('state', '')))
                    ws_dus.cell(r, 8, str(ls.get('pc', '')))
                    ws_dus.cell(r, 9, int(ls.get('done_n', 0)))
                    ws_dus.cell(r, 10, int(ls.get('total_n', 0)))
                    seen_keys.add((model, uid_or_tag))
                r += 1
            # produce_unit 시작했지만 unit_completions 에 없는 = cutoff (sim 종료로 잘림)
            for (model, uid), ls in sorted(live_us.items(),
                                           key=lambda kv: (str(kv[0][0]), kv[0][1])):
                if (model, uid) in seen_keys:
                    continue
                ws_dus.cell(r, 1, str(model))
                ws_dus.cell(r, 2, str(uid))
                ws_dus.cell(r, 3, 'cutoff')
                ws_dus.cell(r, 4, round(menv.env.now/3600, 3))
                ws_dus.cell(r, 5, int(ls.get('done_n', 0)))
                ws_dus.cell(r, 6, int(ls.get('total_n', 0)))
                ws_dus.cell(r, 7, str(ls.get('state', '')))
                ws_dus.cell(r, 8, str(ls.get('pc', '')))
                ws_dus.cell(r, 9, int(ls.get('done_n', 0)))
                ws_dus.cell(r, 10, int(ls.get('total_n', 0)))
                r += 1
            _aw(ws_dus)

            # ── Debug_Process_Coverage ─────────────────────────────
            # 모든 (model, process_code) × 실제 이벤트 개수.
            # 'expected' 는 order_qty 단순 비교 (RMA 경로 unit 은 일부 PACK 만
            # 통과하므로 over/under 가능). missing_count 가 양수면 의심.
            ws_dpc = wb.create_sheet('Debug_Process_Coverage')
            _hdr(ws_dpc, ['model_id', 'process_code', 'process_group',
                          'worker_group', 'expected_qty', 'actual_count',
                          'first_h', 'last_h'])
            r = 2
            for model in menv.order:
                qty = menv.order[model]
                kg_nodes = menv.graphs[model].nodes if model in menv.graphs else {}
                for pc in sorted(kg_nodes.keys()):
                    node = kg_nodes[pc]
                    pc_evs = [e for e in events_all
                              if e.get('mid') == model and e.get('pc') == pc]
                    cnt = len(pc_evs)
                    f_h = round(min((e['start'] for e in pc_evs),
                                    default=0)/3600, 2) if pc_evs else ''
                    l_h = round(max((e['end']   for e in pc_evs),
                                    default=0)/3600, 2) if pc_evs else ''
                    ws_dpc.cell(r, 1, model)
                    ws_dpc.cell(r, 2, pc)
                    ws_dpc.cell(r, 3, node.get('process_group', ''))
                    ws_dpc.cell(r, 4, node.get('worker_group', ''))
                    ws_dpc.cell(r, 5, qty)
                    ws_dpc.cell(r, 6, cnt)
                    ws_dpc.cell(r, 7, f_h)
                    ws_dpc.cell(r, 8, l_h)
                    r += 1
            _aw(ws_dpc)

            # ── Debug_PCB_Flow ─────────────────────────────────────
            # PCB 본 코드별 흐름 카운터: SMT 완성 / 외주 출발 / 외주 복귀 /
            # 외부 발주 트리거(버그 후보) / 외부 발주 도착(버그 후보).
            ws_dpf = wb.create_sheet('Debug_PCB_Flow')
            _hdr(ws_dpf, ['pcb_code', 'role', 'model_hint',
                          'initial_stock', 'final_stock', 'total_consumed',
                          'smt_or_outsource_restore',
                          'outsource_in', 'outsource_returned',
                          'external_order_trigger', 'external_replenish_arrived',
                          'is_bug_candidate'])
            r = 2
            model_for_pcb = {}
            for m, pc_main in PCB_MAP.items():
                model_for_pcb[pc_main] = (m, 'main')
            for m, ths in THT_PCB_BY_MODEL.items():
                for pc_t in ths:
                    model_for_pcb[pc_t] = (m, 'tht')
            for code in sorted(wh._pcb_codes):
                role = model_for_pcb.get(code, ('-', '-'))
                flow = wh.pcb_flow.get(code, {})
                ext_trig = int(flow.get('external_order_trigger', 0))
                ext_arr  = int(flow.get('external_replenish_arrived', 0))
                bug = ext_trig > 0 or ext_arr > 0
                ws_dpf.cell(r, 1, code)
                ws_dpf.cell(r, 2, role[1])
                ws_dpf.cell(r, 3, role[0])
                ws_dpf.cell(r, 4, int(wh._initial_stocks.get(code,
                    wh._bom_init_stock if code in wh._bom_codes else wh._init_stock)))
                ws_dpf.cell(r, 5, int(wh.stock.get(code, 0)))
                ws_dpf.cell(r, 6, int(wh.consumed.get(code, 0)))
                ws_dpf.cell(r, 7, int(flow.get('restore_from_smt_or_outsource', 0)))
                ws_dpf.cell(r, 8, int(flow.get('outsource_in', 0)))
                ws_dpf.cell(r, 9, int(flow.get('outsource_returned', 0)))
                ws_dpf.cell(r, 10, ext_trig)
                ws_dpf.cell(r, 11, ext_arr)
                ws_dpf.cell(r, 12, 'BUG' if bug else '')
                r += 1
            # SMT 라인×모델×PCB 처리량
            r += 1
            ws_dpf.cell(r, 1, '== SMT 처리량 (line × model × pcb) ==')
            r += 1
            ws_dpf.cell(r, 1, 'line')
            ws_dpf.cell(r, 2, 'model')
            ws_dpf.cell(r, 3, 'pcb_code')
            ws_dpf.cell(r, 4, 'restored_qty')
            r += 1
            for (sid, m, pc_code), q in sorted(wh.smt_per_model.items()):
                ws_dpf.cell(r, 1, sid)
                ws_dpf.cell(r, 2, m)
                ws_dpf.cell(r, 3, pc_code)
                ws_dpf.cell(r, 4, int(q))
                r += 1
            _aw(ws_dpf)

            # ── Debug_Outsource_Log ────────────────────────────────
            # THT 외주 발송 → 도착. status='in_flight' 면 sim 종료까지 미복귀.
            ws_dol = wb.create_sheet('Debug_Outsource_Log')
            _hdr(ws_dol, ['pcb_code', 'model_id', 'board_id',
                          'send_time_h', 'return_time_h', 'transit_h',
                          'delay_h', 'status'])
            r = 2
            for entry in wh.outsource_log:
                ret = entry.get('return_time')
                send_h = round(entry.get('send_time', 0)/3600, 3)
                ret_h  = round(ret/3600, 3) if ret is not None else ''
                transit = round((ret - entry.get('send_time', 0))/3600, 3) \
                          if ret is not None else ''
                ws_dol.cell(r, 1, entry.get('pcb_code', ''))
                ws_dol.cell(r, 2, entry.get('model_id', ''))
                ws_dol.cell(r, 3, entry.get('board_id', ''))
                ws_dol.cell(r, 4, send_h)
                ws_dol.cell(r, 5, ret_h)
                ws_dol.cell(r, 6, transit)
                ws_dol.cell(r, 7, round(entry.get('delay_sec', 0)/3600, 3))
                ws_dol.cell(r, 8, entry.get('status', ''))
                r += 1
            _aw(ws_dol)

            # Reorder_Log 시트에 is_pcb 컬럼 추가 — 이미 위에서 작성된 시트라
            # 여기서는 별도의 보조 시트로 PCB 만 따로 모아 보여줌.
            ws_dpr = wb.create_sheet('Debug_PCB_Reorders')
            _hdr(ws_dpr, ['item_code', 'order_h', 'arrive_h',
                          'lot_size', 'incoming', 'stock_at_order'])
            r = 2
            for entry in wh.reorder_log:
                if not entry.get('is_pcb'):
                    continue
                ws_dpr.cell(r, 1, entry.get('item_code', ''))
                ws_dpr.cell(r, 2, round(entry.get('order_time', 0)/3600, 3))
                ws_dpr.cell(r, 3, round(entry.get('arrive_time', 0)/3600, 3))
                ws_dpr.cell(r, 4, int(entry.get('lot_size', 0)))
                ws_dpr.cell(r, 5, int(entry.get('incoming', 0)))
                ws_dpr.cell(r, 6, int(entry.get('stock_at_order', 0)))
                r += 1
            _aw(ws_dpr)

            # ── Debug_Skipped_PCs ─────────────────────────────────
            # produce_unit 의 prow=None 스킵 카운터: events 누락의 직접 증거.
            ws_dsk = wb.create_sheet('Debug_Skipped_PCs')
            _hdr(ws_dsk, ['model_id', 'process_code', 'skip_count',
                          'in_excel_pf', 'in_kg_nodes'])
            r = 2
            for (m, pc), c in sorted(wh.skipped_pcs.items()):
                in_pf = wh.data._pc_map.get(pc) is not None
                in_kg = (m in menv.graphs and
                         pc in menv.graphs[m].nodes)
                ws_dsk.cell(r, 1, m)
                ws_dsk.cell(r, 2, pc)
                ws_dsk.cell(r, 3, int(c))
                ws_dsk.cell(r, 4, 'Y' if in_pf else 'N')
                ws_dsk.cell(r, 5, 'Y' if in_kg else 'N')
                r += 1
            _aw(ws_dsk)

            # ── Debug_SMT_Choices ─────────────────────────────────
            # SMT 라인이 유휴 시 모델을 어떻게 골랐는지 시계열 로그.
            ws_dsm = wb.create_sheet('Debug_SMT_Choices')
            _hdr(ws_dsm, ['time_h', 'chosen_model'])
            r = 2
            for (t, m) in wh.smt_model_choices:
                ws_dsm.cell(r, 1, round(t/3600, 3))
                ws_dsm.cell(r, 2, m)
                r += 1
            _aw(ws_dsm)

            # ── Debug_Safety_Alarms ───────────────────────────────
            # B1·B5 안전 로직 알람. 정상 동작 시엔 빈 시트.
            ws_dsa = wb.create_sheet('Debug_Safety_Alarms')
            _hdr(ws_dsa, ['alarm_type', 'detail_1', 'detail_2',
                          'detail_3', 'time_h', 'extra'])
            r = 2
            for entry in wh.kg_incomplete_log:
                ws_dsa.cell(r, 1, 'B1_kg_incomplete')
                ws_dsa.cell(r, 2, entry['model_id'])
                ws_dsa.cell(r, 3, str(entry['unit_id']))
                ws_dsa.cell(r, 4, '')
                ws_dsa.cell(r, 5, round(entry['time_h'], 3))
                ws_dsa.cell(r, 6, ','.join(entry['missing_pcs']))
                r += 1
            for entry in wh.smt_single_side_log:
                ws_dsa.cell(r, 1, 'B5_single_side_double_pcb')
                ws_dsa.cell(r, 2, entry['pcb_code'])
                ws_dsa.cell(r, 3, entry['model_id'])
                ws_dsa.cell(r, 4, str(entry['board_id']))
                ws_dsa.cell(r, 5, round(entry['time_h'], 3))
                ws_dsa.cell(r, 6, entry.get('reason', ''))
                r += 1
            # B4 BOM 중복 행 충돌 로그
            dup_log = getattr(wh.data, '_bom_dup_merge_log', [])
            for msg in dup_log:
                ws_dsa.cell(r, 1, 'B4_bom_smt_side_conflict')
                ws_dsa.cell(r, 2, str(msg))
                r += 1
            # 재고 차단 timeout fallback 로그
            for entry in wh.stuck_wait_log:
                ws_dsa.cell(r, 1, 'wait_stock_timeout_fallback')
                ws_dsa.cell(r, 2, entry['item_code'])
                ws_dsa.cell(r, 3, str(entry['qty']))
                ws_dsa.cell(r, 4, f'stock={entry["stock_at_end"]}')
                ws_dsa.cell(r, 5, round(entry['wait_end_h'], 3))
                ws_dsa.cell(r, 6, f'wait_for_{entry["wait_end_h"]-entry["wait_start_h"]:.2f}h')
                r += 1
            _aw(ws_dsa)

            # ── Debug_Plogger_Events ──────────────────────────────
            # plogger.events 직접 dump. 간트차트가 사용하는 데이터 그대로.
            # 시각화에 안 보이는 이벤트가 있는지 (wgrp 누락, slot=-1 등) 진단.
            ws_dpe = wb.create_sheet('Debug_Plogger_Events')
            _hdr(ws_dpe, ['pc', 'mid', 'uid', 'grp', 'wgrp', 'slot',
                          'start_h', 'end_h', 'dur_s', 'work_timed'])
            r = 2
            evs = list(plogger.events) if plogger else []
            # PACK·LABEL 만 추출 (다른 그룹은 건수 너무 많아 시트 비대화)
            for e in evs:
                if e.get('grp') != 'PACK':
                    continue
                ws_dpe.cell(r, 1, str(e.get('pc', '')))
                ws_dpe.cell(r, 2, str(e.get('mid', '')))
                ws_dpe.cell(r, 3, int(e.get('uid', 0)))
                ws_dpe.cell(r, 4, str(e.get('grp', '')))
                ws_dpe.cell(r, 5, str(e.get('wgrp', '')))
                ws_dpe.cell(r, 6, int(e.get('slot', -1)))
                ws_dpe.cell(r, 7, round(float(e.get('start', 0))/3600, 3))
                ws_dpe.cell(r, 8, round(float(e.get('end', 0))/3600, 3))
                ws_dpe.cell(r, 9, round(float(e.get('end', 0))-float(e.get('start', 0)), 2))
                ws_dpe.cell(r, 10, str(e.get('work_timed', False)))
                r += 1
            _aw(ws_dpe)

            # PACK render 필터 진단: wgrp 비어있거나 slot<0 인 PACK 이벤트 카운트
            pack_no_wgrp = sum(1 for e in evs
                               if e.get('grp') == 'PACK' and not e.get('wgrp'))
            pack_bad_slot = sum(1 for e in evs
                                if e.get('grp') == 'PACK' and e.get('slot', -1) < 0)
            from collections import Counter
            pack_per_model = Counter(e.get('mid') for e in evs
                                     if e.get('grp') == 'PACK')

            print(f'  디버그 시트 10개 추가 (unit_comps={len(comps)}, '
                  f'outsource_evs={len(wh.outsource_log)}, '
                  f'skipped={sum(wh.skipped_pcs.values())}, '
                  f'smt_choices={len(wh.smt_model_choices)}, '
                  f'kg_incomplete={len(wh.kg_incomplete_log)}, '
                  f'single_side_alarm={len(wh.smt_single_side_log)})')
            print(f'  [PACK 이벤트 진단] no_wgrp={pack_no_wgrp}, bad_slot={pack_bad_slot}, '
                  f'per_model={dict(pack_per_model)}')

            # ── WIP_Timeseries (그룹별 시간별 진행 unit 수) ────────
            wip = menv.wip
            ws_wt = wb.create_sheet('WIP_Timeseries')
            grps = ['SMT', 'MODULE', 'SEMI', 'SET', 'INSP', 'PACK', 'RMA']
            _hdr(ws_wt, ['hour'] + grps)
            # 모든 그룹의 시간 grid 합쳐 정렬
            hours = sorted({int(t // 3600) for g in grps
                            for t, _ in wip.snapshots.get(g, [])})
            # 그룹별 (hour → count) lookup
            lookup = {g: {int(t // 3600): n for t, n in wip.snapshots.get(g, [])}
                      for g in grps}
            r = 2
            for h in hours:
                ws_wt.cell(r, 1, h)
                for ci, g in enumerate(grps, 2):
                    ws_wt.cell(r, ci, lookup[g].get(h, 0))
                r += 1
            _aw(ws_wt)

            # ── Truck_Log (외주 트럭 단위 출발·도착) ──────────────
            pool = getattr(menv, 'outsource_pool', None)
            ws_tl = wb.create_sheet('Truck_Log')
            _hdr(ws_tl, ['dispatch_id', 'send_h', 'eta_h', 'return_h',
                         'transit_h', 'delay_h', 'boards', 'truck_count_eq',
                         'pcb_breakdown'])
            r = 2
            for entry in (pool.truck_log if pool else []):
                ret_h = (entry['return_t'] / 3600) if entry.get('return_t') else None
                send_h = entry['send_t'] / 3600
                eta_h  = entry['eta'] / 3600
                bk = ', '.join(f"{c[-4:]}×{n}"
                               for c, n in entry['breakdown'].items())
                ws_tl.cell(r, 1, entry['dispatch_id'])
                ws_tl.cell(r, 2, round(send_h, 3))
                ws_tl.cell(r, 3, round(eta_h, 3))
                ws_tl.cell(r, 4, round(ret_h, 3) if ret_h else '')
                ws_tl.cell(r, 5, round((ret_h - send_h), 3) if ret_h else '')
                ws_tl.cell(r, 6, round(entry['delay_sec'] / 3600, 3))
                ws_tl.cell(r, 7, entry['size'])
                ws_tl.cell(r, 8, entry['truck_count'])
                ws_tl.cell(r, 9, bk)
                r += 1
            _aw(ws_tl)

            # ── SMT_Stage_Activity (라인별 stage 별 events) ─────
            ws_sa = wb.create_sheet('SMT_Stage_Activity')
            _hdr(ws_sa, ['line', 'stage', 'pcb_code', 'model_id',
                         'board_id', 'is_second', 'start_h', 'end_h', 'dur_s'])
            r = 2
            for sid, line in menv.smt_lines.items():
                for ev in line.stage_events:
                    pc = ev['pc']
                    # 'SMT_LOADER_L1' → stage = 'LOADER', line = 'L1'
                    # 'SMT_AOI'       → stage = 'AOI', line = '공유'
                    if pc == 'SMT_AOI':
                        stage_name = 'AOI'
                        line_name  = '공유'
                    else:
                        parts = pc.split('_')
                        line_name  = parts[-1]
                        stage_name = '_'.join(parts[1:-1])
                    ws_sa.cell(r, 1, line_name)
                    ws_sa.cell(r, 2, stage_name)
                    ws_sa.cell(r, 3, ev['pcb_code'])
                    ws_sa.cell(r, 4, ev['model_id'])
                    ws_sa.cell(r, 5, ev['board_id'])
                    ws_sa.cell(r, 6, 'Y' if ev['is_second'] else 'N')
                    ws_sa.cell(r, 7, round(ev['start'] / 3600, 4))
                    ws_sa.cell(r, 8, round(ev['end'] / 3600, 4))
                    ws_sa.cell(r, 9, round(ev['end'] - ev['start'], 2))
                    r += 1
            _aw(ws_sa)

            print(f'  추가 시트 3개: WIP_Timeseries / Truck_Log / SMT_Stage_Activity')

        wb.save(path)
        print(f'결과 저장: {path}')
        
    def save_figures(self, inference_summary=None, ep_rewards=None,
                     ep_rewards_decomp=None):
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        from matplotlib import font_manager

        # 한글 폰트 설정 (Windows: Malgun Gothic, macOS: AppleGothic,
        # Linux: NanumGothic 순서로 탐색). 설치된 것 중 첫 번째 사용.
        # 안 설정하면 제목·축 라벨·범례가 □□□ 로 렌더링됨.
        _available = {f.name for f in font_manager.fontManager.ttflist}
        for _fname in ('Malgun Gothic', 'AppleGothic', 'NanumGothic',
                       'Noto Sans CJK KR', 'Gulim'):
            if _fname in _available:
                plt.rcParams['font.family'] = _fname
                break
        else:
            print('  [경고] 한글 폰트를 찾지 못함 - 그래프 제목·라벨이 깨질 수 있음')
        plt.rcParams['axes.unicode_minus'] = False  # 마이너스 부호 깨짐 방지

        fig_dir = BASE_DIR
        colors  = {
            'SMT':'#2E86AB', 'MODULE':'#A23B72', 'SEMI':'#F18F01',
            'SET':'#3D9970',  'INSP':'#E84855',  'PACK':'#7B2D8B',
            'RMA':'#888888',  'SMT_SHARED':'#44BBA4',
        }

        # ── Figure 1. 학습 곡선 ──────────────────────────────────
        if ep_rewards and len(ep_rewards) > 1:
            fig, ax = plt.subplots(figsize=(10, 4))
            eps = list(range(1, len(ep_rewards) + 1))
            ax.plot(eps, ep_rewards, color='#AAAAAA', linewidth=0.8,
                    linestyle='--', label='에피소드별 보상')
            window = min(100, len(ep_rewards))
            rolling = [np.mean(ep_rewards[max(0,i-window):i+1])
                       for i in range(len(ep_rewards))]
            ax.plot(eps, rolling, color='#2E86AB', linewidth=1.8,
                    linestyle='-', label=f'이동평균 (window={window})')
            ax.set_xlabel('에피소드')
            ax.set_ylabel('보상')
            ax.set_title('PPO 학습 곡선')
            ax.legend(fontsize=9)
            ax.grid(axis='y', linestyle=':', alpha=0.5)
            plt.tight_layout()
            path1 = os.path.join(fig_dir, 'fig_learning_curve.png')
            plt.savefig(path1, dpi=150, bbox_inches='tight')
            plt.close()
            print(f'학습 곡선 저장: {path1}')

        # ── Figure 1b. 항별 reward 분해 (각 항 평균 차감 — 변동 흐름만 비교) ─
        # r6 가 절대 dominant 라 raw plot 시 다른 항 변화가 안 보임.
        # 각 항 column 의 전체 평균을 빼 zero-centered → 항별 _상대 변화_ 만 표시.
        # 0 라인 = 그 항의 학습 전체 평균 수준.
        if ep_rewards_decomp and len(ep_rewards_decomp) > 1:
            arr = np.array(ep_rewards_decomp)   # (N_ep, 6)
            eps = list(range(1, len(arr) + 1))
            mean_per_term = arr.mean(axis=0)    # (6,)
            arr_centered  = arr - mean_per_term
            labels = [
                'r1 (시간)', 'r2 (전력)', 'r3 (WIP)',
                'r4 (재고)', 'r5 (납기)', 'r6 (유휴)',
            ]
            term_colors = ['#2E86AB', '#A23B72', '#F18F01',
                           '#3D9970', '#E84855', '#7B2D8B']
            fig, ax = plt.subplots(figsize=(10, 4))
            window = min(100, len(arr))
            for i in range(6):
                series = arr_centered[:, i]
                rolling = [np.mean(series[max(0,j-window):j+1])
                           for j in range(len(series))]
                ax.plot(eps, rolling, color=term_colors[i],
                        linewidth=1.6,
                        label=f'{labels[i]} (μ={mean_per_term[i]:+.2f})')
            ax.axhline(0, color='black', linewidth=0.5, alpha=0.3)
            ax.set_xlabel('에피소드')
            ax.set_ylabel(f'(항별 - 평균) 이동평균 window={window}')
            ax.set_title('PPO reward 항별 변동 — 각 항의 전체 평균을 0 으로 centering '
                         '(0 위 = 평균보다 좋아짐 / 아래 = 평균보다 나빠짐)')
            ax.legend(fontsize=8, ncol=3, loc='best')
            ax.grid(axis='y', linestyle=':', alpha=0.5)
            plt.tight_layout()
            path1b = os.path.join(fig_dir, 'fig_learning_decomp.png')
            plt.savefig(path1b, dpi=150, bbox_inches='tight')
            plt.close()
            print(f'학습 항별 분해 저장: {path1b}')

        if inference_summary is None:
            return

        by_grp = inference_summary.get('by_grp_kwh', {})
        if not by_grp:
            return

        # ── Figure 2. 공정그룹별 전력 소비 (수평 막대) ───────────
        fig, ax = plt.subplots(figsize=(8, 4))
        grps = sorted(by_grp, key=by_grp.get, reverse=True)
        kwhs = [by_grp[g] for g in grps]
        bar_colors = [colors.get(g, '#CCCCCC') for g in grps]
        bars = ax.barh(grps, kwhs, color=bar_colors, edgecolor='white')
        for bar, val in zip(bars, kwhs):
            ax.text(bar.get_width() + max(kwhs) * 0.01, bar.get_y() + bar.get_height()/2,
                    f'{val:.3f}', va='center', fontsize=8)
        ax.set_xlabel('kWh')
        ax.set_title('공정그룹별 전력 소비')
        ax.grid(axis='x', linestyle=':', alpha=0.5)
        plt.tight_layout()
        path2 = os.path.join(fig_dir, 'fig_energy_by_group.png')
        plt.savefig(path2, dpi=150, bbox_inches='tight')
        plt.close()
        print(f'전력 분포 저장: {path2}')

        # ── Figure 3. 누적 전력 시계열 (실선) ────────────────────
        # agent 없으면 Figure3 (학습 수렴 곡선)만 skip하고 Figure4로 진행
        agent = getattr(self, '_last_agent', None)

        # 에피소드별 최종 kWh가 없으므로 학습 보상과 함께 전력 추이 표시
        if agent is not None and ep_rewards and len(ep_rewards) > 1:
            fig, ax1 = plt.subplots(figsize=(10, 4))
            eps = list(range(1, len(ep_rewards) + 1))
            window = min(100, len(ep_rewards))
            rolling = [np.mean(ep_rewards[max(0,i-window):i+1])
                       for i in range(len(ep_rewards))]
            ax1.plot(eps, rolling, color='#2E86AB', linewidth=1.8,
                     linestyle='-', label='평균 보상 (실선)')
            ax1.set_xlabel('에피소드')
            ax1.set_ylabel('보상', color='#2E86AB')
            ax1.tick_params(axis='y', labelcolor='#2E86AB')
            ax1.set_title('학습 수렴 곡선')
            ax1.grid(axis='y', linestyle=':', alpha=0.4)
            lines1, labels1 = ax1.get_legend_handles_labels()
            ax1.legend(lines1, labels1, loc='lower right', fontsize=8)
            plt.tight_layout()
            path3 = os.path.join(fig_dir, 'fig_pareto_candidates.png')
            plt.savefig(path3, dpi=150, bbox_inches='tight')
            plt.close()
            print(f'파레토 후보 저장: {path3}')

        # ── Figure 4. 부품 재고 시계열 (TOP 8 소비 부품) + 주요 이벤트 ─────
        # Stock_Timeseries 시트의 시각 요약. 트럭 외주 / 첫 PCB 가용 / 첫 완성 등
        # 흐름 이해에 도움되는 이벤트만 vertical line 으로 표시 (최소화).
        menv = getattr(self, '_last_menv', None)
        if menv is not None and menv.wh.snapshots:
            wh = menv.wh
            top = sorted(wh.consumed.items(), key=lambda kv: -kv[1])[:8]
            top_codes = [c for c, _ in top if c in wh.snapshots and wh.snapshots[c]]
            if top_codes:
                fig, ax = plt.subplots(figsize=(12, 5.5))
                palette = ['#2E86AB', '#A23B72', '#F18F01', '#3D9970',
                           '#E84855', '#7B2D8B', '#44BBA4', '#888888']
                for i, code in enumerate(top_codes):
                    pts = wh.snapshots[code]
                    xs = [t / 3600 for t, _ in pts]
                    ys = [q for _, q in pts]
                    try:
                        name = wh.data.get_item_name(code)
                    except AttributeError:
                        name = ''
                    label = f'{code}' + (f' ({name[:12]})' if name else '')
                    ax.plot(xs, ys, color=palette[i % len(palette)],
                            linewidth=1.3, label=label)

                # ── 주요 이벤트 vertical line ─────────────────────
                # 너무 많으면 시각 오염되므로 주요 milestone 만 표시.
                pool = getattr(menv, 'outsource_pool', None)
                if pool is not None:
                    # 트럭 출발 / 도착 시각
                    for entry in pool.truck_log:
                        send_h = entry['send_t'] / 3600
                        ax.axvline(send_h, color='#FF6B35', linestyle='--',
                                   linewidth=0.8, alpha=0.6)
                        if entry.get('return_t'):
                            ret_h = entry['return_t'] / 3600
                            ax.axvline(ret_h, color='#2ECC71', linestyle='--',
                                       linewidth=0.8, alpha=0.6)
                # 첫 unit 완성 시각
                first_done_h = None
                for k, v in wh.unit_completions.items():
                    if v.get('path') in ('normal', 'rma'):
                        end_h = v.get('end_time', 0) / 3600
                        if first_done_h is None or end_h < first_done_h:
                            first_done_h = end_h
                if first_done_h is not None:
                    ax.axvline(first_done_h, color='#9B59B6', linestyle=':',
                               linewidth=1.2, alpha=0.7)

                # 범례 보조 (이벤트 line 의미)
                from matplotlib.lines import Line2D
                legend_main = ax.legend(fontsize=8, loc='upper right',
                                        title='부품')
                ax.add_artist(legend_main)
                event_handles = [
                    Line2D([0], [0], color='#FF6B35', linestyle='--',
                           label='트럭 출발'),
                    Line2D([0], [0], color='#2ECC71', linestyle='--',
                           label='트럭 도착'),
                    Line2D([0], [0], color='#9B59B6', linestyle=':',
                           label='첫 unit 완성'),
                ]
                ax.legend(handles=event_handles, fontsize=8, loc='lower right',
                          title='이벤트')

                ax.set_xlabel('시간 (h)')
                ax.set_ylabel('재고 수량')
                ax.set_title('부품 재고 시계열 (소비 TOP 8) — 주요 이벤트 표시')
                ax.grid(axis='y', linestyle=':', alpha=0.5)
                plt.tight_layout()
                path4 = os.path.join(fig_dir, 'fig_stock_timeseries.png')
                plt.savefig(path4, dpi=150, bbox_inches='tight')
                plt.close()
                print(f'부품 재고 시계열 저장: {path4}')

        # ── Figures 5 & 6. Dynamic Job Shop Scheduling Gantt ─────
        # Process_Log (plogger.events) 기반. 2026-4-24 재구성 (A안).
        # y축 = 공정 위치 "한 줄" 만 (11 LOCATION). 동시 작업은 한 줄 안에서
        # capacity N 만큼 층(layer)으로 쌓는다 (slot 번호 = 층 위치).
        # Fig 6: y=공정 위치 11행, 색=모델, 라벨=모델/공정번호
        # Fig 5: 모델별 subplot 3장 (Fig 6 구조를 모델마다 그대로), 색=유사공정
        if menv is not None:
            plogger = getattr(menv, 'plogger', None)
            if plogger is not None:
                from matplotlib.patches import Patch
                makespan_s = float(menv.env.now)
                # 미완료(mark_end 안 된) 공정은 현재 시점까지 이어지는 것으로 간주.
                events = list(plogger.events)
                for ev_id, meta in plogger._active.items():
                    events.append({
                        'pc'        : meta['pc'],
                        'mid'       : meta['mid'],
                        'uid'       : meta['uid'],
                        'start'     : meta['start'],
                        'end'       : makespan_s,
                        'grp'       : plogger.groups.get(meta['pc'], ''),
                        'wgrp'      : meta['wgrp'],
                        'slot'      : meta['slot'],
                        'work_timed': meta.get('work_timed', False),
                    })

                def _split_at_worktime(start_s, end_s):
                    # [start, end] 구간을 근무시간만 포함하는 sub-range 로 분할.
                    # 비근무시간(점심·퇴근·야간)은 건너뛴다. work_timed 이벤트에만
                    # 적용 - bare timeout 이벤트는 실제로 pause 한 게 아니라 통짜로
                    # 플롯해야 함.
                    _s = _active_schedule
                    lunch_start = _s['lunch_start_sec']
                    work_end    = _s['work_end_sec']

                    segments = []
                    t = float(start_s)
                    end = float(end_s)
                    while t < end:
                        if not _is_work_time(t):
                            nxt = _next_work_start(t)
                            if nxt >= end:
                                break
                            t = nxt
                            continue
                        tt = t % DAY_SEC
                        if tt < lunch_start:
                            boundary = t - tt + lunch_start
                        else:
                            boundary = t - tt + work_end
                        b = min(end, boundary)
                        if b > t:
                            segments.append((t, b))
                        t = b
                    return segments

                def _event_segments(e):
                    # plot용 (start, end) 세그먼트. work_timed면 경계 분할, 아니면 그대로.
                    if e.get('work_timed'):
                        return _split_at_worktime(e['start'], e['end'])
                    return [(e['start'], e['end'])]

                # 분할 통계 (work_timed 이벤트가 실제 여러 조각으로 쪼개지는지 확인)
                split_hist = defaultdict(lambda: {'total': 0, 'split': 0, 'pieces': 0})
                long_rma = []   # start~end 가 긴 RMA 이벤트 샘플
                for e in events:
                    if e['wgrp'] == 'WORKER_RMA' and len(long_rma) < 5:
                        dur_h = (e['end'] - e['start']) / 3600
                        if dur_h > 0.5:
                            long_rma.append((dur_h, e))
                    if not e.get('work_timed'):
                        continue
                    segs = _event_segments(e)
                    rec = split_hist[e['wgrp']]
                    rec['total'] += 1
                    rec['pieces'] += len(segs)
                    if len(segs) >= 2:
                        rec['split'] += 1
                if split_hist:
                    print('\n[간트차트 work_timed 분할 통계]')
                    for wg, rec in split_hist.items():
                        avg = rec['pieces'] / max(rec['total'], 1)
                        print(f'  {wg:22s} 총 {rec["total"]:4d}건, '
                              f'경계분할 {rec["split"]:4d}건 '
                              f'(평균 {avg:.2f}조각)')
                if long_rma:
                    print('\n[RMA 이벤트 샘플 (dur > 30min)]')
                    for dur_h, e in long_rma:
                        print(f'  start={e["start"]/3600:7.2f}h end={e["end"]/3600:7.2f}h '
                              f'dur={dur_h:6.2f}h wt={e.get("work_timed")} '
                              f'slot={e["slot"]} mid={e["mid"]}')

                def _pc_refno(pc):
                    # 공정번호: 'BT5_42_BOND'->'42', 'VD7_10a_FW'->'10a'.
                    if pc in ('RMA_REPAIR', 'OQC_SAMPLE'):
                        return pc
                    parts = pc.split('_')
                    return parts[1] if len(parts) >= 2 else pc

                def _pc_refno_base(pc):
                    # 유사공정 색상 그룹: suffix 문자 제거. '10a'->'10'.
                    r = _pc_refno(pc)
                    m = re.match(r'(\d+)', r)
                    return m.group(1) if m else r

                # wgrp 없는 이벤트 제외 (SMT 등 plogger 미기록 경로).
                events = [e for e in events
                          if e.get('wgrp') and e.get('slot', -1) >= 0]

                # ── 진단: wgrp 별 이벤트 수 + work_timed 분할 통계 ─────
                wgrp_counts = defaultdict(int)
                wtimed_counts = defaultdict(int)
                split_counts = defaultdict(int)     # wgrp -> N 개 이벤트가 2+조각으로 분할됨
                for e in events:
                    wgrp_counts[e['wgrp']] += 1
                    if e.get('work_timed'):
                        wtimed_counts[e['wgrp']] += 1
                print('\n[간트차트 worker_group 분포]')
                for wg in LOCATION_ORDER + sorted(
                        k for k in menv.wres if k not in LOCATION_ORDER):
                    if wg not in menv.wres:
                        continue
                    cap = int(menv.wres[wg].capacity)
                    cnt = wgrp_counts.get(wg, 0)
                    wt = wtimed_counts.get(wg, 0)
                    flag = '   -> 이벤트 0 (엑셀에 worker_group 미지정?)' \
                           if cnt == 0 and cap > 0 else ''
                    wt_str = f' (work_timed={wt})' if wt > 0 else ''
                    print(f'  {wg:22s} cap={cap:2d}  이벤트={cnt:5d}{wt_str}{flag}')

                if events:
                    from matplotlib.ticker import MultipleLocator, FuncFormatter

                    def _apply_day_axis(ax, xmax_h):
                        # x축 major tick(24h), minor tick(12h), 근무시간대 배경 표시
                        day_max = int(xmax_h / 24) + 2
                        ax.xaxis.set_major_locator(MultipleLocator(24))
                        ax.xaxis.set_minor_locator(MultipleLocator(12))
                        ax.xaxis.set_major_formatter(FuncFormatter(
                            lambda v, _: f'Day{int(v/24)+1}\n({int(v)}h)'))
                        ax.grid(axis='x', which='major', linestyle='-',
                                color='#888888', alpha=0.35)
                        ax.grid(axis='x', which='minor', linestyle=':',
                                color='#BBBBBB', alpha=0.25)
                        # 근무시간 배경 음영 (9시00분-12시00분, 13시00분-18시00분)
                        for d in range(day_max):
                            base = d * 24
                            ax.axvspan(base + 9,  base + 12, color='#E8F4F8',
                                       alpha=0.35, zorder=0)
                            ax.axvspan(base + 13, base + 18, color='#E8F4F8',
                                       alpha=0.35, zorder=0)

                    def _cap_of(wg):
                        observed = plogger.max_slot.get(wg, 0)
                        try:
                            rcap = int(menv.wres[wg].capacity) if wg in menv.wres else 0
                        except Exception:
                            rcap = 0
                        return max(observed, rcap, 1)

                    # y축 11행: LOCATION_ORDER + 이벤트에 등장한 기타 wgrp.
                    loc_order = [wg for wg in LOCATION_ORDER]
                    extra = sorted({e['wgrp'] for e in events
                                    if e['wgrp'] not in LOCATION_ORDER})
                    loc_order.extend(extra)
                    loc_index = {wg: i for i, wg in enumerate(loc_order)}
                    loc_cap   = {wg: _cap_of(wg) for wg in loc_order}
                    loc_label = [f'{LOCATION_LABEL.get(wg, wg)} ({loc_cap[wg]})'
                                 for wg in loc_order]

                    models = sorted({e['mid'] for e in events})
                    cmap6 = plt.get_cmap('tab10', max(len(models), 1))
                    color_by_model = {m: cmap6(i % cmap6.N)
                                      for i, m in enumerate(models)}

                    def _slot_y(row_i, slot_i, cap):
                        # 한 줄(row_i) 내부에서 slot_i 층의 y 중심. cap 층으로 등분.
                        return row_i - 0.5 + (slot_i + 0.5) / max(cap, 1)

                    def _slot_h(cap):
                        # 한 층의 높이 (살짝 gap 넣어 겹치지 않게).
                        return 0.85 / max(cap, 1)

                    # ═════ Figure 6 : y=공정 위치, 색=모델, 층=슬롯 ═════
                    fig_h = max(4.0, 0.55 * len(loc_order))
                    fig, ax = plt.subplots(figsize=(14, fig_h))
                    for e in events:
                        wg = e['wgrp']
                        if wg not in loc_index:
                            continue
                        row_i = loc_index[wg]
                        cap = loc_cap[wg]
                        y = _slot_y(row_i, e['slot'], cap)
                        h = _slot_h(cap)
                        color = color_by_model.get(e['mid'], '#888888')
                        segs = _event_segments(e)
                        longest = max(segs, key=lambda s: s[1] - s[0]) if segs else None
                        for seg_s, seg_e in segs:
                            x0 = seg_s / 3600
                            # 2026-04-30: 최소 막대 폭 1분 → 5분 으로 확장.
                            # 작업시간 37s 짜리 PACK·LABEL 단계가 sub-pixel 로 렌더링돼
                            # 시각적으로 안 보이는 문제 완화. 시각화 전용이라 실제 시간엔
                            # 영향 없음 (이벤트 데이터는 그대로).
                            dur = max((seg_e - seg_s) / 3600, 5/60)
                            ax.barh(y, dur, left=x0, height=h,
                                    color=color, edgecolor='none')
                        # 라벨: 가장 긴 세그먼트 위에만 (바 길이 충분할 때)
                        if longest is not None and h >= 0.15:
                            seg_s, seg_e = longest
                            dur = (seg_e - seg_s) / 3600
                            if dur >= 0.5:
                                ax.text(seg_s / 3600 + dur/2, y,
                                        f"{e['mid']}/{_pc_refno(e['pc'])}",
                                        ha='center', va='center',
                                        fontsize=5, color='white')
                    # 행 구분선
                    for i in range(len(loc_order) + 1):
                        ax.axhline(i - 0.5, color='#CCCCCC',
                                   linewidth=0.5, linestyle='-')
                    ax.set_yticks(range(len(loc_order)))
                    ax.set_yticklabels(loc_label, fontsize=9)
                    ax.set_ylim(len(loc_order) - 0.5, -0.5)   # invert
                    ax.set_xlabel('시간 (h)')
                    ax.set_title('Dynamic Job Shop Gantt '
                                 '(공정 위치 × 슬롯 층 쌓기, 색=모델)')
                    legend_items = [Patch(color=color_by_model[m], label=m)
                                    for m in models]
                    ax.legend(handles=legend_items, loc='lower right', fontsize=8)
                    _apply_day_axis(ax, makespan_s / 3600)
                    plt.tight_layout()
                    path6 = os.path.join(fig_dir, 'fig_gantt_process_model.png')
                    plt.savefig(path6, dpi=200, bbox_inches='tight')
                    plt.close()
                    print(f'간트차트(공정위치×모델) 저장: {path6}')

                    # ═════ Figure 5 : 모델별 subplot 3장, 색=유사공정 ═════
                    # 각 subplot 이 Fig 6 와 동일 11행 구조 (해당 모델 이벤트만).
                    # sharex 로 세로 스택 -> "Fig 6 를 모델로 펼친 느낌" 재현.
                    refno_bases = sorted({_pc_refno_base(e['pc']) for e in events})
                    # tab20 은 dark/light 쌍이라 홀수 index 가 파스텔로 반투명처럼
                    # 보이는 문제가 있어 vivid solid 팔레트 사용. 공정 종류 많을 때
                    # cycling (>27 은 중복). 2026-4-24 교체.
                    VIVID_PALETTE = [
                        '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                        '#8c564b', '#e377c2', '#17becf', '#bcbd22', '#7f7f7f',
                        '#e41a1c', '#377eb8', '#4daf4a', '#984ea3', '#ff7f00',
                        '#a65628', '#f781bf', '#66c2a5', '#fc8d62', '#8da0cb',
                        '#1b9e77', '#d95f02', '#7570b3', '#e7298a', '#66a61e',
                        '#e6ab02', '#a6761d',
                    ]
                    color_by_refno = {b: VIVID_PALETTE[i % len(VIVID_PALETTE)]
                                      for i, b in enumerate(refno_bases)}

                    n_models = len(models)
                    sub_h = max(3.0, 0.50 * len(loc_order))
                    fig, axes = plt.subplots(
                        n_models, 1,
                        figsize=(14, sub_h * n_models),
                        sharex=True)
                    if n_models == 1:
                        axes = [axes]
                    for ax_i, model in enumerate(models):
                        ax = axes[ax_i]
                        m_events = [e for e in events if e['mid'] == model]
                        for e in m_events:
                            wg = e['wgrp']
                            if wg not in loc_index:
                                continue
                            row_i = loc_index[wg]
                            cap = loc_cap[wg]
                            y = _slot_y(row_i, e['slot'], cap)
                            h = _slot_h(cap)
                            color = color_by_refno.get(
                                _pc_refno_base(e['pc']), '#888888')
                            segs = _event_segments(e)
                            longest = max(segs, key=lambda s: s[1] - s[0]) \
                                if segs else None
                            for seg_s, seg_e in segs:
                                x0 = seg_s / 3600
                                # 2026-04-30: 최소 막대 폭 1분 → 5분으로 확장.
                                # 37s PACK 단계가 sub-pixel 로 렌더돼 안 보이는 문제 완화.
                                dur = max((seg_e - seg_s) / 3600, 5/60)
                                ax.barh(y, dur, left=x0, height=h,
                                        color=color, edgecolor='none')
                            if longest is not None and h >= 0.15:
                                seg_s, seg_e = longest
                                dur = (seg_e - seg_s) / 3600
                                if dur >= 0.5:
                                    ax.text(seg_s / 3600 + dur/2, y,
                                            _pc_refno(e['pc']),
                                            ha='center', va='center',
                                            fontsize=5, color='white')
                        for i in range(len(loc_order) + 1):
                            ax.axhline(i - 0.5, color='#CCCCCC',
                                       linewidth=0.5, linestyle='-')
                        ax.set_yticks(range(len(loc_order)))
                        ax.set_yticklabels(loc_label, fontsize=9)
                        ax.set_ylim(len(loc_order) - 0.5, -0.5)
                        ax.set_title(f'{model}', fontsize=11, loc='left')
                        _apply_day_axis(ax, makespan_s / 3600)
                    axes[-1].set_xlabel('시간 (h)')
                    fig.suptitle(
                        'Dynamic Job Shop Gantt (모델별 × 공정 위치 슬롯 층 쌓기, '
                        '색=유사공정)',
                        fontsize=12, y=0.995)
                    plt.tight_layout(rect=[0, 0, 1, 0.985])
                    path5 = os.path.join(fig_dir, 'fig_gantt_model_process.png')
                    plt.savefig(path5, dpi=200, bbox_inches='tight')
                    plt.close()
                    print(f'간트차트(모델별 subplot) 저장: {path5}')


# ══════════════════════════════════════════════════════════
# M16. 진입점
# ══════════════════════════════════════════════════════════

def main():
    # Windows cmd.exe 에서 ANSI escape 코드가 동작하도록 VT 모드 활성화.
    if os.name == 'nt':
        os.system('')

    random.seed(RANDOM_SEED)
    np.random.seed(RANDOM_SEED)
    torch.manual_seed(RANDOM_SEED)

    print('=== CPRO 제조 공정 시뮬레이션 (재고·공정 추적판) ===')

    # ── 데이터 로드 ───────────────────────────────────────
    # SMT 라인 + RMA + MTTR: 정적 fallback (M02 FallbackDataLoader)
    static_data = FallbackDataLoader()
    print(f'  [Fallback] SMT/RMA PROCESS_FLOW:{len(static_data.pf)}공정 | '
          f'MTTR 설비:{len(static_data._mttr)}개')

    # 조립~포장 공정: 모델별 AAS JSON 로더 (단일 출처)
    aas_loaders = {}
    for model_id, json_path in AAS_JSON_PATHS.items():
        loader = AASJsonLoader(model_id, json_path)
        if loader.pf_records:
            aas_loaders[model_id] = loader
            print(f'  [AAS]  {model_id}: {len(loader.pf_records)}공정 | '
                  f'InputBOM:{len(loader.bom_records)}건 | '
                  f'HierarchicalBOM:{len(loader.hs_bom_records)}건')
        else:
            print(f'  [AAS]  {model_id}: JSON 없음 또는 공정 미파싱 — 시뮬에서 제외')

    # 통합 데이터 로더
    data = CombinedDataLoader(static_data, aas_loaders)
    # AAS 에서 읽은 근무 스케줄을 전역 시간 유틸 함수에 반영.
    # work_start/end/break_duration 은 AAS 단일 출처 — 미제공 시 RuntimeError.
    if not data.schedule:
        raise RuntimeError(
            'AAS 가 근무 스케줄을 제공하지 않음. '
            '적어도 1개 모델의 AAS JSON 이 WorkstationWorkerMatchingData 를 포함해야 함.')
    _apply_schedule(data.schedule)
    print(f'  [통합] 총 process_code:{len(data._pc_map)}개 | '
          f'workers:{len(data.workers)}그룹 | '
          f'schedule: {data.schedule.get("work_start_sec",0)//3600:02d}h~'
          f'{data.schedule.get("work_end_sec",0)//3600:02d}h '
          f'(break {data.schedule.get("break_duration_sec",0)//60}min)')

    # AAS 가 로드된 모델만 주문 입력 받기 (JSON 없는 모델은 자동 스킵).
    order = {}
    if not aas_loaders:
        print('  [경고] AAS 로드된 모델이 없어 주문 입력을 건너뜁니다.')
    for m in aas_loaders:
        while True:
            try:
                qty = int(input(f'{m} 주문 수량: '))
                if qty > 0:
                    order[m] = qty
                    break
                print('  1 이상의 정수를 입력하세요.')
            except ValueError:
                print('  숫자를 입력하세요.')

    # 학습 에피소드 개수 입력 (2026-4-22 추가)
    # - 기본 5000 은 논문 상한. 스모크 실행엔 20-100 권장.
    # - 0 입력 시 학습 건너뛰고 기존 ppo_policy.pt 로 곧장 추론.
    max_eps = 5000
    while True:
        s = input('학습 에피소드 최대 개수 (엔터=5000, 0=학습건너뛰기): ').strip()
        if not s:
            max_eps = 5000
            break
        try:
            v = int(s)
            if v >= 0:
                max_eps = v
                break
            print('  0 이상의 정수를 입력하세요.')
        except ValueError:
            print('  숫자를 입력하세요.')

    runner = ExperimentRunner(data, order)
    if max_eps == 0:
        print('\n[학습 생략] 기존 ppo_policy.pt 가 있으면 그대로 추론에 사용합니다.')
        agent = None
    else:
        agent = runner.run_ppo_training(max_episodes=max_eps)

    if agent is not None:
        print(f'\n[학습 완료 요약]')
        print(f'  총 에피소드     : {len(agent.ep_rewards)}')
        if agent.ep_rewards:
            print(f'  최종 보상 평균  : {float(np.mean(agent.ep_rewards[-100:])):.4f}')
            print(f'  최고 에피소드 보상: {max(agent.ep_rewards):.4f}')

    # 학습된 agent를 inference에 직접 전달 (파일 경유 없음)
    s = runner.run_inference(agent=agent)
    runner.save_results(inference_summary=s)
    runner.save_figures(
        inference_summary=s,
        ep_rewards=(agent.ep_rewards if agent is not None else None),
        ep_rewards_decomp=(agent.ep_rewards_decomp if agent is not None else None))


if __name__ == '__main__':
    main()