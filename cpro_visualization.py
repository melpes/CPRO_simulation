# cpro_visualization.py
#
# 책임: ExperimentRunner 의 결과 저장(엑셀) + 그래프 출력 진입점.
#
# - cpro_simulation_ver3 을 직접 import 하지 않는다 (순환 / __main__ 중복 로드 회피).
#   필요한 상수·헬퍼는 keyword 인자로 주입받는다.
# - 도메인 객체(menv·agent·warehouse 등)는 runner 인스턴스를 통해 접근.
#
# 외부 export:
#   save_results(runner, inference_summary, path, *, min_stock,
#                pcb_map, tht_pcb_by_model)
#   save_figures(runner, inference_summary, ep_rewards, *, base_dir,
#                location_order, location_label, day_sec, schedule,
#                is_work_time, next_work_start)
#
# 호출처: cpro_simulation_ver3.ExperimentRunner.save_results / save_figures.

import os
import re
import numpy as np
from collections import defaultdict


# ══════════════════════════════════════════════════════════
# save_results — 엑셀 결과 저장 (Inference + 학습 + 디버그 시트)
# ══════════════════════════════════════════════════════════

def save_results(runner, inference_summary, path, *,
                 min_stock, pcb_map, tht_pcb_by_model):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb       = openpyxl.Workbook()
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='2E4053')

    def _hdr(ws, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font  = hdr_font
            c.fill  = hdr_fill
            c.alignment = Alignment(horizontal='center')

    def _aw(ws):
        for col in ws.columns:
            ml = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+2, 30)

    ws = wb.active
    ws.title = 'Inference'
    if inference_summary:
        _hdr(ws, ['항목', '값'])
        items = [
            ('makespan_hr',  inference_summary.get('makespan_hr', 0)),
            ('total_kwh',    inference_summary.get('total_kwh', 0)),
            ('재고품초과',   inference_summary.get('재고품초과', 0)),
            ('stock_pen',    inference_summary.get('stock_pen', 0)),
            ('total_defect', inference_summary.get('total_defect', 0)),
            ('total_done',   inference_summary.get('total_done', 0)),
            ('total_order',  inference_summary.get('total_order', 0)),
        ]
        for r, (k, v) in enumerate(items, 2):
            ws.cell(r, 1, k)
            ws.cell(r, 2, round(float(v), 6) if isinstance(v, float) else v)
        row_i = len(items) + 3
        ws.cell(row_i, 1, '공정그룹')
        ws.cell(row_i, 2, 'kWh')
        for grp, kwh in inference_summary.get('by_grp_kwh', {}).items():
            row_i += 1
            ws.cell(row_i, 1, grp)
            ws.cell(row_i, 2, round(kwh, 6))
    _aw(ws)

    agent = getattr(runner, '_last_agent', None)
    if agent and agent.ep_rewards:
        ws2 = wb.create_sheet('Training_Rewards')
        _hdr(ws2, ['episode', 'reward', 'rolling_avg_100'])
        for i, r in enumerate(agent.ep_rewards, 1):
            ws2.cell(i+1, 1, i)
            ws2.cell(i+1, 2, round(float(r), 6))
            if i >= 100:
                ws2.cell(i+1, 3, round(float(np.mean(agent.ep_rewards[i-100:i])), 6))
        _aw(ws2)

    # ══════════════════════════════════════════════════════════
    # 부품 재고 · 공정 추적 시트 (2026-4-22 추가)
    # ══════════════════════════════════════════════════════════
    menv = getattr(runner, '_last_menv', None)
    if menv is not None:
        wh = menv.wh
        plogger = menv.plogger
        makespan_s = int(menv.env.now)
        max_hour = max(1, makespan_s // 3600 + 1)
        if max_hour > 720:
            print(f'[경고] makespan {max_hour}h > 720h - Stock_Timeseries 열 수가 많아 '
                  f'openpyxl 쓰기가 오래 걸릴 수 있습니다.')

        # BOM 마스터 전체 item_code (~519개). 한 번도 소비 없던 부품도 포함.
        all_items = set()
        try:
            all_items = wh.data.iter_all_bom_items()
        except AttributeError:
            pass
        all_items.update(wh.stock.keys())
        all_items.update(wh.snapshots.keys())
        all_items = sorted(all_items)

        # ── Stock_Summary ─────────────────────────────────────
        ws_s = wb.create_sheet('Stock_Summary')
        _hdr(ws_s, ['item_code', 'item_name', 'initial_stock',
                    'total_consumed', 'final_stock', 'min_stock_qty',
                    'lot_size', 'violations_count', 'reorder_count'])
        r = 2
        for code in all_items:
            try:
                name = wh.data.get_item_name(code)
            except AttributeError:
                name = ''
            ws_s.cell(r, 1, code)
            ws_s.cell(r, 2, name)
            init_val = wh._initial_stocks.get(code,
                wh._bom_init_stock if code in wh._bom_codes else wh._init_stock)
            ws_s.cell(r, 3, int(init_val))
            ws_s.cell(r, 4, int(wh.consumed.get(code, 0)))
            ws_s.cell(r, 5, int(wh.stock.get(code, init_val)))
            try:
                ws_s.cell(r, 6, float(wh.data.get_min_stock(code)))
            except Exception:
                ws_s.cell(r, 6, min_stock)
            try:
                ws_s.cell(r, 7, int(wh._lot_for(code)))
            except Exception:
                ws_s.cell(r, 7, 0)
            ws_s.cell(r, 8, int(wh.violations.get(code, 0)))
            ws_s.cell(r, 9, int(wh.reorder_count.get(code, 0)))
            r += 1
        _aw(ws_s)

        # ── Stock_Timeseries (wide: 행=부품, 열=시간) ─────────
        ws_t = wb.create_sheet('Stock_Timeseries')
        header = ['item_code', 'item_name'] + [f't={h}h' for h in range(max_hour + 1)]
        _hdr(ws_t, header)
        r = 2
        for code in all_items:
            try:
                name = wh.data.get_item_name(code)
            except AttributeError:
                name = ''
            ws_t.cell(r, 1, code)
            ws_t.cell(r, 2, name)
            by_hour = {}
            for t_sec, q in wh.snapshots.get(code, []):
                by_hour[int(t_sec // 3600)] = int(q)
            prev = int(wh._initial_stocks.get(code,
                wh._bom_init_stock if code in wh._bom_codes else wh._init_stock))
            for h in range(max_hour + 1):
                if h in by_hour:
                    prev = by_hour[h]
                ws_t.cell(r, 3 + h, prev)
            r += 1

        # ── Stock_Events (long: 디버그용, wh.history 전체 덤프) ─
        ws_e = wb.create_sheet('Stock_Events')
        _hdr(ws_e, ['time_sec', 'time_hr', 'item_code', 'stock_after'])
        r = 2
        EVENT_CAP = 60000
        count = 0
        stop = False
        for code in all_items:
            if stop:
                break
            for t_sec, q in wh.history.get(code, []):
                ws_e.cell(r, 1, float(t_sec))
                ws_e.cell(r, 2, round(float(t_sec) / 3600, 3))
                ws_e.cell(r, 3, code)
                ws_e.cell(r, 4, int(q))
                r += 1
                count += 1
                if count >= EVENT_CAP:
                    ws_e.cell(r, 1, '[TRUNCATED]')
                    ws_e.cell(r, 3, f'events > {EVENT_CAP}')
                    stop = True
                    break
        _aw(ws_e)

        # ── Process_Log (wide: 행=process_code, 열=시간) ──────
        ws_p = wb.create_sheet('Process_Log')
        pheader = ['process_code'] + [f't={h}h' for h in range(max_hour + 1)]
        _hdr(ws_p, pheader)
        r = 2
        if plogger is not None:
            summary = plogger.summary()
            for pc, (mid, uid, t0) in plogger.current.items():
                label = f'{mid}/u{uid+1}*'
                h_start = int(t0 // 3600)
                h_end = max_hour
                summary.setdefault(pc, {})
                for h in range(h_start, h_end + 1):
                    existing = summary[pc].get(h, '')
                    summary[pc][h] = (existing + '; ' + label) if existing else label
            for pc in sorted(summary):
                ws_p.cell(r, 1, pc)
                row_map = summary[pc]
                for h in range(max_hour + 1):
                    val = row_map.get(h, '')
                    if val:
                        ws_p.cell(r, 2 + h, val)
                r += 1
        _aw(ws_p)

        # ── Reorder_Log (long) ────────────────────────────────
        ws_r = wb.create_sheet('Reorder_Log')
        _hdr(ws_r, ['item_code', 'order_time_hr', 'arrive_time_hr',
                    'lead_hr', 'lot_size', 'incoming', 'stock_at_order'])
        r = 2
        for entry in wh.reorder_log:
            ot = float(entry.get('order_time', 0)) / 3600
            at = float(entry.get('arrive_time', 0)) / 3600
            ws_r.cell(r, 1, entry.get('item_code', ''))
            ws_r.cell(r, 2, round(ot, 3))
            ws_r.cell(r, 3, round(at, 3))
            ws_r.cell(r, 4, round(at - ot, 3))
            ws_r.cell(r, 5, int(entry.get('lot_size', 0)))
            ws_r.cell(r, 6, int(entry.get('incoming', 0)))
            ws_r.cell(r, 7, int(entry.get('stock_at_order', 0)))
            r += 1
        _aw(ws_r)

        print(f'  재고·공정 시트 5개 추가 (items={len(all_items)}, '
              f'events={count}, reorders={len(wh.reorder_log)})')

        # ══════════════════════════════════════════════════════
        # 디버그 시트 (Q1·Q2·Q3 진단용)
        # ══════════════════════════════════════════════════════

        # ── Debug_Model_Stats ──────────────────────────────────
        ws_dms = wb.create_sheet('Debug_Model_Stats')
        _hdr(ws_dms, ['model_id', 'order_qty', 'stats_done',
                      'normal_completions', 'rma_completions',
                      'blocked_by_quota', 'first_event_h',
                      'last_event_h', 'first_after_24h',
                      'first_after_48h'])
        comps = wh.unit_completions
        events_all = plogger.events if plogger is not None else []
        r = 2
        for model in menv.order:
            qty = menv.order[model]
            stats_done = int(menv.stats.get(f'{model}_done', 0))
            model_comps = [v for k, v in comps.items()
                           if isinstance(k, tuple) and k[0] == model]
            n_normal = sum(1 for c in model_comps if c['path'] == 'normal')
            n_rma    = sum(1 for c in model_comps if c['path'] == 'rma')
            n_blocked = sum(1 for c in model_comps
                            if 'blocked_by_quota' in c['path'])
            m_evs = [e for e in events_all if e.get('mid') == model]
            first_h = round(min((e['start'] for e in m_evs), default=0)/3600, 2)
            last_h  = round(max((e['end']   for e in m_evs), default=0)/3600, 2)
            after_24 = round(min((e['start'] for e in m_evs
                                  if e['start'] >= 24*3600), default=0)/3600, 2)
            after_48 = round(min((e['start'] for e in m_evs
                                  if e['start'] >= 48*3600), default=0)/3600, 2)
            ws_dms.cell(r, 1, model)
            ws_dms.cell(r, 2, qty)
            ws_dms.cell(r, 3, stats_done)
            ws_dms.cell(r, 4, n_normal)
            ws_dms.cell(r, 5, n_rma)
            ws_dms.cell(r, 6, n_blocked)
            ws_dms.cell(r, 7, first_h)
            ws_dms.cell(r, 8, last_h)
            ws_dms.cell(r, 9, after_24)
            ws_dms.cell(r, 10, after_48)
            r += 1
        _aw(ws_dms)

        # ── Debug_Unit_Status ──────────────────────────────────
        ws_dus = wb.create_sheet('Debug_Unit_Status')
        _hdr(ws_dus, ['model_id', 'unit_key', 'completion_path',
                      'end_time_h', 'done_n', 'total_n',
                      'live_state', 'live_pc', 'live_done_n',
                      'live_total_n'])
        r = 2
        live_us = getattr(menv, 'unit_states', {}) or {}
        seen_keys = set()
        for k, c in sorted(comps.items(),
                            key=lambda kv: (str(kv[0][0]), str(kv[0][1]))):
            model = k[0]
            uid_or_tag = k[1]
            ws_dus.cell(r, 1, str(model))
            ws_dus.cell(r, 2, str(uid_or_tag))
            ws_dus.cell(r, 3, c.get('path', ''))
            ws_dus.cell(r, 4, round(c.get('end_time', 0)/3600, 3))
            ws_dus.cell(r, 5, int(c.get('done_n', -1)))
            ws_dus.cell(r, 6, int(c.get('total_n', -1)))
            if isinstance(uid_or_tag, int):
                ls = live_us.get((model, uid_or_tag), {})
                ws_dus.cell(r, 7, str(ls.get('state', '')))
                ws_dus.cell(r, 8, str(ls.get('pc', '')))
                ws_dus.cell(r, 9, int(ls.get('done_n', 0)))
                ws_dus.cell(r, 10, int(ls.get('total_n', 0)))
                seen_keys.add((model, uid_or_tag))
            r += 1
        for (model, uid), ls in sorted(live_us.items(),
                                       key=lambda kv: (str(kv[0][0]), kv[0][1])):
            if (model, uid) in seen_keys:
                continue
            ws_dus.cell(r, 1, str(model))
            ws_dus.cell(r, 2, str(uid))
            ws_dus.cell(r, 3, 'cutoff')
            ws_dus.cell(r, 4, round(menv.env.now/3600, 3))
            ws_dus.cell(r, 5, int(ls.get('done_n', 0)))
            ws_dus.cell(r, 6, int(ls.get('total_n', 0)))
            ws_dus.cell(r, 7, str(ls.get('state', '')))
            ws_dus.cell(r, 8, str(ls.get('pc', '')))
            ws_dus.cell(r, 9, int(ls.get('done_n', 0)))
            ws_dus.cell(r, 10, int(ls.get('total_n', 0)))
            r += 1
        _aw(ws_dus)

        # ── Debug_Process_Coverage ─────────────────────────────
        ws_dpc = wb.create_sheet('Debug_Process_Coverage')
        _hdr(ws_dpc, ['model_id', 'process_code', 'process_group',
                      'worker_group', 'expected_qty', 'actual_count',
                      'first_h', 'last_h'])
        r = 2
        for model in menv.order:
            qty = menv.order[model]
            kg_nodes = menv.graphs[model].nodes if model in menv.graphs else {}
            for pc in sorted(kg_nodes.keys()):
                node = kg_nodes[pc]
                pc_evs = [e for e in events_all
                          if e.get('mid') == model and e.get('pc') == pc]
                cnt = len(pc_evs)
                f_h = round(min((e['start'] for e in pc_evs),
                                default=0)/3600, 2) if pc_evs else ''
                l_h = round(max((e['end']   for e in pc_evs),
                                default=0)/3600, 2) if pc_evs else ''
                ws_dpc.cell(r, 1, model)
                ws_dpc.cell(r, 2, pc)
                ws_dpc.cell(r, 3, node.get('process_group', ''))
                ws_dpc.cell(r, 4, node.get('worker_group', ''))
                ws_dpc.cell(r, 5, qty)
                ws_dpc.cell(r, 6, cnt)
                ws_dpc.cell(r, 7, f_h)
                ws_dpc.cell(r, 8, l_h)
                r += 1
        _aw(ws_dpc)

        # ── Debug_PCB_Flow ─────────────────────────────────────
        ws_dpf = wb.create_sheet('Debug_PCB_Flow')
        _hdr(ws_dpf, ['pcb_code', 'role', 'model_hint',
                      'initial_stock', 'final_stock', 'total_consumed',
                      'smt_or_outsource_restore',
                      'outsource_in', 'outsource_returned',
                      'external_order_trigger', 'external_replenish_arrived',
                      'is_bug_candidate'])
        r = 2
        model_for_pcb = {}
        for m, pc_main in pcb_map.items():
            model_for_pcb[pc_main] = (m, 'main')
        for m, ths in tht_pcb_by_model.items():
            for pc_t in ths:
                model_for_pcb[pc_t] = (m, 'tht')
        for code in sorted(wh._pcb_codes):
            role = model_for_pcb.get(code, ('-', '-'))
            flow = wh.pcb_flow.get(code, {})
            ext_trig = int(flow.get('external_order_trigger', 0))
            ext_arr  = int(flow.get('external_replenish_arrived', 0))
            bug = ext_trig > 0 or ext_arr > 0
            ws_dpf.cell(r, 1, code)
            ws_dpf.cell(r, 2, role[1])
            ws_dpf.cell(r, 3, role[0])
            ws_dpf.cell(r, 4, int(wh._initial_stocks.get(code,
                wh._bom_init_stock if code in wh._bom_codes else wh._init_stock)))
            ws_dpf.cell(r, 5, int(wh.stock.get(code, 0)))
            ws_dpf.cell(r, 6, int(wh.consumed.get(code, 0)))
            ws_dpf.cell(r, 7, int(flow.get('restore_from_smt_or_outsource', 0)))
            ws_dpf.cell(r, 8, int(flow.get('outsource_in', 0)))
            ws_dpf.cell(r, 9, int(flow.get('outsource_returned', 0)))
            ws_dpf.cell(r, 10, ext_trig)
            ws_dpf.cell(r, 11, ext_arr)
            ws_dpf.cell(r, 12, 'BUG' if bug else '')
            r += 1
        r += 1
        ws_dpf.cell(r, 1, '== SMT 처리량 (line × model × pcb) ==')
        r += 1
        ws_dpf.cell(r, 1, 'line')
        ws_dpf.cell(r, 2, 'model')
        ws_dpf.cell(r, 3, 'pcb_code')
        ws_dpf.cell(r, 4, 'restored_qty')
        r += 1
        for (sid, m, pc_code), q in sorted(wh.smt_per_model.items()):
            ws_dpf.cell(r, 1, sid)
            ws_dpf.cell(r, 2, m)
            ws_dpf.cell(r, 3, pc_code)
            ws_dpf.cell(r, 4, int(q))
            r += 1
        _aw(ws_dpf)

        # ── Debug_Outsource_Log ────────────────────────────────
        ws_dol = wb.create_sheet('Debug_Outsource_Log')
        _hdr(ws_dol, ['pcb_code', 'model_id', 'board_id',
                      'send_time_h', 'return_time_h', 'transit_h',
                      'delay_h', 'status'])
        r = 2
        for entry in wh.outsource_log:
            ret = entry.get('return_time')
            send_h = round(entry.get('send_time', 0)/3600, 3)
            ret_h  = round(ret/3600, 3) if ret is not None else ''
            transit = round((ret - entry.get('send_time', 0))/3600, 3) \
                      if ret is not None else ''
            ws_dol.cell(r, 1, entry.get('pcb_code', ''))
            ws_dol.cell(r, 2, entry.get('model_id', ''))
            ws_dol.cell(r, 3, entry.get('board_id', ''))
            ws_dol.cell(r, 4, send_h)
            ws_dol.cell(r, 5, ret_h)
            ws_dol.cell(r, 6, transit)
            ws_dol.cell(r, 7, round(entry.get('delay_sec', 0)/3600, 3))
            ws_dol.cell(r, 8, entry.get('status', ''))
            r += 1
        _aw(ws_dol)

        ws_dpr = wb.create_sheet('Debug_PCB_Reorders')
        _hdr(ws_dpr, ['item_code', 'order_h', 'arrive_h',
                      'lot_size', 'incoming', 'stock_at_order'])
        r = 2
        for entry in wh.reorder_log:
            if not entry.get('is_pcb'):
                continue
            ws_dpr.cell(r, 1, entry.get('item_code', ''))
            ws_dpr.cell(r, 2, round(entry.get('order_time', 0)/3600, 3))
            ws_dpr.cell(r, 3, round(entry.get('arrive_time', 0)/3600, 3))
            ws_dpr.cell(r, 4, int(entry.get('lot_size', 0)))
            ws_dpr.cell(r, 5, int(entry.get('incoming', 0)))
            ws_dpr.cell(r, 6, int(entry.get('stock_at_order', 0)))
            r += 1
        _aw(ws_dpr)

        # ── Debug_Skipped_PCs ─────────────────────────────────
        ws_dsk = wb.create_sheet('Debug_Skipped_PCs')
        _hdr(ws_dsk, ['model_id', 'process_code', 'skip_count',
                      'in_excel_pf', 'in_kg_nodes'])
        r = 2
        for (m, pc), c in sorted(wh.skipped_pcs.items()):
            in_pf = wh.data._pc_map.get(pc) is not None
            in_kg = (m in menv.graphs and
                     pc in menv.graphs[m].nodes)
            ws_dsk.cell(r, 1, m)
            ws_dsk.cell(r, 2, pc)
            ws_dsk.cell(r, 3, int(c))
            ws_dsk.cell(r, 4, 'Y' if in_pf else 'N')
            ws_dsk.cell(r, 5, 'Y' if in_kg else 'N')
            r += 1
        _aw(ws_dsk)

        # ── Debug_SMT_Choices ─────────────────────────────────
        ws_dsm = wb.create_sheet('Debug_SMT_Choices')
        _hdr(ws_dsm, ['time_h', 'chosen_model'])
        r = 2
        for (t, m) in wh.smt_model_choices:
            ws_dsm.cell(r, 1, round(t/3600, 3))
            ws_dsm.cell(r, 2, m)
            r += 1
        _aw(ws_dsm)

        # ── Debug_Safety_Alarms ───────────────────────────────
        ws_dsa = wb.create_sheet('Debug_Safety_Alarms')
        _hdr(ws_dsa, ['alarm_type', 'detail_1', 'detail_2',
                      'detail_3', 'time_h', 'extra'])
        r = 2
        for entry in wh.kg_incomplete_log:
            ws_dsa.cell(r, 1, 'B1_kg_incomplete')
            ws_dsa.cell(r, 2, entry['model_id'])
            ws_dsa.cell(r, 3, str(entry['unit_id']))
            ws_dsa.cell(r, 4, '')
            ws_dsa.cell(r, 5, round(entry['time_h'], 3))
            ws_dsa.cell(r, 6, ','.join(entry['missing_pcs']))
            r += 1
        for entry in wh.smt_single_side_log:
            ws_dsa.cell(r, 1, 'B5_single_side_double_pcb')
            ws_dsa.cell(r, 2, entry['pcb_code'])
            ws_dsa.cell(r, 3, entry['model_id'])
            ws_dsa.cell(r, 4, str(entry['board_id']))
            ws_dsa.cell(r, 5, round(entry['time_h'], 3))
            ws_dsa.cell(r, 6, entry.get('reason', ''))
            r += 1
        dup_log = getattr(wh.data, '_bom_dup_merge_log', [])
        for msg in dup_log:
            ws_dsa.cell(r, 1, 'B4_bom_smt_side_conflict')
            ws_dsa.cell(r, 2, str(msg))
            r += 1
        for entry in wh.stuck_wait_log:
            ws_dsa.cell(r, 1, 'wait_stock_timeout_fallback')
            ws_dsa.cell(r, 2, entry['item_code'])
            ws_dsa.cell(r, 3, str(entry['qty']))
            ws_dsa.cell(r, 4, f'stock={entry["stock_at_end"]}')
            ws_dsa.cell(r, 5, round(entry['wait_end_h'], 3))
            ws_dsa.cell(r, 6, f'wait_for_{entry["wait_end_h"]-entry["wait_start_h"]:.2f}h')
            r += 1
        _aw(ws_dsa)

        # ── Debug_Plogger_Events ──────────────────────────────
        ws_dpe = wb.create_sheet('Debug_Plogger_Events')
        _hdr(ws_dpe, ['pc', 'mid', 'uid', 'grp', 'wgrp', 'slot',
                      'start_h', 'end_h', 'dur_s', 'work_timed'])
        r = 2
        evs = list(plogger.events) if plogger else []
        for e in evs:
            if e.get('grp') != 'PACK':
                continue
            ws_dpe.cell(r, 1, str(e.get('pc', '')))
            ws_dpe.cell(r, 2, str(e.get('mid', '')))
            ws_dpe.cell(r, 3, int(e.get('uid', 0)))
            ws_dpe.cell(r, 4, str(e.get('grp', '')))
            ws_dpe.cell(r, 5, str(e.get('wgrp', '')))
            ws_dpe.cell(r, 6, int(e.get('slot', -1)))
            ws_dpe.cell(r, 7, round(float(e.get('start', 0))/3600, 3))
            ws_dpe.cell(r, 8, round(float(e.get('end', 0))/3600, 3))
            ws_dpe.cell(r, 9, round(float(e.get('end', 0))-float(e.get('start', 0)), 2))
            ws_dpe.cell(r, 10, str(e.get('work_timed', False)))
            r += 1
        _aw(ws_dpe)

        pack_no_wgrp = sum(1 for e in evs
                           if e.get('grp') == 'PACK' and not e.get('wgrp'))
        pack_bad_slot = sum(1 for e in evs
                            if e.get('grp') == 'PACK' and e.get('slot', -1) < 0)
        from collections import Counter
        pack_per_model = Counter(e.get('mid') for e in evs
                                 if e.get('grp') == 'PACK')

        print(f'  디버그 시트 10개 추가 (unit_comps={len(comps)}, '
              f'outsource_evs={len(wh.outsource_log)}, '
              f'skipped={sum(wh.skipped_pcs.values())}, '
              f'smt_choices={len(wh.smt_model_choices)}, '
              f'kg_incomplete={len(wh.kg_incomplete_log)}, '
              f'single_side_alarm={len(wh.smt_single_side_log)})')
        print(f'  [PACK 이벤트 진단] no_wgrp={pack_no_wgrp}, bad_slot={pack_bad_slot}, '
              f'per_model={dict(pack_per_model)}')

        # ── WIP_Timeseries ────────────────────────────────────
        wip = menv.wip
        ws_wt = wb.create_sheet('WIP_Timeseries')
        grps = ['SMT', 'MODULE', 'SEMI', 'SET', 'INSP', 'PACK', 'RMA']
        _hdr(ws_wt, ['hour'] + grps)
        hours = sorted({int(t // 3600) for g in grps
                        for t, _ in wip.snapshots.get(g, [])})
        lookup = {g: {int(t // 3600): n for t, n in wip.snapshots.get(g, [])}
                  for g in grps}
        r = 2
        for h in hours:
            ws_wt.cell(r, 1, h)
            for ci, g in enumerate(grps, 2):
                ws_wt.cell(r, ci, lookup[g].get(h, 0))
            r += 1
        _aw(ws_wt)

        # ── Truck_Log ─────────────────────────────────────────
        pool = getattr(menv, 'outsource_pool', None)
        ws_tl = wb.create_sheet('Truck_Log')
        _hdr(ws_tl, ['dispatch_id', 'send_h', 'eta_h', 'return_h',
                     'transit_h', 'delay_h', 'boards', 'truck_count_eq',
                     'pcb_breakdown'])
        r = 2
        for entry in (pool.truck_log if pool else []):
            ret_h = (entry['return_t'] / 3600) if entry.get('return_t') else None
            send_h = entry['send_t'] / 3600
            eta_h  = entry['eta'] / 3600
            bk = ', '.join(f"{c[-4:]}×{n}"
                           for c, n in entry['breakdown'].items())
            ws_tl.cell(r, 1, entry['dispatch_id'])
            ws_tl.cell(r, 2, round(send_h, 3))
            ws_tl.cell(r, 3, round(eta_h, 3))
            ws_tl.cell(r, 4, round(ret_h, 3) if ret_h else '')
            ws_tl.cell(r, 5, round((ret_h - send_h), 3) if ret_h else '')
            ws_tl.cell(r, 6, round(entry['delay_sec'] / 3600, 3))
            ws_tl.cell(r, 7, entry['size'])
            ws_tl.cell(r, 8, entry['truck_count'])
            ws_tl.cell(r, 9, bk)
            r += 1
        _aw(ws_tl)

        # ── SMT_Stage_Activity ────────────────────────────────
        ws_sa = wb.create_sheet('SMT_Stage_Activity')
        _hdr(ws_sa, ['line', 'stage', 'pcb_code', 'model_id',
                     'board_id', 'is_second', 'start_h', 'end_h', 'dur_s'])
        r = 2
        for sid, line in menv.smt_lines.items():
            for ev in line.stage_events:
                pc = ev['pc']
                if pc == 'SMT_AOI':
                    stage_name = 'AOI'
                    line_name  = '공유'
                else:
                    parts = pc.split('_')
                    line_name  = parts[-1]
                    stage_name = '_'.join(parts[1:-1])
                ws_sa.cell(r, 1, line_name)
                ws_sa.cell(r, 2, stage_name)
                ws_sa.cell(r, 3, ev['pcb_code'])
                ws_sa.cell(r, 4, ev['model_id'])
                ws_sa.cell(r, 5, ev['board_id'])
                ws_sa.cell(r, 6, 'Y' if ev['is_second'] else 'N')
                ws_sa.cell(r, 7, round(ev['start'] / 3600, 4))
                ws_sa.cell(r, 8, round(ev['end'] / 3600, 4))
                ws_sa.cell(r, 9, round(ev['end'] - ev['start'], 2))
                r += 1
        _aw(ws_sa)

        print(f'  추가 시트 3개: WIP_Timeseries / Truck_Log / SMT_Stage_Activity')

    wb.save(path)
    print(f'결과 저장: {path}')


# ══════════════════════════════════════════════════════════
# save_figures — PNG 그래프 출력
# ══════════════════════════════════════════════════════════

def save_figures(runner, inference_summary, ep_rewards, *,
                 base_dir, location_order, location_label,
                 day_sec, schedule, is_work_time, next_work_start):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib import font_manager

    # 한글 폰트 설정 (Windows: Malgun Gothic, macOS: AppleGothic,
    # Linux: NanumGothic 순서로 탐색).
    _available = {f.name for f in font_manager.fontManager.ttflist}
    for _fname in ('Malgun Gothic', 'AppleGothic', 'NanumGothic',
                   'Noto Sans CJK KR', 'Gulim'):
        if _fname in _available:
            plt.rcParams['font.family'] = _fname
            break
    else:
        print('  [경고] 한글 폰트를 찾지 못함 - 그래프 제목·라벨이 깨질 수 있음')
    plt.rcParams['axes.unicode_minus'] = False

    fig_dir = base_dir
    colors  = {
        'SMT': '#2E86AB', 'MODULE': '#A23B72', 'SEMI': '#F18F01',
        'SET': '#3D9970',  'INSP': '#E84855',  'PACK': '#7B2D8B',
        'RMA': '#888888',  'SMT_SHARED': '#44BBA4',
    }

    # ── Figure 1. 학습 곡선 ──────────────────────────────────
    if ep_rewards and len(ep_rewards) > 1:
        fig, ax = plt.subplots(figsize=(10, 4))
        eps = list(range(1, len(ep_rewards) + 1))
        ax.plot(eps, ep_rewards, color='#AAAAAA', linewidth=0.8,
                linestyle='--', label='에피소드별 보상')
        window = min(100, len(ep_rewards))
        rolling = [np.mean(ep_rewards[max(0, i-window):i+1])
                   for i in range(len(ep_rewards))]
        ax.plot(eps, rolling, color='#2E86AB', linewidth=1.8,
                linestyle='-', label=f'이동평균 (window={window})')
        ax.set_xlabel('에피소드')
        ax.set_ylabel('보상')
        ax.set_title('PPO 학습 곡선')
        ax.legend(fontsize=9)
        ax.grid(axis='y', linestyle=':', alpha=0.5)
        plt.tight_layout()
        path1 = os.path.join(fig_dir, 'fig_learning_curve.png')
        plt.savefig(path1, dpi=150, bbox_inches='tight')
        plt.close()
        print(f'학습 곡선 저장: {path1}')

    if inference_summary is None:
        return

    by_grp = inference_summary.get('by_grp_kwh', {})
    if not by_grp:
        return

    # ── Figure 2. 공정그룹별 전력 소비 (수평 막대) ───────────
    fig, ax = plt.subplots(figsize=(8, 4))
    grps = sorted(by_grp, key=by_grp.get, reverse=True)
    kwhs = [by_grp[g] for g in grps]
    bar_colors = [colors.get(g, '#CCCCCC') for g in grps]
    bars = ax.barh(grps, kwhs, color=bar_colors, edgecolor='white')
    for bar, val in zip(bars, kwhs):
        ax.text(bar.get_width() + max(kwhs) * 0.01, bar.get_y() + bar.get_height()/2,
                f'{val:.3f}', va='center', fontsize=8)
    ax.set_xlabel('kWh')
    ax.set_title('공정그룹별 전력 소비')
    ax.grid(axis='x', linestyle=':', alpha=0.5)
    plt.tight_layout()
    path2 = os.path.join(fig_dir, 'fig_energy_by_group.png')
    plt.savefig(path2, dpi=150, bbox_inches='tight')
    plt.close()
    print(f'전력 분포 저장: {path2}')

    # ── Figure 3. 누적 전력 시계열 (실선) ────────────────────
    agent = getattr(runner, '_last_agent', None)
    if agent is not None and ep_rewards and len(ep_rewards) > 1:
        fig, ax1 = plt.subplots(figsize=(10, 4))
        eps = list(range(1, len(ep_rewards) + 1))
        window = min(100, len(ep_rewards))
        rolling = [np.mean(ep_rewards[max(0, i-window):i+1])
                   for i in range(len(ep_rewards))]
        ax1.plot(eps, rolling, color='#2E86AB', linewidth=1.8,
                 linestyle='-', label='평균 보상 (실선)')
        ax1.set_xlabel('에피소드')
        ax1.set_ylabel('보상', color='#2E86AB')
        ax1.tick_params(axis='y', labelcolor='#2E86AB')
        ax1.set_title('학습 수렴 곡선')
        ax1.grid(axis='y', linestyle=':', alpha=0.4)
        lines1, labels1 = ax1.get_legend_handles_labels()
        ax1.legend(lines1, labels1, loc='lower right', fontsize=8)
        plt.tight_layout()
        path3 = os.path.join(fig_dir, 'fig_pareto_candidates.png')
        plt.savefig(path3, dpi=150, bbox_inches='tight')
        plt.close()
        print(f'파레토 후보 저장: {path3}')

    # ── Figure 4. 부품 재고 시계열 (TOP 8) + 주요 이벤트 ─────
    menv = getattr(runner, '_last_menv', None)
    if menv is not None and menv.wh.snapshots:
        wh = menv.wh
        top = sorted(wh.consumed.items(), key=lambda kv: -kv[1])[:8]
        top_codes = [c for c, _ in top if c in wh.snapshots and wh.snapshots[c]]
        if top_codes:
            fig, ax = plt.subplots(figsize=(12, 5.5))
            palette = ['#2E86AB', '#A23B72', '#F18F01', '#3D9970',
                       '#E84855', '#7B2D8B', '#44BBA4', '#888888']
            for i, code in enumerate(top_codes):
                pts = wh.snapshots[code]
                xs = [t / 3600 for t, _ in pts]
                ys = [q for _, q in pts]
                try:
                    name = wh.data.get_item_name(code)
                except AttributeError:
                    name = ''
                label = f'{code}' + (f' ({name[:12]})' if name else '')
                ax.plot(xs, ys, color=palette[i % len(palette)],
                        linewidth=1.3, label=label)

            # ── 주요 이벤트 vertical line ─────────────────────
            pool = getattr(menv, 'outsource_pool', None)
            if pool is not None:
                for entry in pool.truck_log:
                    send_h = entry['send_t'] / 3600
                    ax.axvline(send_h, color='#FF6B35', linestyle='--',
                               linewidth=0.8, alpha=0.6)
                    if entry.get('return_t'):
                        ret_h = entry['return_t'] / 3600
                        ax.axvline(ret_h, color='#2ECC71', linestyle='--',
                                   linewidth=0.8, alpha=0.6)
            first_done_h = None
            for k, v in wh.unit_completions.items():
                if v.get('path') in ('normal', 'rma'):
                    end_h = v.get('end_time', 0) / 3600
                    if first_done_h is None or end_h < first_done_h:
                        first_done_h = end_h
            if first_done_h is not None:
                ax.axvline(first_done_h, color='#9B59B6', linestyle=':',
                           linewidth=1.2, alpha=0.7)

            from matplotlib.lines import Line2D
            legend_main = ax.legend(fontsize=8, loc='upper right',
                                    title='부품')
            ax.add_artist(legend_main)
            event_handles = [
                Line2D([0], [0], color='#FF6B35', linestyle='--',
                       label='트럭 출발'),
                Line2D([0], [0], color='#2ECC71', linestyle='--',
                       label='트럭 도착'),
                Line2D([0], [0], color='#9B59B6', linestyle=':',
                       label='첫 unit 완성'),
            ]
            ax.legend(handles=event_handles, fontsize=8, loc='lower right',
                      title='이벤트')

            ax.set_xlabel('시간 (h)')
            ax.set_ylabel('재고 수량')
            ax.set_title('부품 재고 시계열 (소비 TOP 8) — 주요 이벤트 표시')
            ax.grid(axis='y', linestyle=':', alpha=0.5)
            plt.tight_layout()
            path4 = os.path.join(fig_dir, 'fig_stock_timeseries.png')
            plt.savefig(path4, dpi=150, bbox_inches='tight')
            plt.close()
            print(f'부품 재고 시계열 저장: {path4}')

    # ── Figures 5 & 6. Dynamic Job Shop Scheduling Gantt ─────
    if menv is not None:
        plogger = getattr(menv, 'plogger', None)
        if plogger is not None:
            from matplotlib.patches import Patch
            makespan_s = float(menv.env.now)
            events = list(plogger.events)
            for ev_id, meta in plogger._active.items():
                events.append({
                    'pc'        : meta['pc'],
                    'mid'       : meta['mid'],
                    'uid'       : meta['uid'],
                    'start'     : meta['start'],
                    'end'       : makespan_s,
                    'grp'       : plogger.groups.get(meta['pc'], ''),
                    'wgrp'      : meta['wgrp'],
                    'slot'      : meta['slot'],
                    'work_timed': meta.get('work_timed', False),
                })

            def _split_at_worktime(start_s, end_s):
                # work_timed 이벤트 [start, end] 를 점심·퇴근 boundary 로 분할.
                lunch_start = schedule['lunch_start_sec']
                work_end    = schedule['work_end_sec']

                segments = []
                t = float(start_s)
                end = float(end_s)
                while t < end:
                    if not is_work_time(t):
                        nxt = next_work_start(t)
                        if nxt >= end:
                            break
                        t = nxt
                        continue
                    tt = t % day_sec
                    if tt < lunch_start:
                        boundary = t - tt + lunch_start
                    else:
                        boundary = t - tt + work_end
                    b = min(end, boundary)
                    if b > t:
                        segments.append((t, b))
                    t = b
                return segments

            def _event_segments(e):
                if e.get('work_timed'):
                    return _split_at_worktime(e['start'], e['end'])
                return [(e['start'], e['end'])]

            split_hist = defaultdict(lambda: {'total': 0, 'split': 0, 'pieces': 0})
            long_rma = []
            for e in events:
                if e['wgrp'] == 'WORKER_RMA' and len(long_rma) < 5:
                    dur_h = (e['end'] - e['start']) / 3600
                    if dur_h > 0.5:
                        long_rma.append((dur_h, e))
                if not e.get('work_timed'):
                    continue
                segs = _event_segments(e)
                rec = split_hist[e['wgrp']]
                rec['total'] += 1
                rec['pieces'] += len(segs)
                if len(segs) >= 2:
                    rec['split'] += 1
            if split_hist:
                print('\n[간트차트 work_timed 분할 통계]')
                for wg, rec in split_hist.items():
                    avg = rec['pieces'] / max(rec['total'], 1)
                    print(f'  {wg:22s} 총 {rec["total"]:4d}건, '
                          f'경계분할 {rec["split"]:4d}건 '
                          f'(평균 {avg:.2f}조각)')
            if long_rma:
                print('\n[RMA 이벤트 샘플 (dur > 30min)]')
                for dur_h, e in long_rma:
                    print(f'  start={e["start"]/3600:7.2f}h end={e["end"]/3600:7.2f}h '
                          f'dur={dur_h:6.2f}h wt={e.get("work_timed")} '
                          f'slot={e["slot"]} mid={e["mid"]}')

            def _pc_refno(pc):
                if pc in ('RMA_REPAIR', 'OQC_SAMPLE'):
                    return pc
                parts = pc.split('_')
                return parts[1] if len(parts) >= 2 else pc

            def _pc_refno_base(pc):
                r = _pc_refno(pc)
                m = re.match(r'(\d+)', r)
                return m.group(1) if m else r

            events = [e for e in events
                      if e.get('wgrp') and e.get('slot', -1) >= 0]

            wgrp_counts = defaultdict(int)
            wtimed_counts = defaultdict(int)
            for e in events:
                wgrp_counts[e['wgrp']] += 1
                if e.get('work_timed'):
                    wtimed_counts[e['wgrp']] += 1
            print('\n[간트차트 worker_group 분포]')
            for wg in location_order + sorted(
                    k for k in menv.wres if k not in location_order):
                if wg not in menv.wres:
                    continue
                cap = int(menv.wres[wg].capacity)
                cnt = wgrp_counts.get(wg, 0)
                wt = wtimed_counts.get(wg, 0)
                flag = '   -> 이벤트 0 (엑셀에 worker_group 미지정?)' \
                       if cnt == 0 and cap > 0 else ''
                wt_str = f' (work_timed={wt})' if wt > 0 else ''
                print(f'  {wg:22s} cap={cap:2d}  이벤트={cnt:5d}{wt_str}{flag}')

            if events:
                from matplotlib.ticker import MultipleLocator, FuncFormatter

                def _apply_day_axis(ax, xmax_h):
                    day_max = int(xmax_h / 24) + 2
                    ax.xaxis.set_major_locator(MultipleLocator(24))
                    ax.xaxis.set_minor_locator(MultipleLocator(12))
                    ax.xaxis.set_major_formatter(FuncFormatter(
                        lambda v, _: f'Day{int(v/24)+1}\n({int(v)}h)'))
                    ax.grid(axis='x', which='major', linestyle='-',
                            color='#888888', alpha=0.35)
                    ax.grid(axis='x', which='minor', linestyle=':',
                            color='#BBBBBB', alpha=0.25)
                    for d in range(day_max):
                        base = d * 24
                        ax.axvspan(base + 9,  base + 12, color='#E8F4F8',
                                   alpha=0.35, zorder=0)
                        ax.axvspan(base + 13, base + 18, color='#E8F4F8',
                                   alpha=0.35, zorder=0)

                def _cap_of(wg):
                    observed = plogger.max_slot.get(wg, 0)
                    try:
                        rcap = int(menv.wres[wg].capacity) if wg in menv.wres else 0
                    except Exception:
                        rcap = 0
                    return max(observed, rcap, 1)

                loc_order = [wg for wg in location_order]
                extra = sorted({e['wgrp'] for e in events
                                if e['wgrp'] not in location_order})
                loc_order.extend(extra)
                loc_index = {wg: i for i, wg in enumerate(loc_order)}
                loc_cap   = {wg: _cap_of(wg) for wg in loc_order}
                loc_label = [f'{location_label.get(wg, wg)} ({loc_cap[wg]})'
                             for wg in loc_order]

                models = sorted({e['mid'] for e in events})
                cmap6 = plt.get_cmap('tab10', max(len(models), 1))
                color_by_model = {m: cmap6(i % cmap6.N)
                                  for i, m in enumerate(models)}

                def _slot_y(row_i, slot_i, cap):
                    return row_i - 0.5 + (slot_i + 0.5) / max(cap, 1)

                def _slot_h(cap):
                    return 0.85 / max(cap, 1)

                # ═════ Figure 6 : y=공정 위치, 색=모델, 층=슬롯 ═════
                fig_h = max(4.0, 0.55 * len(loc_order))
                fig, ax = plt.subplots(figsize=(14, fig_h))
                for e in events:
                    wg = e['wgrp']
                    if wg not in loc_index:
                        continue
                    row_i = loc_index[wg]
                    cap = loc_cap[wg]
                    y = _slot_y(row_i, e['slot'], cap)
                    h = _slot_h(cap)
                    color = color_by_model.get(e['mid'], '#888888')
                    segs = _event_segments(e)
                    longest = max(segs, key=lambda s: s[1] - s[0]) if segs else None
                    for seg_s, seg_e in segs:
                        x0 = seg_s / 3600
                        # 시각화 전용: 최소 막대 폭 5분 (실제 시간엔 영향 없음).
                        dur = max((seg_e - seg_s) / 3600, 5/60)
                        ax.barh(y, dur, left=x0, height=h,
                                color=color, edgecolor='none')
                    if longest is not None and h >= 0.15:
                        seg_s, seg_e = longest
                        dur = (seg_e - seg_s) / 3600
                        if dur >= 0.5:
                            ax.text(seg_s / 3600 + dur/2, y,
                                    f"{e['mid']}/{_pc_refno(e['pc'])}",
                                    ha='center', va='center',
                                    fontsize=5, color='white')
                for i in range(len(loc_order) + 1):
                    ax.axhline(i - 0.5, color='#CCCCCC',
                               linewidth=0.5, linestyle='-')
                ax.set_yticks(range(len(loc_order)))
                ax.set_yticklabels(loc_label, fontsize=9)
                ax.set_ylim(len(loc_order) - 0.5, -0.5)
                ax.set_xlabel('시간 (h)')
                ax.set_title('Dynamic Job Shop Gantt '
                             '(공정 위치 × 슬롯 층 쌓기, 색=모델)')
                legend_items = [Patch(color=color_by_model[m], label=m)
                                for m in models]
                ax.legend(handles=legend_items, loc='lower right', fontsize=8)
                _apply_day_axis(ax, makespan_s / 3600)
                plt.tight_layout()
                path6 = os.path.join(fig_dir, 'fig_gantt_process_model.png')
                plt.savefig(path6, dpi=200, bbox_inches='tight')
                plt.close()
                print(f'간트차트(공정위치×모델) 저장: {path6}')

                # ═════ Figure 5 : 모델별 subplot, 색=유사공정 ═════
                refno_bases = sorted({_pc_refno_base(e['pc']) for e in events})
                VIVID_PALETTE = [
                    '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                    '#8c564b', '#e377c2', '#17becf', '#bcbd22', '#7f7f7f',
                    '#e41a1c', '#377eb8', '#4daf4a', '#984ea3', '#ff7f00',
                    '#a65628', '#f781bf', '#66c2a5', '#fc8d62', '#8da0cb',
                    '#1b9e77', '#d95f02', '#7570b3', '#e7298a', '#66a61e',
                    '#e6ab02', '#a6761d',
                ]
                color_by_refno = {b: VIVID_PALETTE[i % len(VIVID_PALETTE)]
                                  for i, b in enumerate(refno_bases)}

                n_models = len(models)
                sub_h = max(3.0, 0.50 * len(loc_order))
                fig, axes = plt.subplots(
                    n_models, 1,
                    figsize=(14, sub_h * n_models),
                    sharex=True)
                if n_models == 1:
                    axes = [axes]
                for ax_i, model in enumerate(models):
                    ax = axes[ax_i]
                    m_events = [e for e in events if e['mid'] == model]
                    for e in m_events:
                        wg = e['wgrp']
                        if wg not in loc_index:
                            continue
                        row_i = loc_index[wg]
                        cap = loc_cap[wg]
                        y = _slot_y(row_i, e['slot'], cap)
                        h = _slot_h(cap)
                        color = color_by_refno.get(
                            _pc_refno_base(e['pc']), '#888888')
                        segs = _event_segments(e)
                        longest = max(segs, key=lambda s: s[1] - s[0]) \
                            if segs else None
                        for seg_s, seg_e in segs:
                            x0 = seg_s / 3600
                            dur = max((seg_e - seg_s) / 3600, 5/60)
                            ax.barh(y, dur, left=x0, height=h,
                                    color=color, edgecolor='none')
                        if longest is not None and h >= 0.15:
                            seg_s, seg_e = longest
                            dur = (seg_e - seg_s) / 3600
                            if dur >= 0.5:
                                ax.text(seg_s / 3600 + dur/2, y,
                                        _pc_refno(e['pc']),
                                        ha='center', va='center',
                                        fontsize=5, color='white')
                    for i in range(len(loc_order) + 1):
                        ax.axhline(i - 0.5, color='#CCCCCC',
                                   linewidth=0.5, linestyle='-')
                    ax.set_yticks(range(len(loc_order)))
                    ax.set_yticklabels(loc_label, fontsize=9)
                    ax.set_ylim(len(loc_order) - 0.5, -0.5)
                    ax.set_title(f'{model}', fontsize=11, loc='left')
                    _apply_day_axis(ax, makespan_s / 3600)
                axes[-1].set_xlabel('시간 (h)')
                fig.suptitle(
                    'Dynamic Job Shop Gantt (모델별 × 공정 위치 슬롯 층 쌓기, '
                    '색=유사공정)',
                    fontsize=12, y=0.995)
                plt.tight_layout(rect=[0, 0, 1, 0.985])
                path5 = os.path.join(fig_dir, 'fig_gantt_model_process.png')
                plt.savefig(path5, dpi=200, bbox_inches='tight')
                plt.close()
                print(f'간트차트(모델별 subplot) 저장: {path5}')
